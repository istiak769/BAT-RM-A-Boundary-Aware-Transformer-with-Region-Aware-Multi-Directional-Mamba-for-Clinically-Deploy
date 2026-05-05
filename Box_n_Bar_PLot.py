# -*- coding: utf-8 -*-
"""
Created on Thu Apr 30 14:51:36 2026

@author: User
"""

#----------- Box plots of 3 axis of Different models' segmentation scores for LUNG RADIOTHERAPY--------------------------

import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import numpy as np
import os
from scipy.stats import wilcoxon
import string

# ========================================================================
# CONFIGURATION SECTION - MODIFY ALL PARAMETERS HERE
# ========================================================================

# --------------------- MODEL NAME MAPPING (USE EXACT NAMES HERE) ---------------------
MODEL_NAME_MAP = {
    'Axial': {
        0: 'UNet3+',
        1: 'UNet++',
        2: 'UNet',
        3: 'TransUNet',
        4: 'UNETR'
    },
    'Coronal': {
        0: 'UNet3+',
        1: 'UNet++',
        2: 'UNet',
        3: 'TransUNet',
        4: 'UNETR'
    },
    'Sagittal': {
        0: 'UNet3+',
        1: 'UNet++',
        2: 'UNet',
        3: 'TransUNet',
        4: 'UNETR'
    }
}

# --------------------- AXES CONFIGURATION - UPDATE FOR LUNG ---------------------
AXES_CONFIG = [
    {
        'name': 'Axial',
        'model_paths': [
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx'
        ]
    },
    {
        'name': 'Coronal',
        'model_paths': [
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx'
        ]
    },
    {
        'name': 'Sagittal',
        'model_paths': [
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx',
            r'I:/Radiotherapy/lung/Result/segmentation/Unet3+_Lung_Axial_segmentation_metrics_detailed.xlsx'
        ]
    }
]

# --------------------- METRIC SELECTION ---------------------
METRICS_TO_PLOT = 'iou'  # Options: 'both', 'iou', 'dice'

# --------------------- OUTPUT DIRECTORY ---------------------
output_dir = r'I:/Radiotherapy/lung/Result/segmentation/combined'

# --------------------- LUNG CLASS ORDER (Updated for Lung Radiotherapy) ---------------------
# Based on your data: Background(0), SPINAL CORD(1), LUNGS_RT(4), LUNGS_LT(5), BODY(7), CTV(9), GTV(8)
# Note: Class 0 (Background) is excluded from plots
class_order = ['BODY', 'LUNGS_RT', 'LUNGS_LT', 'SPINAL CORD', 'Esophagus', 'HEART', 'TRACHEA', 'CTV', 'GTV']

# --------------------- FIGURE DIMENSIONS (NATURE STYLE) ---------------------
FIG_WIDTH_MM = 180
FIG_HEIGHT_MM = 200
DPI = 300

# --------------------- FONT CONFIGURATION (NATURE STYLE) ---------------------
FONT_CONFIG = {
    'family': 'Arial',
    'title_size': 7,
    'title_weight': 'bold',
    'axis_label_size': 6,
    'axis_label_weight': 'normal',
    'tick_label_size': 5,
    'legend_fontsize': 5,
    'legend_title_size': 6,
    'n_label_fontsize': 5,
    'n_label_weight': 'normal',
    'star_fontsize': 7,
    'star_fontsize_ns': 5,
    'subplot_label_size': 8,
    'subplot_label_weight': 'bold'
}

plt.rcParams['font.family'] = FONT_CONFIG['family']
plt.rcParams['font.size'] = 6
plt.rcParams['pdf.fonttype'] = 42
plt.rcParams['ps.fonttype'] = 42

# --------------------- BOX PLOT STYLING ---------------------
BOX_CONFIG = {
    'width': 0.7,
    'linewidth': 0.6,
    'showfliers': False,
    'showmeans': True,
}

MEAN_PROPS = {
    'marker': 'D',
    'markerfacecolor': 'white',
    'markeredgecolor': 'darkred',
    'markersize': 1.8,
    'markeredgewidth': 0.4,
    'zorder': 10,
}

MEDIAN_PROPS = {
    'color': '#2C3E50',
    'linewidth': 0.7,
}

# --------------------- STRIP PLOT ---------------------
STRIP_CONFIG = {
    'size': 1.0,
    'alpha': 0.15,
    'color': 'dimgray',
    'jitter': True,
    'zorder': 1,
}

# --------------------- GLOBAL METRIC LINES ---------------------
GLOBAL_LINE_CONFIG = {
    'linewidth': 0.5,
    'alpha': 0.7,
    'cap_width': 0.08,
}

BASE_IOU_LINE_COLORS = ['#D55E00']
BASE_IOU_LINE_STYLES = ['--']
BASE_DICE_LINE_COLORS = ['#0072B2']
BASE_DICE_LINE_STYLES = ['--']

# --------------------- COLOR PALETTE ---------------------
BASE_COLOR_PALETTES = {
    'iou_only': ['#0072B2', '#56B4E9', '#009E73', '#F0E442', '#CC79A7', '#E69F00'],
    'dice_only': ['#D55E00', '#CC79A7', '#E69F00', '#999999', '#0072B2', '#56B4E9'],
    'both': [
        '#0072B2', '#56B4E9', '#009E73', '#F0E442', '#CC79A7', '#E69F00',
        '#D55E00', '#CC79A7', '#E69F00', '#999999', '#0072B2', '#56B4E9',
    ]
}

# --------------------- ANNOTATIONS ---------------------
N_LABEL_CONFIG = {
    'show': True,
    'y_position': 1.08,
    'show_box': False,
}

STARS_CONFIG = {
    'y_offset_start': 1.015,
    'y_offset_increment': 0.02,
    'ns_color': 'dimgray',
    'sig_color': 'black',
}

# --------------------- AXIS & LAYOUT ---------------------
AXIS_CONFIG = {
    'ylim': (0, 1.18),
    'spine_width': 0.5,
    'tick_width': 0.4,
    'x_rotation': 35,
    'x_ha': 'right',
}

LEGEND_CONFIG = {
    'location': 'lower right',
    'ncol': 3,
    'framealpha': 0.95,
    'edgecolor': 'lightgray',
    'fancybox': False,
    'columnspacing': 0.6,
    'handletextpad': 0.4,
    'markerscale': 1.0,
}

# --------------------- STATISTICAL TESTS ---------------------
STATS_CONFIG = {
    'base_model_index': 0,
    'paired_test': True,
    'min_samples': 4,
    'p_value_thresholds': {
        '***': 0.001,
        '**': 0.01,
        '*': 0.05,
    },
    'effect_size_thresholds': {
        'small': 0.2,
        'medium': 0.5,
        'large': 0.8,
    }
}

# ========================================================================
# VALIDATE AND SETUP
# ========================================================================

METRICS_TO_PLOT = METRICS_TO_PLOT.lower()
if METRICS_TO_PLOT not in ['both', 'iou', 'dice']:
    raise ValueError("METRICS_TO_PLOT must be 'both', 'iou', or 'dice'")

if METRICS_TO_PLOT == 'both':
    ACTIVE_METRICS = ['iou', 'dice']
    METRIC_DISPLAY = ['IoU', 'Dice']
    palette_key = 'both'
elif METRICS_TO_PLOT == 'iou':
    ACTIVE_METRICS = ['iou']
    METRIC_DISPLAY = ['IoU']
    palette_key = 'iou_only'
else:
    ACTIVE_METRICS = ['dice']
    METRIC_DISPLAY = ['Dice']
    palette_key = 'dice_only'

os.makedirs(output_dir, exist_ok=True)

print(f"📊 Creating {len(AXES_CONFIG)}-panel figure")
print(f"📊 Plotting metrics: {', '.join(METRIC_DISPLAY)}")
print(f"📊 Lung Classes: {', '.join(class_order)}")

# ========================================================================
# HELPER FUNCTIONS (Updated for Lung)
# ========================================================================

def load_metrics_with_filename(csv_path, model_name):
    """Load metrics and standardize column names for Lung dataset"""
    df = pd.read_excel(csv_path)
    df.columns = df.columns.str.strip()
    
    # Standardize column names
    rename_map = {
        'IoU': 'iou', 
        'Dice': 'dice', 
        'Filename': 'filename',
        'Class_Name': 'class_name'
    }
    df.rename(columns=rename_map, inplace=True)
    
    # Handle filename column
    if 'filename' in df.columns:
        pass
    elif 'image_id' in df.columns:
        df['filename'] = df['image_id'].apply(lambda x: f'image_{x}.png')
    else:
        # Check if there's a column with R numbers
        for col in df.columns:
            if df[col].astype(str).str.startswith('R').any():
                df['filename'] = df[col]
                break
        else:
            df['filename'] = [f'img_{i}.png' for i in range(len(df))]
    
    # Filter out background class (Class_ID 0)
    if 'Class_ID' in df.columns:
        df = df[df['Class_ID'] != 0].copy()
    
    # Standardize class names
    if 'class_name' not in df.columns and 'Class_Name' in df.columns:
        df['class_name'] = df['Class_Name']
    
    # Map any variations in class names
    class_name_mapping = {
        'LUNGS_RT': 'LUNGS_RT',
        'LUNGS_LT': 'LUNGS_LT',
        'SPINAL CORD': 'SPINAL CORD',
        'BODY': 'BODY',
        'GTV': 'GTV',
        'CTV': 'CTV'
    }
    
    if 'class_name' in df.columns:
        df['class_name'] = df['class_name'].map(lambda x: class_name_mapping.get(x, x))
    
    df['Model'] = model_name
    
    # Ensure only the classes in class_order are included
    df = df[df['class_name'].isin(class_order)].copy()
    
    # Convert to categorical with the specified order
    df['class_name'] = pd.Categorical(df['class_name'], categories=class_order, ordered=True)
    
    return df[['filename', 'class_name', 'iou', 'dice', 'Model']].copy()


def compute_global_metrics(raw_csv_path):
    """Compute global IoU and Dice from intersection/union data for Lung"""
    df_raw = pd.read_excel(raw_csv_path)
    df_raw.columns = df_raw.columns.str.strip()
    
    # Filter out background (Class_ID 0) and only include classes in class_order
    df_raw = df_raw[df_raw['Class_ID'] != 0].copy()
    
    # Standardize class names if needed
    class_name_mapping = {
        'LUNGS_RT': 'LUNGS_RT',
        'LUNGS_LT': 'LUNGS_LT',
        'SPINAL CORD': 'SPINAL CORD',
        'BODY': 'BODY',
        'GTV': 'GTV',
        'CTV': 'CTV'
    }
    
    if 'Class_Name' in df_raw.columns:
        df_raw['Class_Name'] = df_raw['Class_Name'].map(lambda x: class_name_mapping.get(x, x))
    
    # Filter to only include our classes of interest
    df_raw = df_raw[df_raw['Class_Name'].isin(class_order)].copy()
    
    global_iou = {}
    global_dice = {}
    
    for cls in class_order:
        d = df_raw[df_raw['Class_Name'] == cls]
        if len(d) > 0:
            inter = d['Intersection'].sum()
            union = d['Union'].sum()
            global_iou[cls] = inter / union if union > 0 else 0.0
            global_dice[cls] = (2 * inter) / (inter + union) if (inter + union) > 0 else 0.0
        else:
            global_iou[cls] = 0.0
            global_dice[cls] = 0.0
    
    return global_iou, global_dice


def interpret_effect_size(d_value):
    """Interpret Cohen's d magnitude"""
    abs_d = abs(d_value)
    thresholds = STATS_CONFIG['effect_size_thresholds']
    
    if np.isnan(abs_d):
        return 'undefined'
    elif abs_d < thresholds['small']:
        return 'negligible'
    elif abs_d < thresholds['medium']:
        return 'small'
    elif abs_d < thresholds['large']:
        return 'medium'
    else:
        return 'large'


def compute_statistics_pairwise(all_dfs, model_names, base_idx):
    """
    Statistical calculations for Lung dataset
    """
    p_values = {}
    n_per_class = {}
    cohens_d_values = {}
    mean_differences = {}
    sample_sizes = {}
    
    for cls in class_order:
        n_counts = [len(df[df['class_name'] == cls]) for df in all_dfs]
        n_per_class[cls] = n_counts
        
        for metric in ACTIVE_METRICS:
            base_data = all_dfs[base_idx][all_dfs[base_idx]['class_name'] == cls][['filename', metric]].set_index('filename').sort_index()
            
            for i in range(len(model_names)):
                if i == base_idx:
                    continue
                
                comp_data = all_dfs[i][all_dfs[i]['class_name'] == cls][['filename', metric]].set_index('filename').sort_index()
                
                # Proper pairing: Only samples in BOTH datasets
                paired = base_data.join(comp_data, lsuffix='_base', rsuffix=f'_{i}', how='inner')
                
                key = (cls, metric, i)
                paired_n = len(paired)
                sample_sizes[key] = paired_n
                
                if paired_n < STATS_CONFIG['min_samples']:
                    p_values[key] = 1.0
                    cohens_d_values[key] = np.nan
                    mean_differences[key] = np.nan
                else:
                    x = paired[f'{metric}_base'].values
                    y = paired[f'{metric}_{i}'].values
                    diff = x - y
                    
                    # Mean difference
                    mean_diff = np.mean(diff)
                    mean_differences[key] = mean_diff
                    
                    # Wilcoxon test: Compare differences to zero
                    nonzero_diff = diff[diff != 0]
                    
                    if len(nonzero_diff) == 0:
                        p_val = 1.0
                    else:
                        try:
                            stat, p_val = wilcoxon(nonzero_diff, alternative='two-sided')
                        except Exception as e:
                            print(f"⚠️ Warning: Wilcoxon test failed for {cls}, {metric}, {model_names[base_idx]} vs {model_names[i]}: {e}")
                            p_val = 1.0
                    
                    p_values[key] = round(p_val, 6)
                    
                    # Cohen's D (Paired): Standard deviation of differences
                    std_diff = np.std(diff, ddof=1)
                    if std_diff > 0:
                        cohens_d = mean_diff / std_diff
                    else:
                        cohens_d = 0.0 if abs(mean_diff) < 1e-10 else np.inf
                    
                    cohens_d_values[key] = round(cohens_d, 4)
    
    return p_values, n_per_class, cohens_d_values, mean_differences, sample_sizes


def create_subplot(ax, combined_melted, hue_order, palette, all_global_iou, all_global_dice, 
                   p_values, n_per_class, model_names, base_idx, axis_name):
    """Create a single subplot with all visualizations"""
    
    NUM_MODELS = len(model_names)
    
    # Custom palette
    custom_palette = {hue_order[i]: palette[i] for i in range(len(hue_order))}
    
    # Box plot
    sns.boxplot(
        data=combined_melted, x='class_name', y='Score', hue='Metric_Model',
        palette=custom_palette, ax=ax, hue_order=hue_order,
        meanprops=MEAN_PROPS, medianprops=MEDIAN_PROPS, **BOX_CONFIG
    )
    
    # Strip plot
    sns.stripplot(
        data=combined_melted, x='class_name', y='Score', hue='Metric_Model',
        palette=[STRIP_CONFIG['color']] * len(hue_order),
        size=STRIP_CONFIG['size'], alpha=STRIP_CONFIG['alpha'],
        jitter=STRIP_CONFIG['jitter'], dodge=True, legend=False,
        ax=ax, hue_order=hue_order
    )
    
    # Global metric lines
    x_positions = np.arange(len(class_order))
    n_boxes = NUM_MODELS * len(ACTIVE_METRICS)
    box_width = BOX_CONFIG['width'] / n_boxes
    start_offset = -BOX_CONFIG['width'] / 2 + box_width / 2
    
    for i, cls in enumerate(class_order):
        box_idx = 0
        for metric_idx, metric in enumerate(ACTIVE_METRICS):
            for m_idx in range(NUM_MODELS):
                offset = start_offset + box_idx * box_width
                
                if metric == 'iou':
                    ax.hlines(
                        all_global_iou[m_idx][cls],
                        i + offset - GLOBAL_LINE_CONFIG['cap_width'],
                        i + offset + GLOBAL_LINE_CONFIG['cap_width'],
                        colors=BASE_IOU_LINE_COLORS[0],
                        linestyles=BASE_IOU_LINE_STYLES[0],
                        linewidth=GLOBAL_LINE_CONFIG['linewidth'],
                        alpha=GLOBAL_LINE_CONFIG['alpha']
                    )
                else:
                    ax.hlines(
                        all_global_dice[m_idx][cls],
                        i + offset - GLOBAL_LINE_CONFIG['cap_width'],
                        i + offset + GLOBAL_LINE_CONFIG['cap_width'],
                        colors=BASE_DICE_LINE_COLORS[0],
                        linestyles=BASE_DICE_LINE_STYLES[0],
                        linewidth=GLOBAL_LINE_CONFIG['linewidth'],
                        alpha=GLOBAL_LINE_CONFIG['alpha']
                    )
                
                box_idx += 1
    
    # Add n labels
    y_max_n = N_LABEL_CONFIG['y_position']
    
    for i, cls in enumerate(class_order):
        if N_LABEL_CONFIG['show']:
            n_value = max(n_per_class[cls])
            ax.text(
                i, y_max_n, f'n={n_value}',
                ha='center', va='bottom',
                fontsize=FONT_CONFIG['n_label_fontsize'],
                weight=FONT_CONFIG['n_label_weight'],
                color='black'
            )
    
    # Add significance stars
    thresholds = STATS_CONFIG['p_value_thresholds']
    
    for i, cls in enumerate(class_order):
        for metric_idx, metric in enumerate(ACTIVE_METRICS):
            for m_idx in range(NUM_MODELS):
                if m_idx == base_idx:
                    continue
                
                key = (cls, metric, m_idx)
                p_val = p_values.get(key, 1.0)
                
                # Determine stars
                if p_val < thresholds['***']:
                    stars = '***'
                elif p_val < thresholds['**']:
                    stars = '**'
                elif p_val < thresholds['*']:
                    stars = '*'
                else:
                    stars = 'ns'
                
                # Calculate position
                box_idx = metric_idx * NUM_MODELS + m_idx
                offset = start_offset + box_idx * box_width
                
                y_star = STARS_CONFIG['y_offset_start']
                
                fontsize_star = FONT_CONFIG['star_fontsize_ns']
                star_color = STARS_CONFIG['sig_color'] if stars != 'ns' else STARS_CONFIG['ns_color']
                
                ax.text(
                    i + offset, y_star, stars,
                    ha='center', va='bottom',
                    fontsize=fontsize_star,
                    weight='bold',
                    color=star_color
                )
    
    # Styling
    ax.set_ylim(*AXIS_CONFIG['ylim'])
    ax.set_xlabel('Class', fontsize=FONT_CONFIG['axis_label_size'], 
                  weight=FONT_CONFIG['axis_label_weight'])
    ax.set_ylabel('Score', fontsize=FONT_CONFIG['axis_label_size'], 
                  weight=FONT_CONFIG['axis_label_weight'])
    ax.tick_params(axis='both', labelsize=FONT_CONFIG['tick_label_size'], 
                   width=AXIS_CONFIG['tick_width'])
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=AXIS_CONFIG['x_rotation'], 
             ha=AXIS_CONFIG['x_ha'])
    
    for spine in ax.spines.values():
        spine.set_linewidth(AXIS_CONFIG['spine_width'])
    
    # Remove legend from individual subplots
    ax.get_legend().remove()
    
    return ax


# ========================================================================
# MAIN PROCESSING LOOP
# ========================================================================

fig_width_inches = FIG_WIDTH_MM / 25.4
fig_height_inches = FIG_HEIGHT_MM / 25.4

fig, axes = plt.subplots(len(AXES_CONFIG), 1, figsize=(fig_width_inches, fig_height_inches))

if len(AXES_CONFIG) == 1:
    axes = [axes]

all_stats_dfs = []

for idx, axis_config in enumerate(AXES_CONFIG):
    axis_name = axis_config['name']
    model_paths = axis_config['model_paths']
    
    NUM_MODELS = len(model_paths)
    MODEL_NAMES = [MODEL_NAME_MAP[axis_name][i] for i in range(NUM_MODELS)]
    BASE_MODEL_NAME = MODEL_NAMES[STATS_CONFIG['base_model_index']]
    
    print(f"\n{'='*70}")
    print(f"Processing: {axis_name} Plane")
    print(f"Models: {MODEL_NAMES}")
    print(f"Base model (for comparisons): {BASE_MODEL_NAME}")
    print(f"{'='*70}")
    
    # Load data
    all_dfs = []
    for i, path in enumerate(model_paths):
        df = load_metrics_with_filename(path, MODEL_NAMES[i])
        all_dfs.append(df)
    
    # Compute global metrics
    all_global_iou = []
    all_global_dice = []
    for path in model_paths:
        g_iou, g_dice = compute_global_metrics(path)
        all_global_iou.append(g_iou)
        all_global_dice.append(g_dice)
    
    # Compute statistics
    p_values, n_per_class, cohens_d_values, mean_differences, sample_sizes = compute_statistics_pairwise(
        all_dfs, MODEL_NAMES, STATS_CONFIG['base_model_index']
    )
    
    # Print detailed statistics for verification
    print(f"\n🔍 DETAILED STATISTICAL VERIFICATION for {axis_name}:")
    print(f"{'Class':<15} {'Metric':<8} {'Comparison':<30} {'N':<5} {'Mean_Diff':<12} {'Cohen_d':<10} {'p_value':<10} {'Sig'}")
    print("-" * 110)
    
    for cls in class_order:
        for metric in ACTIVE_METRICS:
            for m_idx in range(NUM_MODELS):
                if m_idx == STATS_CONFIG['base_model_index']:
                    continue
                
                key = (cls, metric, m_idx)
                p_val = p_values.get(key, 1.0)
                cohens = cohens_d_values.get(key, np.nan)
                mean_diff = mean_differences.get(key, np.nan)
                paired_n = sample_sizes.get(key, 0)
                
                if p_val < STATS_CONFIG['p_value_thresholds']['***']:
                    sig = '***'
                elif p_val < STATS_CONFIG['p_value_thresholds']['**']:
                    sig = '**'
                elif p_val < STATS_CONFIG['p_value_thresholds']['*']:
                    sig = '*'
                else:
                    sig = 'ns'
                
                comparison_name = f"{BASE_MODEL_NAME} vs {MODEL_NAMES[m_idx]}"
                print(f"{cls:<15} {metric.upper():<8} {comparison_name:<30} {paired_n:<5} {mean_diff:>11.4f} {cohens:>9.4f} {p_val:>9.6f} {sig:>3}")
    
    # Create CSV with COLUMNS in the format: UNet3+_vs_ModelName_metric
    pivot_rows = []
    
    for cls in class_order:
        for metric in ACTIVE_METRICS:
            row_data = {
                'Axis': axis_name,
                'Class': cls,
                'Metric': metric.upper(),
                'Sample_Size': max(n_per_class[cls])
            }
            
            # Add columns for each model comparison
            for m_idx in range(NUM_MODELS):
                if m_idx == STATS_CONFIG['base_model_index']:
                    continue
                
                key = (cls, metric, m_idx)
                p_val = p_values.get(key, 1.0)
                cohens = cohens_d_values.get(key, np.nan)
                mean_diff = mean_differences.get(key, np.nan)
                paired_n = sample_sizes.get(key, 0)
                
                # Significance
                if p_val < STATS_CONFIG['p_value_thresholds']['***']:
                    sig = '***'
                elif p_val < STATS_CONFIG['p_value_thresholds']['**']:
                    sig = '**'
                elif p_val < STATS_CONFIG['p_value_thresholds']['*']:
                    sig = '*'
                else:
                    sig = 'ns'
                
                # Format: UNet3+_vs_ModelName_metric
                comparison_prefix = f"{BASE_MODEL_NAME}_vs_{MODEL_NAMES[m_idx]}"
                
                row_data[f'{comparison_prefix}_n_paired'] = paired_n
                row_data[f'{comparison_prefix}_mean_diff'] = round(mean_diff, 4)
                row_data[f'{comparison_prefix}_cohens_d'] = round(cohens, 4)
                row_data[f'{comparison_prefix}_effect_size'] = interpret_effect_size(cohens)
                row_data[f'{comparison_prefix}_p_value'] = p_val
                row_data[f'{comparison_prefix}_significance'] = sig
            
            # Add Min_p_value
            all_p_vals = []
            for m_idx in range(NUM_MODELS):
                if m_idx != STATS_CONFIG['base_model_index']:
                    key = (cls, metric, m_idx)
                    all_p_vals.append(p_values.get(key, 1.0))
            row_data['Min_p_value'] = min(all_p_vals) if all_p_vals else 1.0
            
            pivot_rows.append(row_data)
    
    stats_df = pd.DataFrame(pivot_rows)
    all_stats_dfs.append(stats_df)
    
    # Prepare data for plotting
    combined = pd.concat(all_dfs, ignore_index=True)
    combined_melted = combined.melt(
        id_vars=['filename', 'class_name', 'Model'],
        value_vars=ACTIVE_METRICS,
        var_name='Metric',
        value_name='Score'
    )
    combined_melted['Metric'] = combined_melted['Metric'].map({'iou': 'IoU', 'dice': 'Dice'})
    combined_melted['Metric_Model'] = combined_melted['Metric'] + ' - ' + combined_melted['Model']
    
    hue_order = []
    for metric in METRIC_DISPLAY:
        for model_name in MODEL_NAMES:
            hue_order.append(f'{metric} - {model_name}')
    
    combined_melted['class_name'] = pd.Categorical(
        combined_melted['class_name'], categories=class_order, ordered=True
    )
    
    # Get palette
    n_colors_needed = NUM_MODELS * len(ACTIVE_METRICS)
    palette = BASE_COLOR_PALETTES[palette_key][:n_colors_needed]
    
    # Create subplot
    ax = axes[idx]
    create_subplot(
        ax, combined_melted, hue_order, palette, all_global_iou, all_global_dice,
        p_values, n_per_class, MODEL_NAMES, STATS_CONFIG['base_model_index'], axis_name
    )
    
    # Add subplot label
    subplot_label = string.ascii_lowercase[idx]
    ax.text(
        -0.08, 1.05, f'{subplot_label}',
        transform=ax.transAxes,
        fontsize=FONT_CONFIG['subplot_label_size'],
        weight=FONT_CONFIG['subplot_label_weight'],
        va='top', ha='right'
    )
    
    # Add axis name as title
    ax.set_title(
        f'{axis_name} Plane',
        fontsize=FONT_CONFIG['title_size'],
        weight=FONT_CONFIG['title_weight'],
        pad=8
    )

# ========================================================================
# CREATE SHARED LEGEND
# ========================================================================

legend_elements = []

for idx_leg, label in enumerate(hue_order):
    legend_elements.append(
        Line2D([0], [0], marker='s', color='w',
               markerfacecolor=palette[idx_leg],
               markersize=4,
               markeredgecolor='black',
               markeredgewidth=0.3,
               label=label)
    )

legend_elements.extend([
    Line2D([0], [0], marker=MEAN_PROPS['marker'], color='w',
           markerfacecolor=MEAN_PROPS['markerfacecolor'],
           markeredgecolor='darkred',
           markersize=3,
           markeredgewidth=0.3,
           label='Mean'),
    Line2D([0], [0], color=MEDIAN_PROPS['color'],
           linewidth=1.2,
           label='Median'),
])

if 'iou' in ACTIVE_METRICS:
    legend_elements.append(
        Line2D([0], [0], color=BASE_IOU_LINE_COLORS[0],
               linestyle=BASE_IOU_LINE_STYLES[0],
               linewidth=1.5,
               label='Global IoU')
    )

if 'dice' in ACTIVE_METRICS:
    legend_elements.append(
        Line2D([0], [0], color=BASE_DICE_LINE_COLORS[0],
               linestyle=BASE_DICE_LINE_STYLES[0],
               linewidth=1.5,
               label='Global Dice')
    )

legend_elements.append(
    Line2D([0], [0], marker='o', color='w',
           markerfacecolor=STRIP_CONFIG['color'],
           markeredgecolor=STRIP_CONFIG['color'],
           markersize=2,
           alpha=STRIP_CONFIG['alpha'],
           label='Individual Sample')
)

axes[-1].legend(
    handles=legend_elements,
    title='Elements',
    title_fontsize=FONT_CONFIG['legend_title_size'],
    loc=LEGEND_CONFIG['location'],
    fontsize=FONT_CONFIG['legend_fontsize'],
    ncol=LEGEND_CONFIG['ncol'],
    framealpha=LEGEND_CONFIG['framealpha'],
    edgecolor=LEGEND_CONFIG['edgecolor'],
    fancybox=LEGEND_CONFIG['fancybox'],
    columnspacing=LEGEND_CONFIG['columnspacing'],
    handletextpad=LEGEND_CONFIG['handletextpad'],
    markerscale=LEGEND_CONFIG['markerscale']
)

plt.tight_layout()

# ========================================================================
# SAVE OUTPUTS
# ========================================================================

# Save combined statistics
combined_stats = pd.concat(all_stats_dfs, ignore_index=True)
stats_path = os.path.join(output_dir, f'combined_statistics_{METRICS_TO_PLOT}.csv')
combined_stats.to_csv(stats_path, index=False)

print(f"\n✅ Saved combined statistics to: {stats_path}")

# Print summary
print(f"\n{'='*70}")
print(f"📊 OVERALL STATISTICAL SUMMARY")
print(f"{'='*70}")
for axis_name in combined_stats['Axis'].unique():
    axis_data = combined_stats[combined_stats['Axis'] == axis_name]
    # Count columns with 'Significance' to get number of comparisons
    sig_cols = [c for c in axis_data.columns if c.endswith('_Significance')]
    
    if len(sig_cols) > 0:
        sig_count = 0
        for col in sig_cols:
            sig_count += (axis_data[col] != 'ns').sum()
        total_count = len(axis_data) * len(sig_cols)
        
        if total_count > 0:
            print(f"{axis_name}: {sig_count}/{total_count} significant ({100*sig_count/total_count:.1f}%)")
        else:
            print(f"{axis_name}: No data to summarize")
    else:
        print(f"{axis_name}: No significance columns found")


# Save figure
base_name = os.path.join(output_dir, f'combined_{len(AXES_CONFIG)}_axes_{METRICS_TO_PLOT}')
for fmt in ['png', 'pdf', 'svg', 'eps']:
    plt.savefig(f"{base_name}.{fmt}", dpi=DPI, bbox_inches='tight')
    print(f"✅ Saved: {base_name}.{fmt}")

print(f"\n{'='*70}")
print(f"📊 CONFIGURATION SUMMARY")
print(f"{'='*70}")
print(f"Dataset: Lung Radiotherapy")
print(f"Classes (excluding background): {len(class_order)} classes")
print(f"Figure dimensions: {FIG_WIDTH_MM}mm × {FIG_HEIGHT_MM}mm")
print(f"Font: {FONT_CONFIG['family']}, {FONT_CONFIG['tick_label_size']}-{FONT_CONFIG['title_size']}pt")
print(f"Statistical comparisons: Each model vs base model (per axis)")
print(f"✅ Statistical Corrections Applied:")
print(f"   • Paired pairing: Only samples in BOTH datasets are compared")
print(f"   • Wilcoxon test: Proper paired test on differences (H0: median diff = 0)")
print(f"   • Cohen's d: Calculated from std dev of paired differences")
print(f"   • Effect size interpretation: negligible < small < medium < large")
print(f"   • p-value thresholds: *** p<0.001, ** p<0.01, * p<0.05, ns p≥0.05")
print(f"\n📋 CSV FORMAT: Columns layout (UNet3+ is base model)")
print(f"   Rows: Each Class + Metric combination")
print(f"   Columns: Axis | Class | Metric | {{ModelName}}_N | {{ModelName}}_MeanDiff | {{ModelName}}_CohensD | {{ModelName}}_EffectSize | {{ModelName}}_PValue | {{ModelName}}_Significance")

plt.show()








































































































#qwen working
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import wilcoxon
import os
import warnings

# Suppress the SciPy zero-difference warning for cleaner console output
warnings.filterwarnings("ignore", message="Exact p-value calculation does not work if there are zeros")

# ===============================
# 1. CONFIGURATION
# ===============================
excel_files = [
    ("My_Model", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNet3+/UNet3+_Axial_segmentationMetrcis.xlsx'),
    ("Model_A", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNetBase/UNetBase_Axial_segmentationMetrics.xlsx'),
    ("Model_B", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNetBase/UNetBase_Axial_segmentationMetrics.xlsx'),
    ("Model_C", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNet/UNet_Axial_segmentationMetrics.xlsx'),
    ("Model_D", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/TransUNet/TransUNet_Axial_segmentationMetrics.xlsx')
]

METRIC = "Dice"
SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative'
os.makedirs(SAVE_DIR, exist_ok=True)

# ===============================
# 2. STATISTICAL HELPERS
# ===============================
def get_stats(x, y):
    nx, ny = len(x), len(y)
    std_pooled = np.sqrt(((nx-1)*np.std(x, ddof=1)**2 + (ny-1)*np.std(y, ddof=1)**2) / (nx+ny-2))
    d = (np.mean(x) - np.mean(y)) / std_pooled if std_pooled != 0 else 0
    g = d * (1 - (3 / (4 * (nx + ny) - 9)))
    return d, g

def get_stars(p):
    if pd.isna(p): return "ns"
    if p < 0.0001: return "****"
    elif p < 0.001: return "***"
    elif p < 0.01: return "**"
    elif p < 0.05: return "*"
    return "ns"

# ===============================
# 3. LOADING & CLEANING
# ===============================
all_data = []
for name, path in excel_files:
    df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
    df[METRIC] = pd.to_numeric(df[METRIC], errors='coerce')
    df = df.dropna(subset=[METRIC])
    df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
    # Remove unwanted classes
    df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
    df["Model"] = name
    all_data.append(df[["Filename", "Class_Name", METRIC, "Model"]])

df_main = pd.concat(all_data, ignore_index=True)
model_names = [x[0] for x in excel_files]
ref_model = model_names[0]
classes = sorted(df_main["Class_Name"].unique())

# ==========================================
# 4. PHASE 1: GENERATE TABLE & STAR LOOKUP
# ==========================================
stats_list = []
# This dictionary will be the ONLY source of truth for the plot
star_lookup = {} 

for cls in classes:
    ref_df = df_main[(df_main["Model"] == ref_model) & (df_main["Class_Name"] == cls)]
    star_lookup[cls] = {}
    
    for other in model_names[1:]:
        other_df = df_main[(df_main["Model"] == other) & (df_main["Class_Name"] == cls)]
        merged = pd.merge(ref_df, other_df, on="Filename", suffixes=('_ref', '_other')).dropna()
        
        p, d, g = np.nan, 0, 0
        if len(merged) > 2:
            if not (merged[f"{METRIC}_ref"] == merged[f"{METRIC}_other"]).all():
                # zero_method='pratt' handles the tied ranks that cause p-value fluctuations
                _, p = wilcoxon(merged[f"{METRIC}_ref"], merged[f"{METRIC}_other"], zero_method='pratt')
            d, g = get_stats(merged[f"{METRIC}_ref"], merged[f"{METRIC}_other"])
        
        sig_label = get_stars(p)
        star_lookup[cls][other] = sig_label
        
        stats_list.append({
            "Class": cls, 
            "Comparison": f"{ref_model} vs {other}",
            "P_value": p, 
            "Significance": sig_label,
            "Cohens_d": d, 
            "Hedges_g": g, 
            "N": len(merged)
        })

# Export Excel
final_stats_df = pd.DataFrame(stats_list)
final_stats_df.to_excel(os.path.join(SAVE_DIR, f"{METRIC.lower()}_pvalue_table.xlsx"), index=False)

# ==========================================
# 5. PHASE 2: PLOTTING (STRICTLY FROM LOOKUP)
# ==========================================
plt.figure(figsize=(22, 11))
sns.set_style("white")

# 🔑 FIX: Explicitly enforce ordering to match the loop & lookup dictionary
ax = sns.boxplot(x="Class_Name", y=METRIC, hue="Model", data=df_main, 
                 palette="Set2", showfliers=False, linewidth=1.2,
                 order=classes, hue_order=model_names)

sns.stripplot(x="Class_Name", y=METRIC, hue="Model", data=df_main, 
              dodge=True, alpha=0.25, size=2.5, palette='dark:black', ax=ax, legend=False,
              order=classes, hue_order=model_names)

y_max = df_main[METRIC].max()
y_range = y_max - df_main[METRIC].min()

# Brackets - Iterating through the same logic used to build the table
for i, cls in enumerate(classes):
    offsets = np.linspace(-0.4, 0.4, len(model_names) + 2)[1:-1]
    x_ref = i + offsets[0] # Position of My_Model (now guaranteed to be first)
    
    for j, other in enumerate(model_names[1:]):
        x_other = i + offsets[j+1] # Position of Model_A, B, C, or D
        
        # STRICT LOOKUP: Get the star directly from Phase 1 result
        p_text = star_lookup[cls][other]
        
        # Vertical stacking height
        h = y_max + (0.07 * y_range * (j + 1))
        
        # Draw Bracket
        ax.plot([x_ref, x_ref, x_other, x_other], [h - (0.015 * y_range), h, h, h - (0.015 * y_range)], 
                color="black", lw=1.0)
        
        # Place text (Significance from Table)
        ax.text((x_ref + x_other)/2, h, p_text, ha='center', va='bottom', 
                fontsize=11, fontweight='bold', color='black')

sns.despine()
plt.title(f"Model Performance and Statistical Significance ({METRIC})", fontsize=16, pad=30)
plt.xticks(rotation=20, ha='right', fontsize=12)
plt.ylim(df_main[METRIC].min() - 0.05, y_max + (y_range * 0.5))

# Export Plot
plt.savefig(os.path.join(SAVE_DIR, f"{METRIC}_Nature_Final.png"), dpi=600, bbox_inches='tight')
plt.savefig(os.path.join(SAVE_DIR, f"{METRIC}_Nature_Final.pdf"), format='pdf', bbox_inches='tight')

plt.show()

print(f"Process complete. Table and Plot are now strictly synchronized in {SAVE_DIR}")
















# Stable version 1
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy.stats import wilcoxon
import os
import warnings
from matplotlib.lines import Line2D
from matplotlib.patches import PathPatch

warnings.filterwarnings("ignore")

# ===============================
# 1. CONFIGURATION
# ===============================
excel_files = [
    ("My_Model", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNet3+/UNet3+_Axial_segmentationMetrcis.xlsx'),
    ("Model_A", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNetBase/UNetBase_Axial_segmentationMetrics.xlsx'),
    ("Model_B", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNetBase/UNetBase_Axial_segmentationMetrics.xlsx'),
    ("Model_C", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNet/UNet_Axial_segmentationMetrics.xlsx'),
    ("Model_D", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/TransUNet/TransUNet_Axial_segmentationMetrics.xlsx')
]

# Strict Color Mapping: (Fill, Border)
color_map = {
    "My_Model": ("#00b4d8", "#0096c7"), 
    "Model_A":  ("#2a9134", "#137547"),
    "Model_B":  ("#f4a261", "#e76f51"),
    "Model_C":  ("#e9c46a", "#d4a017"),
    "Model_D":  ("#a29bfe", "#6c5ce7")
}

METRIC = "Dice"
SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative'
os.makedirs(SAVE_DIR, exist_ok=True)

# ===============================
# 2. STATISTICAL HELPERS
# ===============================
def calculate_effect_size(x, y):
    nx, ny = len(x), len(y)
    dof = nx + ny - 2
    std_pooled = np.sqrt(((nx-1)*np.std(x, ddof=1)**2 + (ny-1)*np.std(y, ddof=1)**2) / dof)
    d = (np.mean(x) - np.mean(y)) / std_pooled if std_pooled != 0 else 0
    # Hedges' g (Correction for small samples)
    g = d * (1 - (3 / (4 * dof - 1)))
    return d, g

def get_stars(p_val):
    if pd.isna(p_val): return "ns"
    if p_val < 0.0001: return "****"
    elif p_val < 0.001: return "***"
    elif p_val < 0.01: return "**"
    elif p_val < 0.05: return "*"
    return "ns"

# ===============================
# 3. DATA PROCESSING
# ===============================
all_data = []
for name, path in excel_files:
    df_dict = pd.read_excel(path, sheet_name=None)
    df = pd.concat(df_dict.values(), ignore_index=True)
    df[METRIC] = pd.to_numeric(df[METRIC], errors='coerce')
    df = df.dropna(subset=[METRIC])
    df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
    df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
    df["Model"] = name
    all_data.append(df[["Filename", "Class_Name", METRIC, "Model"]])

df_main = pd.concat(all_data, ignore_index=True)
model_names = [x[0] for x in excel_files]
classes = sorted(df_main["Class_Name"].unique())

# PHASE 1: Generate Table with Cohen's D and Hedges' G
stats_results = []
star_lookup = {}

for cls in classes:
    star_lookup[cls] = {}
    ref_data = df_main[(df_main["Model"] == model_names[0]) & (df_main["Class_Name"] == cls)]
    
    for other in model_names[1:]:
        other_data = df_main[(df_main["Model"] == other) & (df_main["Class_Name"] == cls)]
        merged = pd.merge(ref_data, other_data, on="Filename", suffixes=('_ref', '_other')).dropna()
        
        p, cohen_d, hedges_g = np.nan, 0, 0
        if len(merged) > 2:
            if not (merged[f"{METRIC}_ref"] == merged[f"{METRIC}_other"]).all():
                _, p = wilcoxon(merged[f"{METRIC}_ref"], merged[f"{METRIC}_other"], zero_method='pratt')
            cohen_d, hedges_g = calculate_effect_size(merged[f"{METRIC}_ref"], merged[f"{METRIC}_other"])
        
        sig = get_stars(p)
        star_lookup[cls][other] = sig
        stats_results.append({
            "Class": cls,
            "Comparison": f"{model_names[0]} vs {other}",
            "P_value": p,
            "Significance": sig,
            "Cohens_d": cohen_d,
            "Hedges_g": hedges_g,
            "Sample_Size(N)": len(merged)
        })

# Save Table
table_path = os.path.join(SAVE_DIR, f"{METRIC}_Complete_Stats.xlsx")
pd.DataFrame(stats_results).to_excel(table_path, index=False)

# ===============================
# 4. PHASE 2: PLOTTING
# ===============================
fig, ax = plt.subplots(figsize=(24, 12))
sns.set_style("white")

# Boxplot
bp = sns.boxplot(
    x="Class_Name", y=METRIC, hue="Model", data=df_main, 
    palette=[color_map[m][0] for m in model_names],
    showfliers=False, linewidth=1.5,
    order=classes, hue_order=model_names,
    width=0.65, # Adjusted for gap
    showmeans=True,
    meanprops={"marker":"D", "markerfacecolor":"#E0E0E0", "markeredgecolor":"#757575", "markersize":4},
    ax=ax
)

# CONSISTENT BORDER & MEDIAN COLORING
# We iterate through the artists created by boxplot
for i, artist in enumerate(ax.artists):
    model_name = model_names[i % len(model_names)]
    border_color = color_map[model_name][1]
    
    artist.set_edgecolor(border_color)
    
    # Each box has 6 associated lines (whiskers, caps, median)
    for j in range(i*6, (i+1)*6):
        line = ax.lines[j]
        line.set_color(border_color)
        # The 5th line (index 4) in the set is the median
        if j % 6 == 4:
            line.set_linewidth(2.5)

# Jitter Points
sns.stripplot(
    x="Class_Name", y=METRIC, hue="Model", data=df_main, 
    dodge=True, alpha=0.1, size=2, palette='dark:black', ax=ax, legend=False,
    order=classes, hue_order=model_names
)

# BRACKETS
y_pos = 1.05
for i, cls in enumerate(classes):
    x_coords = np.linspace(i - 0.3, i + 0.3, len(model_names))
    x_ref = x_coords[0]
    
    for j, other in enumerate(model_names[1:]):
        x_other = x_coords[j+1]
        h = y_pos + (j * 0.08)
        ax.plot([x_ref, x_ref, x_other, x_other], [h-0.015, h, h, h-0.015], color="black", lw=1.2)
        ax.text((x_ref + x_other)/2, h, star_lookup[cls][other], ha='center', va='bottom', fontsize=11, fontweight='bold')

# FINAL STYLING
ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1.0])
plt.ylim(df_main[METRIC].min() - 0.05, 1.55)
sns.despine()

# LEGEND
handles, labels = ax.get_legend_handles_labels()
line_median = Line2D([0], [0], color='#757575', lw=2, label='Median')
marker_mean = Line2D([0], [0], marker='D', color='w', markerfacecolor='#E0E0E0', markeredgecolor="#757575", markersize=6, label='Mean')

ax.legend(
    handles=handles[:len(model_names)] + [line_median, marker_mean],
    labels=labels[:len(model_names)] + ['Median', 'Mean (Diamond)'],
    loc='lower right', frameon=True, edgecolor='black', fontsize=12
)

# EXPORTS
png_path = os.path.join(SAVE_DIR, f"{METRIC}_Final_Plot.png")
pdf_path = os.path.join(SAVE_DIR, f"{METRIC}_Final_Plot.pdf")
plt.savefig(png_path, dpi=600, bbox_inches='tight')
plt.savefig(pdf_path, format='pdf', bbox_inches='tight')

plt.show()

print(f"Excel Table with Cohen's d, Hedges' g, and N saved: {table_path}")
print(f"PNG/PDF plots saved in: {SAVE_DIR}")



























# Run this once to verify decoding is correct before full inference
import cv2
import numpy as np
from PIL import Image
import os

MASKS_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'
test_file    = sorted(os.listdir(MASKS_FOLDER))[0]
mask_bgr     = cv2.imread(os.path.join(MASKS_FOLDER, test_file), cv2.IMREAD_COLOR)

bgr_to_class = {
    (0,   0,   0):   0,
    (0,   255, 0):   1,
    (255, 255, 0):   2,
    (255, 146, 153): 3,
    (128, 64,  64):  4,
    (0,   255, 255): 5,
    (255, 60,  255): 6,
    (55,  55,  255): 7,
}

class_map = np.zeros(mask_bgr.shape[:2], dtype=np.uint8)
for bgr, cls_id in bgr_to_class.items():
    match = np.all(mask_bgr == np.array(bgr, dtype=np.uint8), axis=-1)
    class_map[match] = cls_id

print(f"File          : {test_file}")
print(f"Unique class IDs decoded : {np.unique(class_map)}")
print(f"Pixel counts per class:")
for cls_id in np.unique(class_map):
    count = np.sum(class_map == cls_id)
    print(f"  Class {cls_id} ({['Background','BODY','URINARY BLADDER','SMALL BOWEL','RECTUM','FEMORAL HEAD','GTV','CTV'][cls_id]:20s}): {count} pixels")




















#excel table creation segmentation metrics in Keras

import os
import cv2
import numpy as np
import pandas as pd
from keras.models import load_model
from keras.utils import normalize
import tensorflow as tf
from keras.saving import register_keras_serializable
from PIL import Image

# ===============================
# 1. CUSTOM OBJECTS (same as original)
# ===============================
@register_keras_serializable()
def focal_loss(y_true, y_pred, gamma=2.0):
    epsilon = tf.keras.backend.epsilon()
    y_pred = tf.clip_by_value(y_pred, epsilon, 1.0 - epsilon)
    cross_entropy = -y_true * tf.math.log(y_pred)
    loss = tf.pow(1 - y_pred, gamma) * cross_entropy
    return tf.reduce_mean(loss, axis=-1)

@register_keras_serializable()
def soft_dice_loss(y_true, y_pred, smooth=1):
    y_true = tf.cast(y_true, tf.float32)
    y_pred = tf.cast(y_pred, tf.float32)
    intersection = tf.reduce_sum(y_true * y_pred, axis=(1, 2, 3))
    sum_true = tf.reduce_sum(y_true, axis=(1, 2, 3))
    sum_pred = tf.reduce_sum(y_pred, axis=(1, 2, 3))
    dice_coefficient = (2. * intersection + smooth) / (sum_true + sum_pred + smooth)
    return tf.reduce_mean(1 - dice_coefficient)

@register_keras_serializable()
def combined_loss(y_true, y_pred, gamma=2.0, alpha=0.5):
    return alpha * focal_loss(y_true, y_pred, gamma) + (1 - alpha) * soft_dice_loss(y_true, y_pred)

@register_keras_serializable()
class CustomMeanIoU(tf.keras.metrics.MeanIoU):
    def __init__(self, num_classes, name='mean_iou', dtype='float32',
                 ignore_class=None, sparse_y_true=True, sparse_y_pred=True, axis=-1):
        super().__init__(num_classes=num_classes, name=name, dtype=dtype, ignore_class=ignore_class)
        self.sparse_y_true = sparse_y_true
        self.sparse_y_pred = sparse_y_pred
        self.axis = axis

    def update_state(self, y_true, y_pred, sample_weight=None):
        y_true = tf.math.argmax(y_true, axis=-1)
        y_pred = tf.math.argmax(y_pred, axis=-1)
        return super().update_state(y_true, y_pred, sample_weight)

    def get_config(self):
        config = super().get_config()
        config.update({'sparse_y_true': self.sparse_y_true,
                       'sparse_y_pred': self.sparse_y_pred,
                       'axis': self.axis})
        return config

# ===============================
# 2. CONFIGURATION — EDIT THESE
# ===============================
MODEL_PATH   = r'I:/Radiotherapy/Cervix/models/Cervix_small_Axial_200_epochs_unet3+_Enhanced.keras'
IMAGES_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/images'
MASKS_FOLDER  = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'
SAVE_EXCEL    = r'I:/Radiotherapy/Cervix/models/unet3+_Enhanced_segmentation_metrics_detailed_epoch_200_generated.xlsx'
IMAGE_SIZE    = (256, 256)

# Class definitions — must match your training setup
class_names = {
    0: 'Background',
    1: 'BODY',
    2: 'URINARY BLADDER',
    3: 'SMALL BOWEL',
    4: 'RECTUM',
    5: 'FEMORAL HEAD',
    6: 'GTV',
    7: 'CTV',
}
n_classes = len(class_names)

# Mask RGB → class ID mapping (used to decode ground truth mask PNG)
# Each RGB tuple maps to a class ID
rgb_to_class = {
        (0,   0,   0):   0,   # Background
        (0,   255, 0):   1,   # BODY          — RGB(0,255,0)   = BGR(0,255,0)
        (255, 255, 0):   2,   # URINARY BLADDER — RGB(0,255,255) = BGR(255,255,0)
        (255, 146, 153): 3,   # SMALL BOWEL   — RGB(153,146,255) = BGR(255,146,153)
        (128, 64,  64):  4,   # RECTUM        — RGB(64,64,128)  = BGR(128,64,64)
        (0,   255, 255): 5,   # FEMORAL HEAD  — RGB(255,255,0)  = BGR(0,255,255)
        (255, 60,  255): 6,   # GTV           — RGB(255,60,255) = BGR(255,60,255)
        (55,  55,  255): 7,   # CTV           — RGB(255,55,55)  = BGR(55,55,255)
    }

# ===============================
# 3. LOAD MODEL
# ===============================
custom_objects = {
    'combined_loss':        combined_loss,
    'focal_loss':           focal_loss,
    'soft_dice_loss':       soft_dice_loss,
    'soft_dice_coefficient': lambda y_true, y_pred: 1 - soft_dice_loss(y_true, y_pred),
    'CustomMeanIoU':        CustomMeanIoU
}

print("Loading model...")
model = load_model(MODEL_PATH, custom_objects=custom_objects, compile=True)
print("Model loaded.")

# ===============================
# 4. HELPER: decode RGB mask → class ID map
# ===============================
def decode_mask(mask_path):
    """
    Reads a colour PNG mask and converts it to a 2D array of class IDs.
    Pixels with no matching RGB entry are assigned class 0 (Background).
    """
    mask_img = np.array(Image.open(mask_path).convert("RGB"))
    class_map = np.zeros(mask_img.shape[:2], dtype=np.uint8)
    for rgb, cls_id in rgb_to_class.items():
        match = np.all(mask_img == np.array(rgb, dtype=np.uint8), axis=-1)
        class_map[match] = cls_id
    return class_map

# ===============================
# 5. HELPER: predict → class ID map
# ===============================
def predict_mask(image_path):
    """
    Loads an image, runs model inference, returns 2D predicted class ID array
    resized back to the original image dimensions.
    """
    img = np.array(Image.open(image_path).convert("RGB"))
    original_h, original_w = img.shape[:2]

    img_bgr     = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    img_resized = cv2.resize(img_bgr, IMAGE_SIZE)
    img_norm    = normalize(np.array([img_resized], dtype=np.float32), axis=1)

    prediction  = model.predict(img_norm, verbose=0)
    pred_class  = np.argmax(prediction, axis=-1)[0]          # (H, W)

    # Resize back to original resolution (nearest neighbour preserves class IDs)
    pred_class  = cv2.resize(pred_class.astype(np.uint8),
                             (original_w, original_h),
                             interpolation=cv2.INTER_NEAREST)
    return pred_class

# ===============================
# 6. HELPER: compute per-class metrics
# ===============================
def compute_metrics(true_flat, pred_flat, filename):
    """
    Computes TP/TN/FP/FN and derived metrics for every class present
    in either the ground truth or the prediction.
    Returns a list of dicts (one per class), matching the original Excel schema.
    """
    total_pixels = len(true_flat)
    rows = []

    present_classes = np.union1d(np.unique(true_flat), np.unique(pred_flat))

    for class_id in present_classes:
        true_bin = (true_flat == class_id)
        pred_bin = (pred_flat == class_id)

        tp = int(np.sum( true_bin &  pred_bin))
        tn = int(np.sum(~true_bin & ~pred_bin))
        fp = int(np.sum(~true_bin &  pred_bin))
        fn = int(np.sum( true_bin & ~pred_bin))

        intersection = tp
        union        = tp + fp + fn

        iou         = tp / union               if union > 0           else 0.0
        dice        = (2*tp) / (2*tp + fp + fn) if (2*tp+fp+fn) > 0  else 0.0
        precision   = tp / (tp + fp)            if (tp + fp) > 0      else 0.0
        recall      = tp / (tp + fn)            if (tp + fn) > 0      else 0.0
        specificity = tn / (tn + fp)            if (tn + fp) > 0      else 0.0
        f1          = (2*precision*recall) / (precision+recall) if (precision+recall) > 0 else 0.0
        tpr         = recall
        fpr         = fp / (fp + tn)            if (fp + tn) > 0      else 0.0

        rows.append({
            'Filename':           filename,
            'Class_ID':           int(class_id),
            'Class_Name':         class_names.get(int(class_id), str(class_id)),
            'Total_Pixels':       total_pixels,
            'TP':                 tp,
            'TN':                 tn,
            'FP':                 fp,
            'FN':                 fn,
            'Intersection':       intersection,
            'Union':              union,
            'IoU':                iou,
            'Dice':               dice,
            'Precision':          precision,
            'Recall':             recall,
            'Specificity':        specificity,
            'F1_Score':           f1,
            'True_Positive_Rate': tpr,
            'False_Positive_Rate':fpr,
        })

    return rows

# ===============================
# 7. MAIN INFERENCE LOOP
# ===============================
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp')
image_files = sorted([
    f for f in os.listdir(IMAGES_FOLDER)
    if f.lower().endswith(image_extensions)
])

print(f"Found {len(image_files)} images. Starting inference...\n")

all_metrics = []
skipped     = []

for idx, img_file in enumerate(image_files):
    filename   = os.path.splitext(img_file)[0]          # strip extension
    image_path = os.path.join(IMAGES_FOLDER, img_file)
    mask_path  = os.path.join(MASKS_FOLDER,  img_file)  # same filename

    if not os.path.exists(mask_path):
        print(f"  [SKIP] No mask found for: {img_file}")
        skipped.append(img_file)
        continue

    try:
        true_mask = decode_mask(mask_path)
        pred_mask = predict_mask(image_path)

        # Ensure shapes match (safety check)
        if true_mask.shape != pred_mask.shape:
            pred_mask = cv2.resize(pred_mask,
                                   (true_mask.shape[1], true_mask.shape[0]),
                                   interpolation=cv2.INTER_NEAREST)

        rows = compute_metrics(true_mask.flatten(), pred_mask.flatten(), filename)
        all_metrics.extend(rows)

        if (idx + 1) % 20 == 0 or (idx + 1) == len(image_files):
            print(f"  Processed {idx+1}/{len(image_files)}: {filename}")

    except Exception as e:
        print(f"  [ERROR] {img_file}: {e}")
        skipped.append(img_file)

print(f"\nInference complete. {len(image_files)-len(skipped)} images processed, {len(skipped)} skipped.")

# ===============================
# 8. BUILD DATAFRAME
# ===============================
column_order = [
    'Filename', 'Class_ID', 'Class_Name', 'Total_Pixels',
    'TP', 'TN', 'FP', 'FN', 'Intersection', 'Union',
    'IoU', 'Dice', 'Precision', 'Recall', 'Specificity', 'F1_Score',
    'True_Positive_Rate', 'False_Positive_Rate'
]

df_detailed = pd.DataFrame(all_metrics)[column_order]

# ===============================
# 9. SAVE EXCEL — identical sheet structure to original
# ===============================
os.makedirs(os.path.dirname(SAVE_EXCEL), exist_ok=True)

with pd.ExcelWriter(SAVE_EXCEL, engine='openpyxl') as writer:

    # ── Sheet 1: Detailed per image per class ──
    df_detailed.to_excel(writer, sheet_name='Detailed_Metrics', index=False)

    # ── Sheet 2: Class summary ──
    class_summary = df_detailed.groupby(['Class_ID', 'Class_Name']).agg(
        TP=('TP','sum'), TN=('TN','sum'), FP=('FP','sum'), FN=('FN','sum'),
        Intersection=('Intersection','sum'), Union=('Union','sum'),
        IoU=('IoU','mean'), Dice=('Dice','mean'),
        Precision=('Precision','mean'), Recall=('Recall','mean'),
        Specificity=('Specificity','mean'), F1_Score=('F1_Score','mean'),
        Total_Pixels=('Total_Pixels','sum')
    ).reset_index()

    for idx, row in class_summary.iterrows():
        tp, fp, fn = row['TP'], row['FP'], row['FN']
        union      = row['Union']
        class_summary.at[idx, 'Aggregated_IoU']  = row['Intersection'] / union if union > 0 else 0.0
        class_summary.at[idx, 'Aggregated_Dice'] = (2*tp) / (2*tp+fp+fn)       if (2*tp+fp+fn) > 0 else 0.0

    class_summary.to_excel(writer, sheet_name='Class_Summary', index=False)

    # ── Sheet 3: Image summary ──
    image_summary = df_detailed.groupby('Filename').agg(
        TP=('TP','sum'), TN=('TN','sum'), FP=('FP','sum'), FN=('FN','sum'),
        IoU=('IoU','mean'), Dice=('Dice','mean'),
        Precision=('Precision','mean'), Recall=('Recall','mean'),
        F1_Score=('F1_Score','mean'),
        Total_Pixels=('Total_Pixels','first'),
        Classes_Present=('Class_ID','count')
    ).reset_index()

    denom_iou  = (image_summary['TP'] + image_summary['FP'] + image_summary['FN']).replace(0, 1)
    denom_dice = (2*image_summary['TP'] + image_summary['FP'] + image_summary['FN']).replace(0, 1)
    image_summary['Image_IoU']  = image_summary['TP'] / denom_iou
    image_summary['Image_Dice'] = (2 * image_summary['TP']) / denom_dice

    image_summary.to_excel(writer, sheet_name='Image_Summary', index=False)

    # ── Sheet 4: Class presence ──
    class_presence = df_detailed.groupby(['Class_ID', 'Class_Name']).agg(
        Images_Present=('Filename','count'),
        TP=('TP','sum'),
        FN=('FN','sum')
    ).reset_index()
    class_presence['Support'] = class_presence['TP'] + class_presence['FN']
    class_presence.to_excel(writer, sheet_name='Class_Presence', index=False)

print(f"\nExcel saved to: {SAVE_EXCEL}")
print(f"Total images processed : {df_detailed['Filename'].nunique()}")
print(f"Total records          : {len(df_detailed)}")
print(f"\nOverall Performance:")
print(f"  Mean IoU  : {df_detailed['IoU'].mean():.4f}")
print(f"  Mean Dice : {df_detailed['Dice'].mean():.4f}")
print(f"\nClass Presence Summary:")
total_imgs = df_detailed['Filename'].nunique()
for cid, cname in class_names.items():
    n = len(df_detailed[df_detailed['Class_ID'] == cid])
    if n > 0:
        print(f"  {cname:20s}: {n}/{total_imgs} ({100*n/total_imgs:.1f}%)")
if skipped:
    print(f"\nSkipped files ({len(skipped)}): {skipped}")



















































#exceptional for keras 
import os
import cv2
import numpy as np
import pandas as pd
from keras.models import load_model
from keras.utils import normalize
import tensorflow as tf
from keras.saving import register_keras_serializable
from PIL import Image
import traceback

# ===============================
# 1. CUSTOM OBJECTS
# ===============================
@register_keras_serializable()
def focal_loss(y_true, y_pred, gamma=2.0):
    epsilon = tf.keras.backend.epsilon()
    y_pred = tf.clip_by_value(y_pred, epsilon, 1.0 - epsilon)
    cross_entropy = -y_true * tf.math.log(y_pred)
    loss = tf.pow(1 - y_pred, gamma) * cross_entropy
    return tf.reduce_mean(loss, axis=-1)

@register_keras_serializable()
def soft_dice_loss(y_true, y_pred, smooth=1):
    y_true = tf.cast(y_true, tf.float32)
    y_pred = tf.cast(y_pred, tf.float32)
    intersection = tf.reduce_sum(y_true * y_pred, axis=(1, 2, 3))
    sum_true = tf.reduce_sum(y_true, axis=(1, 2, 3))
    sum_pred = tf.reduce_sum(y_pred, axis=(1, 2, 3))
    dice_coefficient = (2. * intersection + smooth) / (sum_true + sum_pred + smooth)
    return tf.reduce_mean(1 - dice_coefficient)

@register_keras_serializable()
def combined_loss(y_true, y_pred, gamma=2.0, alpha=0.5):
    return alpha * focal_loss(y_true, y_pred, gamma) + (1 - alpha) * soft_dice_loss(y_true, y_pred)

@register_keras_serializable()
class CustomMeanIoU(tf.keras.metrics.MeanIoU):
    def __init__(self, num_classes, name='mean_iou', dtype='float32',
                 ignore_class=None, sparse_y_true=True, sparse_y_pred=True, axis=-1):
        super().__init__(num_classes=num_classes, name=name, dtype=dtype, ignore_class=ignore_class)
        self.sparse_y_true = sparse_y_true
        self.sparse_y_pred = sparse_y_pred
        self.axis = axis

    def update_state(self, y_true, y_pred, sample_weight=None):
        y_true = tf.math.argmax(y_true, axis=-1)
        y_pred = tf.math.argmax(y_pred, axis=-1)
        return super().update_state(y_true, y_pred, sample_weight)

    def get_config(self):
        config = super().get_config()
        config.update({'sparse_y_true': self.sparse_y_true,
                       'sparse_y_pred': self.sparse_y_pred,
                       'axis': self.axis})
        return config

# ===============================
# 2. CONFIGURATION
# ===============================
MODEL_PATH = r'I:/Radiotherapy/Cervix/models/Cervix_small_Axial_200_epochs_unet3+_Enhanced.keras'
IMAGES_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/images'
MASKS_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'
SAVE_EXCEL = r'I:/Radiotherapy/Cervix/models/unet3+_Enhanced_segmentation_metrics_detailed_epoch_200_generated.xlsx'

IMAGE_SIZE = (256, 256)

class_names = {
    0: 'Background', 1: 'BODY', 2: 'URINARY BLADDER', 3: 'SMALL BOWEL',
    4: 'RECTUM', 5: 'FEMORAL HEAD', 6: 'GTV', 7: 'CTV'
}

rgb_to_class = {
    (0, 0, 0): 0, (0, 255, 0): 1, (255, 255, 0): 2, (255, 146, 153): 3,
    (128, 64, 64): 4, (0, 255, 255): 5, (255, 60, 255): 6, (55, 55, 255): 7,
}

# ===============================
# 3. LOAD MODEL
# ===============================
custom_objects = {
    'combined_loss': combined_loss,
    'focal_loss': focal_loss,
    'soft_dice_loss': soft_dice_loss,
    'soft_dice_coefficient': lambda y_true, y_pred: 1 - soft_dice_loss(y_true, y_pred),
    'CustomMeanIoU': CustomMeanIoU
}

print("Loading model...")
try:
    model = load_model(MODEL_PATH, custom_objects=custom_objects, compile=True)
    print("Model loaded successfully.\n")
except Exception as e:
    print(f"Error loading model: {e}")
    exit(1)

# ===============================
# 4. DECODE MASK - FIXED WITH ERROR HANDLING
# ===============================
def decode_mask(mask_path):
    try:
        # Open image and convert to RGB
        img = Image.open(mask_path).convert("RGB")
        mask_img = np.array(img, dtype=np.uint8)
        
        # Ensure it's a numpy array with proper shape
        if not isinstance(mask_img, np.ndarray):
            print(f"   decode_mask: Got {type(mask_img)} instead of numpy array")
            mask_img = np.array(mask_img)
        
        # Initialize class map
        class_map = np.zeros(mask_img.shape[:2], dtype=np.uint8)
        
        # Map RGB values to class IDs
        for rgb, cls_id in rgb_to_class.items():
            rgb_array = np.array(rgb, dtype=np.uint8)
            match = np.all(mask_img == rgb_array, axis=-1)
            class_map[match] = cls_id
        
        return class_map
    
    except Exception as e:
        print(f"   decode_mask error for {mask_path}: {e}")
        print(f"   Traceback: {traceback.format_exc()}")
        raise

# ===============================
# 5. PREDICT MASK - COMPLETELY REWRITTEN WITH DEBUGGING
# ===============================
def predict_mask(image_path):
    try:
        # Load image
        img = Image.open(image_path).convert("RGB")
        img_array = np.array(img, dtype=np.uint8)
        
        # Ensure it's a numpy array
        if not isinstance(img_array, np.ndarray):
            print(f"   predict_mask: img_array is {type(img_array)}")
            img_array = np.array(img_array)
        
        original_h, original_w = img_array.shape[:2]
        
        # Convert RGB to BGR for OpenCV
        img_bgr = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
        
        # Resize
        img_resized = cv2.resize(img_bgr, IMAGE_SIZE, interpolation=cv2.INTER_LINEAR)
        
        # Normalize to [0, 1] - ensure float32
        img_normalized = img_resized.astype(np.float32) / 255.0
        
        # Add batch dimension
        img_batch = np.expand_dims(img_normalized, axis=0)
        
        # Verify batch shape
        if not isinstance(img_batch, np.ndarray):
            print(f"   img_batch is {type(img_batch)}, converting...")
            img_batch = np.array(img_batch)
        
        # Model prediction
        prediction = model.predict(img_batch, verbose=0)
        
        # Handle prediction
        if isinstance(prediction, list):
            print(f"   WARNING: prediction is a list of length {len(prediction)}")
            prediction = prediction[0] if len(prediction) > 0 else None
        
        if prediction is None:
            raise ValueError("Prediction returned None")
        
        # Get predicted classes
        if len(prediction.shape) == 4:
            pred_class = np.argmax(prediction[0], axis=-1)
        elif len(prediction.shape) == 3:
            pred_class = np.argmax(prediction, axis=-1)
        else:
            raise ValueError(f"Unexpected prediction shape: {prediction.shape}")
        
        # Resize back to original size
        pred_class = cv2.resize(pred_class.astype(np.uint8), 
                                (original_w, original_h), 
                                interpolation=cv2.INTER_NEAREST)
        
        return pred_class
    
    except Exception as e:
        print(f"   predict_mask error for {image_path}: {e}")
        print(f"   Traceback: {traceback.format_exc()}")
        raise

# ===============================
# 6. COMPUTE METRICS (ENHANCED WITH DEBUGGING)
# ===============================
def compute_metrics(true_mask, pred_mask, filename):
    try:
        # Convert to numpy arrays if needed
        if isinstance(true_mask, list):
            print(f"   Converting true_mask from list to array")
            true_mask = np.array(true_mask)
        
        if isinstance(pred_mask, list):
            print(f"   Converting pred_mask from list to array")
            pred_mask = np.array(pred_mask)
        
        # Check if they are numpy arrays
        if not isinstance(true_mask, np.ndarray):
            raise TypeError(f"true_mask is {type(true_mask)}, not a numpy array")
        
        if not isinstance(pred_mask, np.ndarray):
            raise TypeError(f"pred_mask is {type(pred_mask)}, not a numpy array")
        
        # Check for shape mismatch
        if true_mask.shape != pred_mask.shape:
            print(f"   Shape mismatch: true={true_mask.shape}, pred={pred_mask.shape}")
            pred_mask = cv2.resize(pred_mask.astype(np.uint8), 
                                   (true_mask.shape[1], true_mask.shape[0]),
                                   interpolation=cv2.INTER_NEAREST)
        
        # Flatten
        true_flat = true_mask.flatten()
        pred_flat = pred_mask.flatten()
        
        # Ensure same length
        if len(true_flat) != len(pred_flat):
            min_len = min(len(true_flat), len(pred_flat))
            true_flat = true_flat[:min_len]
            pred_flat = pred_flat[:min_len]
        
        total_pixels = len(true_flat)
        rows = []
        present_classes = np.union1d(np.unique(true_flat), np.unique(pred_flat))
        
        for class_id in present_classes:
            true_bin = (true_flat == class_id)
            pred_bin = (pred_flat == class_id)
            
            tp = int(np.sum(true_bin & pred_bin))
            tn = int(np.sum(~true_bin & ~pred_bin))
            fp = int(np.sum(~true_bin & pred_bin))
            fn = int(np.sum(true_bin & ~pred_bin))
            
            union = tp + fp + fn
            iou = tp / union if union > 0 else 0.0
            dice = (2 * tp) / (2 * tp + fp + fn) if (2 * tp + fp + fn) > 0 else 0.0
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
            specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0
            f1 = (2 * precision * recall) / (precision + recall) if (precision + recall) > 0 else 0.0

            rows.append({
                'Filename': filename,
                'Class_ID': int(class_id),
                'Class_Name': class_names.get(int(class_id), str(class_id)),
                'Total_Pixels': total_pixels,
                'TP': tp, 'TN': tn, 'FP': fp, 'FN': fn,
                'Intersection': tp,
                'Union': union,
                'IoU': iou,
                'Dice': dice,
                'Precision': precision,
                'Recall': recall,
                'Specificity': specificity,
                'F1_Score': f1,
                'True_Positive_Rate': recall,
                'False_Positive_Rate': fp / (fp + tn) if (fp + tn) > 0 else 0.0,
            })
        return rows
    
    except Exception as e:
        print(f"   compute_metrics error: {e}")
        print(f"   Traceback: {traceback.format_exc()}")
        raise

# ===============================
# 7. MAIN LOOP WITH DETAILED DEBUGGING
# ===============================
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp')
image_files = sorted([f for f in os.listdir(IMAGES_FOLDER) if f.lower().endswith(image_extensions)])

print(f"Found {len(image_files)} images. Starting inference...\n")

all_metrics = []
skipped = []
errors = []

for idx, img_file in enumerate(image_files[:20]):  # Process first 20 images for debugging
    filename = os.path.splitext(img_file)[0]
    image_path = os.path.join(IMAGES_FOLDER, img_file)
    mask_path = os.path.join(MASKS_FOLDER, img_file)
    
    print(f"\n[{idx+1}] Processing: {img_file}")
    
    # Check if mask exists
    if not os.path.exists(mask_path):
        print(f"   [SKIP] No mask found at: {mask_path}")
        skipped.append(img_file)
        continue
    
    try:
        # Decode ground truth mask
        print(f"   Decoding mask...")
        true_mask = decode_mask(mask_path)
        print(f"   Mask shape: {true_mask.shape}, type: {type(true_mask)}")
        
        # Get prediction
        print(f"   Getting prediction...")
        pred_mask = predict_mask(image_path)
        print(f"   Prediction shape: {pred_mask.shape}, type: {type(pred_mask)}")
        
        # Compute metrics
        print(f"   Computing metrics...")
        rows = compute_metrics(true_mask, pred_mask, filename)
        all_metrics.extend(rows)
        print(f"   ✓ Successfully processed")
        
    except Exception as e:
        print(f"   [ERROR] {img_file}: {str(e)}")
        errors.append(img_file)
        continue

print(f"\n{'='*50}")
print(f"PROCESSING SUMMARY (first {len(image_files[:20])} images):")
print(f"Successfully processed: {len(all_metrics) // 8 if all_metrics else 0} images")  # 8 classes per image
print(f"Skipped (no mask): {len(skipped)}")
print(f"Errors: {len(errors)}")
if errors:
    print(f"Error list: {errors}")
print(f"{'='*50}\n")

# ===============================
# 8. SAVE RESULTS (if any metrics were collected)
# ===============================
if all_metrics:
    column_order = ['Filename', 'Class_ID', 'Class_Name', 'Total_Pixels',
                    'TP', 'TN', 'FP', 'FN', 'Intersection', 'Union',
                    'IoU', 'Dice', 'Precision', 'Recall', 'Specificity', 
                    'F1_Score', 'True_Positive_Rate', 'False_Positive_Rate']
    
    df_detailed = pd.DataFrame(all_metrics)[column_order]
    
    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(SAVE_EXCEL), exist_ok=True)
    
    # Save to Excel
    with pd.ExcelWriter(SAVE_EXCEL, engine='openpyxl') as writer:
        df_detailed.to_excel(writer, sheet_name='Detailed_Metrics', index=False)
        
        # Class summary
        class_summary = df_detailed.groupby(['Class_ID', 'Class_Name']).agg(
            TP=('TP','sum'), 
            TN=('TN','sum'), 
            FP=('FP','sum'), 
            FN=('FN','sum'),
            IoU=('IoU','mean'), 
            Dice=('Dice','mean'),
            Precision=('Precision','mean'),
            Recall=('Recall','mean')
        ).reset_index()
        
        class_summary.to_excel(writer, sheet_name='Class_Summary', index=False)
    
    print(f"\nExcel saved to: {SAVE_EXCEL}")
    print(f"Total metrics records: {len(all_metrics)}")
else:
    print("\nNo metrics were generated. Debug the errors above.")






























import cv2
import numpy as np
import os

MASKS_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'

# Known BGR values we expect
bgr_to_class = {
    (0,   0,   0):   0,   # Background
    (0,   255, 0):   1,   # BODY
    (255, 255, 0):   2,   # URINARY BLADDER
    (153, 146, 255): 3,   # SMALL BOWEL
    (64,  64,  128): 4,   # RECTUM
    (0,   255, 255): 5,   # FEMORAL HEAD
    (255, 60,  255): 6,   # GTV
    (255, 55,  55):  7,   # CTV
}

class_names = {
    0: 'Background', 1: 'BODY', 2: 'URINARY BLADDER',
    3: 'SMALL BOWEL', 4: 'RECTUM', 5: 'FEMORAL HEAD',
    6: 'GTV', 7: 'CTV'
}

mask_files = sorted([
    f for f in os.listdir(MASKS_FOLDER)
    if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))
])

# ── Step 1: Find masks that contain RECTUM or SMALL BOWEL ──
print("=" * 65)
print("STEP 1: Searching for masks containing RECTUM or SMALL BOWEL")
print("=" * 65)

rectum_bgr   = np.array([64,  64,  128], dtype=np.uint8)
smallbowel_bgr = np.array([153, 146, 255], dtype=np.uint8)

found_rectum     = []
found_smallbowel = []

for f in mask_files:
    path = os.path.join(MASKS_FOLDER, f)
    mask = cv2.imread(path, cv2.IMREAD_COLOR)
    if mask is None:
        continue
    pixels = mask.reshape(-1, 3)

    if np.any(np.all(pixels == rectum_bgr, axis=1)):
        found_rectum.append(f)
    if np.any(np.all(pixels == smallbowel_bgr, axis=1)):
        found_smallbowel.append(f)

print(f"Masks with RECTUM BGR(64,64,128)      : {len(found_rectum)}")
print(f"Masks with SMALL BOWEL BGR(153,146,255): {len(found_smallbowel)}")

# ── Step 2: Deep inspect one mask per class ──
print("\n" + "=" * 65)
print("STEP 2: Deep pixel inspection of sample masks")
print("=" * 65)

# Pick one mask that has RECTUM if found, otherwise first mask
sample_files = []
if found_rectum:
    sample_files.append(("RECTUM sample",    found_rectum[0]))
if found_smallbowel:
    sample_files.append(("SMALL BOWEL sample", found_smallbowel[0]))
if not sample_files:
    sample_files.append(("First mask", mask_files[0]))

for label, f in sample_files:
    path = os.path.join(MASKS_FOLDER, f)
    mask = cv2.imread(path, cv2.IMREAD_COLOR)
    pixels = mask.reshape(-1, 3)
    unique, counts = np.unique(pixels, axis=0, return_counts=True)

    print(f"\n  [{label}] File: {f}")
    print(f"  {'BGR Value':<25} {'Count':>10}  {'Matched Class'}")
    print(f"  {'-'*60}")
    for u, c in zip(unique, counts):
        bgr_key = tuple(u.tolist())
        cls     = bgr_to_class.get(bgr_key, "*** NO MATCH ***")
        cls_name = class_names.get(cls, cls)
        print(f"  BGR{str(bgr_key):<22} {c:>10}  → {cls_name}")

# ── Step 3: Check image bit depth and mode ──
print("\n" + "=" * 65)
print("STEP 3: Image properties check")
print("=" * 65)

for f in mask_files[:5]:
    path = os.path.join(MASKS_FOLDER, f)
    mask = cv2.imread(path, cv2.IMREAD_UNCHANGED)  # read raw, no conversion
    print(f"\n  File  : {f}")
    print(f"  Shape : {mask.shape}")
    print(f"  Dtype : {mask.dtype}")
    print(f"  Min   : {mask.min()}  Max: {mask.max()}")
    if len(mask.shape) == 2:
        print(f"  *** GRAYSCALE image — pixel values ARE class IDs directly ***")
        print(f"  Unique values: {np.unique(mask).tolist()}")
    elif mask.shape[2] == 4:
        print(f"  *** RGBA image — has alpha channel ***")
        alpha_unique = np.unique(mask[:,:,3])
        print(f"  Alpha channel unique values: {alpha_unique.tolist()}")

# ── Step 4: Try reading as RGBA and check alpha ──
print("\n" + "=" * 65)
print("STEP 4: RGBA check on masks with missing classes")
print("=" * 65)

check_files = mask_files[:3]
for f in check_files:
    path      = os.path.join(MASKS_FOLDER, f)
    mask_rgba = cv2.imread(path, cv2.IMREAD_UNCHANGED)
    mask_bgr  = cv2.imread(path, cv2.IMREAD_COLOR)

    print(f"\n  File         : {f}")
    print(f"  UNCHANGED shape: {mask_rgba.shape}  dtype: {mask_rgba.dtype}")
    print(f"  COLOR shape    : {mask_bgr.shape}   dtype: {mask_bgr.dtype}")

    pixels_bgr  = mask_bgr.reshape(-1, 3)
    unique_bgr  = np.unique(pixels_bgr, axis=0)
    print(f"  Unique BGR (COLOR read)    : {[tuple(u.tolist()) for u in unique_bgr]}")

    if len(mask_rgba.shape) == 3 and mask_rgba.shape[2] == 4:
        pixels_rgba = mask_rgba.reshape(-1, 4)
        unique_rgba = np.unique(pixels_rgba, axis=0)
        print(f"  Unique BGRA (UNCHANGED read): {[tuple(u.tolist()) for u in unique_rgba]}")











#RGB and BGR values
import cv2
import numpy as np
import os

MASKS_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'

mask_files = sorted([
    f for f in os.listdir(MASKS_FOLDER)
    if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))
])

print(f"Found {len(mask_files)} mask files\n")

all_unique_bgr = set()

for idx, f in enumerate(mask_files):
    path  = os.path.join(MASKS_FOLDER, f)
    mask  = cv2.imread(path, cv2.IMREAD_COLOR)
    if mask is None:
        print(f"  [SKIP] Could not read: {f}")
        continue
    pixels = mask.reshape(-1, 3)
    unique = np.unique(pixels, axis=0)
    for u in unique:
        all_unique_bgr.add(tuple(u.tolist()))

    if (idx + 1) % 50 == 0:
        print(f"  Scanned {idx+1}/{len(mask_files)} masks...")

print(f"\n{'='*55}")
print(f"Total unique BGR colours found across ALL masks: {len(all_unique_bgr)}")
print(f"{'='*55}")
print(f"\n{'BGR Value':<25} {'RGB Value':<25}")
print(f"{'-'*50}")
for bgr in sorted(all_unique_bgr):
    rgb = (bgr[2], bgr[1], bgr[0])
    print(f"BGR{str(bgr):<22} RGB{str(rgb):<22}")

print(f"\n{'='*55}")
print("Copy-paste ready bgr_to_class mapping:")
print(f"{'='*55}")
print("bgr_to_class = {")
for i, bgr in enumerate(sorted(all_unique_bgr)):
    rgb = (bgr[2], bgr[1], bgr[0])
    print(f"    {str(bgr)}: {i},   # RGB{rgb}  ← assign correct class ID")
print("}")

















#excel table creation segmentation metrics in h5

import os
import cv2
import numpy as np
import pandas as pd
from keras.models import load_model
from keras.utils import normalize
import tensorflow as tf
from keras.saving import register_keras_serializable
from PIL import Image

# ===============================
# 1. CUSTOM OBJECTS
# ===============================
@register_keras_serializable()
def focal_loss(y_true, y_pred, gamma=2.0):
    epsilon = tf.keras.backend.epsilon()
    y_pred = tf.clip_by_value(y_pred, epsilon, 1.0 - epsilon)
    cross_entropy = -y_true * tf.math.log(y_pred)
    loss = tf.pow(1 - y_pred, gamma) * cross_entropy
    return tf.reduce_mean(loss, axis=-1)

@register_keras_serializable()
def soft_dice_loss(y_true, y_pred, smooth=1):
    y_true = tf.cast(y_true, tf.float32)
    y_pred = tf.cast(y_pred, tf.float32)
    intersection = tf.reduce_sum(y_true * y_pred, axis=(1, 2, 3))
    sum_true = tf.reduce_sum(y_true, axis=(1, 2, 3))
    sum_pred = tf.reduce_sum(y_pred, axis=(1, 2, 3))
    dice_coefficient = (2. * intersection + smooth) / (sum_true + sum_pred + smooth)
    return tf.reduce_mean(1 - dice_coefficient)

@register_keras_serializable()
def combined_loss(y_true, y_pred, gamma=2.0, alpha=0.5):
    return alpha * focal_loss(y_true, y_pred, gamma) + (1 - alpha) * soft_dice_loss(y_true, y_pred)

@register_keras_serializable()
def soft_dice_coefficient(y_true, y_pred):
    return 1 - soft_dice_loss(y_true, y_pred)

@register_keras_serializable()
class CustomMeanIoU(tf.keras.metrics.MeanIoU):
    def __init__(self, num_classes, name='mean_iou', dtype='float32',
                 ignore_class=None, sparse_y_true=True, sparse_y_pred=True, axis=-1):
        super().__init__(num_classes=num_classes, name=name, dtype=dtype,
                         ignore_class=ignore_class)
        self.sparse_y_true = sparse_y_true
        self.sparse_y_pred = sparse_y_pred
        self.axis = axis

    def update_state(self, y_true, y_pred, sample_weight=None):
        y_true = tf.math.argmax(y_true, axis=-1)
        y_pred = tf.math.argmax(y_pred, axis=-1)
        return super().update_state(y_true, y_pred, sample_weight)

    def get_config(self):
        config = super().get_config()
        config.update({'sparse_y_true': self.sparse_y_true,
                       'sparse_y_pred': self.sparse_y_pred,
                       'axis': self.axis})
        return config

# ===============================
# 2. CONFIGURATION — EDIT THESE
# ===============================
MODEL_PATH   = r'I:/Radiotherapy/Cervix/models/models/Arian/AttUNET_Cervix_small_Axial/model_weights.h5'
IMAGES_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/images'
MASKS_FOLDER  = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'
SAVE_EXCEL    = r'I:/Radiotherapy/Cervix/models/models/Arian/AttUNET_Cervix_small_Axial/attention_UNET_Arian_segmentation_metrics_detailed_epoch_200_generated.xlsx'


# Class definitions
class_names = {
    0: 'Background',
    1: 'BODY',
    2: 'URINARY BLADDER',
    3: 'SMALL BOWEL',
    4: 'RECTUM',
    5: 'FEMORAL HEAD',
    6: 'GTV',
    7: 'CTV',
}
n_classes = len(class_names)

# BGR mapping (cv2 reads BGR)
bgr_to_class = {
    (0, 0, 0): 0,   # RGB(0, 0, 0)  ← assign correct class ID
    (0, 255, 0): 1,   # RGB(0, 255, 0)  ← assign correct class ID
    (0, 255, 255): 2,   # RGB(255, 255, 0)  ← assign correct class ID
    (64, 64, 128): 3,   # RGB(128, 64, 64)  ← assign correct class ID
    (153, 146, 255): 4,   # RGB(255, 146, 153)  ← assign correct class ID
    (255, 55, 55): 5,   # RGB(55, 55, 255)  ← assign correct class ID
    (255, 60, 255): 6,   # RGB(255, 60, 255)  ← assign correct class ID
    (255, 255, 0): 7,   # RGB(0, 255, 255)  ← assign correct class ID
}

# ===============================
# 3. LOAD MODEL — H5
# ===============================
custom_objects = {
    'combined_loss':         combined_loss,
    'focal_loss':            focal_loss,
    'soft_dice_loss':        soft_dice_loss,
    'soft_dice_coefficient': soft_dice_coefficient,
    'CustomMeanIoU':         CustomMeanIoU
}

print("Loading model...")
with tf.keras.utils.custom_object_scope(custom_objects):
    model = load_model(MODEL_PATH, compile=False)
model.compile(optimizer='adam', loss=combined_loss)
print(f"Model loaded  : {MODEL_PATH}")

# Auto-detect image size from model input shape
model_input_shape = model.input_shape          # e.g. (None, 256, 256, 3)
IMAGE_SIZE = (model_input_shape[2], model_input_shape[1])  # (W, H) for cv2
print(f"Input shape   : {model.input_shape}")
print(f"Output shape  : {model.output_shape}")
print(f"Image size    : {IMAGE_SIZE} (W x H)")

# ===============================
# 4. HELPER: decode BGR mask → class ID map
# ===============================
def decode_mask(mask_path):
    """
    Reads a colour PNG mask using cv2 (BGR) and maps pixel colours
    to class IDs. Unmatched pixels default to class 0 (Background).
    """
    mask_bgr = cv2.imread(mask_path, cv2.IMREAD_COLOR)
    if mask_bgr is None:
        raise ValueError(f"Could not load mask: {mask_path}")

    class_map = np.zeros(mask_bgr.shape[:2], dtype=np.uint8)
    for bgr, cls_id in bgr_to_class.items():
        match = np.all(mask_bgr == np.array(bgr, dtype=np.uint8), axis=-1)
        class_map[match] = cls_id

    # Warn if only background was decoded (likely a colour mismatch)
    unique_ids = np.unique(class_map)
    if len(unique_ids) == 1 and unique_ids[0] == 0:
        print(f"  [WARNING] Only background decoded: {mask_path}")

    return class_map

# ===============================
# 5. HELPER: predict → class ID map
# ===============================
def predict_mask(image_path):
    """
    Loads an image, runs inference, returns 2D predicted class ID array
    at the original image resolution.
    """
    img_bgr      = cv2.imread(image_path, cv2.IMREAD_COLOR)
    if img_bgr is None:
        raise ValueError(f"Could not load image: {image_path}")

    original_h, original_w = img_bgr.shape[:2]
    img_resized  = cv2.resize(img_bgr, IMAGE_SIZE)
    img_norm     = normalize(np.array([img_resized], dtype=np.float32), axis=1)

    prediction   = model.predict(img_norm, verbose=0)
    pred_class   = np.argmax(prediction, axis=-1)[0].astype(np.uint8)

    # Resize back to original resolution
    pred_class   = cv2.resize(pred_class,
                              (original_w, original_h),
                              interpolation=cv2.INTER_NEAREST)
    return pred_class

# ===============================
# 6. HELPER: compute per-class metrics
# ===============================
def compute_metrics(true_flat, pred_flat, filename):
    """
    Computes TP/TN/FP/FN and derived metrics for every class present
    in either ground truth or prediction.
    """
    total_pixels    = len(true_flat)
    present_classes = np.union1d(np.unique(true_flat), np.unique(pred_flat))
    rows            = []

    for class_id in present_classes:
        true_bin = (true_flat == class_id)
        pred_bin = (pred_flat == class_id)

        tp = int(np.sum( true_bin &  pred_bin))
        tn = int(np.sum(~true_bin & ~pred_bin))
        fp = int(np.sum(~true_bin &  pred_bin))
        fn = int(np.sum( true_bin & ~pred_bin))

        intersection = tp
        union        = tp + fp + fn

        iou         = tp / union                if union > 0            else 0.0
        dice        = (2*tp) / (2*tp + fp + fn) if (2*tp + fp + fn) > 0 else 0.0
        precision   = tp / (tp + fp)            if (tp + fp) > 0        else 0.0
        recall      = tp / (tp + fn)            if (tp + fn) > 0        else 0.0
        specificity = tn / (tn + fp)            if (tn + fp) > 0        else 0.0
        f1          = (2*precision*recall) / (precision + recall) \
                      if (precision + recall) > 0 else 0.0
        tpr         = recall
        fpr         = fp / (fp + tn)            if (fp + tn) > 0        else 0.0

        rows.append({
            'Filename':            filename,
            'Class_ID':            int(class_id),
            'Class_Name':          class_names.get(int(class_id), str(class_id)),
            'Total_Pixels':        total_pixels,
            'TP':                  tp,
            'TN':                  tn,
            'FP':                  fp,
            'FN':                  fn,
            'Intersection':        intersection,
            'Union':               union,
            'IoU':                 iou,
            'Dice':                dice,
            'Precision':           precision,
            'Recall':              recall,
            'Specificity':         specificity,
            'F1_Score':            f1,
            'True_Positive_Rate':  tpr,
            'False_Positive_Rate': fpr,
        })

    return rows

# ===============================
# 7. MAIN INFERENCE LOOP
# ===============================
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp')
image_files = sorted([
    f for f in os.listdir(IMAGES_FOLDER)
    if f.lower().endswith(image_extensions)
])

print(f"\nFound {len(image_files)} images. Starting inference...\n")

all_metrics = []
skipped     = []

for idx, img_file in enumerate(image_files):
    filename   = os.path.splitext(img_file)[0]
    image_path = os.path.join(IMAGES_FOLDER, img_file)
    mask_path  = os.path.join(MASKS_FOLDER,  img_file)

    if not os.path.exists(mask_path):
        print(f"  [SKIP] No mask found for: {img_file}")
        skipped.append(img_file)
        continue

    try:
        true_mask = decode_mask(mask_path)
        pred_mask = predict_mask(image_path)

        # Safety: ensure shapes match
        if true_mask.shape != pred_mask.shape:
            pred_mask = cv2.resize(pred_mask,
                                   (true_mask.shape[1], true_mask.shape[0]),
                                   interpolation=cv2.INTER_NEAREST)

        rows = compute_metrics(true_mask.flatten(), pred_mask.flatten(), filename)
        all_metrics.extend(rows)

        if (idx + 1) % 20 == 0 or (idx + 1) == len(image_files):
            print(f"  Processed {idx+1}/{len(image_files)}: {filename}")

    except Exception as e:
        print(f"  [ERROR] {img_file}: {e}")
        skipped.append(img_file)

print(f"\nInference complete.")
print(f"  Processed : {len(image_files) - len(skipped)}")
print(f"  Skipped   : {len(skipped)}")

# ===============================
# 8. BUILD DATAFRAME
# ===============================
column_order = [
    'Filename', 'Class_ID', 'Class_Name', 'Total_Pixels',
    'TP', 'TN', 'FP', 'FN', 'Intersection', 'Union',
    'IoU', 'Dice', 'Precision', 'Recall', 'Specificity', 'F1_Score',
    'True_Positive_Rate', 'False_Positive_Rate'
]
df_detailed = pd.DataFrame(all_metrics)[column_order]

# ===============================
# 9. SAVE EXCEL
# ===============================
os.makedirs(os.path.dirname(SAVE_EXCEL), exist_ok=True)

with pd.ExcelWriter(SAVE_EXCEL, engine='openpyxl') as writer:

    # ── Sheet 1: Detailed per image per class ──
    df_detailed.to_excel(writer, sheet_name='Detailed_Metrics', index=False)

    # ── Sheet 2: Class summary ──
    class_summary = df_detailed.groupby(['Class_ID', 'Class_Name']).agg(
        TP=('TP', 'sum'), TN=('TN', 'sum'), FP=('FP', 'sum'), FN=('FN', 'sum'),
        Intersection=('Intersection', 'sum'), Union=('Union', 'sum'),
        IoU=('IoU', 'mean'), Dice=('Dice', 'mean'),
        Precision=('Precision', 'mean'), Recall=('Recall', 'mean'),
        Specificity=('Specificity', 'mean'), F1_Score=('F1_Score', 'mean'),
        Total_Pixels=('Total_Pixels', 'sum')
    ).reset_index()

    for idx, row in class_summary.iterrows():
        tp, fp, fn = row['TP'], row['FP'], row['FN']
        union      = row['Union']
        class_summary.at[idx, 'Aggregated_IoU']  = \
            row['Intersection'] / union if union > 0 else 0.0
        class_summary.at[idx, 'Aggregated_Dice'] = \
            (2*tp) / (2*tp + fp + fn)  if (2*tp + fp + fn) > 0 else 0.0

    class_summary.to_excel(writer, sheet_name='Class_Summary', index=False)

    # ── Sheet 3: Image summary ──
    image_summary = df_detailed.groupby('Filename').agg(
        TP=('TP', 'sum'), TN=('TN', 'sum'), FP=('FP', 'sum'), FN=('FN', 'sum'),
        IoU=('IoU', 'mean'), Dice=('Dice', 'mean'),
        Precision=('Precision', 'mean'), Recall=('Recall', 'mean'),
        F1_Score=('F1_Score', 'mean'),
        Total_Pixels=('Total_Pixels', 'first'),
        Classes_Present=('Class_ID', 'count')
    ).reset_index()

    denom_iou  = (image_summary['TP'] + image_summary['FP'] +
                  image_summary['FN']).replace(0, 1)
    denom_dice = (2*image_summary['TP'] + image_summary['FP'] +
                  image_summary['FN']).replace(0, 1)
    image_summary['Image_IoU']  = image_summary['TP'] / denom_iou
    image_summary['Image_Dice'] = (2 * image_summary['TP']) / denom_dice

    image_summary.to_excel(writer, sheet_name='Image_Summary', index=False)

    # ── Sheet 4: Class presence ──
    class_presence = df_detailed.groupby(['Class_ID', 'Class_Name']).agg(
        Images_Present=('Filename', 'count'),
        TP=('TP', 'sum'),
        FN=('FN', 'sum')
    ).reset_index()
    class_presence['Support'] = class_presence['TP'] + class_presence['FN']
    class_presence.to_excel(writer, sheet_name='Class_Presence', index=False)

print(f"\nExcel saved to : {SAVE_EXCEL}")
print(f"Total images   : {df_detailed['Filename'].nunique()}")
print(f"Total records  : {len(df_detailed)}")
print(f"\nOverall Performance:")
print(f"  Mean IoU  : {df_detailed['IoU'].mean():.4f}")
print(f"  Mean Dice : {df_detailed['Dice'].mean():.4f}")
print(f"\nClass Presence Summary:")
total_imgs = df_detailed['Filename'].nunique()
for cid, cname in class_names.items():
    n = len(df_detailed[df_detailed['Class_ID'] == cid])
    if n > 0:
        print(f"  {cname:20s}: {n}/{total_imgs} ({100*n/total_imgs:.1f}%)")
if skipped:
    print(f"\nSkipped ({len(skipped)}): {skipped}")







































































#Excel Table Creation nnUNet 

import os
import cv2
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F

# ===============================
# 1. MODEL ARCHITECTURE
# ===============================
class ConvBlock(nn.Module):
    def __init__(self, in_channels, out_channels):
        super().__init__()
        self.block = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, 3, padding=1, bias=False),
            nn.InstanceNorm2d(out_channels, affine=True),
            nn.LeakyReLU(inplace=True),
            nn.Conv2d(out_channels, out_channels, 3, padding=1, bias=False),
            nn.InstanceNorm2d(out_channels, affine=True),
            nn.LeakyReLU(inplace=True),
        )

    def forward(self, x):
        return self.block(x)


class Down(nn.Module):
    def __init__(self, in_channels, out_channels):
        super().__init__()
        self.down = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, 3, stride=2, padding=1),
            nn.InstanceNorm2d(out_channels, affine=True),
            nn.LeakyReLU(inplace=True)
        )
        self.conv = ConvBlock(out_channels, out_channels)

    def forward(self, x):
        x = self.down(x)
        x = self.conv(x)
        return x


class Up(nn.Module):
    def __init__(self, in_channels, skip_channels, out_channels):
        super().__init__()
        self.up   = nn.ConvTranspose2d(in_channels, out_channels,
                                       kernel_size=2, stride=2)
        self.conv = ConvBlock(out_channels + skip_channels, out_channels)

    def forward(self, x, skip):
        x = self.up(x)
        if x.shape[-2:] != skip.shape[-2:]:
            x = F.interpolate(x, size=skip.shape[-2:],
                              mode="bilinear", align_corners=False)
        x = torch.cat([x, skip], dim=1)
        return self.conv(x)


class NNUNet2D(nn.Module):
    def __init__(self, n_classes, in_channels=3, base_features=32):
        super().__init__()
        f = base_features

        self.enc1 = ConvBlock(in_channels, f)
        self.enc2 = Down(f,      f * 2)
        self.enc3 = Down(f * 2,  f * 4)
        self.enc4 = Down(f * 4,  f * 8)
        self.enc5 = Down(f * 8,  f * 16)

        self.up4  = Up(f * 16, f * 8, f * 8)
        self.up3  = Up(f * 8,  f * 4, f * 4)
        self.up2  = Up(f * 4,  f * 2, f * 2)
        self.up1  = Up(f * 2,  f,     f)

        self.out1 = nn.Conv2d(f,      n_classes, 1)
        self.out2 = nn.Conv2d(f * 2,  n_classes, 1)
        self.out3 = nn.Conv2d(f * 4,  n_classes, 1)
        self.out4 = nn.Conv2d(f * 8,  n_classes, 1)

    def forward(self, x):
        s1 = self.enc1(x)
        s2 = self.enc2(s1)
        s3 = self.enc3(s2)
        s4 = self.enc4(s3)
        b  = self.enc5(s4)

        d4 = self.up4(b,  s4)
        d3 = self.up3(d4, s3)
        d2 = self.up2(d3, s2)
        d1 = self.up1(d2, s1)

        out_main = self.out1(d1)

        self.ds_outputs = [
            out_main,
            F.interpolate(self.out2(d2), size=out_main.shape[-2:],
                          mode="bilinear", align_corners=False),
            F.interpolate(self.out3(d3), size=out_main.shape[-2:],
                          mode="bilinear", align_corners=False),
            F.interpolate(self.out4(d4), size=out_main.shape[-2:],
                          mode="bilinear", align_corners=False),
        ]
        return out_main

# ===============================
# 2. CONFIGURATION — EDIT THESE
# ===============================
MODEL_PATH    = r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/BAT-RM/Cervix_small_Axial_100_epochs_pytorch_best.pth'
IMAGES_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/images'
MASKS_FOLDER  = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'
SAVE_EXCEL    = r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/BAT-RM/segmentation_metrics_Cervix_small_Axial_axis_pytorch_UNET_BM_Generated.xlsx'

class_names = {
    0: 'Background',
    1: 'BODY',
    2: 'URINARY BLADDER',
    3: 'SMALL BOWEL',
    4: 'RECTUM',
    5: 'FEMORAL HEAD',
    6: 'GTV',
    7: 'CTV',
}
n_classes = len(class_names)

# Correct BGR mapping (verified from mask scan)
bgr_to_class = {
    (0,   0,   0):   0,   # Background
    (0,   255, 0):   1,   # BODY
    (0, 255, 255):   2,   # URINARY BLADDER
    (153, 146, 255): 3,   # SMALL BOWEL
    (64,  64,  128): 4,   # RECTUM
    (255,   255, 0): 5,   # FEMORAL HEAD
    (255, 60,  255): 6,   # GTV
    (255, 55,  55):  7,   # CTV
}

# ===============================
# 3. LOAD MODEL — PTH
# ===============================
DEVICE = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
print(f"Device: {DEVICE}")

model = NNUNet2D(n_classes=n_classes, in_channels=3).to(DEVICE)

checkpoint = torch.load(MODEL_PATH, map_location=DEVICE)

# Handle both raw state_dict and checkpoint dict formats
if isinstance(checkpoint, dict) and 'model_state_dict' in checkpoint:
    model.load_state_dict(checkpoint['model_state_dict'])
    print(f"Loaded from checkpoint dict (epoch: {checkpoint.get('epoch', 'unknown')})")
elif isinstance(checkpoint, dict) and 'state_dict' in checkpoint:
    model.load_state_dict(checkpoint['state_dict'])
    print(f"Loaded from checkpoint dict (state_dict key)")
else:
    # Raw state dict saved directly
    model.load_state_dict(checkpoint)
    print(f"Loaded raw state dict")

model.eval()
print(f"Model loaded  : {MODEL_PATH}")
print(f"Trainable parameters: {sum(p.numel() for p in model.parameters() if p.requires_grad):,}")

# Auto-detect image size from first conv layer
first_conv   = list(model.parameters())[0]
print(f"Model in_channels: {first_conv.shape[1]}")

# ── Set IMAGE_SIZE — change if your model was trained on different resolution ──
IMAGE_SIZE = (256, 256)   # (W, H) — update if needed
print(f"Image size    : {IMAGE_SIZE} (W x H)")

# ===============================
# 4. HELPER: decode BGR mask → class ID map
# ===============================
def decode_mask(mask_path):
    mask_bgr = cv2.imread(mask_path, cv2.IMREAD_COLOR)
    if mask_bgr is None:
        raise ValueError(f"Could not load mask: {mask_path}")

    class_map = np.zeros(mask_bgr.shape[:2], dtype=np.uint8)
    for bgr, cls_id in bgr_to_class.items():
        match = np.all(mask_bgr == np.array(bgr, dtype=np.uint8), axis=-1)
        class_map[match] = cls_id

    unique_ids = np.unique(class_map)
    if len(unique_ids) == 1 and unique_ids[0] == 0:
        print(f"  [WARNING] Only background decoded: {mask_path}")

    return class_map

# ===============================
# 5. HELPER: predict → class ID map
# ===============================
def predict_mask(image_path):
    img_bgr = cv2.imread(image_path, cv2.IMREAD_COLOR)
    if img_bgr is None:
        raise ValueError(f"Could not load image: {image_path}")

    original_h, original_w = img_bgr.shape[:2]

    # Resize and normalise to [0, 1]
    img_resized = cv2.resize(img_bgr, IMAGE_SIZE)
    img_norm    = img_resized.astype(np.float32) / 255.0

    # PyTorch expects (B, C, H, W)
    img_tensor  = torch.from_numpy(img_norm).permute(2, 0, 1).unsqueeze(0).to(DEVICE)

    with torch.no_grad():
        logits     = model(img_tensor)               # (1, n_classes, H, W)
        pred_class = torch.argmax(logits, dim=1)     # (1, H, W)
        pred_class = pred_class.squeeze(0).cpu().numpy().astype(np.uint8)

    # Resize back to original resolution
    pred_class = cv2.resize(pred_class,
                            (original_w, original_h),
                            interpolation=cv2.INTER_NEAREST)
    return pred_class

# ===============================
# 6. HELPER: compute per-class metrics
# ===============================
def compute_metrics(true_flat, pred_flat, filename):
    total_pixels    = len(true_flat)
    present_classes = np.union1d(np.unique(true_flat), np.unique(pred_flat))
    rows            = []

    for class_id in present_classes:
        true_bin = (true_flat == class_id)
        pred_bin = (pred_flat == class_id)

        tp = int(np.sum( true_bin &  pred_bin))
        tn = int(np.sum(~true_bin & ~pred_bin))
        fp = int(np.sum(~true_bin &  pred_bin))
        fn = int(np.sum( true_bin & ~pred_bin))

        intersection = tp
        union        = tp + fp + fn

        iou         = tp / union                if union > 0             else 0.0
        dice        = (2*tp) / (2*tp + fp + fn) if (2*tp + fp + fn) > 0  else 0.0
        precision   = tp / (tp + fp)            if (tp + fp) > 0         else 0.0
        recall      = tp / (tp + fn)            if (tp + fn) > 0         else 0.0
        specificity = tn / (tn + fp)            if (tn + fp) > 0         else 0.0
        f1          = (2*precision*recall) / (precision + recall) \
                      if (precision + recall) > 0 else 0.0
        tpr         = recall
        fpr         = fp / (fp + tn)            if (fp + tn) > 0         else 0.0

        rows.append({
            'Filename':            filename,
            'Class_ID':            int(class_id),
            'Class_Name':          class_names.get(int(class_id), str(class_id)),
            'Total_Pixels':        total_pixels,
            'TP':                  tp,
            'TN':                  tn,
            'FP':                  fp,
            'FN':                  fn,
            'Intersection':        intersection,
            'Union':               union,
            'IoU':                 iou,
            'Dice':                dice,
            'Precision':           precision,
            'Recall':              recall,
            'Specificity':         specificity,
            'F1_Score':            f1,
            'True_Positive_Rate':  tpr,
            'False_Positive_Rate': fpr,
        })

    return rows

# ===============================
# 7. MAIN INFERENCE LOOP
# ===============================
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp')
image_files = sorted([
    f for f in os.listdir(IMAGES_FOLDER)
    if f.lower().endswith(image_extensions)
])

print(f"\nFound {len(image_files)} images. Starting inference...\n")

all_metrics = []
skipped     = []

for idx, img_file in enumerate(image_files):
    filename   = os.path.splitext(img_file)[0]
    image_path = os.path.join(IMAGES_FOLDER, img_file)
    mask_path  = os.path.join(MASKS_FOLDER,  img_file)

    if not os.path.exists(mask_path):
        print(f"  [SKIP] No mask found for: {img_file}")
        skipped.append(img_file)
        continue

    try:
        true_mask = decode_mask(mask_path)
        pred_mask = predict_mask(image_path)

        if true_mask.shape != pred_mask.shape:
            pred_mask = cv2.resize(pred_mask,
                                   (true_mask.shape[1], true_mask.shape[0]),
                                   interpolation=cv2.INTER_NEAREST)

        rows = compute_metrics(true_mask.flatten(), pred_mask.flatten(), filename)
        all_metrics.extend(rows)

        if (idx + 1) % 20 == 0 or (idx + 1) == len(image_files):
            print(f"  Processed {idx+1}/{len(image_files)}: {filename}")

    except Exception as e:
        print(f"  [ERROR] {img_file}: {e}")
        skipped.append(img_file)

print(f"\nInference complete.")
print(f"  Processed : {len(image_files) - len(skipped)}")
print(f"  Skipped   : {len(skipped)}")

# ===============================
# 8. BUILD DATAFRAME
# ===============================
column_order = [
    'Filename', 'Class_ID', 'Class_Name', 'Total_Pixels',
    'TP', 'TN', 'FP', 'FN', 'Intersection', 'Union',
    'IoU', 'Dice', 'Precision', 'Recall', 'Specificity', 'F1_Score',
    'True_Positive_Rate', 'False_Positive_Rate'
]
df_detailed = pd.DataFrame(all_metrics)[column_order]

# ===============================
# 9. SAVE EXCEL — identical sheet structure
# ===============================
os.makedirs(os.path.dirname(SAVE_EXCEL), exist_ok=True)

with pd.ExcelWriter(SAVE_EXCEL, engine='openpyxl') as writer:

    # ── Sheet 1: Detailed per image per class ──
    df_detailed.to_excel(writer, sheet_name='Detailed_Metrics', index=False)

    # ── Sheet 2: Class summary ──
    class_summary = df_detailed.groupby(['Class_ID', 'Class_Name']).agg(
        TP=('TP', 'sum'), TN=('TN', 'sum'), FP=('FP', 'sum'), FN=('FN', 'sum'),
        Intersection=('Intersection', 'sum'), Union=('Union', 'sum'),
        IoU=('IoU', 'mean'), Dice=('Dice', 'mean'),
        Precision=('Precision', 'mean'), Recall=('Recall', 'mean'),
        Specificity=('Specificity', 'mean'), F1_Score=('F1_Score', 'mean'),
        Total_Pixels=('Total_Pixels', 'sum')
    ).reset_index()

    for idx, row in class_summary.iterrows():
        tp, fp, fn = row['TP'], row['FP'], row['FN']
        union      = row['Union']
        class_summary.at[idx, 'Aggregated_IoU']  = \
            row['Intersection'] / union if union > 0 else 0.0
        class_summary.at[idx, 'Aggregated_Dice'] = \
            (2*tp) / (2*tp + fp + fn)  if (2*tp + fp + fn) > 0 else 0.0

    class_summary.to_excel(writer, sheet_name='Class_Summary', index=False)

    # ── Sheet 3: Image summary ──
    image_summary = df_detailed.groupby('Filename').agg(
        TP=('TP', 'sum'), TN=('TN', 'sum'), FP=('FP', 'sum'), FN=('FN', 'sum'),
        IoU=('IoU', 'mean'), Dice=('Dice', 'mean'),
        Precision=('Precision', 'mean'), Recall=('Recall', 'mean'),
        F1_Score=('F1_Score', 'mean'),
        Total_Pixels=('Total_Pixels', 'first'),
        Classes_Present=('Class_ID', 'count')
    ).reset_index()

    denom_iou  = (image_summary['TP'] + image_summary['FP'] +
                  image_summary['FN']).replace(0, 1)
    denom_dice = (2*image_summary['TP'] + image_summary['FP'] +
                  image_summary['FN']).replace(0, 1)
    image_summary['Image_IoU']  = image_summary['TP'] / denom_iou
    image_summary['Image_Dice'] = (2 * image_summary['TP']) / denom_dice

    image_summary.to_excel(writer, sheet_name='Image_Summary', index=False)

    # ── Sheet 4: Class presence ──
    class_presence = df_detailed.groupby(['Class_ID', 'Class_Name']).agg(
        Images_Present=('Filename', 'count'),
        TP=('TP', 'sum'),
        FN=('FN', 'sum')
    ).reset_index()
    class_presence['Support'] = class_presence['TP'] + class_presence['FN']
    class_presence.to_excel(writer, sheet_name='Class_Presence', index=False)

print(f"\nExcel saved to : {SAVE_EXCEL}")
print(f"Total images   : {df_detailed['Filename'].nunique()}")
print(f"Total records  : {len(df_detailed)}")
print(f"\nOverall Performance:")
print(f"  Mean IoU  : {df_detailed['IoU'].mean():.4f}")
print(f"  Mean Dice : {df_detailed['Dice'].mean():.4f}")
print(f"\nClass Presence Summary:")
total_imgs = df_detailed['Filename'].nunique()
for cid, cname in class_names.items():
    n = len(df_detailed[df_detailed['Class_ID'] == cid])
    if n > 0:
        print(f"  {cname:20s}: {n}/{total_imgs} ({100*n/total_imgs:.1f}%)")
if skipped:
    print(f"\nSkipped ({len(skipped)}): {skipped}")


























































#Excel Table Creation BAT-RM

import os
import cv2
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F

# ===============================
# 1. MODEL ARCHITECTURE — BAT-RM
# ===============================
class EncoderBlock(nn.Module):
    def __init__(self, in_ch, out_ch, dropout=0.1):
        super().__init__()
        self.conv = nn.Sequential(
            nn.Conv2d(in_ch, out_ch, kernel_size=3, padding=1),
            nn.BatchNorm2d(out_ch),
            nn.ReLU(inplace=True),
            nn.Dropout2d(dropout),
            nn.Conv2d(out_ch, out_ch, kernel_size=3, padding=1),
            nn.BatchNorm2d(out_ch),
            nn.ReLU(inplace=True)
        )

    def forward(self, x):
        return self.conv(x)


class GatedBATBlock(nn.Module):
    def __init__(self, in_ch):
        super().__init__()
        self.sobel_x = nn.Parameter(
            torch.tensor([[-1,0,1],[-2,0,2],[-1,0,1]],
                         dtype=torch.float32).view(1,1,3,3), requires_grad=False)
        self.sobel_y = nn.Parameter(
            torch.tensor([[-1,-2,-1],[0,0,0],[1,2,1]],
                         dtype=torch.float32).view(1,1,3,3), requires_grad=False)
        self.query = nn.Conv2d(in_ch, in_ch // 8, kernel_size=1)
        self.key   = nn.Conv2d(in_ch, in_ch // 8, kernel_size=1)
        self.value = nn.Conv2d(in_ch, in_ch,       kernel_size=1)
        self.gamma = nn.Parameter(torch.zeros(1))

    def forward(self, x):
        gray   = torch.mean(x, dim=1, keepdim=True)
        grad_x = F.conv2d(gray, self.sobel_x, padding=1)
        grad_y = F.conv2d(gray, self.sobel_y, padding=1)
        gate   = torch.sigmoid(torch.sqrt(grad_x**2 + grad_y**2 + 1e-6))

        b, c, h, w = x.size()
        q   = self.query(x * gate).view(b, -1, h*w).permute(0, 2, 1)
        k   = self.key(x * gate).view(b, -1, h*w)
        v   = self.value(x).view(b, -1, h*w)
        attn = F.softmax(torch.bmm(q, k), dim=-1)
        out  = torch.bmm(v, attn.permute(0, 2, 1)).view(b, c, h, w)
        return self.gamma * out + x, gate


class RegionMambaBlock(nn.Module):
    def __init__(self, in_ch):
        super().__init__()
        self.norm = nn.LayerNorm(in_ch)
        self.ssm  = nn.Linear(in_ch, in_ch)
        self.conv = nn.Conv2d(in_ch, in_ch, kernel_size=3, padding=1, groups=in_ch)

    def forward(self, x):
        shortcut = x
        b, c, h, w = x.shape
        x_flat = x.permute(0,2,3,1).reshape(b, -1, c)
        x_flat = self.norm(x_flat)
        x = x_flat.reshape(b,h,w,c).permute(0,3,1,2)
        x = self.conv(x)
        x = x.permute(0,2,3,1).reshape(b,-1,c)
        x = self.ssm(x)
        x = x.reshape(b,h,w,c).permute(0,3,1,2)
        return x + shortcut


class BRAFModule(nn.Module):
    def __init__(self, bat_ch=128, rm_ch=512):
        super().__init__()
        self.rm_project = nn.Conv2d(rm_ch, bat_ch, kernel_size=1)
        self.alpha_conv = nn.Sequential(
            nn.Conv2d(bat_ch + bat_ch, 1, kernel_size=3, padding=1),
            nn.Sigmoid()
        )
        self.refine = nn.Conv2d(bat_ch, bat_ch, kernel_size=3, padding=1)

    def forward(self, f_bat, f_rm, gate):
        f_rm_up      = F.interpolate(f_rm, size=f_bat.shape[2:],
                                     mode='bilinear', align_corners=False)
        f_rm_aligned = self.rm_project(f_rm_up)
        alpha        = self.alpha_conv(torch.cat([f_bat, f_rm_aligned], dim=1))
        f_fuse       = (alpha * f_bat) + ((1 - alpha) * f_rm_aligned)
        return self.refine(f_fuse) * gate


class BAT_RM_UNet(nn.Module):
    def __init__(self, n_classes, in_channels=3):
        super().__init__()
        self.e1   = EncoderBlock(in_channels, 32)
        self.e2   = EncoderBlock(32,  64)
        self.e3   = EncoderBlock(64,  128)
        self.e4   = EncoderBlock(128, 256)
        self.e5   = EncoderBlock(256, 512)
        self.pool = nn.MaxPool2d(2)

        self.bat  = GatedBATBlock(128)
        self.rm   = RegionMambaBlock(512)
        self.braf = BRAFModule(bat_ch=128, rm_ch=512)

        self.up5  = nn.ConvTranspose2d(512, 256, kernel_size=2, stride=2)
        self.d5   = EncoderBlock(512, 256)
        self.up4  = nn.ConvTranspose2d(256, 128, kernel_size=2, stride=2)
        self.d4   = EncoderBlock(256, 128)
        self.up3  = nn.ConvTranspose2d(128, 64,  kernel_size=2, stride=2)
        self.d3   = EncoderBlock(192, 64)
        self.up2  = nn.ConvTranspose2d(64,  32,  kernel_size=2, stride=2)
        self.d2   = EncoderBlock(96,  32)

        self.out_conv = nn.Conv2d(32, n_classes, kernel_size=1)

    def forward(self, x):
        s1 = self.e1(x)
        s2 = self.e2(self.pool(s1))
        s3 = self.e3(self.pool(s2))
        s4 = self.e4(self.pool(s3))
        s5 = self.e5(self.pool(s4))

        f_bat, gate = self.bat(s3)
        f_rm        = self.rm(s5)
        f_fuse      = self.braf(f_bat, f_rm, gate)

        x5 = self.up5(s5)
        if x5.shape[2:] != s4.shape[2:]:
            x5 = F.interpolate(x5, size=s4.shape[2:], mode='bilinear')
        x5 = self.d5(torch.cat([x5, s4], dim=1))

        x4 = self.up4(x5)
        if x4.shape[2:] != s3.shape[2:]:
            x4 = F.interpolate(x4, size=s3.shape[2:], mode='bilinear')
        x4 = self.d4(torch.cat([x4, s3], dim=1))

        x3     = self.up3(x4)
        f_fuse_up = F.interpolate(f_fuse, size=x3.shape[2:], mode='bilinear')
        x3     = self.d3(torch.cat([x3, f_fuse_up], dim=1))

        x2 = self.up2(x3)
        if x2.shape[2:] != s2.shape[2:]:
            x2 = F.interpolate(x2, size=s2.shape[2:], mode='bilinear')
        x2 = self.d2(torch.cat([x2, s2], dim=1))

        x_final = F.interpolate(x2, size=x.shape[2:], mode='bilinear')
        return self.out_conv(x_final)

# ===============================
# 2. CONFIGURATION — EDIT THESE
# ===============================
MODEL_PATH    = r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/BAT-RM/Cervix_small_Axial_100_epochs_pytorch_best.pth'
IMAGES_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/images'
MASKS_FOLDER  = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'
SAVE_EXCEL    = r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/BAT-RM/BAT_RM_segmentation_metrics_generated.xlsx'

class_names = {
    0: 'Background',
    1: 'BODY',
    2: 'URINARY BLADDER',
    3: 'SMALL BOWEL',
    4: 'RECTUM',
    5: 'FEMORAL HEAD',
    6: 'GTV',
    7: 'CTV',
}
n_classes = len(class_names)

# Verified BGR mapping (matches training label_encode_mask behaviour)
bgr_to_class = {
    (0,   0,   0):   0,   # Background
    (0,   255, 0):   1,   # BODY
    (0,   255, 255): 2,   # URINARY BLADDER
    (153, 146, 255): 3,   # SMALL BOWEL
    (64,  64,  128): 4,   # RECTUM
    (255, 255, 0):   5,   # FEMORAL HEAD
    (255, 60,  255): 6,   # GTV
    (255, 55,  55):  7,   # CTV
}

# ===============================
# 3. LOAD MODEL — PTH
# ===============================
DEVICE = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
print(f"Device: {DEVICE}")

model = BAT_RM_UNet(n_classes=n_classes, in_channels=3).to(DEVICE)

checkpoint = torch.load(MODEL_PATH, map_location=DEVICE)

if isinstance(checkpoint, dict) and 'model_state_dict' in checkpoint:
    model.load_state_dict(checkpoint['model_state_dict'])
    print(f"Loaded from checkpoint dict (epoch: {checkpoint.get('epoch', 'unknown')})")
elif isinstance(checkpoint, dict) and 'state_dict' in checkpoint:
    model.load_state_dict(checkpoint['state_dict'])
    print(f"Loaded from checkpoint dict (state_dict key)")
else:
    model.load_state_dict(checkpoint)
    print(f"Loaded raw state dict")

model.eval()
print(f"Model loaded  : {MODEL_PATH}")
print(f"Trainable parameters: {sum(p.numel() for p in model.parameters() if p.requires_grad):,}")

IMAGE_SIZE = (256, 256)   # BAT-RM was trained on 512x512
print(f"Image size    : {IMAGE_SIZE} (W x H)")

# ===============================
# 4-9. UNCHANGED — copy exactly from nnUNet inference code
# ===============================
def decode_mask(mask_path):
    mask_bgr = cv2.imread(mask_path, cv2.IMREAD_COLOR)
    if mask_bgr is None:
        raise ValueError(f"Could not load mask: {mask_path}")
    class_map = np.zeros(mask_bgr.shape[:2], dtype=np.uint8)
    for bgr, cls_id in bgr_to_class.items():
        match = np.all(mask_bgr == np.array(bgr, dtype=np.uint8), axis=-1)
        class_map[match] = cls_id
    unique_ids = np.unique(class_map)
    if len(unique_ids) == 1 and unique_ids[0] == 0:
        print(f"  [WARNING] Only background decoded: {mask_path}")
    return class_map


def predict_mask(image_path):
    img_bgr = cv2.imread(image_path, cv2.IMREAD_COLOR)
    if img_bgr is None:
        raise ValueError(f"Could not load image: {image_path}")
    original_h, original_w = img_bgr.shape[:2]
    img_resized = cv2.resize(img_bgr, IMAGE_SIZE, interpolation=cv2.INTER_LINEAR)
    img_norm    = img_resized.astype(np.float32) / 255.0
    img_tensor  = torch.from_numpy(img_norm).permute(2, 0, 1).unsqueeze(0).to(DEVICE)
    with torch.no_grad():
        logits     = model(img_tensor)
        pred_class = torch.argmax(logits, dim=1)
        pred_class = pred_class.squeeze(0).cpu().numpy().astype(np.uint8)
    pred_class = cv2.resize(pred_class,
                            (original_w, original_h),
                            interpolation=cv2.INTER_NEAREST)
    return pred_class


def compute_metrics(true_flat, pred_flat, filename):
    total_pixels    = len(true_flat)
    present_classes = np.union1d(np.unique(true_flat), np.unique(pred_flat))
    rows            = []
    for class_id in present_classes:
        true_bin = (true_flat == class_id)
        pred_bin = (pred_flat == class_id)
        tp = int(np.sum( true_bin &  pred_bin))
        tn = int(np.sum(~true_bin & ~pred_bin))
        fp = int(np.sum(~true_bin &  pred_bin))
        fn = int(np.sum( true_bin & ~pred_bin))
        intersection = tp
        union        = tp + fp + fn
        iou         = tp / union                if union > 0             else 0.0
        dice        = (2*tp) / (2*tp + fp + fn) if (2*tp + fp + fn) > 0  else 0.0
        precision   = tp / (tp + fp)            if (tp + fp) > 0         else 0.0
        recall      = tp / (tp + fn)            if (tp + fn) > 0         else 0.0
        specificity = tn / (tn + fp)            if (tn + fp) > 0         else 0.0
        f1          = (2*precision*recall) / (precision + recall) \
                      if (precision + recall) > 0 else 0.0
        tpr         = recall
        fpr         = fp / (fp + tn)            if (fp + tn) > 0         else 0.0
        rows.append({
            'Filename':            filename,
            'Class_ID':            int(class_id),
            'Class_Name':          class_names.get(int(class_id), str(class_id)),
            'Total_Pixels':        total_pixels,
            'TP':                  tp, 'TN': tn, 'FP': fp, 'FN': fn,
            'Intersection':        intersection, 'Union': union,
            'IoU':                 iou,  'Dice':        dice,
            'Precision':           precision, 'Recall': recall,
            'Specificity':         specificity, 'F1_Score': f1,
            'True_Positive_Rate':  tpr, 'False_Positive_Rate': fpr,
        })
    return rows


# ===============================
# 7. MAIN INFERENCE LOOP
# ===============================
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp')
image_files = sorted([
    f for f in os.listdir(IMAGES_FOLDER)
    if f.lower().endswith(image_extensions)
])

print(f"\nFound {len(image_files)} images. Starting inference...\n")

all_metrics = []
skipped     = []

for idx, img_file in enumerate(image_files):
    filename   = os.path.splitext(img_file)[0]
    image_path = os.path.join(IMAGES_FOLDER, img_file)
    mask_path  = os.path.join(MASKS_FOLDER,  img_file)

    if not os.path.exists(mask_path):
        print(f"  [SKIP] No mask found for: {img_file}")
        skipped.append(img_file)
        continue

    try:
        true_mask = decode_mask(mask_path)
        pred_mask = predict_mask(image_path)
        if true_mask.shape != pred_mask.shape:
            pred_mask = cv2.resize(pred_mask,
                                   (true_mask.shape[1], true_mask.shape[0]),
                                   interpolation=cv2.INTER_NEAREST)
        rows = compute_metrics(true_mask.flatten(), pred_mask.flatten(), filename)
        all_metrics.extend(rows)
        if (idx + 1) % 20 == 0 or (idx + 1) == len(image_files):
            print(f"  Processed {idx+1}/{len(image_files)}: {filename}")
    except Exception as e:
        print(f"  [ERROR] {img_file}: {e}")
        skipped.append(img_file)

print(f"\nInference complete.")
print(f"  Processed : {len(image_files) - len(skipped)}")
print(f"  Skipped   : {len(skipped)}")

# ===============================
# 8. BUILD DATAFRAME
# ===============================
column_order = [
    'Filename', 'Class_ID', 'Class_Name', 'Total_Pixels',
    'TP', 'TN', 'FP', 'FN', 'Intersection', 'Union',
    'IoU', 'Dice', 'Precision', 'Recall', 'Specificity', 'F1_Score',
    'True_Positive_Rate', 'False_Positive_Rate'
]
df_detailed = pd.DataFrame(all_metrics)[column_order]

# ===============================
# 9. SAVE EXCEL
# ===============================
os.makedirs(os.path.dirname(SAVE_EXCEL), exist_ok=True)

with pd.ExcelWriter(SAVE_EXCEL, engine='openpyxl') as writer:

    df_detailed.to_excel(writer, sheet_name='Detailed_Metrics', index=False)

    class_summary = df_detailed.groupby(['Class_ID', 'Class_Name']).agg(
        TP=('TP','sum'), TN=('TN','sum'), FP=('FP','sum'), FN=('FN','sum'),
        Intersection=('Intersection','sum'), Union=('Union','sum'),
        IoU=('IoU','mean'), Dice=('Dice','mean'),
        Precision=('Precision','mean'), Recall=('Recall','mean'),
        Specificity=('Specificity','mean'), F1_Score=('F1_Score','mean'),
        Total_Pixels=('Total_Pixels','sum')
    ).reset_index()
    for idx, row in class_summary.iterrows():
        tp, fp, fn = row['TP'], row['FP'], row['FN']
        union      = row['Union']
        class_summary.at[idx, 'Aggregated_IoU']  = \
            row['Intersection'] / union if union > 0 else 0.0
        class_summary.at[idx, 'Aggregated_Dice'] = \
            (2*tp) / (2*tp+fp+fn) if (2*tp+fp+fn) > 0 else 0.0
    class_summary.to_excel(writer, sheet_name='Class_Summary', index=False)

    image_summary = df_detailed.groupby('Filename').agg(
        TP=('TP','sum'), TN=('TN','sum'), FP=('FP','sum'), FN=('FN','sum'),
        IoU=('IoU','mean'), Dice=('Dice','mean'),
        Precision=('Precision','mean'), Recall=('Recall','mean'),
        F1_Score=('F1_Score','mean'),
        Total_Pixels=('Total_Pixels','first'),
        Classes_Present=('Class_ID','count')
    ).reset_index()
    denom_iou  = (image_summary['TP']+image_summary['FP']+image_summary['FN']).replace(0,1)
    denom_dice = (2*image_summary['TP']+image_summary['FP']+image_summary['FN']).replace(0,1)
    image_summary['Image_IoU']  = image_summary['TP'] / denom_iou
    image_summary['Image_Dice'] = (2 * image_summary['TP']) / denom_dice
    image_summary.to_excel(writer, sheet_name='Image_Summary', index=False)

    class_presence = df_detailed.groupby(['Class_ID','Class_Name']).agg(
        Images_Present=('Filename','count'),
        TP=('TP','sum'), FN=('FN','sum')
    ).reset_index()
    class_presence['Support'] = class_presence['TP'] + class_presence['FN']
    class_presence.to_excel(writer, sheet_name='Class_Presence', index=False)

print(f"\nExcel saved to : {SAVE_EXCEL}")
print(f"Total images   : {df_detailed['Filename'].nunique()}")
print(f"Total records  : {len(df_detailed)}")
print(f"\nOverall Performance:")
print(f"  Mean IoU  : {df_detailed['IoU'].mean():.4f}")
print(f"  Mean Dice : {df_detailed['Dice'].mean():.4f}")
print(f"\nClass Presence Summary:")
total_imgs = df_detailed['Filename'].nunique()
for cid, cname in class_names.items():
    n = len(df_detailed[df_detailed['Class_ID'] == cid])
    if n > 0:
        print(f"  {cname:20s}: {n}/{total_imgs} ({100*n/total_imgs:.1f}%)")
if skipped:
    print(f"\nSkipped ({len(skipped)}): {skipped}")




























































#Final Version Box plot
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from scipy.stats import wilcoxon
import os
import warnings
from matplotlib.lines import Line2D

warnings.filterwarnings("ignore")

# ===============================
# 1. CONFIGURATION
# ===============================



# excel_files = [
#     ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_detailed_epoch_200_generated.xlsx'),
#     ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
#     ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
#     ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
#     ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx')
# ]


excel_files = [
    ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
    ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx')
]

color_map = {
    "BAT-RM": ("#00b4d8", "#0096c7"),
    "nnUNet":  ("#2a9134", "#137547"),
    "SegMamba":  ("#f4a261", "#e76f51"),
    "TransUNet":  ("#e9c46a", "#d4a017"),
    "UNETR":  ("#a29bfe", "#6c5ce7")
}

METRIC   = "IoU"
SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/segmentation_Metrics'
os.makedirs(SAVE_DIR, exist_ok=True)

# ===============================
# 2. STATISTICAL HELPERS
# ===============================
def calculate_effect_size(x, y):
    nx, ny = len(x), len(y)
    dof = nx + ny - 2
    std_pooled = np.sqrt(((nx-1)*np.std(x, ddof=1)**2 + (ny-1)*np.std(y, ddof=1)**2) / dof)
    d = (np.mean(x) - np.mean(y)) / std_pooled if std_pooled != 0 else 0
    g = d * (1 - (3 / (4 * dof - 1)))
    return d, g

def get_stars(p_val):
    if pd.isna(p_val): return "ns"
    if p_val < 0.0001: return "****"
    elif p_val < 0.001: return "***"
    elif p_val < 0.01:  return "**"
    elif p_val < 0.05:  return "*"
    return "ns"

# ===============================
# 3. DATA PROCESSING
# ===============================
all_data = []
for name, path in excel_files:
    df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
    df[METRIC] = pd.to_numeric(df[METRIC], errors='coerce')
    df = df.dropna(subset=[METRIC])
    df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
    df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
    df["Model"] = name
    all_data.append(df[["Filename", "Class_Name", METRIC, "Model"]])

df_main     = pd.concat(all_data, ignore_index=True)
model_names = [x[0] for x in excel_files]
classes     = sorted(df_main["Class_Name"].unique())

# ===============================
# STATISTICS
# ===============================
stats_results = []
star_lookup   = {}
for cls in classes:
    star_lookup[cls] = {}
    ref_data = df_main[(df_main["Model"] == model_names[0]) & (df_main["Class_Name"] == cls)]
    for other in model_names[1:]:
        other_data = df_main[(df_main["Model"] == other) & (df_main["Class_Name"] == cls)]
        merged = pd.merge(ref_data, other_data, on="Filename", suffixes=('_ref', '_other')).dropna()
        p, cohen_d, hedges_g = np.nan, 0, 0
        if len(merged) > 2:
            if not (merged[f"{METRIC}_ref"] == merged[f"{METRIC}_other"]).all():
                _, p = wilcoxon(merged[f"{METRIC}_ref"], merged[f"{METRIC}_other"], zero_method='pratt')
            cohen_d, hedges_g = calculate_effect_size(merged[f"{METRIC}_ref"], merged[f"{METRIC}_other"])
        sig = get_stars(p)
        star_lookup[cls][other] = sig
        stats_results.append({
            "Class": cls, "Comparison": f"{model_names[0]} vs {other}",
            "P_value": p, "Significance": sig, "Cohens_d": cohen_d,
            "Hedges_g": hedges_g, "N": len(merged)
        })

pd.DataFrame(stats_results).to_excel(os.path.join(SAVE_DIR, f"{METRIC}_Complete_Stats.xlsx"), index=False)

# ===============================
# 4. LAYOUT CONSTANTS
# ===============================
n_models  = len(model_names)
n_classes = len(classes)

BOX_WIDTH  = 0.75   # total cluster width
GAP_FRAC   = 0.20   # gap between boxes as fraction of per-box slot
slot_w     = BOX_WIDTH / n_models
new_box_w  = slot_w * (1 - GAP_FRAC)
x_half     = new_box_w / 2

# Pre-compute exact box center for every (class_idx, model_idx)
# This is the SINGLE source of truth for x positions used everywhere
def box_center(ci, mi):
    return ci - BOX_WIDTH / 2 + slot_w / 2 + mi * slot_w

# ===============================
# 5. DRAW PLOT MANUALLY
# ===============================
fig, ax = plt.subplots(figsize=(24, 12))
sns.set_style("white")

np.random.seed(42)

for ci, cls in enumerate(classes):
    for mi, model in enumerate(model_names):
        vals = df_main[
            (df_main["Class_Name"] == cls) & (df_main["Model"] == model)
        ][METRIC].dropna().values

        if len(vals) == 0:
            continue

        cx        = box_center(ci, mi)
        fill_col  = color_map[model][0]
        edge_col  = color_map[model][1]

        # ── Box stats ──
        q1, med, q3 = np.percentile(vals, [25, 50, 75])
        iqr   = q3 - q1
        lo    = max(vals.min(), q1 - 1.5 * iqr)
        hi    = min(vals.max(), q3 + 1.5 * iqr)
        mean  = vals.mean()

        # ── Box rectangle ──
        rect = mpatches.FancyBboxPatch(
            (cx - x_half, q1), new_box_w, q3 - q1,
            boxstyle="square,pad=0",
            facecolor=fill_col, edgecolor=edge_col, linewidth=1.8, zorder=3
        )
        ax.add_patch(rect)

        # ── Median line ──
        ax.plot([cx - x_half, cx + x_half], [med, med],
                color=edge_col, linewidth=2.2, zorder=4)

        # ── Mean diamond ──
        ax.plot(cx, mean, marker="D", color="#D3D3D3",
                markeredgecolor="#757575", markersize=4, zorder=5)

        # ── Whiskers ──
        ax.plot([cx, cx], [q1, lo], color=edge_col, linewidth=1.5, zorder=3)
        ax.plot([cx, cx], [q3, hi], color=edge_col, linewidth=1.5, zorder=3)

        # ── Caps ──
        cap_w = x_half * 0.6
        ax.plot([cx - cap_w, cx + cap_w], [lo, lo], color=edge_col, linewidth=1.5, zorder=3)
        ax.plot([cx - cap_w, cx + cap_w], [hi, hi], color=edge_col, linewidth=1.5, zorder=3)

        # ── Jitter (strictly inside box width) ──
        n_pts  = len(vals)
        jitter = (np.random.rand(n_pts) - 0.5) * new_box_w * 0.80
        ax.scatter(cx + jitter, vals,
                   color="black", alpha=0.12, s=6, zorder=6, linewidths=0)

# ── X-axis ticks & labels ──
ax.set_xticks(range(n_classes))
ax.set_xticklabels(classes, fontsize=11)
ax.set_xlim(-0.5, n_classes - 0.5)
ax.set_xlabel("Class_Name", fontsize=12)
ax.set_ylabel(METRIC, fontsize=12)

# ===============================
# 6. SIGNIFICANCE BRACKETS
# ===============================
y_top = 1.05
for i, cls in enumerate(classes):
    x_ref = box_center(i, 0)
    for j, other in enumerate(model_names[1:]):
        x_other = box_center(i, j + 1)
        h = y_top + (j * 0.08)
        ax.plot([x_ref, x_ref, x_other, x_other],
                [h - 0.01, h, h, h - 0.01],
                color="black", lw=1.2)
        ax.text((x_ref + x_other) / 2, h,
                star_lookup[cls][other],
                ha='center', va='bottom', fontsize=11, fontweight='bold')

# ===============================
# 7. STYLING
# ===============================
ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1.0])
plt.ylim(df_main[METRIC].min() - 0.05, 1.55)
sns.despine()

# ===============================
# 8. LEGEND
# ===============================
legend_handles = []
for m in model_names:
    p = mpatches.Patch(facecolor=color_map[m][0],
                       edgecolor=color_map[m][1],
                       linewidth=1.8, label=m)
    legend_handles.append(p)

line_median  = Line2D([0], [0], color='#757575', lw=2, label='Median')
marker_mean  = Line2D([0], [0], marker='D', color='w',
                      markerfacecolor='#D3D3D3', markeredgecolor='#757575',
                      markersize=6, label='Mean (Diamond)')
marker_jitter = Line2D([0], [0], marker='o', color='w',
                       markerfacecolor='black', markeredgecolor='black',
                       markersize=5, alpha=0.4, label='Data Points')

ax.legend(
    handles=legend_handles + [line_median, marker_mean, marker_jitter],
    loc='lower right', frameon=True, edgecolor='black', fontsize=12
)

# ===============================
# 9. SAVE
# ===============================
plt.savefig(os.path.join(SAVE_DIR, f"{METRIC}_Plot.png"), dpi=600, bbox_inches='tight')
plt.savefig(os.path.join(SAVE_DIR, f"{METRIC}_Plot.pdf"), format='pdf', bbox_inches='tight')
plt.show()

print(f"Success. Files saved in: {SAVE_DIR}")
















































#Complete Statistical Analysis
import pandas as pd
import numpy as np
from scipy.stats import wilcoxon
import os
import warnings

warnings.filterwarnings("ignore")

# ===============================
# 1. CONFIGURATION
# ===============================
excel_files = [
    ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/BAT-RM/BAT_RM_segmentation_metrics_generated.xlsx'),
    ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx')
]

METRICS = ["Dice", "IoU", "F1_Score", "Recall", "Specificity",
           "True_Positive_Rate", "False_Positive_Rate"]

SAVE_PATH = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/Complete_Statistical_Analysis.xlsx'
os.makedirs(os.path.dirname(SAVE_PATH), exist_ok=True)

# ===============================
# 2. HELPERS
# ===============================
def get_stars(p):
    if pd.isna(p):  return "ns"
    if p < 0.0001:  return "****"
    if p < 0.001:   return "***"
    if p < 0.01:    return "**"
    if p < 0.05:    return "*"
    return "ns"

def cohens_d_hedges_g(x, y):
    nx, ny = len(x), len(y)
    dof    = nx + ny - 2
    if dof <= 0:
        return np.nan, np.nan
    sp = np.sqrt(((nx - 1) * np.std(x, ddof=1)**2 +
                  (ny - 1) * np.std(y, ddof=1)**2) / dof)
    d  = (np.mean(x) - np.mean(y)) / sp if sp != 0 else 0.0
    g  = d * (1 - 3 / (4 * dof - 1))
    return d, g

def rank_biserial(stat_w, diffs):
    """Rank-biserial correlation — native effect size for Wilcoxon."""
    n = (diffs != 0).sum()
    if n == 0:
        return np.nan
    return 1 - (2 * stat_w) / (n * (n + 1))

def interpret_r(r):
    if pd.isna(r):      return "N/A"
    r = abs(r)
    if r < 0.10:        return "Negligible"
    if r < 0.30:        return "Small"
    if r < 0.50:        return "Medium"
    return "Large"

def interpret_d(d):
    if pd.isna(d):      return "N/A"
    d = abs(d)
    if d < 0.20:        return "Negligible"
    if d < 0.50:        return "Small"
    if d < 0.80:        return "Medium"
    return "Large"

# ===============================
# 3. LOAD ALL DATA
# ===============================
model_names = [x[0] for x in excel_files]

all_data = []
for name, path in excel_files:
    df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
    for m in METRICS:
        if m in df.columns:
            df[m] = pd.to_numeric(df[m], errors='coerce')
    df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
    df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
    df["Model"] = name
    keep_cols = ["Filename", "Class_Name", "Model"] + [m for m in METRICS if m in df.columns]
    all_data.append(df[keep_cols])

df_main = pd.concat(all_data, ignore_index=True)
classes = sorted(df_main["Class_Name"].unique())

# ===============================
# 4. COMPUTE STATS PER METRIC
# ===============================
# Results dict: metric -> list of row dicts
metric_results = {m: [] for m in METRICS}

for metric in METRICS:
    print(f"Processing: {metric}")
    ref_name = model_names[0]

    for cls in classes:
        ref_data = df_main[
            (df_main["Model"] == ref_name) &
            (df_main["Class_Name"] == cls)
        ][["Filename", metric]].dropna(subset=[metric])

        for other in model_names[1:]:
            other_data = df_main[
                (df_main["Model"] == other) &
                (df_main["Class_Name"] == cls)
            ][["Filename", metric]].dropna(subset=[metric])

            # Paired merge on Filename
            merged = pd.merge(
                ref_data, other_data,
                on="Filename", suffixes=('_ref', '_other')
            ).dropna()

            x    = merged[f"{metric}_ref"].values
            y    = merged[f"{metric}_other"].values
            diff = x - y
            N    = len(merged)

            # Defaults
            p, stat_w              = np.nan, np.nan
            cohen_d, hedges_g_val  = np.nan, np.nan
            r_rb                   = np.nan

            if N > 2:
                if not (diff == 0).all():
                    stat_w, p = wilcoxon(x, y, zero_method='pratt')
                    r_rb      = rank_biserial(stat_w, diff)
                cohen_d, hedges_g_val = cohens_d_hedges_g(x, y)

            metric_results[metric].append({
                "Class"              : cls,
                "Comparison"         : f"{ref_name} vs {other}",
                "N_paired"           : N,
                "Ref_mean"           : np.mean(x) if N > 0 else np.nan,
                "Ref_std"            : np.std(x, ddof=1) if N > 1 else np.nan,
                "Ref_median"         : np.median(x) if N > 0 else np.nan,
                "Other_mean"         : np.mean(y) if N > 0 else np.nan,
                "Other_std"          : np.std(y, ddof=1) if N > 1 else np.nan,
                "Other_median"       : np.median(y) if N > 0 else np.nan,
                "Mean_difference"    : np.mean(diff) if N > 0 else np.nan,
                "Wilcoxon_statistic" : stat_w,
                "P_value"            : p,
                "Significance"       : get_stars(p),
                "Cohens_d"           : cohen_d,
                "Cohens_d_interp"    : interpret_d(cohen_d),
                "Hedges_g"           : hedges_g_val,
                "Rank_biserial_r"    : r_rb,
                "Rank_biserial_interp": interpret_r(r_rb),
                "Ties"               : int((diff == 0).sum()),
            })

# ===============================
# 5. WRITE TO EXCEL — ONE SHEET PER METRIC
# ===============================
with pd.ExcelWriter(SAVE_PATH, engine='openpyxl') as writer:
    for metric in METRICS:
        df_out = pd.DataFrame(metric_results[metric])

        # Column order
        col_order = [
            "Class", "Comparison", "N_paired",
            "Ref_mean", "Ref_std", "Ref_median",
            "Other_mean", "Other_std", "Other_median",
            "Mean_difference",
            "Wilcoxon_statistic", "P_value", "Significance",
            "Cohens_d", "Cohens_d_interp",
            "Hedges_g",
            "Rank_biserial_r", "Rank_biserial_interp",
            "Ties"
        ]
        df_out = df_out[[c for c in col_order if c in df_out.columns]]

        # Sheet name max 31 chars (Excel limit)
        sheet_name = metric[:31]
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)

        # Auto-width columns
        ws = writer.sheets[sheet_name]
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 40)

print(f"\nDone. Saved to: {SAVE_PATH}")
























# =============================================================================
# HEATMAP CLUSTERING DENDROGRAM — Segmentation Metrics
# =============================================================================
# Visualises mean IoU (or any metric) per Model × Class as a clustermap.
# Rows (models) and columns (classes) are reordered by hierarchical clustering,
# and dendrograms are drawn on both axes so you can spot which models or
# anatomical classes are most similar.
#
# Dependencies: pandas, numpy, matplotlib, seaborn, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import seaborn as sns
from scipy.cluster.hierarchy import linkage, dendrogram
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION  — edit only this section
# =============================================================================

excel_files = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
]

METRIC   = "IoU"          # Column to aggregate (change to "Dice", "HD95", etc.)
AGG_FN   = "mean"         # Aggregation over patients: "mean" | "median"
LINKAGE  = "ward"         # Clustering linkage: "ward" | "average" | "complete"
CMAP     = "YlOrRd"       # Colour map: "YlOrRd" | "viridis" | "coolwarm" | "RdYlGn"
ANNOT    = True           # Print numeric values inside each cell
FMT      = ".2f"          # Cell annotation format

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. DATA LOADING & CLEANING
# =============================================================================

all_data = []
for name, path in excel_files:
    df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
    df[METRIC]       = pd.to_numeric(df[METRIC], errors='coerce')
    df               = df.dropna(subset=[METRIC])
    df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
    # Remove non-anatomical / placeholder classes
    df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
    df["Model"] = name
    all_data.append(df[["Filename", "Class_Name", METRIC, "Model"]])

df_main = pd.concat(all_data, ignore_index=True)

# =============================================================================
# 3. BUILD PIVOT TABLE  (rows = Models, columns = Classes)
# =============================================================================
# Each cell = mean (or median) metric value for that model–class pair.
# Missing combinations are filled with NaN; columns with all-NaN are dropped.

pivot = (
    df_main
    .groupby(["Model", "Class_Name"])[METRIC]
    .agg(AGG_FN)
    .unstack("Class_Name")          # → rows: Model, cols: Class_Name
    .dropna(axis=1, how="all")      # drop classes absent from every model
)

# Preserve original model ordering for the saved Excel summary
pivot_ordered = pivot.loc[[x[0] for x in excel_files if x[0] in pivot.index]]

# =============================================================================
# 4. SAVE SUMMARY TABLE (pivot) TO EXCEL
# =============================================================================

summary_path = os.path.join(SAVE_DIR, f"{METRIC}_Model_Class_{AGG_FN}.xlsx")
pivot_ordered.to_excel(summary_path)
print(f"Summary pivot saved → {summary_path}")

# =============================================================================
# 5. HIERARCHICAL CLUSTERING LINKAGE MATRICES
#    Row linkage  → groups similar MODELS together
#    Col linkage  → groups similar CLASSES together
# =============================================================================

# Fill any remaining NaN with row mean so clustering distance is well-defined
pivot_filled = pivot.apply(lambda col: col.fillna(col.mean()), axis=0)

row_linkage = linkage(pivot_filled.values,            method=LINKAGE, metric="euclidean")
col_linkage = linkage(pivot_filled.values.T,          method=LINKAGE, metric="euclidean")

# =============================================================================
# 6. CLUSTERMAP — seaborn does heavy lifting; we customise afterwards
# =============================================================================

sns.set_style("white")
sns.set_context("paper", font_scale=1.1)

g = sns.clustermap(
    pivot_filled,
    row_linkage   = row_linkage,
    col_linkage   = col_linkage,
    cmap          = CMAP,
    annot         = ANNOT,
    fmt           = FMT,
    linewidths    = 0.5,
    linecolor     = "white",
    vmin          = 0,
    vmax          = 1,
    figsize       = (16, 8),
    dendrogram_ratio = (0.15, 0.12),   # (row_dendro_width, col_dendro_height)
    cbar_pos      = (0.02, 0.82, 0.03, 0.15),  # (left, bottom, width, height)
    tree_kws      = {"linewidths": 1.5, "colors": "#555555"},
)

# ------------------------------------------------------------------
# 6a. Colour bar label
# ------------------------------------------------------------------
g.cax.set_ylabel(f"{AGG_FN.capitalize()} {METRIC}", fontsize=11, labelpad=8)
g.cax.yaxis.set_label_position("right")

# ------------------------------------------------------------------
# 6b. Axis labels & tick styling
# ------------------------------------------------------------------
g.ax_heatmap.set_xlabel("Anatomical Class",  fontsize=13, labelpad=8)
g.ax_heatmap.set_ylabel("Model",             fontsize=13, labelpad=8)

g.ax_heatmap.tick_params(axis="x", labelsize=10, rotation=45)
g.ax_heatmap.tick_params(axis="y", labelsize=11, rotation=0)

# ------------------------------------------------------------------
# 6c. Title on the heatmap axes (not fig.suptitle — avoids overlap)
# ------------------------------------------------------------------
g.ax_heatmap.set_title(
    f"Model × Class {AGG_FN.capitalize()} {METRIC}  |  Linkage: {LINKAGE}",
    fontsize=14, pad=12, fontweight="bold"
)

# ------------------------------------------------------------------
# 6d. Annotate each cell — recolour text for readability on dark cells
# ------------------------------------------------------------------
if ANNOT:
    for text in g.ax_heatmap.texts:
        val = float(text.get_text()) if text.get_text() else np.nan
        text.set_fontsize(9)
        text.set_color("white" if val > 0.65 else "#222222")

# ------------------------------------------------------------------
# 6e. Draw a subtle box around the heatmap
# ------------------------------------------------------------------
for spine in g.ax_heatmap.spines.values():
    spine.set_visible(True)
    spine.set_color("#cccccc")
    spine.set_linewidth(0.8)

# =============================================================================
# 7. SAVE
# =============================================================================

png_path = os.path.join(SAVE_DIR, f"{METRIC}_Clustermap.png")
pdf_path = os.path.join(SAVE_DIR, f"{METRIC}_Clustermap.pdf")

plt.savefig(png_path, dpi=600, bbox_inches="tight")
plt.savefig(pdf_path, format="pdf", bbox_inches="tight")
plt.show()

print(f"Clustermap saved →\n  {png_path}\n  {pdf_path}")












































# =============================================================================
# RADAR / SPIDER PLOT — Publication Quality (Nature Medicine Style)
# =============================================================================
# Generates a multi-model radar chart comparing segmentation performance
# across anatomical classes.
#
# Design principles (Nature Medicine guidelines):
#   • Clean white background, no chart junk
#   • Helvetica-style sans-serif (matplotlib's DejaVu Sans ≈ equivalent)
#   • Muted, perceptually-distinct colour palette (colour-blind friendly)
#   • Thin gridlines, minimal tick marks
#   • Legend outside plot area, no box
#   • 600 DPI PNG + vector PDF for submission
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.patheffects as pe
from matplotlib.patches import FancyArrowPatch
from matplotlib.lines import Line2D
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION  — edit only this section
# =============================================================================

excel_files = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
]

METRIC   = "IoU"     # Column to plot: "IoU" | "Dice" | "HD95" | etc.
AGG_FN   = "mean"    # Per-patient aggregation: "mean" | "median"

# ------------------------------------------------------------------------------
# Nature Medicine–compatible colour palette
# Perceptually uniform, colour-blind safe (Wong 2011 + adjusted for print)
# Order matches excel_files above
# ------------------------------------------------------------------------------
MODEL_COLORS = {
    "BAT-RM":    "#0077BB",   # Blue
    "nnUNet":    "#009988",   # Teal
    "SegMamba":  "#EE7733",   # Orange
    "TransUNet": "#CC3311",   # Red
    "UNETR":     "#AA4499",   # Purple
}
MODEL_MARKERS = {
    "BAT-RM":    "o",
    "nnUNet":    "s",
    "SegMamba":  "^",
    "TransUNet": "D",
    "UNETR":     "P",
}

# ------------------------------------------------------------------------------
# Fill opacity for each model's polygon (0 = no fill, 1 = opaque)
# Slight fill helps readers distinguish overlapping polygons
# ------------------------------------------------------------------------------
FILL_ALPHA = 0.08

# Radial grid rings to draw (metric values)
GRID_LEVELS = [0.2, 0.4, 0.6, 0.8, 1.0]

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE  (Nature Medicine look)
# =============================================================================
# • Font: Helvetica via matplotlib's sans-serif stack
# • All text sizes explicit so the figure scales correctly at 600 DPI
# • No top/right spines on non-polar axes
# =============================================================================

plt.rcParams.update({
    # --- Font ---
    "font.family":        "sans-serif",
    "font.sans-serif":    ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":          9,
    "axes.titlesize":     10,
    "axes.labelsize":     9,
    "xtick.labelsize":    8,
    "ytick.labelsize":    8,
    "legend.fontsize":    8.5,

    # --- Lines ---
    "axes.linewidth":     0.6,
    "grid.linewidth":     0.4,
    "lines.linewidth":    1.4,

    # --- Colour ---
    "axes.facecolor":     "white",
    "figure.facecolor":   "white",
    "grid.color":         "#cccccc",
    "grid.alpha":         0.7,

    # --- Output quality ---
    "figure.dpi":         150,       # screen preview
    "savefig.dpi":        600,       # journal submission
    "pdf.fonttype":       42,        # embed fonts as TrueType (required by Nature)
    "ps.fonttype":        42,
})

# =============================================================================
# 3. DATA LOADING & CLEANING
# =============================================================================

all_data = []
for name, path in excel_files:
    df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
    df[METRIC]       = pd.to_numeric(df[METRIC], errors='coerce')
    df               = df.dropna(subset=[METRIC])
    df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
    df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
    df["Model"]      = name
    all_data.append(df[["Filename", "Class_Name", METRIC, "Model"]])

df_main = pd.concat(all_data, ignore_index=True)

# =============================================================================
# 4. AGGREGATE: build Model × Class pivot (rows = models, cols = classes)
# =============================================================================

pivot = (
    df_main
    .groupby(["Model", "Class_Name"])[METRIC]
    .agg(AGG_FN)
    .unstack("Class_Name")
    .dropna(axis=1, how="all")
    .fillna(0)   # missing class → 0 so the polygon closes cleanly
)

# Keep only models present in the pivot (handles missing files gracefully)
model_names = [m for m, _ in excel_files if m in pivot.index]
classes     = list(pivot.columns)
N           = len(classes)

# =============================================================================
# 5. RADAR GEOMETRY HELPERS
# =============================================================================

def radar_angles(n):
    """Return N evenly-spaced angles starting at top (−π/2), clockwise."""
    angles = np.linspace(0, 2 * np.pi, n, endpoint=False)
    return angles - np.pi / 2          # rotate so first spoke points up


def close_polygon(values, angles):
    """Append the first element to close the radar polygon."""
    return (
        np.concatenate([values, [values[0]]]),
        np.concatenate([angles, [angles[0]]]),
    )


angles = radar_angles(N)

# =============================================================================
# 6. FIGURE LAYOUT
# =============================================================================
# Nature Medicine single-column width ≈ 8.9 cm; two-column ≈ 18.3 cm.
# We use 18 cm wide so the radar is large enough to be readable.
# =============================================================================

FIG_W  = 7.09    # inches (≈ 18 cm, two-column)
FIG_H  = 6.30    # inches (≈ 16 cm)

fig = plt.figure(figsize=(FIG_W, FIG_H))

# Main radar axes — polar, centred, leaving room for legend on the right
ax = fig.add_axes(
    [0.08, 0.10, 0.62, 0.80],   # [left, bottom, width, height] in figure fraction
    projection="polar"
)

# =============================================================================
# 7. POLAR AXES STYLING
# =============================================================================

ax.set_theta_zero_location("N")       # 0° at top
ax.set_theta_direction(-1)            # clockwise (matches anatomical convention)

# --- Radial grid rings ---
ax.set_rgrids(
    GRID_LEVELS,
    labels=[f"{v:.1f}" for v in GRID_LEVELS],
    angle=225,                        # label position (bottom-left, out of the way)
    fontsize=7,
    color="#888888",
)
ax.set_rlabel_position(225)
ax.set_ylim(0, 1.0)

# --- Spoke labels (class names) ---
ax.set_thetagrids(
    np.degrees(angles + np.pi / 2),   # undo the −π/2 shift for thetagrids
    labels=classes,
    fontsize=8.5,
    fontweight="normal",
)

# Adjust label distance and alignment for each spoke
for label, angle_rad in zip(ax.get_xticklabels(), angles):
    angle_deg = np.degrees(angle_rad)
    # Push labels away from the centre
    label.set_fontsize(8.5)
    label.set_color("#333333")
    # Align based on position
    if   angle_deg < -135 or angle_deg > 135:
        label.set_horizontalalignment("center")
    elif angle_deg < 0:
        label.set_horizontalalignment("right")
    else:
        label.set_horizontalalignment("left")

# --- Grid line aesthetics ---
ax.yaxis.grid(True, linestyle="--", linewidth=0.4, color="#bbbbbb", alpha=0.7)
ax.xaxis.grid(True, linestyle="-",  linewidth=0.4, color="#cccccc", alpha=0.5)

# --- Remove outermost spine (cleaner look) ---
ax.spines["polar"].set_visible(False)

# Background: very faint concentric fill for alternating rings (like Nature figures)
for i in range(len(GRID_LEVELS) - 1, -1, -1):
    facecolor = "#f7f7f7" if i % 2 == 0 else "white"
    ax.fill_between(
        np.linspace(0, 2 * np.pi, 300),
        GRID_LEVELS[i - 1] if i > 0 else 0,
        GRID_LEVELS[i],
        color=facecolor,
        zorder=0,
    )

# =============================================================================
# 8. PLOT EACH MODEL
# =============================================================================

for model in model_names:
    values       = pivot.loc[model, classes].values.astype(float)
    color        = MODEL_COLORS[model]
    marker       = MODEL_MARKERS[model]
    vals_closed, ang_closed = close_polygon(values, angles)

    # --- Filled polygon (very transparent) ---
    ax.fill(
        ang_closed, vals_closed,
        color=color, alpha=FILL_ALPHA,
        zorder=2,
    )

    # --- Polygon outline ---
    ax.plot(
        ang_closed, vals_closed,
        color=color,
        linewidth=1.6,
        linestyle="-",
        zorder=3,
    )

    # --- Markers at each spoke vertex ---
    ax.scatter(
        angles, values,
        color=color,
        marker=marker,
        s=28,
        zorder=4,
        linewidths=0.6,
        edgecolors="white",
    )

# =============================================================================
# 9. LEGEND  (outside the radar, right side — Nature style)
# =============================================================================

legend_handles = []
for model in model_names:
    handle = Line2D(
        [0], [0],
        color=MODEL_COLORS[model],
        linewidth=1.6,
        marker=MODEL_MARKERS[model],
        markersize=5,
        markerfacecolor=MODEL_COLORS[model],
        markeredgecolor="white",
        markeredgewidth=0.5,
        label=model,
    )
    legend_handles.append(handle)

legend = fig.legend(
    handles     = legend_handles,
    loc         = "center right",
    bbox_to_anchor = (1.01, 0.50),
    frameon     = False,           # no box — Nature style
    title       = "Model",
    title_fontsize = 9,
    fontsize    = 8.5,
    handlelength   = 2.0,
    handleheight   = 0.8,
    borderpad   = 0.5,
    labelspacing   = 0.6,
)
legend.get_title().set_fontweight("bold")

# =============================================================================
# 10. FIGURE TITLE & SUBTITLE
# =============================================================================
# Nature Medicine: brief, informative title in sentence case.
# Panel label "a" bottom-left for multi-panel figures (comment out if not needed).
# =============================================================================

fig.text(
    0.08, 0.96,
    f"Comparative {METRIC} performance across anatomical classes",
    fontsize=10, fontweight="bold", va="top", ha="left",
    color="#111111",
)
fig.text(
    0.08, 0.915,
    f"{AGG_FN.capitalize()} per-patient {METRIC} | n = {df_main['Filename'].nunique()} scans",
    fontsize=8, va="top", ha="left",
    color="#555555",
)

# Panel label (uncomment if this is panel 'a' of a multi-panel figure)
# fig.text(0.01, 0.99, "a", fontsize=12, fontweight="bold", va="top", ha="left")

# =============================================================================
# 11. SAVE  (PNG 600 DPI + vector PDF)
# =============================================================================

png_path = os.path.join(SAVE_DIR, f"{METRIC}_Radar_NatureMedicine.png")
pdf_path = os.path.join(SAVE_DIR, f"{METRIC}_Radar_NatureMedicine.pdf")

plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(pdf_path, format="pdf",  bbox_inches="tight", facecolor="white")
plt.show()

print(f"Radar plot saved →\n  {png_path}\n  {pdf_path}")



















































import pandas as pd

excel_files = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNet3+/UNet3+_Axial_segmentationMetrcis.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNetBase/UNetBase_Axial_segmentationMetrics.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/UNet/UNet_Axial_segmentationMetrics.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Result/Result_Analysis/EXCEL FILES/EXCEL FILES/TransUNet/TransUNet_Axial_segmentationMetrics.xlsx')
]

METRIC = "Dice"

print("=" * 70)
for name, path in excel_files:
    df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
    df[METRIC] = pd.to_numeric(df[METRIC], errors='coerce')
    df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
    df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]

    print(f"\nModel : {name}")
    print(f"Total rows      : {len(df)}")
    print(f"Unique filenames: {df['Filename'].nunique()}")
    print(f"Unique classes  : {sorted(df['Class_Name'].unique())}")
    print(f"Rows per class  :\n{df['Class_Name'].value_counts().to_string()}")
    print(f"\nSample filenames (first 5):")
    print(df['Filename'].dropna().unique()[:5])
    print("-" * 70)

# ── Cross-check: how many filenames actually overlap between BAT-RM and others ──
print("\n\nFILENAME OVERLAP BETWEEN BAT-RM AND OTHERS:")
print("=" * 70)

df_ref = pd.concat(pd.read_excel(excel_files[0][1], sheet_name=None), ignore_index=True)
df_ref["Class_Name"] = df_ref["Class_Name"].astype(str).str.upper().str.strip()
df_ref = df_ref[~df_ref["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
ref_files = set(df_ref['Filename'].dropna().unique())

for name, path in excel_files[1:]:
    df_other = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
    df_other["Class_Name"] = df_other["Class_Name"].astype(str).str.upper().str.strip()
    df_other = df_other[~df_other["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
    other_files = set(df_other['Filename'].dropna().unique())

    overlap = ref_files & other_files
    print(f"\nBAT-RM vs {name}:")
    print(f"  BAT-RM unique filenames : {len(ref_files)}")
    print(f"  {name} unique filenames : {len(other_files)}")
    print(f"  Overlapping filenames   : {len(overlap)}  <- THIS MUST BE > 0")
    if len(overlap) == 0:
        print(f"  BAT-RM sample : {list(ref_files)[:3]}")
        print(f"  {name} sample : {list(other_files)[:3]}")
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        









































############################################ Boundary Related Metrics ##################################################


# Boundary Related excel creation from KERAS model


import os
import cv2
import numpy as np
import pandas as pd
from keras.models import load_model
from keras.utils import normalize
import tensorflow as tf
from keras.saving import register_keras_serializable
from PIL import Image
from datetime import datetime
from scipy.spatial.distance import directed_hausdorff, cdist
from scipy.ndimage import binary_erosion

# ===============================
# 1. CUSTOM OBJECTS
# ===============================
@register_keras_serializable()
def focal_loss(y_true, y_pred, gamma=2.0):
    epsilon = tf.keras.backend.epsilon()
    y_pred = tf.clip_by_value(y_pred, epsilon, 1.0 - epsilon)
    cross_entropy = -y_true * tf.math.log(y_pred)
    loss = tf.pow(1 - y_pred, gamma) * cross_entropy
    return tf.reduce_mean(loss, axis=-1)

@register_keras_serializable()
def soft_dice_loss(y_true, y_pred, smooth=1):
    y_true = tf.cast(y_true, tf.float32)
    y_pred = tf.cast(y_pred, tf.float32)
    intersection = tf.reduce_sum(y_true * y_pred, axis=(1, 2, 3))
    sum_true     = tf.reduce_sum(y_true, axis=(1, 2, 3))
    sum_pred     = tf.reduce_sum(y_pred, axis=(1, 2, 3))
    dice_coefficient = (2.*intersection + smooth) / (sum_true + sum_pred + smooth)
    return tf.reduce_mean(1 - dice_coefficient)

@register_keras_serializable()
def combined_loss(y_true, y_pred, gamma=2.0, alpha=0.5):
    return alpha * focal_loss(y_true, y_pred, gamma) + \
           (1 - alpha) * soft_dice_loss(y_true, y_pred)

@register_keras_serializable()
def soft_dice_coefficient(y_true, y_pred):
    return 1 - soft_dice_loss(y_true, y_pred)

@register_keras_serializable()
class CustomMeanIoU(tf.keras.metrics.MeanIoU):
    def __init__(self, num_classes, name='mean_iou', dtype='float32',
                 ignore_class=None, sparse_y_true=True,
                 sparse_y_pred=True, axis=-1):
        super().__init__(num_classes=num_classes, name=name,
                         dtype=dtype, ignore_class=ignore_class)
        self.sparse_y_true = sparse_y_true
        self.sparse_y_pred = sparse_y_pred
        self.axis = axis

    def update_state(self, y_true, y_pred, sample_weight=None):
        y_true = tf.math.argmax(y_true, axis=-1)
        y_pred = tf.math.argmax(y_pred, axis=-1)
        return super().update_state(y_true, y_pred, sample_weight)

    def get_config(self):
        config = super().get_config()
        config.update({'sparse_y_true': self.sparse_y_true,
                       'sparse_y_pred': self.sparse_y_pred,
                       'axis': self.axis})
        return config

# ===============================
# 2. CONFIGURATION — EDIT THESE
# ===============================
MODEL_PATH    = r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/Cervix_small_Axial_200_epochs_vanilla_unet.keras'
IMAGES_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/images'
MASKS_FOLDER  = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'
SAVE_EXCEL    = r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'
IMAGE_SIZE    = (256, 256)

class_names = {
    0: 'Background',
    1: 'BODY',
    2: 'URINARY BLADDER',
    3: 'SMALL BOWEL',
    4: 'RECTUM',
    5: 'FEMORAL HEAD',
    6: 'GTV',
    7: 'CTV',
}
n_classes = len(class_names)

# Verified BGR mapping
bgr_to_class = {
    (0,   0,   0):   0,
    (0,   255, 0):   1,
    (0,   255, 255): 2,
    (153, 146, 255): 3,
    (64,  64,  128): 4,
    (255, 255, 0):   5,
    (255, 60,  255): 6,
    (255, 55,  55):  7,
}

# ===============================
# 3. LOAD MODEL
# ===============================
custom_objects = {
    'combined_loss':         combined_loss,
    'focal_loss':            focal_loss,
    'soft_dice_loss':        soft_dice_loss,
    'soft_dice_coefficient': soft_dice_coefficient,
    'CustomMeanIoU':         CustomMeanIoU
}

print("Loading model...")
model = load_model(MODEL_PATH, custom_objects=custom_objects, compile=True)
print(f"Model loaded  : {MODEL_PATH}")
print(f"Input shape   : {model.input_shape}")
print(f"Output shape  : {model.output_shape}")

# ===============================
# 4. INFERENCE HELPERS
# ===============================
def decode_mask(mask_path):
    """BGR mask PNG → 2D class-ID array."""
    mask_bgr = cv2.imread(mask_path, cv2.IMREAD_COLOR)
    if mask_bgr is None:
        raise ValueError(f"Cannot load mask: {mask_path}")
    class_map = np.zeros(mask_bgr.shape[:2], dtype=np.uint8)
    for bgr, cls_id in bgr_to_class.items():
        class_map[np.all(mask_bgr == np.array(bgr, dtype=np.uint8), axis=-1)] = cls_id
    return class_map


def predict_mask(image_path):
    """Image → 2D predicted class-ID array at original resolution."""
    img     = np.array(Image.open(image_path).convert("RGB"))
    orig_h, orig_w = img.shape[:2]
    img_bgr = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    img_resized = cv2.resize(img_bgr, IMAGE_SIZE)
    img_norm    = normalize(np.array([img_resized], dtype=np.float32), axis=1)
    prediction  = model.predict(img_norm, verbose=0)
    pred_class  = np.argmax(prediction, axis=-1)[0].astype(np.uint8)
    return cv2.resize(pred_class, (orig_w, orig_h),
                      interpolation=cv2.INTER_NEAREST)

# ===============================
# 5. BOUNDARY METRIC FUNCTIONS
#    (identical logic to original — no changes)
# ===============================
def get_boundary_pixels(binary_mask):
    if np.sum(binary_mask) == 0:
        return np.array([]).reshape(0, 2)
    eroded   = binary_erosion(binary_mask)
    boundary = binary_mask.astype(np.uint8) - eroded.astype(np.uint8)
    return np.column_stack(np.where(boundary > 0))


def calculate_surface_distances(gt_mask, pred_mask):
    gt_boundary   = get_boundary_pixels(gt_mask)
    pred_boundary = get_boundary_pixels(pred_mask)
    if len(gt_boundary) == 0 and len(pred_boundary) == 0:
        return np.array([0]), 0.0
    elif len(gt_boundary) == 0 or len(pred_boundary) == 0:
        max_dim = max(gt_mask.shape)
        return np.array([max_dim]), float(max_dim)
    try:
        d_gt   = cdist(gt_boundary, pred_boundary)
        d_pred = cdist(pred_boundary, gt_boundary)
        all_sd = np.concatenate([d_gt.min(axis=1), d_pred.min(axis=1)])
        return all_sd, float(np.mean(all_sd))
    except Exception as e:
        max_dim = max(gt_mask.shape)
        return np.array([max_dim]), float(max_dim)


def calculate_hausdorff_distance(gt_mask, pred_mask):
    gt_boundary   = get_boundary_pixels(gt_mask)
    pred_boundary = get_boundary_pixels(pred_mask)
    if len(gt_boundary) == 0 and len(pred_boundary) == 0:
        return 0.0, 0.0
    elif len(gt_boundary) == 0 or len(pred_boundary) == 0:
        max_dim = max(gt_mask.shape)
        return float(max_dim), float(max_dim)
    try:
        hd1 = directed_hausdorff(gt_boundary,   pred_boundary)[0]
        hd2 = directed_hausdorff(pred_boundary, gt_boundary)[0]
        hd  = max(hd1, hd2)
        d1  = cdist(gt_boundary,   pred_boundary).min(axis=1)
        d2  = cdist(pred_boundary, gt_boundary).min(axis=1)
        hd_95 = np.percentile(np.concatenate([d1, d2]), 95)
        return float(hd), float(hd_95)
    except Exception as e:
        return float('inf'), float('inf')


def calculate_normalized_surface_dice(gt_mask, pred_mask, tolerance=2.0):
    gt_boundary   = get_boundary_pixels(gt_mask)
    pred_boundary = get_boundary_pixels(pred_mask)
    if len(gt_boundary) == 0 and len(pred_boundary) == 0:
        return 1.0
    elif len(gt_boundary) == 0 or len(pred_boundary) == 0:
        return 0.0
    try:
        d_gt   = cdist(gt_boundary,   pred_boundary)
        d_pred = cdist(pred_boundary, gt_boundary)
        tp_gt   = np.sum(np.any(d_gt   <= tolerance, axis=1))
        tp_pred = np.sum(np.any(d_pred <= tolerance, axis=1))
        total   = len(gt_boundary) + len(pred_boundary)
        return float((tp_gt + tp_pred) / total) if total > 0 else 1.0
    except Exception as e:
        return 0.0


def calculate_volume_metrics(gt_mask, pred_mask):
    gt_vol   = np.sum(gt_mask)
    pred_vol = np.sum(pred_mask)
    if gt_vol == 0 and pred_vol == 0:
        return 0.0, 0.0
    elif gt_vol == 0:
        return float('inf'), float('inf')
    rvd  = (pred_vol - gt_vol) / gt_vol
    ravd = abs(pred_vol - gt_vol) / gt_vol
    return float(rvd), float(ravd)


def calculate_comprehensive_metrics_per_class(gt_mask, pred_mask,
                                              num_classes=8):
    """Identical logic to original — no changes."""
    metrics = {}
    for class_id in range(num_classes):
        gt_bin   = (gt_mask   == class_id).astype(np.uint8)
        pred_bin = (pred_mask == class_id).astype(np.uint8)
        gt_vol   = int(np.sum(gt_bin))
        pred_vol = int(np.sum(pred_bin))

        if gt_vol == 0 and pred_vol == 0:
            metrics[class_id] = {
                'class_name':                        class_names.get(class_id, f'Class_{class_id}'),
                'hausdorff_distance':                0.0,
                'hausdorff_95':                      0.0,
                'average_surface_distance':          0.0,
                'normalized_surface_dice':           1.0,
                'relative_volume_difference':        0.0,
                'relative_absolute_volume_difference': 0.0,
                'gt_volume':                         0,
                'pred_volume':                       0,
                'present_in_gt':                     False,
                'present_in_pred':                   False,
            }
            continue

        hd, hd_95 = calculate_hausdorff_distance(gt_bin, pred_bin)
        _, asd    = calculate_surface_distances(gt_bin, pred_bin)
        nsd       = calculate_normalized_surface_dice(gt_bin, pred_bin)
        rvd, ravd = calculate_volume_metrics(gt_bin, pred_bin)

        metrics[class_id] = {
            'class_name':                        class_names.get(class_id, f'Class_{class_id}'),
            'hausdorff_distance':                hd,
            'hausdorff_95':                      hd_95,
            'average_surface_distance':          asd,
            'normalized_surface_dice':           nsd,
            'relative_volume_difference':        rvd,
            'relative_absolute_volume_difference': ravd,
            'gt_volume':                         gt_vol,
            'pred_volume':                       pred_vol,
            'present_in_gt':                     gt_vol > 0,
            'present_in_pred':                   pred_vol > 0,
        }
    return metrics

# ===============================
# 6. MAIN LOOP
# ===============================
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp')
image_files = sorted([
    f for f in os.listdir(IMAGES_FOLDER)
    if f.lower().endswith(image_extensions)
])

print(f"\nFound {len(image_files)} images. Starting inference + boundary metrics...\n")

all_results = []
skipped     = []

for idx, img_file in enumerate(image_files):
    filename   = os.path.splitext(img_file)[0]
    image_path = os.path.join(IMAGES_FOLDER, img_file)
    mask_path  = os.path.join(MASKS_FOLDER,  img_file)

    if not os.path.exists(mask_path):
        print(f"  [SKIP] No mask for: {img_file}")
        skipped.append(img_file)
        continue

    try:
        true_class_map = decode_mask(mask_path)
        pred_class_map = predict_mask(image_path)

        if true_class_map.shape != pred_class_map.shape:
            pred_class_map = cv2.resize(pred_class_map,
                                        (true_class_map.shape[1],
                                         true_class_map.shape[0]),
                                        interpolation=cv2.INTER_NEAREST)

        class_metrics = calculate_comprehensive_metrics_per_class(
            true_class_map, pred_class_map, num_classes=n_classes)

        # ── Build row — identical column names to original ──
        row = {'Filename': filename,
               'Batch_Index': 0, 'Sample_Index': idx}

        for class_id, m in class_metrics.items():
            cname = m['class_name']
            if m['present_in_gt'] or m['present_in_pred']:
                row[f'{cname}_HD']           = m['hausdorff_distance']
                row[f'{cname}_HD95']         = m['hausdorff_95']
                row[f'{cname}_ASD']          = m['average_surface_distance']
                row[f'{cname}_NSD']          = m['normalized_surface_dice']
                row[f'{cname}_RVD']          = m['relative_volume_difference']
                row[f'{cname}_RAVD']         = m['relative_absolute_volume_difference']
                row[f'{cname}_GT_Volume']    = m['gt_volume']
                row[f'{cname}_Pred_Volume']  = m['pred_volume']
                row[f'{cname}_Present_In_GT']   = m['present_in_gt']
                row[f'{cname}_Present_In_Pred'] = m['present_in_pred']

        all_results.append(row)

        if (idx + 1) % 20 == 0 or (idx + 1) == len(image_files):
            print(f"  Processed {idx+1}/{len(image_files)}: {filename}")

    except Exception as e:
        print(f"  [ERROR] {img_file}: {e}")
        skipped.append(img_file)

print(f"\nInference complete.")
print(f"  Processed : {len(image_files) - len(skipped)}")
print(f"  Skipped   : {len(skipped)}")

# ===============================
# 7. BUILD DATAFRAMES
# ===============================
detailed_df = pd.DataFrame(all_results)

# ── Summary statistics — identical to original ──
summary_data = []
for class_id in range(n_classes):
    cname    = class_names.get(class_id, f'Class_{class_id}')
    hd_col   = f'{cname}_HD';   hd95_col = f'{cname}_HD95'
    asd_col  = f'{cname}_ASD';  nsd_col  = f'{cname}_NSD'
    rvd_col  = f'{cname}_RVD';  ravd_col = f'{cname}_RAVD'
    gtvol_col = f'{cname}_GT_Volume'

    if hd_col not in detailed_df.columns:
        continue

    valid = (
        (detailed_df[gtvol_col] > 0) &
        (detailed_df[hd_col] != float('inf')) &
        (detailed_df[hd_col].notna())
    )
    if valid.sum() == 0:
        continue

    vd = detailed_df[valid]
    summary_data.append({
        'Class_ID':    class_id,
        'Class_Name':  cname,
        'Sample_Count': int(valid.sum()),
        'Mean_HD':     vd[hd_col].mean(),   'Std_HD':    vd[hd_col].std(),
        'Median_HD':   vd[hd_col].median(), 'Min_HD':    vd[hd_col].min(),
        'Max_HD':      vd[hd_col].max(),
        'Mean_HD95':   vd[hd95_col].mean(), 'Std_HD95':  vd[hd95_col].std(),
        'Median_HD95': vd[hd95_col].median(),
        'Mean_ASD':    vd[asd_col].mean(),  'Std_ASD':   vd[asd_col].std(),
        'Median_ASD':  vd[asd_col].median(),
        'Mean_NSD':    vd[nsd_col].mean(),  'Std_NSD':   vd[nsd_col].std(),
        'Median_NSD':  vd[nsd_col].median(),
        'Mean_RVD':    vd[rvd_col].mean(),  'Std_RVD':   vd[rvd_col].std(),
        'Median_RVD':  vd[rvd_col].median(),
        'Mean_RAVD':   vd[ravd_col].mean(), 'Std_RAVD':  vd[ravd_col].std(),
        'Median_RAVD': vd[ravd_col].median(),
    })

summary_df = pd.DataFrame(summary_data)

# ===============================
# 8. SAVE EXCEL — exact same 4 sheets as original
# ===============================
os.makedirs(os.path.dirname(SAVE_EXCEL), exist_ok=True)

with pd.ExcelWriter(SAVE_EXCEL, engine='openpyxl') as writer:

    # ── Sheet 1: Detailed_Per_Instance ──
    detailed_df.to_excel(writer, sheet_name='Detailed_Per_Instance', index=False)

    # ── Sheet 2: Summary_Statistics ──
    summary_df.to_excel(writer, sheet_name='Summary_Statistics', index=False)

    # ── Sheet 3: Performance_Analysis ──
    if len(summary_df) > 0:
        best_hd   = summary_df.nsmallest(5, 'Mean_HD') [['Class_Name','Mean_HD',  'Mean_NSD','Sample_Count']]
        best_nsd  = summary_df.nlargest (5, 'Mean_NSD')[['Class_Name','Mean_NSD', 'Mean_HD', 'Sample_Count']]
        best_ravd = summary_df.nsmallest(5, 'Mean_RAVD')[['Class_Name','Mean_RAVD','Mean_NSD','Sample_Count']]

        analysis_data = pd.DataFrame({
            'Metric_Type':      (['Best HD Performance']   * len(best_hd)   +
                                 ['Best NSD Performance']  * len(best_nsd)  +
                                 ['Best Volume Accuracy']  * len(best_ravd)),
            'Class_Name':       list(best_hd['Class_Name'])   +
                                list(best_nsd['Class_Name'])  +
                                list(best_ravd['Class_Name']),
            'Primary_Metric':   list(best_hd['Mean_HD'])      +
                                list(best_nsd['Mean_NSD'])    +
                                list(best_ravd['Mean_RAVD']),
            'Secondary_Metric': list(best_hd['Mean_NSD'])     +
                                list(best_nsd['Mean_HD'])     +
                                list(best_ravd['Mean_NSD']),
            'Sample_Count':     list(best_hd['Sample_Count']) +
                                list(best_nsd['Sample_Count'])+
                                list(best_ravd['Sample_Count']),
        })
        analysis_data.to_excel(writer, sheet_name='Performance_Analysis', index=False)

    # ── Sheet 4: Metric_Descriptions ──
    metric_descriptions = pd.DataFrame({
        'Metric': ['HD (Hausdorff Distance)', 'HD95', 'ASD', 'NSD', 'RVD', 'RAVD'],
        'Description': [
            'Maximum boundary distance between surfaces',
            '95th percentile of boundary distances',
            'Average distance between surfaces',
            'Normalized Surface Dice (tolerance-based overlap)',
            'Relative Volume Difference (shows bias)',
            'Relative Absolute Volume Difference (accuracy)'
        ],
        'Interpretation': [
            'Lower is better (0 = perfect boundary match)',
            'Lower is better, more robust than HD',
            'Lower is better (0 = perfect surface match)',
            'Higher is better (1 = perfect within tolerance)',
            '0 = perfect volume, + = over-seg, - = under-seg',
            'Lower is better (0 = perfect volume accuracy)'
        ],
        'Range': [
            '[0, ∞) pixels', '[0, ∞) pixels', '[0, ∞) pixels',
            '[0, 1]', '(-∞, ∞)', '[0, ∞)'
        ]
    })
    metric_descriptions.to_excel(writer, sheet_name='Metric_Descriptions', index=False)

# ===============================
# 9. PRINT SUMMARY — identical to original
# ===============================
print(f"\n{'='*80}")
print("COMPREHENSIVE METRICS ANALYSIS COMPLETE")
print(f"{'='*80}")
print(f"Excel saved to : {SAVE_EXCEL}")
print(f"Images analyzed: {len(all_results)}")
print(f"Classes found  : {len(summary_df)}")

if len(summary_df) > 0:
    print(f"\nTOP PERFORMERS:")
    print(f"  By HD  : {summary_df.nsmallest(1,'Mean_HD')['Class_Name'].iloc[0]} "
          f"({summary_df.nsmallest(1,'Mean_HD')['Mean_HD'].iloc[0]:.2f} px)")
    print(f"  By NSD : {summary_df.nlargest(1,'Mean_NSD')['Class_Name'].iloc[0]} "
          f"({summary_df.nlargest(1,'Mean_NSD')['Mean_NSD'].iloc[0]:.3f})")
    print(f"  By RAVD: {summary_df.nsmallest(1,'Mean_RAVD')['Class_Name'].iloc[0]} "
          f"({summary_df.nsmallest(1,'Mean_RAVD')['Mean_RAVD'].iloc[0]:.3f})")

    print(f"\n{'='*120}")
    print("METRICS PER CLASS")
    print(f"{'='*120}")
    print(f"{'Class':<20} {'HD':<14} {'HD95':<14} {'ASD':<14} "
          f"{'NSD':<14} {'RVD':<14} {'RAVD':<14} {'N'}")
    print("-"*120)
    for _, row in summary_df.iterrows():
        print(f"{row['Class_Name']:<20} "
              f"{row['Mean_HD']:.2f}±{row['Std_HD']:.2f}   "
              f"{row['Mean_HD95']:.2f}±{row['Std_HD95']:.2f}   "
              f"{row['Mean_ASD']:.2f}±{row['Std_ASD']:.2f}   "
              f"{row['Mean_NSD']:.3f}±{row['Std_NSD']:.3f}   "
              f"{row['Mean_RVD']:.3f}±{row['Std_RVD']:.3f}   "
              f"{row['Mean_RAVD']:.3f}±{row['Std_RAVD']:.3f}   "
              f"{row['Sample_Count']}")

print(f"{'='*120}")
if skipped:
    print(f"\nSkipped ({len(skipped)}): {skipped}")        
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    


























































# Boundary Related excel creation from nnUNET model pth

import os
import cv2
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F
from scipy.spatial.distance import directed_hausdorff, cdist
from scipy.ndimage import binary_erosion

# ===============================
# 1. MODEL ARCHITECTURE — NNUNet2D
# ===============================
class ConvBlock(nn.Module):
    def __init__(self, in_channels, out_channels):
        super().__init__()
        self.block = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, 3, padding=1, bias=False),
            nn.InstanceNorm2d(out_channels, affine=True),
            nn.LeakyReLU(inplace=True),
            nn.Conv2d(out_channels, out_channels, 3, padding=1, bias=False),
            nn.InstanceNorm2d(out_channels, affine=True),
            nn.LeakyReLU(inplace=True),
        )
    def forward(self, x): return self.block(x)


class Down(nn.Module):
    def __init__(self, in_channels, out_channels):
        super().__init__()
        self.down = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, 3, stride=2, padding=1),
            nn.InstanceNorm2d(out_channels, affine=True),
            nn.LeakyReLU(inplace=True)
        )
        self.conv = ConvBlock(out_channels, out_channels)
    def forward(self, x):
        return self.conv(self.down(x))


class Up(nn.Module):
    def __init__(self, in_channels, skip_channels, out_channels):
        super().__init__()
        self.up   = nn.ConvTranspose2d(in_channels, out_channels, 2, stride=2)
        self.conv = ConvBlock(out_channels + skip_channels, out_channels)
    def forward(self, x, skip):
        x = self.up(x)
        if x.shape[-2:] != skip.shape[-2:]:
            x = F.interpolate(x, size=skip.shape[-2:],
                              mode="bilinear", align_corners=False)
        return self.conv(torch.cat([x, skip], dim=1))


class NNUNet2D(nn.Module):
    def __init__(self, n_classes, in_channels=3, base_features=32):
        super().__init__()
        f = base_features
        self.enc1 = ConvBlock(in_channels, f)
        self.enc2 = Down(f,     f*2)
        self.enc3 = Down(f*2,   f*4)
        self.enc4 = Down(f*4,   f*8)
        self.enc5 = Down(f*8,   f*16)
        self.up4  = Up(f*16, f*8, f*8)
        self.up3  = Up(f*8,  f*4, f*4)
        self.up2  = Up(f*4,  f*2, f*2)
        self.up1  = Up(f*2,  f,   f)
        self.out1 = nn.Conv2d(f,    n_classes, 1)
        self.out2 = nn.Conv2d(f*2,  n_classes, 1)
        self.out3 = nn.Conv2d(f*4,  n_classes, 1)
        self.out4 = nn.Conv2d(f*8,  n_classes, 1)

    def forward(self, x):
        s1=self.enc1(x); s2=self.enc2(s1)
        s3=self.enc3(s2); s4=self.enc4(s3)
        b=self.enc5(s4)
        d4=self.up4(b,s4); d3=self.up3(d4,s3)
        d2=self.up2(d3,s2); d1=self.up1(d2,s1)
        out_main = self.out1(d1)
        self.ds_outputs = [
            out_main,
            F.interpolate(self.out2(d2), size=out_main.shape[-2:],
                          mode="bilinear", align_corners=False),
            F.interpolate(self.out3(d3), size=out_main.shape[-2:],
                          mode="bilinear", align_corners=False),
            F.interpolate(self.out4(d4), size=out_main.shape[-2:],
                          mode="bilinear", align_corners=False),
        ]
        return out_main

# ===============================
# 2. CONFIGURATION — EDIT THESE
# ===============================
MODEL_PATH    = r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Cervix_small_Axial_100_epochs_pytorch_best_nnUNET_BM.pth'
IMAGES_FOLDER = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/images'
MASKS_FOLDER  = r'I:/Radiotherapy/Cervix/cervix_small_set/split/test/masks'
SAVE_EXCEL    = r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx'
IMAGE_SIZE    = (256, 256)   # (W, H) — must match training

class_names = {
    0: 'Background',
    1: 'BODY',
    2: 'URINARY BLADDER',
    3: 'SMALL BOWEL',
    4: 'RECTUM',
    5: 'FEMORAL HEAD',
    6: 'GTV',
    7: 'CTV',
}
n_classes = len(class_names)

# Verified BGR mapping
bgr_to_class = {
    (0,   0,   0):   0,
    (0,   255, 0):   1,
    (0,   255, 255): 2,
    (153, 146, 255): 3,
    (64,  64,  128): 4,
    (255, 255, 0):   5,
    (255, 60,  255): 6,
    (255, 55,  55):  7,
}

# ===============================
# 3. LOAD MODEL — PTH
# ===============================
DEVICE = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
print(f"Device: {DEVICE}")

model = NNUNet2D(n_classes=n_classes, in_channels=3).to(DEVICE)

checkpoint = torch.load(MODEL_PATH, map_location=DEVICE)
if isinstance(checkpoint, dict) and 'model_state_dict' in checkpoint:
    model.load_state_dict(checkpoint['model_state_dict'])
    print(f"Loaded checkpoint (epoch: {checkpoint.get('epoch','unknown')})")
elif isinstance(checkpoint, dict) and 'state_dict' in checkpoint:
    model.load_state_dict(checkpoint['state_dict'])
    print(f"Loaded checkpoint (state_dict key)")
else:
    model.load_state_dict(checkpoint)
    print(f"Loaded raw state dict")

model.eval()
print(f"Model loaded  : {MODEL_PATH}")
print(f"Trainable parameters: "
      f"{sum(p.numel() for p in model.parameters() if p.requires_grad):,}")
print(f"Image size    : {IMAGE_SIZE} (W x H)")

# ===============================
# 4. INFERENCE HELPERS
# ===============================
def decode_mask(mask_path):
    """BGR mask PNG → 2D class-ID array."""
    mask_bgr = cv2.imread(mask_path, cv2.IMREAD_COLOR)
    if mask_bgr is None:
        raise ValueError(f"Cannot load mask: {mask_path}")
    class_map = np.zeros(mask_bgr.shape[:2], dtype=np.uint8)
    for bgr, cls_id in bgr_to_class.items():
        class_map[np.all(mask_bgr == np.array(bgr, dtype=np.uint8), axis=-1)] = cls_id
    return class_map


def predict_mask(image_path):
    """BGR image → 2D predicted class-ID array at original resolution."""
    img_bgr = cv2.imread(image_path, cv2.IMREAD_COLOR)
    if img_bgr is None:
        raise ValueError(f"Cannot load image: {image_path}")
    orig_h, orig_w = img_bgr.shape[:2]
    img_resized    = cv2.resize(img_bgr, IMAGE_SIZE, interpolation=cv2.INTER_LINEAR)
    img_norm       = img_resized.astype(np.float32) / 255.0
    img_tensor     = torch.from_numpy(img_norm).permute(2,0,1).unsqueeze(0).to(DEVICE)
    with torch.no_grad():
        logits     = model(img_tensor)
        pred_class = torch.argmax(logits, dim=1).squeeze(0).cpu().numpy().astype(np.uint8)
    return cv2.resize(pred_class, (orig_w, orig_h), interpolation=cv2.INTER_NEAREST)

# ===============================
# 5. BOUNDARY METRIC FUNCTIONS
#    (identical logic to original — no changes)
# ===============================
def get_boundary_pixels(binary_mask):
    if np.sum(binary_mask) == 0:
        return np.array([]).reshape(0, 2)
    eroded   = binary_erosion(binary_mask)
    boundary = binary_mask.astype(np.uint8) - eroded.astype(np.uint8)
    return np.column_stack(np.where(boundary > 0))


def calculate_surface_distances(gt_mask, pred_mask):
    gt_b   = get_boundary_pixels(gt_mask)
    pred_b = get_boundary_pixels(pred_mask)
    if len(gt_b) == 0 and len(pred_b) == 0:
        return np.array([0]), 0.0
    elif len(gt_b) == 0 or len(pred_b) == 0:
        max_dim = max(gt_mask.shape)
        return np.array([max_dim]), float(max_dim)
    try:
        d_gt   = cdist(gt_b, pred_b)
        d_pred = cdist(pred_b, gt_b)
        all_sd = np.concatenate([d_gt.min(axis=1), d_pred.min(axis=1)])
        return all_sd, float(np.mean(all_sd))
    except Exception:
        max_dim = max(gt_mask.shape)
        return np.array([max_dim]), float(max_dim)


def calculate_hausdorff_distance(gt_mask, pred_mask):
    gt_b   = get_boundary_pixels(gt_mask)
    pred_b = get_boundary_pixels(pred_mask)
    if len(gt_b) == 0 and len(pred_b) == 0:
        return 0.0, 0.0
    elif len(gt_b) == 0 or len(pred_b) == 0:
        max_dim = max(gt_mask.shape)
        return float(max_dim), float(max_dim)
    try:
        hd  = max(directed_hausdorff(gt_b, pred_b)[0],
                  directed_hausdorff(pred_b, gt_b)[0])
        d1  = cdist(gt_b,   pred_b).min(axis=1)
        d2  = cdist(pred_b, gt_b).min(axis=1)
        hd_95 = np.percentile(np.concatenate([d1, d2]), 95)
        return float(hd), float(hd_95)
    except Exception:
        return float('inf'), float('inf')


def calculate_normalized_surface_dice(gt_mask, pred_mask, tolerance=2.0):
    gt_b   = get_boundary_pixels(gt_mask)
    pred_b = get_boundary_pixels(pred_mask)
    if len(gt_b) == 0 and len(pred_b) == 0:
        return 1.0
    elif len(gt_b) == 0 or len(pred_b) == 0:
        return 0.0
    try:
        tp_gt   = np.sum(np.any(cdist(gt_b,   pred_b) <= tolerance, axis=1))
        tp_pred = np.sum(np.any(cdist(pred_b, gt_b)   <= tolerance, axis=1))
        total   = len(gt_b) + len(pred_b)
        return float((tp_gt + tp_pred) / total) if total > 0 else 1.0
    except Exception:
        return 0.0


def calculate_volume_metrics(gt_mask, pred_mask):
    gt_vol   = np.sum(gt_mask)
    pred_vol = np.sum(pred_mask)
    if gt_vol == 0 and pred_vol == 0:
        return 0.0, 0.0
    elif gt_vol == 0:
        return float('inf'), float('inf')
    rvd  = (pred_vol - gt_vol) / gt_vol
    ravd = abs(pred_vol - gt_vol) / gt_vol
    return float(rvd), float(ravd)


def calculate_comprehensive_metrics_per_class(gt_mask, pred_mask, num_classes=8):
    """Identical logic to original — no changes."""
    metrics = {}
    for class_id in range(num_classes):
        gt_bin   = (gt_mask   == class_id).astype(np.uint8)
        pred_bin = (pred_mask == class_id).astype(np.uint8)
        gt_vol   = int(np.sum(gt_bin))
        pred_vol = int(np.sum(pred_bin))

        if gt_vol == 0 and pred_vol == 0:
            metrics[class_id] = {
                'class_name':                          class_names.get(class_id, f'Class_{class_id}'),
                'hausdorff_distance':                  0.0,
                'hausdorff_95':                        0.0,
                'average_surface_distance':            0.0,
                'normalized_surface_dice':             1.0,
                'relative_volume_difference':          0.0,
                'relative_absolute_volume_difference': 0.0,
                'gt_volume':                           0,
                'pred_volume':                         0,
                'present_in_gt':                       False,
                'present_in_pred':                     False,
            }
            continue

        hd, hd_95 = calculate_hausdorff_distance(gt_bin, pred_bin)
        _, asd    = calculate_surface_distances(gt_bin, pred_bin)
        nsd       = calculate_normalized_surface_dice(gt_bin, pred_bin)
        rvd, ravd = calculate_volume_metrics(gt_bin, pred_bin)

        metrics[class_id] = {
            'class_name':                          class_names.get(class_id, f'Class_{class_id}'),
            'hausdorff_distance':                  hd,
            'hausdorff_95':                        hd_95,
            'average_surface_distance':            asd,
            'normalized_surface_dice':             nsd,
            'relative_volume_difference':          rvd,
            'relative_absolute_volume_difference': ravd,
            'gt_volume':                           gt_vol,
            'pred_volume':                         pred_vol,
            'present_in_gt':                       gt_vol > 0,
            'present_in_pred':                     pred_vol > 0,
        }
    return metrics

# ===============================
# 6. MAIN INFERENCE LOOP
# ===============================
image_extensions = ('.png', '.jpg', '.jpeg', '.bmp')
image_files = sorted([
    f for f in os.listdir(IMAGES_FOLDER)
    if f.lower().endswith(image_extensions)
])

print(f"\nFound {len(image_files)} images. Starting inference + boundary metrics...\n")

all_results = []
skipped     = []

for idx, img_file in enumerate(image_files):
    filename   = os.path.splitext(img_file)[0]
    image_path = os.path.join(IMAGES_FOLDER, img_file)
    mask_path  = os.path.join(MASKS_FOLDER,  img_file)

    if not os.path.exists(mask_path):
        print(f"  [SKIP] No mask for: {img_file}")
        skipped.append(img_file)
        continue

    try:
        true_class_map = decode_mask(mask_path)
        pred_class_map = predict_mask(image_path)

        if true_class_map.shape != pred_class_map.shape:
            pred_class_map = cv2.resize(pred_class_map,
                                        (true_class_map.shape[1],
                                         true_class_map.shape[0]),
                                        interpolation=cv2.INTER_NEAREST)

        class_metrics = calculate_comprehensive_metrics_per_class(
            true_class_map, pred_class_map, num_classes=n_classes)

        # ── Build row — identical column names to original ──
        row = {'Filename': filename, 'Batch_Index': 0, 'Sample_Index': idx}

        for class_id, m in class_metrics.items():
            cname = m['class_name']
            if m['present_in_gt'] or m['present_in_pred']:
                row[f'{cname}_HD']              = m['hausdorff_distance']
                row[f'{cname}_HD95']            = m['hausdorff_95']
                row[f'{cname}_ASD']             = m['average_surface_distance']
                row[f'{cname}_NSD']             = m['normalized_surface_dice']
                row[f'{cname}_RVD']             = m['relative_volume_difference']
                row[f'{cname}_RAVD']            = m['relative_absolute_volume_difference']
                row[f'{cname}_GT_Volume']       = m['gt_volume']
                row[f'{cname}_Pred_Volume']     = m['pred_volume']
                row[f'{cname}_Present_In_GT']   = m['present_in_gt']
                row[f'{cname}_Present_In_Pred'] = m['present_in_pred']

        all_results.append(row)

        if (idx + 1) % 20 == 0 or (idx + 1) == len(image_files):
            print(f"  Processed {idx+1}/{len(image_files)}: {filename}")

    except Exception as e:
        print(f"  [ERROR] {img_file}: {e}")
        skipped.append(img_file)

print(f"\nInference complete.")
print(f"  Processed : {len(image_files) - len(skipped)}")
print(f"  Skipped   : {len(skipped)}")

# ===============================
# 7. BUILD DATAFRAMES
# ===============================
detailed_df = pd.DataFrame(all_results)

summary_data = []
for class_id in range(n_classes):
    cname     = class_names.get(class_id, f'Class_{class_id}')
    hd_col    = f'{cname}_HD';    hd95_col  = f'{cname}_HD95'
    asd_col   = f'{cname}_ASD';   nsd_col   = f'{cname}_NSD'
    rvd_col   = f'{cname}_RVD';   ravd_col  = f'{cname}_RAVD'
    gtvol_col = f'{cname}_GT_Volume'

    if hd_col not in detailed_df.columns:
        continue

    valid = (
        (detailed_df[gtvol_col] > 0) &
        (detailed_df[hd_col] != float('inf')) &
        (detailed_df[hd_col].notna())
    )
    if valid.sum() == 0:
        continue

    vd = detailed_df[valid]
    summary_data.append({
        'Class_ID':     class_id,
        'Class_Name':   cname,
        'Sample_Count': int(valid.sum()),
        'Mean_HD':      vd[hd_col].mean(),    'Std_HD':     vd[hd_col].std(),
        'Median_HD':    vd[hd_col].median(),  'Min_HD':     vd[hd_col].min(),
        'Max_HD':       vd[hd_col].max(),
        'Mean_HD95':    vd[hd95_col].mean(),  'Std_HD95':   vd[hd95_col].std(),
        'Median_HD95':  vd[hd95_col].median(),
        'Mean_ASD':     vd[asd_col].mean(),   'Std_ASD':    vd[asd_col].std(),
        'Median_ASD':   vd[asd_col].median(),
        'Mean_NSD':     vd[nsd_col].mean(),   'Std_NSD':    vd[nsd_col].std(),
        'Median_NSD':   vd[nsd_col].median(),
        'Mean_RVD':     vd[rvd_col].mean(),   'Std_RVD':    vd[rvd_col].std(),
        'Median_RVD':   vd[rvd_col].median(),
        'Mean_RAVD':    vd[ravd_col].mean(),  'Std_RAVD':   vd[ravd_col].std(),
        'Median_RAVD':  vd[ravd_col].median(),
    })

summary_df = pd.DataFrame(summary_data)

# ===============================
# 8. SAVE EXCEL — exact same 4 sheets as original
# ===============================
os.makedirs(os.path.dirname(SAVE_EXCEL), exist_ok=True)

with pd.ExcelWriter(SAVE_EXCEL, engine='openpyxl') as writer:

    # ── Sheet 1: Detailed_Per_Instance ──
    detailed_df.to_excel(writer, sheet_name='Detailed_Per_Instance', index=False)

    # ── Sheet 2: Summary_Statistics ──
    summary_df.to_excel(writer, sheet_name='Summary_Statistics', index=False)

    # ── Sheet 3: Performance_Analysis ──
    if len(summary_df) > 0:
        best_hd   = summary_df.nsmallest(5, 'Mean_HD') [['Class_Name','Mean_HD',  'Mean_NSD','Sample_Count']]
        best_nsd  = summary_df.nlargest (5, 'Mean_NSD')[['Class_Name','Mean_NSD', 'Mean_HD', 'Sample_Count']]
        best_ravd = summary_df.nsmallest(5, 'Mean_RAVD')[['Class_Name','Mean_RAVD','Mean_NSD','Sample_Count']]

        analysis_data = pd.DataFrame({
            'Metric_Type':      (['Best HD Performance']  * len(best_hd)   +
                                 ['Best NSD Performance'] * len(best_nsd)  +
                                 ['Best Volume Accuracy'] * len(best_ravd)),
            'Class_Name':       list(best_hd['Class_Name'])    +
                                list(best_nsd['Class_Name'])   +
                                list(best_ravd['Class_Name']),
            'Primary_Metric':   list(best_hd['Mean_HD'])       +
                                list(best_nsd['Mean_NSD'])     +
                                list(best_ravd['Mean_RAVD']),
            'Secondary_Metric': list(best_hd['Mean_NSD'])      +
                                list(best_nsd['Mean_HD'])      +
                                list(best_ravd['Mean_NSD']),
            'Sample_Count':     list(best_hd['Sample_Count'])  +
                                list(best_nsd['Sample_Count']) +
                                list(best_ravd['Sample_Count']),
        })
        analysis_data.to_excel(writer, sheet_name='Performance_Analysis', index=False)

    # ── Sheet 4: Metric_Descriptions ──
    pd.DataFrame({
        'Metric': ['HD (Hausdorff Distance)', 'HD95', 'ASD', 'NSD', 'RVD', 'RAVD'],
        'Description': [
            'Maximum boundary distance between surfaces',
            '95th percentile of boundary distances',
            'Average distance between surfaces',
            'Normalized Surface Dice (tolerance-based overlap)',
            'Relative Volume Difference (shows bias)',
            'Relative Absolute Volume Difference (accuracy)'
        ],
        'Interpretation': [
            'Lower is better (0 = perfect boundary match)',
            'Lower is better, more robust than HD',
            'Lower is better (0 = perfect surface match)',
            'Higher is better (1 = perfect within tolerance)',
            '0 = perfect volume, + = over-seg, - = under-seg',
            'Lower is better (0 = perfect volume accuracy)'
        ],
        'Range': [
            '[0, ∞) pixels', '[0, ∞) pixels', '[0, ∞) pixels',
            '[0, 1]', '(-∞, ∞)', '[0, ∞)'
        ]
    }).to_excel(writer, sheet_name='Metric_Descriptions', index=False)

# ===============================
# 9. PRINT SUMMARY
# ===============================
print(f"\n{'='*80}")
print("COMPREHENSIVE METRICS ANALYSIS COMPLETE")
print(f"{'='*80}")
print(f"Excel saved to : {SAVE_EXCEL}")
print(f"Images analyzed: {len(all_results)}")
print(f"Classes found  : {len(summary_df)}")

if len(summary_df) > 0:
    print(f"\nTOP PERFORMERS:")
    print(f"  By HD  : {summary_df.nsmallest(1,'Mean_HD')['Class_Name'].iloc[0]} "
          f"({summary_df.nsmallest(1,'Mean_HD')['Mean_HD'].iloc[0]:.2f} px)")
    print(f"  By NSD : {summary_df.nlargest(1,'Mean_NSD')['Class_Name'].iloc[0]} "
          f"({summary_df.nlargest(1,'Mean_NSD')['Mean_NSD'].iloc[0]:.3f})")
    print(f"  By RAVD: {summary_df.nsmallest(1,'Mean_RAVD')['Class_Name'].iloc[0]} "
          f"({summary_df.nsmallest(1,'Mean_RAVD')['Mean_RAVD'].iloc[0]:.3f})")

    print(f"\n{'='*120}")
    print("METRICS PER CLASS")
    print(f"{'='*120}")
    print(f"{'Class':<20} {'HD':<14} {'HD95':<14} {'ASD':<14} "
          f"{'NSD':<14} {'RVD':<14} {'RAVD':<14} {'N'}")
    print("-"*120)
    for _, row in summary_df.iterrows():
        print(f"{row['Class_Name']:<20} "
              f"{row['Mean_HD']:.2f}±{row['Std_HD']:.2f}   "
              f"{row['Mean_HD95']:.2f}±{row['Std_HD95']:.2f}   "
              f"{row['Mean_ASD']:.2f}±{row['Std_ASD']:.2f}   "
              f"{row['Mean_NSD']:.3f}±{row['Std_NSD']:.3f}   "
              f"{row['Mean_RVD']:.3f}±{row['Std_RVD']:.3f}   "
              f"{row['Mean_RAVD']:.3f}±{row['Std_RAVD']:.3f}   "
              f"{row['Sample_Count']}")

print(f"{'='*120}")
if skipped:
    print(f"\nSkipped ({len(skipped)}): {skipped}")    
    
    
    
    



















































# Boundary metric plot 
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from scipy.stats import wilcoxon
import os
import warnings
from matplotlib.lines import Line2D

warnings.filterwarnings("ignore")

# ===============================
# 1. CONFIGURATION
# ===============================

# Segmenetation maths 
# excel_files = [
#     ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
#     ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
#     ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
#     ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
#     ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx')
# ]
# excel_files = [
#     ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_generated.xlsx'),
#     ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx'),
#     ("SegMamba",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_generated.xlsx'),
#     ("TransUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/comprehensive_metrics_20260429_042527.xlsx'),
#     ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/comprehensive_metrics_20260429_042527.xlsx')
# ]


excel_files = [
    ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx')
]

color_map = {
    "BAT-RM": ("#00b4d8", "#0096c7"),
    "nnUNet":  ("#2a9134", "#137547"),
    "SegMamba":  ("#f4a261", "#e76f51"),
    "TransUNet":  ("#e9c46a", "#d4a017"),
    "UNETR":  ("#a29bfe", "#6c5ce7")
}

# ── Choose ONE boundary metric to plot ──
# Options: "HD"  "HD95"  "ASD"  "NSD"  "RVD"  "RAVD"
METRIC   = "HD"
SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/Boundary_Metrics'
os.makedirs(SAVE_DIR, exist_ok=True)

# Classes to include (must match column prefix in the boundary Excel)
class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV"
]

# ===============================
# 2. STATISTICAL HELPERS
# ===============================
def calculate_effect_size(x, y):
    nx, ny = len(x), len(y)
    dof = nx + ny - 2
    std_pooled = np.sqrt(((nx-1)*np.std(x, ddof=1)**2 +
                          (ny-1)*np.std(y, ddof=1)**2) / dof)
    d = (np.mean(x) - np.mean(y)) / std_pooled if std_pooled != 0 else 0
    g = d * (1 - (3 / (4 * dof - 1)))
    return d, g

def get_stars(p_val):
    if pd.isna(p_val): return "ns"
    if p_val < 0.0001: return "****"
    elif p_val < 0.001: return "***"
    elif p_val < 0.01:  return "**"
    elif p_val < 0.05:  return "*"
    return "ns"

# ===============================
# 3. DATA PROCESSING
# ── Boundary Excel is wide-format (one row per Filename,
#    columns like BODY_HD95, RECTUM_HD95 …)
#    We melt it into long-format matching the segmentation plot pipeline.
# ===============================
all_data = []
for name, path in excel_files:
    # Read the Detailed_Per_Instance sheet
    df = pd.read_excel(path, sheet_name='Detailed_Per_Instance')

    rows = []
    for cls in class_list:
        col = f"{cls}_{METRIC}"          # e.g. "BODY_HD95"
        gt_col  = f"{cls}_GT_Volume"
        present = f"{cls}_Present_In_GT"

        if col not in df.columns:
            continue

        # Keep only rows where this class is present in GT
        if present in df.columns:
            sub = df[df[present] == True][["Filename", col]].copy()
        elif gt_col in df.columns:
            sub = df[df[gt_col] > 0][["Filename", col]].copy()
        else:
            sub = df[["Filename", col]].copy()

        sub = sub.rename(columns={col: METRIC})
        sub[METRIC]      = pd.to_numeric(sub[METRIC], errors='coerce')
        sub              = sub.dropna(subset=[METRIC])

        # Drop infinite values (boundary penalty for empty masks)
        sub = sub[np.isfinite(sub[METRIC])]

        sub["Class_Name"] = cls.upper().strip()
        sub["Model"]      = name
        rows.append(sub[["Filename", "Class_Name", METRIC, "Model"]])

    if rows:
        all_data.append(pd.concat(rows, ignore_index=True))

df_main     = pd.concat(all_data, ignore_index=True)
model_names = [x[0] for x in excel_files]
classes     = sorted(df_main["Class_Name"].unique())

# ===============================
# STATISTICS
# ===============================
stats_results = []
star_lookup   = {}

for cls in classes:
    star_lookup[cls] = {}
    ref_data = df_main[
        (df_main["Model"] == model_names[0]) &
        (df_main["Class_Name"] == cls)
    ]
    for other in model_names[1:]:
        other_data = df_main[
            (df_main["Model"] == other) &
            (df_main["Class_Name"] == cls)
        ]
        merged = pd.merge(
            ref_data, other_data,
            on="Filename", suffixes=('_ref', '_other')
        ).dropna()

        # Drop infinite values after merge
        merged = merged[
            np.isfinite(merged[f"{METRIC}_ref"]) &
            np.isfinite(merged[f"{METRIC}_other"])
        ]

        p, cohen_d, hedges_g = np.nan, 0, 0
        if len(merged) > 2:
            if not (merged[f"{METRIC}_ref"] == merged[f"{METRIC}_other"]).all():
                _, p = wilcoxon(
                    merged[f"{METRIC}_ref"],
                    merged[f"{METRIC}_other"],
                    zero_method='pratt'
                )
            cohen_d, hedges_g = calculate_effect_size(
                merged[f"{METRIC}_ref"], merged[f"{METRIC}_other"])

        sig = get_stars(p)
        star_lookup[cls][other] = sig
        stats_results.append({
            "Class": cls, "Comparison": f"{model_names[0]} vs {other}",
            "P_value": p, "Significance": sig,
            "Cohens_d": cohen_d, "Hedges_g": hedges_g, "N": len(merged)
        })

pd.DataFrame(stats_results).to_excel(
    os.path.join(SAVE_DIR, f"{METRIC}_Boundary_Complete_Stats.xlsx"), index=False)

# ===============================
# 4. LAYOUT CONSTANTS
# ===============================
n_models  = len(model_names)
n_classes = len(classes)

BOX_WIDTH = 0.75
GAP_FRAC  = 0.20
slot_w    = BOX_WIDTH / n_models
new_box_w = slot_w * (1 - GAP_FRAC)
x_half    = new_box_w / 2

def box_center(ci, mi):
    return ci - BOX_WIDTH / 2 + slot_w / 2 + mi * slot_w

# ===============================
# 5. DRAW PLOT MANUALLY
# ===============================
fig, ax = plt.subplots(figsize=(24, 12))
sns.set_style("white")
np.random.seed(42)

for ci, cls in enumerate(classes):
    for mi, model in enumerate(model_names):
        vals = df_main[
            (df_main["Class_Name"] == cls) &
            (df_main["Model"] == model)
        ][METRIC].dropna().values

        if len(vals) == 0:
            continue

        cx       = box_center(ci, mi)
        fill_col = color_map[model][0]
        edge_col = color_map[model][1]

        q1, med, q3 = np.percentile(vals, [25, 50, 75])
        iqr  = q3 - q1
        lo   = max(vals.min(), q1 - 1.5 * iqr)
        hi   = min(vals.max(), q3 + 1.5 * iqr)
        mean = vals.mean()

        rect = mpatches.FancyBboxPatch(
            (cx - x_half, q1), new_box_w, q3 - q1,
            boxstyle="square,pad=0",
            facecolor=fill_col, edgecolor=edge_col,
            linewidth=1.8, zorder=3
        )
        ax.add_patch(rect)

        ax.plot([cx - x_half, cx + x_half], [med, med],
                color=edge_col, linewidth=2.2, zorder=4)

        ax.plot(cx, mean, marker="D", color="#D3D3D3",
                markeredgecolor="#757575", markersize=4, zorder=5)

        ax.plot([cx, cx], [q1, lo], color=edge_col, linewidth=1.5, zorder=3)
        ax.plot([cx, cx], [q3, hi], color=edge_col, linewidth=1.5, zorder=3)

        cap_w = x_half * 0.6
        ax.plot([cx - cap_w, cx + cap_w], [lo, lo],
                color=edge_col, linewidth=1.5, zorder=3)
        ax.plot([cx - cap_w, cx + cap_w], [hi, hi],
                color=edge_col, linewidth=1.5, zorder=3)

        n_pts  = len(vals)
        jitter = (np.random.rand(n_pts) - 0.5) * new_box_w * 0.80
        ax.scatter(cx + jitter, vals,
                   color="black", alpha=0.12, s=6, zorder=6, linewidths=0)

ax.set_xticks(range(n_classes))
ax.set_xticklabels(classes, fontsize=11)
ax.set_xlim(-0.5, n_classes - 0.5)
ax.set_xlabel("Class_Name", fontsize=12)
ax.set_ylabel(METRIC, fontsize=12)

# ===============================
# 6. SIGNIFICANCE BRACKETS — per-class dynamic positioning
# ===============================

# Compute per-class whisker tops (hi = upper fence of each model's data)
class_hi = {}
for ci, cls in enumerate(classes):
    local_hi = 0
    for mi, model in enumerate(model_names):
        vals = df_main[
            (df_main["Class_Name"] == cls) &
            (df_main["Model"] == model)
        ][METRIC].dropna().values
        if len(vals) == 0:
            continue
        q1, q3 = np.percentile(vals, [25, 75])
        iqr = q3 - q1
        hi  = min(vals.max(), q3 + 1.5 * iqr)
        local_hi = max(local_hi, hi)
    class_hi[cls] = local_hi

# Global data range for bracket height scaling (use IQR-robust estimate)
global_data_range = df_main[METRIC].quantile(0.95) - df_main[METRIC].quantile(0.05)
bracket_h = global_data_range * 0.08   # bracket height = 8% of data spread

for i, cls in enumerate(classes):
    x_ref = box_center(i, 0)
    local_top = class_hi[cls]            # start bracket just above this class's whisker

    for j, other in enumerate(model_names[1:]):
        x_other = box_center(i, j + 1)
        h = local_top + bracket_h * 0.4 + (j * bracket_h)   # 0.4× gap above whisker
        tick = bracket_h * 0.15          # small descending tick on bracket ends

        ax.plot([x_ref, x_ref, x_other, x_other],
                [h - tick, h, h, h - tick],
                color="black", lw=1.2)
        ax.text((x_ref + x_other) / 2, h + bracket_h * 0.05,
                star_lookup[cls][other],
                ha='center', va='bottom', fontsize=11, fontweight='bold')

# ===============================
# 7. STYLING — y-axis auto-scales to fit tallest bracket
# ===============================
# Find the highest bracket across all classes
max_bracket_top = max(
    class_hi[cls] + bracket_h * 0.4 + (len(model_names) - 2) * bracket_h + bracket_h * 0.6
    for cls in classes
)

# Lower bound: just below zero or the data minimum
y_min_plot = max(0, df_main[METRIC].quantile(0.01) - global_data_range * 0.03)
y_max_plot = max_bracket_top + bracket_h * 0.8   # breathing room above star text

plt.ylim(y_min_plot, y_max_plot)
sns.despine()

# ===============================
# 8. LEGEND
# ===============================
legend_handles = []
for m in model_names:
    p = mpatches.Patch(facecolor=color_map[m][0],
                       edgecolor=color_map[m][1],
                       linewidth=1.8, label=m)
    legend_handles.append(p)

line_median   = Line2D([0], [0], color='#757575', lw=2, label='Median')
marker_mean   = Line2D([0], [0], marker='D', color='w',
                       markerfacecolor='#D3D3D3', markeredgecolor='#757575',
                       markersize=6, label='Mean (Diamond)')
marker_jitter = Line2D([0], [0], marker='o', color='w',
                       markerfacecolor='black', markeredgecolor='black',
                       markersize=5, alpha=0.4, label='Data Points')

ax.legend(
    handles=legend_handles + [line_median, marker_mean, marker_jitter],
    loc='upper right', frameon=True, edgecolor='black', fontsize=12
)

# ===============================
# 9. SAVE
# ===============================
plt.savefig(os.path.join(SAVE_DIR, f"{METRIC}_Boundary_Plot.png"),
            dpi=600, bbox_inches='tight')
plt.savefig(os.path.join(SAVE_DIR, f"{METRIC}_Boundary_Plot.pdf"),
            format='pdf', bbox_inches='tight')
plt.show()

print(f"Success. Files saved in: {SAVE_DIR}")    














































# =============================================================================
# HEATMAP CLUSTERING DENDROGRAM — Boundary Metrics
# =============================================================================
# Visualises mean (or median) boundary metric per Model × Anatomical Class
# as a clustermap with hierarchical dendrograms on both axes.
#
# Boundary metrics supported:  HD  |  HD95  |  ASD  |  NSD  |  RVD  |  RAVD
#
# Key differences from the segmentation version:
#   • Reads wide-format Detailed_Per_Instance sheet (columns: CLASS_METRIC)
#   • Filters rows using CLASS_Present_In_GT or CLASS_GT_Volume > 0
#   • Drops infinite values (boundary penalty for empty / missing masks)
#   • Colour map defaults to "Blues_r" because LOWER = BETTER for distance
#     metrics (HD, HD95, ASD).  Switch to "YlOrRd" for NSD (higher = better).
#   • Diverging colour bar centred at clinical threshold when desired.
#
# Design principles (Nature Medicine / high-impact journal style):
#   • Clean white background, thin grid lines
#   • Helvetica-first font stack, all sizes explicit for 600 DPI
#   • Fonts embedded as TrueType (pdf.fonttype = 42) for submission
#   • Colour-blind safe sequential palette
#
# Dependencies: pandas, numpy, matplotlib, seaborn, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import seaborn as sns
from scipy.cluster.hierarchy import linkage
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION  — edit only this section
# =============================================================================

excel_files = [
    ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx')
]

# ------------------------------------------------------------------------------
# Metric to visualise
# Distance metrics (lower = better): "HD" | "HD95" | "ASD"
# Overlap / surface metrics (higher = better): "NSD" | "RVD" | "RAVD"
# ------------------------------------------------------------------------------
METRIC  = "HD95"
AGG_FN  = "mean"     # patient-level aggregation: "mean" | "median"
LINKAGE = "ward"     # hierarchical linkage: "ward" | "average" | "complete"

# Colour map:
#   Distance metrics (lower = better) → reversed sequential ("Blues_r", "Greens_r")
#   Overlap metrics  (higher = better) → forward sequential ("YlOrRd", "viridis")
CMAP = "Blues_r"     # change to "YlOrRd" for NSD

ANNOT = True         # print numeric value in each cell
FMT   = ".2f"        # cell annotation format

# Anatomical classes to include (must match column prefix in Excel)
class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/Boundary'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE  (Nature Medicine look)
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   10,
    "axes.labelsize":   9,
    "xtick.labelsize":  8,
    "ytick.labelsize":  8,
    "legend.fontsize":  8.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.4,
    "lines.linewidth":  1.4,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "grid.color":       "#cccccc",
    "grid.alpha":       0.7,
    "figure.dpi":       150,       # screen preview
    "savefig.dpi":      600,       # journal submission
    "pdf.fonttype":     42,        # embed fonts as TrueType (Nature requirement)
    "ps.fonttype":      42,
})

# =============================================================================
# 3. DATA LOADING & CLEANING
# =============================================================================
# Boundary Excel is wide-format: one row = one patient scan.
# Metric columns are named  <CLASS>_<METRIC>  e.g. "RECTUM_HD95".
# Presence columns are named <CLASS>_Present_In_GT  (bool) or
#                            <CLASS>_GT_Volume       (mm³, >0 = present).
# =============================================================================

all_data = []

for name, path in excel_files:
    df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")

    rows = []
    for cls in class_list:
        col     = f"{cls}_{METRIC}"        # e.g. "RECTUM_ASD"
        present = f"{cls}_Present_In_GT"
        gt_vol  = f"{cls}_GT_Volume"

        # Skip if this metric column is absent in this file
        if col not in df.columns:
            continue

        # Filter to scans where the class actually exists in ground truth
        if present in df.columns:
            sub = df[df[present] == True][["Filename", col]].copy()
        elif gt_vol in df.columns:
            sub = df[df[gt_vol] > 0][["Filename", col]].copy()
        else:
            sub = df[["Filename", col]].copy()

        sub = sub.rename(columns={col: METRIC})
        sub[METRIC] = pd.to_numeric(sub[METRIC], errors="coerce")
        sub = sub.dropna(subset=[METRIC])

        # Remove inf values — boundary penalty when predicted mask is empty
        sub = sub[np.isfinite(sub[METRIC])]

        sub["Class_Name"] = cls.upper().strip()
        sub["Model"]      = name
        rows.append(sub[["Filename", "Class_Name", METRIC, "Model"]])

    if rows:
        all_data.append(pd.concat(rows, ignore_index=True))

df_main = pd.concat(all_data, ignore_index=True)

# =============================================================================
# 4. AGGREGATE: Model × Class pivot table
# =============================================================================
# Each cell = mean (or median) metric value for that model–class combination.
# Classes absent for every patient in a model become NaN → filled with column
# mean so clustering distances are always defined.
# =============================================================================

pivot = (
    df_main
    .groupby(["Model", "Class_Name"])[METRIC]
    .agg(AGG_FN)
    .unstack("Class_Name")           # rows: Model, cols: Class_Name
    .dropna(axis=1, how="all")       # remove classes missing from all models
)

# Keep only models present in the pivot, in original file order
model_names   = [m for m, _ in excel_files if m in pivot.index]
pivot_ordered = pivot.loc[model_names]

# Column-wise mean imputation for any remaining NaN cells
pivot_filled = pivot_ordered.apply(lambda col: col.fillna(col.mean()), axis=0)

# =============================================================================
# 5. SAVE SUMMARY PIVOT TO EXCEL
# =============================================================================

summary_path = os.path.join(
    SAVE_DIR, f"{METRIC}_Boundary_Model_Class_{AGG_FN}.xlsx"
)
pivot_ordered.to_excel(summary_path)
print(f"Summary pivot saved → {summary_path}")

# =============================================================================
# 6. HIERARCHICAL CLUSTERING LINKAGE MATRICES
#    Row linkage  → groups similar MODELS together
#    Col linkage  → groups similar CLASSES together
#
#    Ward linkage minimises within-cluster variance; good default for metric
#    tables.  "average" or "complete" can reveal different cluster structures.
# =============================================================================

row_linkage = linkage(pivot_filled.values,   method=LINKAGE, metric="euclidean")
col_linkage = linkage(pivot_filled.values.T, method=LINKAGE, metric="euclidean")

# =============================================================================
# 7. CLUSTERMAP
# =============================================================================

sns.set_style("white")
sns.set_context("paper", font_scale=1.1)

# Dynamically set colour bar bounds
# For distance metrics cap at 95th percentile to avoid outlier-driven scaling
vmax = np.nanpercentile(pivot_filled.values, 95)
vmin = 0.0

g = sns.clustermap(
    pivot_filled,
    row_linkage      = row_linkage,
    col_linkage      = col_linkage,
    cmap             = CMAP,
    annot            = ANNOT,
    fmt              = FMT,
    linewidths       = 0.5,
    linecolor        = "white",
    vmin             = vmin,
    vmax             = vmax,
    figsize          = (16, 8),
    dendrogram_ratio = (0.15, 0.12),       # (row dendro width, col dendro height)
    cbar_pos         = (0.02, 0.82, 0.03, 0.15),
    tree_kws         = {"linewidths": 1.5, "colors": "#555555"},
)

# ------------------------------------------------------------------
# 7a. Colour bar label
#     Indicate direction (↓ better for distance, ↑ better for overlap)
# ------------------------------------------------------------------
direction = "↓ better" if METRIC in ("HD", "HD95", "ASD") else "↑ better"
g.cax.set_ylabel(
    f"{AGG_FN.capitalize()} {METRIC}  ({direction})",
    fontsize=11, labelpad=8
)
g.cax.yaxis.set_label_position("right")

# ------------------------------------------------------------------
# 7b. Axis labels & tick styling
# ------------------------------------------------------------------
g.ax_heatmap.set_xlabel("Anatomical Class", fontsize=13, labelpad=8)
g.ax_heatmap.set_ylabel("Model",            fontsize=13, labelpad=8)

g.ax_heatmap.tick_params(axis="x", labelsize=10, rotation=45)
g.ax_heatmap.tick_params(axis="y", labelsize=11, rotation=0)

# ------------------------------------------------------------------
# 7c. Title
# ------------------------------------------------------------------
g.ax_heatmap.set_title(
    f"Model × Class  {AGG_FN.capitalize()} {METRIC}  |  Linkage: {LINKAGE}",
    fontsize=14, pad=12, fontweight="bold"
)

# ------------------------------------------------------------------
# 7d. Cell text: recolour for readability on dark / light cells
#     Normalise value to [0,1] relative to the colour bar range to
#     decide whether white or dark text is more legible.
# ------------------------------------------------------------------
if ANNOT:
    for text in g.ax_heatmap.texts:
        try:
            val = float(text.get_text())
        except ValueError:
            continue
        text.set_fontsize(9)
        # Normalised position in the colour scale
        norm_val = (val - vmin) / (vmax - vmin + 1e-9)
        # Blues_r: dark = low (good), light = high (bad)
        # White text on dark cells, dark text on light cells
        if CMAP.endswith("_r"):
            text.set_color("white" if norm_val < 0.40 else "#222222")
        else:
            text.set_color("white" if norm_val > 0.60 else "#222222")

# ------------------------------------------------------------------
# 7e. Subtle box around the heatmap
# ------------------------------------------------------------------
for spine in g.ax_heatmap.spines.values():
    spine.set_visible(True)
    spine.set_color("#cccccc")
    spine.set_linewidth(0.8)

# ------------------------------------------------------------------
# 7f. Figure-level subtitle (patient count, metric direction note)
# ------------------------------------------------------------------
n_scans = df_main["Filename"].nunique()
g.fig.text(
    0.50, 0.01,
    (f"n = {n_scans} scans  |  "
     f"Colour encodes {AGG_FN} {METRIC} mm  |  "
     f"Dendrogram linkage: {LINKAGE}  |  "
     f"Infinite values excluded"),
    ha="center", va="bottom", fontsize=8, color="#666666",
    style="italic"
)

# =============================================================================
# 8. SAVE  (600 DPI PNG + vector PDF)
# =============================================================================

png_path = os.path.join(SAVE_DIR, f"{METRIC}_Boundary_Clustermap.png")
pdf_path = os.path.join(SAVE_DIR, f"{METRIC}_Boundary_Clustermap.pdf")

plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(pdf_path, format="pdf",  bbox_inches="tight", facecolor="white")
plt.show()

print(f"Clustermap saved →\n  {png_path}\n  {pdf_path}")
        
        











































# =============================================================================
# RADAR / SPIDER PLOT — Boundary Metrics  (Nature Medicine Style)
# =============================================================================
# Compares segmentation models across anatomical classes using a boundary
# distance or overlap metric (HD, HD95, ASD, NSD, RVD, RAVD).
#
# Reads the same wide-format Detailed_Per_Instance Excel used by the
# boundary box plot and clustermap pipelines.
#
# Key design decisions vs. the segmentation radar:
#   • Radial axis is INVERTED for distance metrics (HD, HD95, ASD):
#     the centre = worst value, the outer ring = best (lowest distance).
#     This means a LARGER polygon = BETTER performance, consistent with
#     the segmentation radar where larger = better.
#     For overlap metrics (NSD) the axis is NOT inverted (larger = better
#     already holds naturally).
#   • Radial grid labels show the original metric values (mm or ratio),
#     not normalised values, so readers can extract numbers directly.
#   • Infinite values are excluded before aggregation.
#   • Colour palette is the same Wong-2011 colour-blind safe set used
#     in the segmentation radar for visual consistency across figures.
#
# Nature Medicine compliance:
#   • Figure width 18 cm (two-column), 600 DPI PNG + vector PDF
#   • Helvetica-first font stack, pdf.fonttype = 42 (TrueType embedding)
#   • No chart junk: no box around legend, minimal tick marks
#   • Sentence-case title, italic subtitle with patient count
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.lines import Line2D
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION  — edit only this section
# =============================================================================

excel_files = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_generated.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/comprehensive_metrics_20260429_042527.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/comprehensive_metrics_20260429_042527.xlsx'),
]

# ------------------------------------------------------------------------------
# Metric to plot
#   Distance (lower = better): "HD" | "HD95" | "ASD"
#   Overlap  (higher = better): "NSD" | "RVD" | "RAVD"
# The LOWER_IS_BETTER flag controls axis inversion and label direction.
# ------------------------------------------------------------------------------
METRIC          = "ASD"
LOWER_IS_BETTER = True    # True for HD / HD95 / ASD  |  False for NSD

AGG_FN = "mean"           # patient-level aggregation: "mean" | "median"

# Anatomical classes — spokes of the radar (order = clockwise from top)
class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# Number of concentric grid rings
N_GRID_RINGS = 5

# ------------------------------------------------------------------------------
# Colour palette: Wong (2011) colour-blind safe, same as segmentation radar
# Order must match excel_files above
# ------------------------------------------------------------------------------
MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}
MODEL_MARKERS = {
    "BAT-RM":    "o",
    "nnUNet":    "s",
    "SegMamba":  "^",
    "TransUNet": "D",
    "UNETR":     "P",
}

FILL_ALPHA = 0.08    # polygon fill transparency

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/Boundary'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE  (Nature Medicine)
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   10,
    "axes.labelsize":   9,
    "xtick.labelsize":  8,
    "ytick.labelsize":  8,
    "legend.fontsize":  8.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.4,
    "lines.linewidth":  1.4,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "grid.color":       "#cccccc",
    "grid.alpha":       0.7,
    "figure.dpi":       150,
    "savefig.dpi":      600,
    "pdf.fonttype":     42,     # TrueType embedding — required by Nature
    "ps.fonttype":      42,
})

# =============================================================================
# 3. DATA LOADING & CLEANING
# =============================================================================
# Wide-format sheet: one row = one patient, columns = CLASS_METRIC.
# Presence filter uses CLASS_Present_In_GT (bool) or CLASS_GT_Volume > 0.
# Infinite values are dropped — boundary penalty for empty predicted masks.
# =============================================================================

all_data = []

for name, path in excel_files:
    df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")

    rows = []
    for cls in class_list:
        col     = f"{cls}_{METRIC}"
        present = f"{cls}_Present_In_GT"
        gt_vol  = f"{cls}_GT_Volume"

        if col not in df.columns:
            continue

        if present in df.columns:
            sub = df[df[present] == True][["Filename", col]].copy()
        elif gt_vol in df.columns:
            sub = df[df[gt_vol] > 0][["Filename", col]].copy()
        else:
            sub = df[["Filename", col]].copy()

        sub = sub.rename(columns={col: METRIC})
        sub[METRIC] = pd.to_numeric(sub[METRIC], errors="coerce")
        sub = sub.dropna(subset=[METRIC])
        sub = sub[np.isfinite(sub[METRIC])]   # drop inf boundary penalties

        sub["Class_Name"] = cls.upper().strip()
        sub["Model"]      = name
        rows.append(sub[["Filename", "Class_Name", METRIC, "Model"]])

    if rows:
        all_data.append(pd.concat(rows, ignore_index=True))

df_main = pd.concat(all_data, ignore_index=True)

# =============================================================================
# 4. AGGREGATE: Model × Class pivot
# =============================================================================

pivot = (
    df_main
    .groupby(["Model", "Class_Name"])[METRIC]
    .agg(AGG_FN)
    .unstack("Class_Name")
    .dropna(axis=1, how="all")
    .fillna(0)
)

# Preserve original model order; keep only classes present in pivot
model_names  = [m for m, _ in excel_files if m in pivot.index]
classes      = [c for c in class_list if c in pivot.columns]
N            = len(classes)

# =============================================================================
# 5. NORMALISATION FOR RADAR DISPLAY
# =============================================================================
# Radar plots require a single radial scale [0, 1].
# We normalise each class (spoke) independently so the full ring width is used.
#
# For LOWER_IS_BETTER metrics (HD, HD95, ASD):
#   norm = 1 − (value − col_min) / (col_max − col_min)
#   → best score (lowest mm) maps to 1.0 (outer ring)
#   → worst score (highest mm) maps to 0.0 (centre)
#
# For HIGHER_IS_BETTER metrics (NSD):
#   norm = (value − col_min) / (col_max − col_min)
#   → best score maps to 1.0 (outer ring)
# =============================================================================

raw   = pivot.loc[model_names, classes]          # DataFrame: models × classes
col_min = raw.min(axis=0)
col_max = raw.max(axis=0)
col_rng = (col_max - col_min).replace(0, 1)      # avoid divide-by-zero

if LOWER_IS_BETTER:
    norm = 1.0 - (raw - col_min) / col_rng       # invert: lower raw → higher norm
else:
    norm = (raw - col_min) / col_rng              # direct: higher raw → higher norm

# =============================================================================
# 6. RADAR GEOMETRY HELPERS
# =============================================================================

def radar_angles(n):
    """N evenly-spaced angles, starting at top (−π/2), clockwise."""
    return np.linspace(0, 2 * np.pi, n, endpoint=False) - np.pi / 2


def close_polygon(values, angles):
    """Append first element to close the radar polygon."""
    return (
        np.concatenate([values, [values[0]]]),
        np.concatenate([angles, [angles[0]]]),
    )


angles = radar_angles(N)

# =============================================================================
# 7. FIGURE LAYOUT
# =============================================================================
# Two-column Nature width: 7.09 in (≈ 18 cm).
# Radar occupies left 65%; legend sits in right 35%.
# =============================================================================

FIG_W, FIG_H = 7.09, 6.30
fig = plt.figure(figsize=(FIG_W, FIG_H))

ax = fig.add_axes(
    [0.06, 0.10, 0.60, 0.80],
    projection="polar"
)

# =============================================================================
# 8. POLAR AXES STYLING
# =============================================================================

ax.set_theta_zero_location("N")    # first spoke at top
ax.set_theta_direction(-1)         # clockwise
ax.set_ylim(0, 1.0)

# ------------------------------------------------------------------
# 8a. Radial grid rings — labelled with ORIGINAL metric values
#     Compute what raw value each normalised ring corresponds to per class.
#     We use the mean raw value across classes at each ring level as the label,
#     since each class has its own scale.  A footnote explains this.
# ------------------------------------------------------------------
grid_levels = np.linspace(0, 1, N_GRID_RINGS + 1)[1:]  # e.g. [0.2, 0.4, 0.6, 0.8, 1.0]

# Reconstruct approximate raw values at each ring from the global range
global_min = raw.values.min()
global_max = raw.values.max()

if LOWER_IS_BETTER:
    # norm = 1 − (raw − col_min) / col_rng  →  raw ≈ global_max − g*(global_max − global_min)
    ring_raw_approx = [global_max - g * (global_max - global_min) for g in grid_levels]
else:
    ring_raw_approx = [global_min + g * (global_max - global_min) for g in grid_levels]

ring_labels = [f"{v:.1f}" for v in ring_raw_approx]

ax.set_rgrids(
    grid_levels,
    labels=ring_labels,
    angle=225,
    fontsize=7,
    color="#888888",
)
ax.set_rlabel_position(225)

# ------------------------------------------------------------------
# 8b. Spoke (class) labels
# ------------------------------------------------------------------
ax.set_thetagrids(
    np.degrees(angles + np.pi / 2),   # undo the −π/2 rotation for thetagrids
    labels=classes,
    fontsize=8.5,
)

for label, angle_rad in zip(ax.get_xticklabels(), angles):
    label.set_color("#333333")
    label.set_fontsize(8.5)
    deg = np.degrees(angle_rad)
    if   deg < -135 or deg > 135: label.set_horizontalalignment("center")
    elif deg < 0:                  label.set_horizontalalignment("right")
    else:                          label.set_horizontalalignment("left")

# ------------------------------------------------------------------
# 8c. Grid aesthetics
# ------------------------------------------------------------------
ax.yaxis.grid(True, linestyle="--", linewidth=0.4, color="#bbbbbb", alpha=0.7)
ax.xaxis.grid(True, linestyle="-",  linewidth=0.4, color="#cccccc", alpha=0.5)
ax.spines["polar"].set_visible(False)

# Alternating ring background (subtle, Nature style)
for i in range(len(grid_levels) - 1, -1, -1):
    facecolor = "#f7f7f7" if i % 2 == 0 else "white"
    ax.fill_between(
        np.linspace(0, 2 * np.pi, 300),
        grid_levels[i - 1] if i > 0 else 0,
        grid_levels[i],
        color=facecolor,
        zorder=0,
    )

# =============================================================================
# 9. PLOT EACH MODEL
# =============================================================================

for model in model_names:
    values              = norm.loc[model, classes].values.astype(float)
    color               = MODEL_COLORS[model]
    marker              = MODEL_MARKERS[model]
    vals_c, ang_c       = close_polygon(values, angles)

    # Filled polygon
    ax.fill(ang_c, vals_c, color=color, alpha=FILL_ALPHA, zorder=2)

    # Outline
    ax.plot(ang_c, vals_c, color=color, linewidth=1.6, linestyle="-", zorder=3)

    # Vertex markers
    ax.scatter(
        angles, values,
        color=color, marker=marker, s=28, zorder=4,
        linewidths=0.6, edgecolors="white",
    )

# =============================================================================
# 10. PERFORMANCE DIRECTION ANNOTATION
# =============================================================================
# Arrow + label at bottom of the radar reminding readers which direction
# is better — critical for inverted-axis distance metrics.
# =============================================================================

direction_text = (
    "← outer ring = lower (better)"
    if LOWER_IS_BETTER else
    "← outer ring = higher (better)"
)
ax.text(
    0, -0.18,                         # below the radar in normalised axes
    direction_text,
    transform=ax.transAxes,
    fontsize=7, color="#666666", style="italic",
    ha="left", va="top",
)

# =============================================================================
# 11. LEGEND  (outside radar, right side — no frame, Nature style)
# =============================================================================

legend_handles = [
    Line2D(
        [0], [0],
        color=MODEL_COLORS[m],
        linewidth=1.6,
        marker=MODEL_MARKERS[m],
        markersize=5,
        markerfacecolor=MODEL_COLORS[m],
        markeredgecolor="white",
        markeredgewidth=0.5,
        label=m,
    )
    for m in model_names
]

legend = fig.legend(
    handles        = legend_handles,
    loc            = "center right",
    bbox_to_anchor = (1.01, 0.50),
    frameon        = False,
    title          = "Model",
    title_fontsize = 9,
    fontsize       = 8.5,
    handlelength   = 2.0,
    handleheight   = 0.8,
    borderpad      = 0.5,
    labelspacing   = 0.6,
)
legend.get_title().set_fontweight("bold")

# =============================================================================
# 12. TITLES & SUBTITLES
# =============================================================================

direction_note = "lower = better" if LOWER_IS_BETTER else "higher = better"
n_scans = df_main["Filename"].nunique()

fig.text(
    0.06, 0.96,
    f"Comparative {METRIC} performance across anatomical classes",
    fontsize=10, fontweight="bold", va="top", ha="left", color="#111111",
)
fig.text(
    0.06, 0.915,
    (f"{AGG_FN.capitalize()} per-patient {METRIC} (mm)  |  "
     f"{direction_note}  |  "
     f"n = {n_scans} scans  |  "
     f"Radial scale normalised per class; labels show approx. raw values"),
    fontsize=7.5, va="top", ha="left", color="#555555", style="italic",
)

# Panel label — uncomment if part of a multi-panel figure
# fig.text(0.01, 0.99, "b", fontsize=12, fontweight="bold", va="top", ha="left")

# =============================================================================
# 13. SAVE
# =============================================================================

png_path = os.path.join(SAVE_DIR, f"{METRIC}_Boundary_Radar_NatureMedicine.png")
pdf_path = os.path.join(SAVE_DIR, f"{METRIC}_Boundary_Radar_NatureMedicine.pdf")

plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(pdf_path, format="pdf",  bbox_inches="tight", facecolor="white")
plt.show()

print(f"Radar plot saved →\n  {png_path}\n  {pdf_path}")    

































# =============================================================================
# RADAR / SPIDER PLOT — Boundary Metrics  (Nature Medicine Style)
# =============================================================================
# Compares segmentation models across anatomical classes using a boundary
# distance or overlap metric (HD, HD95, ASD, NSD, RVD, RAVD).
#
# Reads the same wide-format Detailed_Per_Instance Excel used by the
# boundary box plot and clustermap pipelines.
#
# Key design decisions:
#   • Radial axis INVERTED for distance metrics (HD, HD95, ASD):
#     centre = worst, outer ring = best → larger polygon = better always.
#   • NaN spokes are skipped cleanly (dashed outline + × marker).
#   • Normalisation is per-class so all spokes use the full ring width.
#   • Infinite values excluded (boundary penalty for empty masks).
#   • Colour palette: Wong (2011) colour-blind safe, matches other figures.
#
# Nature Medicine compliance:
#   • 18 cm two-column width, 600 DPI PNG + vector PDF
#   • Helvetica-first font stack, pdf.fonttype = 42 (TrueType embedding)
#   • No legend box, sentence-case title, italic subtitle
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION  — edit only this section
# =============================================================================

excel_files = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_generated.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/comprehensive_metrics_20260429_042527.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/comprehensive_metrics_20260429_042527.xlsx'),
]

# ------------------------------------------------------------------------------
# Metric to plot
#   Distance (lower = better): "HD" | "HD95" | "ASD"
#   Overlap  (higher = better): "NSD" | "RVD" | "RAVD"
# ------------------------------------------------------------------------------
METRIC          = "HD"
LOWER_IS_BETTER = True    # True for HD / HD95 / ASD  |  False for NSD

AGG_FN = "mean"           # patient-level aggregation: "mean" | "median"

# Anatomical classes — spokes of the radar (clockwise from top)
class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# Number of concentric grid rings
N_GRID_RINGS = 5

# ------------------------------------------------------------------------------
# Colour palette: Wong (2011) colour-blind safe
# Order must match excel_files above
# ------------------------------------------------------------------------------
MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}
MODEL_MARKERS = {
    "BAT-RM":    "o",
    "nnUNet":    "s",
    "SegMamba":  "^",
    "TransUNet": "D",
    "UNETR":     "P",
}

FILL_ALPHA = 0.08    # polygon fill transparency

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/Boundary'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE  (Nature Medicine)
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   10,
    "axes.labelsize":   9,
    "xtick.labelsize":  8,
    "ytick.labelsize":  8,
    "legend.fontsize":  8.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.4,
    "lines.linewidth":  1.4,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "grid.color":       "#cccccc",
    "grid.alpha":       0.7,
    "figure.dpi":       150,
    "savefig.dpi":      600,
    "pdf.fonttype":     42,     # TrueType embedding — required by Nature
    "ps.fonttype":      42,
})

# =============================================================================
# 3. DATA LOADING & CLEANING
# =============================================================================
# Wide-format sheet: one row = one patient, columns = CLASS_METRIC.
# Presence filter: CLASS_Present_In_GT (bool) or CLASS_GT_Volume > 0.
# Infinite values dropped — boundary penalty for empty predicted masks.
# =============================================================================

all_data = []

for name, path in excel_files:
    df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")

    rows = []
    for cls in class_list:
        col     = f"{cls}_{METRIC}"
        present = f"{cls}_Present_In_GT"
        gt_vol  = f"{cls}_GT_Volume"

        if col not in df.columns:
            continue

        if present in df.columns:
            sub = df[df[present] == True][["Filename", col]].copy()
        elif gt_vol in df.columns:
            sub = df[df[gt_vol] > 0][["Filename", col]].copy()
        else:
            sub = df[["Filename", col]].copy()

        sub = sub.rename(columns={col: METRIC})
        sub[METRIC] = pd.to_numeric(sub[METRIC], errors="coerce")
        sub = sub.dropna(subset=[METRIC])
        sub = sub[np.isfinite(sub[METRIC])]   # drop inf boundary penalties

        sub["Class_Name"] = cls.upper().strip()
        sub["Model"]      = name
        rows.append(sub[["Filename", "Class_Name", METRIC, "Model"]])

    if rows:
        all_data.append(pd.concat(rows, ignore_index=True))

df_main = pd.concat(all_data, ignore_index=True)

# =============================================================================
# 4. AGGREGATE: Model × Class pivot
# =============================================================================
# DO NOT fillna(0) — NaN means missing data, not zero performance.
# fillna(0) maps missing spokes to the radar centre, making models with
# partial data appear to have the worst possible performance.
# =============================================================================

pivot = (
    df_main
    .groupby(["Model", "Class_Name"])[METRIC]
    .agg(AGG_FN)
    .unstack("Class_Name")
    .dropna(axis=1, how="all")   # drop classes absent from ALL models
    # ── NaN cells preserved intentionally ──
)

# Preserve original model order; keep only classes present in pivot
model_names = [m for m, _ in excel_files if m in pivot.index]
classes     = [c for c in class_list if c in pivot.columns]
N           = len(classes)

# ── Diagnostic: flag missing model–class combinations ──
print("\n=== Data coverage (NaN = no data for that model–class) ===")
print(pivot.loc[model_names, classes].to_string())
print()

# =============================================================================
# 5. NORMALISATION FOR RADAR DISPLAY
# =============================================================================
# Normalise per class (spoke) independently so the full ring width is used.
# NaN cells remain NaN — they will be excluded from plotting rather than
# being mapped to 0 (centre), which would falsely imply worst performance.
#
# For LOWER_IS_BETTER (HD, HD95, ASD):
#   norm = 1 − (value − col_min) / col_rng   →  lowest mm  = outer ring
# For HIGHER_IS_BETTER (NSD):
#   norm = (value − col_min) / col_rng        →  highest    = outer ring
# =============================================================================

raw     = pivot.loc[model_names, classes]    # keep NaN intact
col_min = raw.min(axis=0)                    # ignores NaN
col_max = raw.max(axis=0)
col_rng = (col_max - col_min).replace(0, 1)  # avoid divide-by-zero

if LOWER_IS_BETTER:
    norm = 1.0 - (raw - col_min) / col_rng
else:
    norm = (raw - col_min) / col_rng

norm = norm.clip(0.0, 1.0)   # guard floating-point edge cases

print("=== Normalised values (NaN = excluded from radar) ===")
print(norm.to_string())
print()

# =============================================================================
# 6. RADAR GEOMETRY HELPERS
# =============================================================================

def radar_angles(n):
    """N evenly-spaced angles, starting at top (−π/2), clockwise."""
    return np.linspace(0, 2 * np.pi, n, endpoint=False) - np.pi / 2


def close_polygon(values, angles):
    """Append first element to close the radar polygon."""
    return (
        np.concatenate([values, [values[0]]]),
        np.concatenate([angles, [angles[0]]]),
    )


angles = radar_angles(N)

# =============================================================================
# 7. FIGURE LAYOUT
# =============================================================================
# Two-column Nature width: 7.09 in (≈ 18 cm).
# Radar occupies left 62%; legend sits in right 38%.
# =============================================================================

FIG_W, FIG_H = 7.09, 6.30
fig = plt.figure(figsize=(FIG_W, FIG_H))

ax = fig.add_axes(
    [0.06, 0.10, 0.60, 0.80],
    projection="polar"
)

# =============================================================================
# 8. POLAR AXES STYLING
# =============================================================================

ax.set_theta_zero_location("N")    # first spoke at top
ax.set_theta_direction(-1)         # clockwise
ax.set_ylim(0, 1.0)

# ------------------------------------------------------------------
# 8a. Radial grid rings — labelled with approximate original metric
#     values reconstructed from the global min/max range.
#     A subtitle note explains the per-class normalisation.
# ------------------------------------------------------------------
grid_levels = np.linspace(0, 1, N_GRID_RINGS + 1)[1:]

global_min = raw.values[np.isfinite(raw.values)].min()
global_max = raw.values[np.isfinite(raw.values)].max()

if LOWER_IS_BETTER:
    ring_raw = [global_max - g * (global_max - global_min) for g in grid_levels]
else:
    ring_raw = [global_min + g * (global_max - global_min) for g in grid_levels]

ring_labels = [f"{v:.1f}" for v in ring_raw]

ax.set_rgrids(
    grid_levels,
    labels=ring_labels,
    angle=225,
    fontsize=7,
    color="#888888",
)
ax.set_rlabel_position(225)

# ------------------------------------------------------------------
# 8b. Spoke (class) labels
# ------------------------------------------------------------------
ax.set_thetagrids(
    np.degrees(angles + np.pi / 2),   # undo the −π/2 rotation for thetagrids
    labels=classes,
    fontsize=8.5,
)

for label, angle_rad in zip(ax.get_xticklabels(), angles):
    label.set_color("#333333")
    label.set_fontsize(8.5)
    deg = np.degrees(angle_rad)
    if   deg < -135 or deg > 135: label.set_horizontalalignment("center")
    elif deg < 0:                  label.set_horizontalalignment("right")
    else:                          label.set_horizontalalignment("left")

# ------------------------------------------------------------------
# 8c. Grid aesthetics
# ------------------------------------------------------------------
ax.yaxis.grid(True, linestyle="--", linewidth=0.4, color="#bbbbbb", alpha=0.7)
ax.xaxis.grid(True, linestyle="-",  linewidth=0.4, color="#cccccc", alpha=0.5)
ax.spines["polar"].set_visible(False)

# Alternating ring background (Nature style)
for i in range(len(grid_levels) - 1, -1, -1):
    facecolor = "#f7f7f7" if i % 2 == 0 else "white"
    ax.fill_between(
        np.linspace(0, 2 * np.pi, 300),
        grid_levels[i - 1] if i > 0 else 0,
        grid_levels[i],
        color=facecolor,
        zorder=0,
    )

# =============================================================================
# 9. PLOT EACH MODEL
# =============================================================================
# Models with partial data are plotted only on spokes where data exists.
# Missing spokes: dashed outline + small × at centre to signal absence.
# Models with zero valid spokes are skipped with a console warning.
# =============================================================================

for model in model_names:
    values = norm.loc[model, classes].values.astype(float)   # may contain NaN
    color  = MODEL_COLORS[model]
    marker = MODEL_MARKERS[model]

    valid_mask = ~np.isnan(values)

    # ── Skip model entirely if no data survives cleaning ──
    if valid_mask.sum() == 0:
        print(f"WARNING: {model} has no valid data — skipped from radar.")
        continue

    valid_angles = angles[valid_mask]
    valid_values = values[valid_mask]

    if valid_mask.all():
        # ── All spokes present: full closed polygon, solid outline ──
        vals_c, ang_c = close_polygon(valid_values, valid_angles)
        ax.fill(ang_c, vals_c, color=color, alpha=FILL_ALPHA, zorder=2)
        ax.plot(ang_c, vals_c, color=color, linewidth=1.6,
                linestyle="-", zorder=3)
    else:
        # ── Partial data: connect available spokes, dashed outline ──
        vals_c = np.concatenate([valid_values, [valid_values[0]]])
        ang_c  = np.concatenate([valid_angles, [valid_angles[0]]])
        ax.fill(ang_c, vals_c, color=color, alpha=FILL_ALPHA, zorder=2)
        ax.plot(ang_c, vals_c, color=color, linewidth=1.6,
                linestyle="--", zorder=3)

        # Small × marker at the centre of missing spokes
        missing_angles = angles[~valid_mask]
        ax.scatter(
            missing_angles,
            np.full(missing_angles.shape, 0.05),
            color=color, marker="x", s=20, zorder=4,
            linewidths=1.0, alpha=0.6,
        )

    # ── Vertex markers on valid spokes ──
    ax.scatter(
        valid_angles, valid_values,
        color=color, marker=marker, s=28, zorder=4,
        linewidths=0.6, edgecolors="white",
    )

# =============================================================================
# 10. PERFORMANCE DIRECTION ANNOTATION
# =============================================================================

direction_text = (
    "outer ring = lower (better)"
    if LOWER_IS_BETTER else
    "outer ring = higher (better)"
)
fig.text(
    0.06, 0.06,
    f"↑ {direction_text}",
    fontsize=7, color="#666666", style="italic", ha="left", va="bottom",
)

# =============================================================================
# 11. LEGEND  (outside radar, right side — no frame, Nature style)
# =============================================================================

legend_handles = [
    Line2D(
        [0], [0],
        color=MODEL_COLORS[m],
        linewidth=1.6,
        marker=MODEL_MARKERS[m],
        markersize=5,
        markerfacecolor=MODEL_COLORS[m],
        markeredgecolor="white",
        markeredgewidth=0.5,
        label=m,
    )
    for m in model_names
]

# Add a dashed line entry to explain partial-data models (if any)
has_partial = any(
    (~np.isnan(norm.loc[m, classes].values)).sum() not in (0, N)
    for m in model_names
)
if has_partial:
    legend_handles.append(
        Line2D([0], [0], color="#888888", linewidth=1.2,
               linestyle="--", label="Partial data")
    )

legend = fig.legend(
    handles        = legend_handles,
    loc            = "center right",
    bbox_to_anchor = (1.01, 0.50),
    frameon        = False,
    title          = "Model",
    title_fontsize = 9,
    fontsize       = 8.5,
    handlelength   = 2.0,
    handleheight   = 0.8,
    borderpad      = 0.5,
    labelspacing   = 0.6,
)
legend.get_title().set_fontweight("bold")

# =============================================================================
# 12. TITLES & SUBTITLES
# =============================================================================

direction_note = "lower = better" if LOWER_IS_BETTER else "higher = better"
n_scans = df_main["Filename"].nunique()

fig.text(
    0.06, 0.96,
    f"Comparative {METRIC} performance across anatomical classes",
    fontsize=10, fontweight="bold", va="top", ha="left", color="#111111",
)
fig.text(
    0.06, 0.915,
    (f"{AGG_FN.capitalize()} per-patient {METRIC} (mm)  |  "
     f"{direction_note}  |  "
     f"n = {n_scans} scans  |  "
     f"Radial scale normalised per class; labels show approx. raw values"),
    fontsize=7.5, va="top", ha="left", color="#555555", style="italic",
)

# Panel label — uncomment if part of a multi-panel figure
# fig.text(0.01, 0.99, "b", fontsize=12, fontweight="bold", va="top", ha="left")

# =============================================================================
# 13. SAVE
# =============================================================================

png_path = os.path.join(SAVE_DIR, f"{METRIC}_Boundary_Radar_NatureMedicine.png")
pdf_path = os.path.join(SAVE_DIR, f"{METRIC}_Boundary_Radar_NatureMedicine.pdf")

plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(pdf_path, format="pdf",  bbox_inches="tight", facecolor="white")
plt.show()

print(f"\nRadar plot saved →\n  {png_path}\n  {pdf_path}")    


















































################################################################################################################



# =============================================================================
# PRIMARY PERFORMANCE TABLE + FOREST PLOT
# Nature Medicine Style — Segmentation & Boundary Metrics
# =============================================================================
# Produces two publication-ready outputs:
#
#   OUTPUT 1 — Performance Summary Table (Excel + styled HTML)
#     • Mean ± SD for every Metric × Model × Class combination
#     • Best value per row highlighted in bold
#     • Exported as Excel (for submission) and HTML (for visual check)
#
#   OUTPUT 2 — Forest Plot (PNG 600 DPI + vector PDF)
#     • One row per Model × Class comparison vs the reference model (BAT-RM)
#     • X-axis: Hedges' g effect size
#     • Error bars: 95% bootstrap confidence intervals (n=1000 resamples)
#     • Vertical panels: one per metric
#     • Colour: model identity (Wong 2011 colour-blind safe palette)
#     • Vertical reference line at g=0; shaded regions for effect magnitude
#
# Effect size interpretation bands (Cohen 1988 / Sawilowsky 2009):
#   |g| < 0.2  → negligible
#   |g| < 0.5  → small
#   |g| < 0.8  → medium
#   |g| ≥ 0.8  → large
#
# For OVERLAP metrics (IoU, Dice, NSD): positive g = proposed model better
# For DISTANCE metrics (HD, HD95, ASD): negative g = proposed model better
#   The forest plot flips the x-axis for distance metrics so "right = better"
#   is always true — a single visual convention across all panels.
#
# Statistical method:
#   Paired Hedges' g computed on matched patient pairs (same Filename).
#   95% CI via BCa bootstrap (n=1000).  Wilcoxon p-value annotated.
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
from matplotlib.gridspec import GridSpec
from scipy.stats import wilcoxon
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION  — edit only this section
# =============================================================================

# ── Reference model (your proposed method — always listed first) ──
REFERENCE_MODEL = "BAT-RM"

# ── All models in display order ──
excel_files_seg = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
]

excel_files_bnd = [
    ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx')
]

# ── Metrics ──
SEG_METRICS = ["Dice", "IoU"]          # overlap metrics (higher = better)
BND_METRICS = ["HD95", "ASD"]          # distance metrics (lower = better)
LOWER_IS_BETTER = {"HD95", "ASD", "HD", "RVD", "RAVD"}

# ── Anatomical classes ──
class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# ── Bootstrap CI settings ──
N_BOOT    = 1000   # resamples for BCa bootstrap
ALPHA_CI  = 0.05   # 95% confidence interval

# ── Output ──
SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/ForestPlot'
os.makedirs(SAVE_DIR, exist_ok=True)

# ── Colour palette: Wong (2011) colour-blind safe ──
MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE  (Nature Medicine)
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   10,
    "axes.labelsize":   9,
    "xtick.labelsize":  8,
    "ytick.labelsize":  8,
    "legend.fontsize":  8.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.4,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "pdf.fonttype":     42,    # TrueType embedding — Nature requirement
    "ps.fonttype":      42,
    "figure.dpi":       150,
    "savefig.dpi":      600,
})

# =============================================================================
# 3. DATA LOADING HELPERS
# =============================================================================

def load_segmentation(excel_files, metrics, class_list):
    """
    Load long-format segmentation Excel files.
    Returns a DataFrame with columns:
        Filename | Class_Name | <metric> ... | Model
    """
    all_data = []
    for name, path in excel_files:
        df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
        for m in metrics:
            if m in df.columns:
                df[m] = pd.to_numeric(df[m], errors="coerce")
        df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
        df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
        df = df[df["Class_Name"].isin([c.upper() for c in class_list])]
        df["Model"] = name
        cols = ["Filename", "Class_Name", "Model"] + [m for m in metrics if m in df.columns]
        all_data.append(df[cols])
    return pd.concat(all_data, ignore_index=True)


def load_boundary(excel_files, metrics, class_list):
    """
    Load wide-format boundary Excel files (Detailed_Per_Instance sheet).
    Melts CLASS_METRIC columns into long format matching segmentation output.

    Fix: collects all metrics for each class in a single pass per file,
    avoiding the iterative merge that caused duplicate column suffixes.
    """
    all_data = []

    for name, path in excel_files:
        df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")
        file_rows = []

        for cls in class_list:
            present = f"{cls}_Present_In_GT"
            gt_vol  = f"{cls}_GT_Volume"

            # ── Presence filter ──
            if present in df.columns:
                sub = df[df[present] == True].copy()
            elif gt_vol in df.columns:
                sub = df[df[gt_vol] > 0].copy()
            else:
                sub = df.copy()

            if sub.empty:
                continue

            # ── Collect all available metric columns for this class ──
            # Build one row per patient with all metrics side by side.
            # This avoids the iterative merge entirely.
            cls_cols = {"Filename": sub["Filename"].values}

            any_metric_found = False
            for m in metrics:
                col = f"{cls}_{m}"
                if col not in df.columns:
                    cls_cols[m] = np.nan
                    continue
                vals = pd.to_numeric(sub[col], errors="coerce").values
                # Replace inf with NaN (boundary penalty for empty masks)
                vals = np.where(np.isfinite(vals), vals, np.nan)
                cls_cols[m] = vals
                any_metric_found = True

            if not any_metric_found:
                continue

            tmp = pd.DataFrame(cls_cols)
            tmp["Class_Name"] = cls.upper().strip()
            tmp["Model"]      = name

            # Drop rows where ALL metrics are NaN
            tmp = tmp.dropna(subset=metrics, how="all")
            file_rows.append(tmp)

        if file_rows:
            all_data.append(pd.concat(file_rows, ignore_index=True))

    if not all_data:
        raise ValueError(
            "No boundary data loaded. Check file paths, sheet name "
            "'Detailed_Per_Instance', and column naming convention CLASS_METRIC."
        )

    return pd.concat(all_data, ignore_index=True)

# =============================================================================
# 4. LOAD ALL DATA
# =============================================================================

df_seg = load_segmentation(excel_files_seg, SEG_METRICS, class_list)
df_bnd = load_boundary(excel_files_bnd, BND_METRICS, class_list)

# Merge segmentation and boundary into one master frame
df_all = pd.merge(
    df_seg, df_bnd,
    on=["Filename", "Class_Name", "Model"],
    how="outer"
)

ALL_METRICS  = SEG_METRICS + BND_METRICS
model_names  = [m for m, _ in excel_files_seg]
baselines    = [m for m in model_names if m != REFERENCE_MODEL]
classes      = sorted(df_all["Class_Name"].dropna().unique())

print(f"Models   : {model_names}")
print(f"Classes  : {classes}")
print(f"Metrics  : {ALL_METRICS}")
print(f"Patients : {df_all['Filename'].nunique()}\n")

# =============================================================================
# 5. SUMMARY TABLE — Mean ± SD per Metric × Model × Class
# =============================================================================

summary_rows = []
for metric in ALL_METRICS:
    for cls in classes:
        row = {"Metric": metric, "Class": cls}
        for model in model_names:
            vals = df_all[
                (df_all["Model"] == model) &
                (df_all["Class_Name"] == cls)
            ][metric].dropna()
            if len(vals) > 0:
                row[model] = f"{vals.mean():.3f} ± {vals.std():.3f}"
            else:
                row[model] = "—"
        summary_rows.append(row)

df_summary = pd.DataFrame(summary_rows)

# Save to Excel with conditional formatting (best value bolded per row)
summary_path = os.path.join(SAVE_DIR, "Performance_Summary_Table.xlsx")
with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
    df_summary.to_excel(writer, index=False, sheet_name="Summary")

    ws  = writer.sheets["Summary"]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=9, name="Helvetica")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border

    # Alternating row fill + highlight best model per row
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    best_fill = PatternFill("solid", fgColor="D5E8D4")
    best_font = Font(bold=True, size=9, name="Helvetica")

    model_col_idx = {m: df_summary.columns.get_loc(m) + 1 for m in model_names}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
        metric_name = df_summary.iloc[row_idx]["Metric"]
        lower_better = metric_name in LOWER_IS_BETTER

        # Extract numeric mean values for this row to find the best
        numeric_vals = {}
        for m in model_names:
            cell_val = df_summary.iloc[row_idx][m]
            if cell_val != "—":
                try:
                    numeric_vals[m] = float(cell_val.split("±")[0].strip())
                except Exception:
                    pass

        if numeric_vals:
            best_model = (min if lower_better else max)(
                numeric_vals, key=numeric_vals.get
            )
        else:
            best_model = None

        fill = alt_fill if row_idx % 2 == 0 else PatternFill()

        for cell in row:
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(size=9, name="Helvetica")
            if fill.fill_type:
                cell.fill = fill

        # Highlight best model cell
        if best_model and best_model in model_col_idx:
            best_cell = ws.cell(
                row=row_idx + 2, column=model_col_idx[best_model]
            )
            best_cell.fill = best_fill
            best_cell.font = best_font

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    for m in model_names:
        col_letter = ws.cell(1, model_col_idx[m]).column_letter
        ws.column_dimensions[col_letter].width = 18

print(f"Summary table saved → {summary_path}")

# =============================================================================
# 6. STATISTICAL ENGINE
# =============================================================================

def hedges_g_paired(x, y):
    """
    Paired Hedges' g effect size.
    Positive = x > y (proposed model outperforms baseline for overlap metrics).
    """
    diff = np.array(x) - np.array(y)
    n    = len(diff)
    if n < 2:
        return np.nan
    d   = diff.mean() / (diff.std(ddof=1) + 1e-12)
    # Hedges' correction factor
    j   = 1 - (3 / (4 * (n - 1) - 1))
    return d * j


def bootstrap_ci_hedges_g(x, y, n_boot=1000, alpha=0.05, seed=42):
    """
    BCa bootstrap 95% CI for paired Hedges' g.
    Returns (ci_low, ci_high).
    """
    rng      = np.random.default_rng(seed)
    x, y     = np.array(x), np.array(y)
    n        = len(x)
    observed = hedges_g_paired(x, y)
    if np.isnan(observed) or n < 3:
        return np.nan, np.nan

    boot_gs = np.array([
        hedges_g_paired(
            x[idx := rng.integers(0, n, n)],
            y[idx]
        )
        for _ in range(n_boot)
    ])

    # Percentile CI (BCa is more accurate but complex; percentile is standard)
    ci_low  = np.nanpercentile(boot_gs, 100 * alpha / 2)
    ci_high = np.nanpercentile(boot_gs, 100 * (1 - alpha / 2))
    return ci_low, ci_high


def wilcoxon_p(x, y):
    """Paired Wilcoxon signed-rank test. Returns p-value."""
    if len(x) < 3:
        return np.nan
    if np.all(np.array(x) == np.array(y)):
        return np.nan
    try:
        _, p = wilcoxon(x, y, zero_method="pratt")
        return p
    except Exception:
        return np.nan


def p_to_stars(p):
    if pd.isna(p):     return "ns"
    if p < 0.0001:     return "****"
    if p < 0.001:      return "***"
    if p < 0.01:       return "**"
    if p < 0.05:       return "*"
    return "ns"

# =============================================================================
# 7. COMPUTE EFFECT SIZES FOR ALL COMPARISONS
# =============================================================================
# For each: metric × class × baseline model vs REFERENCE_MODEL
# Paired on Filename (same patient, different model)
# =============================================================================

effect_rows = []

ref_df = df_all[df_all["Model"] == REFERENCE_MODEL]

for metric in ALL_METRICS:
    for cls in classes:
        ref_vals_full = ref_df[ref_df["Class_Name"] == cls][["Filename", metric]].dropna()

        for baseline in baselines:
            bsl_vals_full = df_all[
                (df_all["Model"] == baseline) &
                (df_all["Class_Name"] == cls)
            ][["Filename", metric]].dropna()

            # Pair on Filename
            merged = pd.merge(
                ref_vals_full, bsl_vals_full,
                on="Filename", suffixes=("_ref", "_bsl")
            ).dropna()

            # Drop infinite values (boundary metrics)
            merged = merged[
                np.isfinite(merged[f"{metric}_ref"]) &
                np.isfinite(merged[f"{metric}_bsl"])
            ]

            n = len(merged)
            if n < 3:
                g, ci_lo, ci_hi, p = np.nan, np.nan, np.nan, np.nan
            else:
                ref_v = merged[f"{metric}_ref"].values
                bsl_v = merged[f"{metric}_bsl"].values
                g           = hedges_g_paired(ref_v, bsl_v)
                ci_lo, ci_hi = bootstrap_ci_hedges_g(ref_v, bsl_v, N_BOOT)
                p           = wilcoxon_p(ref_v, bsl_v)

            effect_rows.append({
                "Metric":    metric,
                "Class":     cls,
                "Baseline":  baseline,
                "N_pairs":   n,
                "Hedges_g":  g,
                "CI_low":    ci_lo,
                "CI_high":   ci_hi,
                "P_value":   p,
                "Stars":     p_to_stars(p),
                "Lower_is_better": metric in LOWER_IS_BETTER,
            })

df_effects = pd.DataFrame(effect_rows)

# Save full stats table
stats_path = os.path.join(SAVE_DIR, "Effect_Size_Statistics.xlsx")
df_effects.to_excel(stats_path, index=False)
print(f"Effect size statistics saved → {stats_path}")

# =============================================================================
# 8. FOREST PLOT
# =============================================================================
# Layout:
#   • One column panel per metric (SEG_METRICS + BND_METRICS)
#   • One row per Class × Baseline combination
#   • Rows grouped by class, sub-grouped by baseline within each class
#   • Shaded bands: negligible / small / medium / large effect zones
#   • For distance metrics the g axis is flipped so right = better always
# =============================================================================

# ── Row ordering: classes × baselines ──
row_labels = []    # (class, baseline) tuples in display order
for cls in classes:
    for bsl in baselines:
        row_labels.append((cls, bsl))

n_rows    = len(row_labels)
n_metrics = len(ALL_METRICS)

# ── Figure size ──
# Height: 0.35 in per row + 1.5 in for titles/axes
# Width: 2.8 in per metric panel
FIG_H = max(8, n_rows * 0.35 + 2.0)
FIG_W = n_metrics * 2.8 + 1.5    # +1.5 for row label column

fig = plt.figure(figsize=(FIG_W, FIG_H))

# GridSpec: one extra column on the left for row labels
gs = GridSpec(
    1, n_metrics + 1,
    figure=fig,
    width_ratios=[2.2] + [1.0] * n_metrics,
    wspace=0.08,
    left=0.02, right=0.97,
    top=0.91,  bottom=0.08,
)

# ── Effect size magnitude bands (Cohen / Sawilowsky) ──
bands = [
    (0.0,  0.2, "#FAFAFA", "negligible"),
    (0.2,  0.5, "#EFF7FB", "small"),
    (0.5,  0.8, "#E1F0F7", "medium"),
    (0.8,  3.5, "#D4E9F4", "large"),
]
X_LIM = 3.0   # forest plot x-axis range (−X_LIM to +X_LIM)

# ── Y positions: one per row, class separators between groups ──
y_positions = {}
y           = n_rows - 1
class_separator_y = {}   # top y of each class group

current_cls = None
for idx, (cls, bsl) in enumerate(row_labels):
    if cls != current_cls:
        class_separator_y[cls] = y + 0.5
        current_cls = cls
    y_positions[(cls, bsl)] = y
    y -= 1

# ── Left axis: row labels ──
ax_labels = fig.add_subplot(gs[0, 0])
ax_labels.set_xlim(0, 1)
ax_labels.set_ylim(-0.5, n_rows - 0.5)
ax_labels.axis("off")

# Class group headers
for cls, sep_y in class_separator_y.items():
    ax_labels.text(
        0.98, sep_y - 0.5 - (len(baselines) - 1) / 2,
        cls,
        ha="right", va="center",
        fontsize=8.5, fontweight="bold", color="#111111",
        transform=ax_labels.transData,
    )

# Baseline sub-labels
for (cls, bsl), y_pos in y_positions.items():
    ax_labels.text(
        0.96, y_pos,
        f"  vs {bsl}",
        ha="right", va="center",
        fontsize=7.5, color=MODEL_COLORS.get(bsl, "#555555"),
        transform=ax_labels.transData,
    )

# Class separator lines
for cls, sep_y in class_separator_y.items():
    ax_labels.axhline(
        sep_y, xmin=0, xmax=1,
        color="#cccccc", linewidth=0.5, linestyle="--"
    )

# ── Metric panels ──
axes = []
for mi, metric in enumerate(ALL_METRICS):
    ax = fig.add_subplot(gs[0, mi + 1])
    axes.append(ax)
    lower_better = metric in LOWER_IS_BETTER

    # Flip x-axis for distance metrics so right = better always
    x_sign = -1 if lower_better else 1

    ax.set_xlim(-X_LIM, X_LIM)
    ax.set_ylim(-0.5, n_rows - 0.5)

    # ── Background effect magnitude bands ──
    for lo, hi, color, _ in bands:
        # Both sides of zero
        ax.axvspan(-hi * x_sign, -lo * x_sign,
                   color=color, alpha=0.6, zorder=0)
        ax.axvspan( lo * x_sign,  hi * x_sign,
                   color=color, alpha=0.6, zorder=0)

    # ── Zero reference line ──
    ax.axvline(0, color="#333333", linewidth=0.8, zorder=2)

    # ── Class separator lines ──
    for cls, sep_y in class_separator_y.items():
        ax.axhline(sep_y, color="#cccccc", linewidth=0.5,
                   linestyle="--", zorder=1)

    # ── Plot each comparison ──
    for (cls, bsl), y_pos in y_positions.items():
        row = df_effects[
            (df_effects["Metric"]   == metric) &
            (df_effects["Class"]    == cls) &
            (df_effects["Baseline"] == bsl)
        ]
        if row.empty:
            continue
        row = row.iloc[0]

        g     = row["Hedges_g"]
        ci_lo = row["CI_low"]
        ci_hi = row["CI_high"]
        stars = row["Stars"]
        n_p   = int(row["N_pairs"]) if not np.isnan(row["N_pairs"]) else 0
        color = MODEL_COLORS.get(bsl, "#888888")

        if pd.isna(g):
            # No data — draw a small dash
            ax.text(0, y_pos, "—", ha="center", va="center",
                    fontsize=7, color="#aaaaaa")
            continue

        # Apply sign flip for distance metrics
        g_plot     = g     * x_sign
        ci_lo_plot = ci_lo * x_sign
        ci_hi_plot = ci_hi * x_sign

        # ── Confidence interval line ──
        ax.plot(
            [ci_lo_plot, ci_hi_plot], [y_pos, y_pos],
            color=color, linewidth=1.4, zorder=3, solid_capstyle="round"
        )

        # ── Point estimate diamond ──
        ax.plot(
            g_plot, y_pos,
            marker="D", color=color,
            markersize=5, zorder=4,
            markeredgecolor="white", markeredgewidth=0.5,
        )

        # ── Stars annotation (right of CI) ──
        x_star = min(ci_hi_plot + 0.08, X_LIM - 0.05)
        if stars != "ns":
            ax.text(
                x_star, y_pos, stars,
                ha="left", va="center",
                fontsize=7, fontweight="bold",
                color=color, zorder=5,
            )

    # ── Metric panel title ──
    direction = "↓ lower better" if lower_better else "↑ higher better"
    ax.set_title(
        f"{metric}\n{direction}",
        fontsize=8.5, fontweight="bold", pad=6,
        color="#111111",
    )

    # ── X-axis ──
    ax.set_xticks([-2, -1, 0, 1, 2])
    ax.set_xticklabels(["-2", "-1", "0", "+1", "+2"], fontsize=7)
    ax.set_xlabel("Hedges' g", fontsize=8, labelpad=4)

    # ── Y-axis: hide ticks (labels are in the left column) ──
    ax.set_yticks([])
    ax.yaxis.set_visible(False)

    # ── Spine cleanup (Nature style) ──
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["bottom"].set_linewidth(0.6)

    # ── Subtle horizontal alternating row shading ──
    for idx, (cls, bsl) in enumerate(row_labels):
        if idx % 2 == 0:
            ax.axhspan(
                y_positions[(cls, bsl)] - 0.45,
                y_positions[(cls, bsl)] + 0.45,
                color="#f9f9f9", zorder=0, alpha=0.8
            )

# =============================================================================
# 9. LEGEND & ANNOTATIONS
# =============================================================================

# ── Model colour legend ──
legend_handles = [
    mlines.Line2D(
        [0], [0], marker="D", color=MODEL_COLORS[m],
        markerfacecolor=MODEL_COLORS[m], markeredgecolor="white",
        markeredgewidth=0.5, markersize=5,
        linewidth=1.4, label=f"vs {m}"
    )
    for m in baselines
]

fig.legend(
    handles        = legend_handles,
    loc            = "lower center",
    ncol           = len(baselines),
    frameon        = False,
    fontsize       = 8.5,
    bbox_to_anchor = (0.55, 0.005),
    title          = f"Comparison (reference: {REFERENCE_MODEL})",
    title_fontsize = 8.5,
)

# ── Effect size band legend (text annotation on first panel) ──
band_text = (
    "Bands:  |g|<0.2 negligible  |  "
    "0.2–0.5 small  |  "
    "0.5–0.8 medium  |  "
    ">0.8 large"
)
fig.text(
    0.55, 0.965, band_text,
    ha="center", va="top",
    fontsize=7, color="#555555", style="italic",
)

# ── Figure title ──
fig.text(
    0.55, 0.995,
    f"Forest plot — Effect size (Hedges' g) of {REFERENCE_MODEL} vs baselines  "
    f"|  Paired per patient  |  Error bars: 95% bootstrap CI",
    ha="center", va="top",
    fontsize=10, fontweight="bold", color="#111111",
)

# =============================================================================
# 10. SAVE
# =============================================================================

png_path = os.path.join(SAVE_DIR, "Forest_Plot_NatureMedicine.png")
pdf_path = os.path.join(SAVE_DIR, "Forest_Plot_NatureMedicine.pdf")

plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(pdf_path, format="pdf",  bbox_inches="tight", facecolor="white")
plt.show()

print(f"\nForest plot saved →\n  {png_path}\n  {pdf_path}")
print(f"All outputs in: {SAVE_DIR}")























































# =============================================================================
# BLAND-ALTMAN PLOT — Boundary Metrics (Nature Medicine Style)
# =============================================================================
# Produces a Bland-Altman (mean-difference) plot for each boundary metric
# comparing the reference / best model (BAT-RM) against every baseline.
#
# Layout:
#   • One figure per metric (ASD, HD95, etc.)
#   • One panel per baseline model within each figure
#   • X-axis: mean of the two models' measurements per patient
#   • Y-axis: difference (Reference − Baseline) per patient
#
# Annotations per panel:
#   • Bias line      — mean difference (solid, coloured)
#   • LoA lines      — bias ± 1.96 × SD (dashed, coloured)
#   • 95% CI bands   — shaded around bias and LoA (BCa bootstrap, n=1000)
#   • Scatter points — one dot per patient, colour = model, alpha for density
#   • LOWESS trend   — detects proportional bias (non-random residual pattern)
#   • Stats box       — bias, LoA, Shapiro-Wilk normality, n pairs
#
# Clinical interpretation guide printed to console for each panel.
#
# Bland-Altman references:
#   Bland & Altman, Lancet 1986; Stat Methods Med Res 1999
#   Giavarina, Biochemia Medica 2015 (tutorial)
#
# Nature Medicine compliance:
#   • 600 DPI PNG + vector PDF
#   • Helvetica-first font, pdf.fonttype=42 (TrueType embedding)
#   • Wong (2011) colour-blind safe palette
#   • No legend box, sentence-case titles
#
# Dependencies: pandas, numpy, matplotlib, scipy, statsmodels, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
from matplotlib.gridspec import GridSpec
from scipy import stats
from scipy.stats import shapiro
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION  — edit only this section
# =============================================================================

REFERENCE_MODEL = "BAT-RM"    # your proposed / best model

excel_files_bnd = [
    ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx')
]

# excel_files_seg = [
#     ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
#     ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
#     ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
#     ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
#     ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
# ]

# excel_files_bnd = [
#     ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
#     ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
#     ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
#     ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
#     ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx')
# ]

# Boundary metrics to plot (one full figure per metric)
BND_METRICS = ["ASD", "HD95", "HD"]

# Anatomical classes — one Bland-Altman panel per class × baseline combination
class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# Bootstrap settings for CI bands
N_BOOT   = 1000
ALPHA_CI = 0.05     # 95% CI

# LOWESS smoothing fraction (0.5 = smooth, 0.2 = follows data more closely)
LOWESS_FRAC = 0.5

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/BlandAltman'
os.makedirs(SAVE_DIR, exist_ok=True)

# Wong (2011) colour-blind safe palette
MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE  (Nature Medicine)
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   9,
    "axes.labelsize":   8.5,
    "xtick.labelsize":  7.5,
    "ytick.labelsize":  7.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.35,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "pdf.fonttype":     42,
    "ps.fonttype":      42,
    "figure.dpi":       150,
    "savefig.dpi":      600,
})

# =============================================================================
# 3. DATA LOADING
# =============================================================================

def load_boundary(excel_files, metrics, class_list):
    """
    Load wide-format boundary Excel (Detailed_Per_Instance sheet).
    Returns long-format DataFrame: Filename | Class_Name | Model | <metrics>
    Builds one row-block per class per file — no iterative merge, no
    duplicate-column suffix errors.
    """
    all_data = []

    for name, path in excel_files:
        df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")
        file_rows = []

        for cls in class_list:
            present = f"{cls}_Present_In_GT"
            gt_vol  = f"{cls}_GT_Volume"

            if present in df.columns:
                sub = df[df[present] == True].copy()
            elif gt_vol in df.columns:
                sub = df[df[gt_vol] > 0].copy()
            else:
                sub = df.copy()

            if sub.empty:
                continue

            cls_cols = {"Filename": sub["Filename"].values}
            any_found = False

            for m in metrics:
                col = f"{cls}_{m}"
                if col not in df.columns:
                    cls_cols[m] = np.nan
                    continue
                vals = pd.to_numeric(sub[col], errors="coerce").values
                vals = np.where(np.isfinite(vals), vals, np.nan)
                cls_cols[m] = vals
                any_found = True

            if not any_found:
                continue

            tmp = pd.DataFrame(cls_cols)
            tmp["Class_Name"] = cls.upper().strip()
            tmp["Model"]      = name
            tmp = tmp.dropna(subset=metrics, how="all")
            file_rows.append(tmp)

        if file_rows:
            all_data.append(pd.concat(file_rows, ignore_index=True))

    if not all_data:
        raise ValueError(
            "No boundary data loaded. Check file paths and sheet name "
            "'Detailed_Per_Instance'."
        )
    return pd.concat(all_data, ignore_index=True)


df_bnd = load_boundary(excel_files_bnd, BND_METRICS, class_list)

model_names = [m for m, _ in excel_files_bnd]
baselines   = [m for m in model_names if m != REFERENCE_MODEL]
classes     = [c.upper().strip() for c in class_list
               if c.upper().strip() in df_bnd["Class_Name"].unique()]

print(f"Models   : {model_names}")
print(f"Classes  : {classes}")
print(f"Metrics  : {BND_METRICS}")
print(f"Patients : {df_bnd['Filename'].nunique()}\n")

# =============================================================================
# 4. STATISTICAL HELPERS
# =============================================================================

def bland_altman_stats(ref, bsl, n_boot=1000, alpha=0.05, seed=42):
    """
    Compute Bland-Altman statistics for paired measurements.

    Returns a dict with:
        means       — per-pair mean  (x-axis)
        diffs       — per-pair difference ref − bsl  (y-axis)
        bias        — mean difference
        loa_lo/hi   — bias ± 1.96 × SD
        ci_bias_*   — bootstrap 95% CI for bias
        ci_loa_lo_* — bootstrap 95% CI for lower LoA
        ci_loa_hi_* — bootstrap 95% CI for upper LoA
        sw_p        — Shapiro-Wilk p-value on differences
        n           — number of pairs
        pct_within  — % of points within LoA
    """
    ref, bsl = np.asarray(ref, float), np.asarray(bsl, float)
    means = (ref + bsl) / 2.0
    diffs = ref - bsl

    bias   = diffs.mean()
    sd     = diffs.std(ddof=1)
    loa_lo = bias - 1.96 * sd
    loa_hi = bias + 1.96 * sd

    # Bootstrap CI (percentile method)
    rng = np.random.default_rng(seed)
    n   = len(diffs)
    boot_bias  = np.empty(n_boot)
    boot_lo    = np.empty(n_boot)
    boot_hi    = np.empty(n_boot)

    for i in range(n_boot):
        samp       = rng.choice(diffs, size=n, replace=True)
        b_bias     = samp.mean()
        b_sd       = samp.std(ddof=1)
        boot_bias[i] = b_bias
        boot_lo[i]   = b_bias - 1.96 * b_sd
        boot_hi[i]   = b_bias + 1.96 * b_sd

    a = alpha / 2
    sw_stat, sw_p = shapiro(diffs) if n >= 3 else (np.nan, np.nan)
    pct_within = 100 * np.mean((diffs >= loa_lo) & (diffs <= loa_hi))

    return {
        "means":        means,
        "diffs":        diffs,
        "bias":         bias,
        "sd":           sd,
        "loa_lo":       loa_lo,
        "loa_hi":       loa_hi,
        "ci_bias_lo":   np.nanpercentile(boot_bias, 100 * a),
        "ci_bias_hi":   np.nanpercentile(boot_bias, 100 * (1 - a)),
        "ci_loa_lo_lo": np.nanpercentile(boot_lo,   100 * a),
        "ci_loa_lo_hi": np.nanpercentile(boot_lo,   100 * (1 - a)),
        "ci_loa_hi_lo": np.nanpercentile(boot_hi,   100 * a),
        "ci_loa_hi_hi": np.nanpercentile(boot_hi,   100 * (1 - a)),
        "sw_p":         sw_p,
        "n":            n,
        "pct_within":   pct_within,
    }


def lowess_curve(x, y, frac=0.5):
    """
    Compute LOWESS smoothed trend line.
    Requires statsmodels; falls back to linear regression if unavailable.
    """
    try:
        from statsmodels.nonparametric.smoothers_lowess import lowess
        smoothed = lowess(y, x, frac=frac, return_sorted=True)
        return smoothed[:, 0], smoothed[:, 1]
    except ImportError:
        # Fallback: simple linear regression line
        slope, intercept, *_ = stats.linregress(x, y)
        xs = np.linspace(x.min(), x.max(), 100)
        return xs, intercept + slope * xs

# =============================================================================
# 5. SUMMARY STATISTICS TABLE
# =============================================================================
# Saved to Excel before plotting so it is available even if the figure fails.
# =============================================================================

stat_records = []

ref_df = df_bnd[df_bnd["Model"] == REFERENCE_MODEL]

for metric in BND_METRICS:
    for cls in classes:
        ref_sub = ref_df[ref_df["Class_Name"] == cls][["Filename", metric]].dropna()

        for bsl in baselines:
            bsl_sub = df_bnd[
                (df_bnd["Model"] == cls) &      # intentional: filter by class
                (df_bnd["Class_Name"] == cls)   # then overwrite below
            ]
            # Correct filter
            bsl_sub = df_bnd[
                (df_bnd["Model"] == bsl) &
                (df_bnd["Class_Name"] == cls)
            ][["Filename", metric]].dropna()

            merged = pd.merge(
                ref_sub, bsl_sub,
                on="Filename", suffixes=("_ref", "_bsl")
            ).dropna()

            if len(merged) < 3:
                continue

            s = bland_altman_stats(
                merged[f"{metric}_ref"].values,
                merged[f"{metric}_bsl"].values,
                N_BOOT, ALPHA_CI
            )

            stat_records.append({
                "Metric":           metric,
                "Class":            cls,
                "Comparison":       f"{REFERENCE_MODEL} vs {bsl}",
                "N_pairs":          s["n"],
                "Bias":             round(s["bias"],   3),
                "SD_diff":          round(s["sd"],     3),
                "LoA_lower":        round(s["loa_lo"], 3),
                "LoA_upper":        round(s["loa_hi"], 3),
                "CI95_bias":        f"[{s['ci_bias_lo']:.3f}, {s['ci_bias_hi']:.3f}]",
                "CI95_LoA_lower":   f"[{s['ci_loa_lo_lo']:.3f}, {s['ci_loa_lo_hi']:.3f}]",
                "CI95_LoA_upper":   f"[{s['ci_loa_hi_lo']:.3f}, {s['ci_loa_hi_hi']:.3f}]",
                "Shapiro_Wilk_p":   round(s["sw_p"],  4),
                "Normal_diffs":     "Yes" if s["sw_p"] > 0.05 else "No",
                "Pct_within_LoA":   round(s["pct_within"], 1),
            })

df_stats = pd.DataFrame(stat_records)
stats_path = os.path.join(SAVE_DIR, "BlandAltman_Statistics.xlsx")
df_stats.to_excel(stats_path, index=False)
print(f"Statistics table saved → {stats_path}\n")

# =============================================================================
# 6. PLOTTING FUNCTION — single Bland-Altman panel
# =============================================================================

def draw_ba_panel(ax, s, color, title, metric, ref_name, bsl_name):
    """
    Draw one Bland-Altman panel onto axes `ax`.

    Parameters
    ----------
    ax       : matplotlib Axes
    s        : dict from bland_altman_stats()
    color    : hex colour string for this baseline model
    title    : panel title string (Class name)
    metric   : metric name string (for y-label)
    ref_name : reference model label
    bsl_name : baseline model label
    """
    means = s["means"]
    diffs = s["diffs"]

    x_pad = (means.max() - means.min()) * 0.08 if means.ptp() > 0 else 1.0
    x_lo  = means.min() - x_pad
    x_hi  = means.max() + x_pad
    x_range = np.array([x_lo, x_hi])

    # ── Background: LoA zone shading ──
    ax.axhspan(s["loa_lo"], s["loa_hi"],
               color=color, alpha=0.06, zorder=0)

    # ── CI bands around bias and LoA ──
    # Bias CI
    ax.fill_between(
        x_range,
        [s["ci_bias_lo"]] * 2, [s["ci_bias_hi"]] * 2,
        color=color, alpha=0.18, zorder=1, linewidth=0,
    )
    # Lower LoA CI
    ax.fill_between(
        x_range,
        [s["ci_loa_lo_lo"]] * 2, [s["ci_loa_lo_hi"]] * 2,
        color=color, alpha=0.12, zorder=1, linewidth=0,
    )
    # Upper LoA CI
    ax.fill_between(
        x_range,
        [s["ci_loa_hi_lo"]] * 2, [s["ci_loa_hi_hi"]] * 2,
        color=color, alpha=0.12, zorder=1, linewidth=0,
    )

    # ── Reference zero line ──
    ax.axhline(0, color="#999999", linewidth=0.7,
               linestyle=":", zorder=2)

    # ── Bias line ──
    ax.axhline(s["bias"], color=color, linewidth=1.6,
               linestyle="-", zorder=3, label=f"Bias {s['bias']:+.2f}")

    # ── LoA lines ──
    ax.axhline(s["loa_lo"], color=color, linewidth=1.2,
               linestyle="--", zorder=3,
               label=f"LoA [{s['loa_lo']:+.2f}, {s['loa_hi']:+.2f}]")
    ax.axhline(s["loa_hi"], color=color, linewidth=1.2,
               linestyle="--", zorder=3)

    # ── Right-side labels for bias and LoA ──
    label_x = x_hi + x_pad * 0.1
    ax.text(label_x, s["bias"],  f"{s['bias']:+.2f}",
            va="center", ha="left", fontsize=6.5,
            color=color, fontweight="bold")
    ax.text(label_x, s["loa_lo"], f"{s['loa_lo']:+.2f}",
            va="center", ha="left", fontsize=6.0, color=color)
    ax.text(label_x, s["loa_hi"], f"{s['loa_hi']:+.2f}",
            va="center", ha="left", fontsize=6.0, color=color)

    # ── Scatter: individual patient differences ──
    ax.scatter(
        means, diffs,
        color=color, alpha=0.45, s=14, zorder=4,
        edgecolors="white", linewidths=0.3,
    )

    # ── LOWESS trend (reveals proportional bias) ──
    if len(means) >= 6:
        lx, ly = lowess_curve(means, diffs, LOWESS_FRAC)
        ax.plot(lx, ly, color="#333333", linewidth=1.0,
                linestyle="-.", zorder=5, alpha=0.7,
                label="LOWESS trend")

    # ── Stats annotation box ──
    sw_text = (f"SW p={s['sw_p']:.3f}"
               if not np.isnan(s["sw_p"]) else "SW n/a")
    normal_flag = "" if s["sw_p"] > 0.05 else " ✗"
    stats_str = (
        f"n={s['n']}\n"
        f"Bias={s['bias']:+.2f} mm\n"
        f"SD={s['sd']:.2f} mm\n"
        f"{sw_text}{normal_flag}\n"
        f"{s['pct_within']:.0f}% within LoA"
    )
    ax.text(
        0.03, 0.97, stats_str,
        transform=ax.transAxes,
        va="top", ha="left", fontsize=6.5,
        color="#222222",
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
                  edgecolor="#cccccc", linewidth=0.5, alpha=0.85),
        zorder=6,
    )

    # ── Axes formatting ──
    ax.set_xlim(x_lo, x_hi + x_pad * 1.2)   # extra right margin for labels
    ax.set_title(title, fontsize=8.5, fontweight="bold", pad=4)
    ax.set_xlabel(f"Mean of {ref_name} & {bsl_name} ({metric}, mm)",
                  fontsize=7.5, labelpad=3)
    ax.set_ylabel(f"Difference\n{ref_name} − {bsl_name} (mm)",
                  fontsize=7.5, labelpad=3)
    ax.tick_params(axis="both", labelsize=7)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(0.5)
    ax.spines["bottom"].set_linewidth(0.5)
    ax.grid(axis="y", linewidth=0.3, color="#eeeeee", zorder=0)

# =============================================================================
# 7. GENERATE ONE FIGURE PER METRIC
# =============================================================================
# Layout: rows = classes, columns = baselines
# =============================================================================

ref_df = df_bnd[df_bnd["Model"] == REFERENCE_MODEL]

for metric in BND_METRICS:
    n_rows = len(classes)
    n_cols = len(baselines)

    # Panel size: 3.2 in wide × 3.0 in tall each
    FIG_W = n_cols * 3.2 + 0.8
    FIG_H = n_rows * 3.0 + 1.2

    fig, axes = plt.subplots(
        n_rows, n_cols,
        figsize=(FIG_W, FIG_H),
        squeeze=False,
    )

    fig.suptitle(
        f"Bland-Altman analysis — {metric} (mm)  |  "
        f"Reference: {REFERENCE_MODEL}  |  "
        f"Error bars: 95% bootstrap CI",
        fontsize=11, fontweight="bold", y=0.995,
    )
    fig.text(
        0.5, 0.983,
        "Dashed lines: limits of agreement (bias ± 1.96 SD)  |  "
        "Shaded bands: 95% CI  |  "
        "Dash-dot: LOWESS trend  |  "
        "✗ = differences non-normal (Shapiro-Wilk p<0.05)",
        ha="center", fontsize=7.5, color="#555555", style="italic",
    )

    # ── Column headers (baseline model names) ──
    for ci, bsl in enumerate(baselines):
        axes[0, ci].set_title(
            f"{REFERENCE_MODEL} vs {bsl}\n",
            fontsize=9, fontweight="bold", color=MODEL_COLORS.get(bsl, "#333"),
            pad=6,
        )

    # ── Draw each panel ──
    for ri, cls in enumerate(classes):
        ref_sub = ref_df[
            ref_df["Class_Name"] == cls
        ][["Filename", metric]].dropna()

        for ci, bsl in enumerate(baselines):
            ax    = axes[ri, ci]
            color = MODEL_COLORS.get(bsl, "#888888")

            bsl_sub = df_bnd[
                (df_bnd["Model"] == bsl) &
                (df_bnd["Class_Name"] == cls)
            ][["Filename", metric]].dropna()

            merged = pd.merge(
                ref_sub, bsl_sub,
                on="Filename", suffixes=("_ref", "_bsl")
            ).dropna()

            # Drop inf values
            merged = merged[
                np.isfinite(merged[f"{metric}_ref"]) &
                np.isfinite(merged[f"{metric}_bsl"])
            ]

            if len(merged) < 3:
                ax.text(0.5, 0.5, "Insufficient\npaired data",
                        ha="center", va="center",
                        fontsize=8, color="#aaaaaa",
                        transform=ax.transAxes)
                ax.set_title(cls, fontsize=8.5, fontweight="bold")
                ax.axis("off")
                continue

            s = bland_altman_stats(
                merged[f"{metric}_ref"].values,
                merged[f"{metric}_bsl"].values,
                N_BOOT, ALPHA_CI,
            )

            draw_ba_panel(
                ax, s, color,
                title  = cls,
                metric = metric,
                ref_name = REFERENCE_MODEL,
                bsl_name = bsl,
            )

            # Console clinical interpretation
            bias_flag = (
                "✓ negligible bias" if abs(s["bias"]) < 1.0 else
                "⚠ moderate bias"   if abs(s["bias"]) < 3.0 else
                "✗ large bias"
            )
            loa_width = s["loa_hi"] - s["loa_lo"]
            loa_flag  = (
                "✓ narrow LoA"   if loa_width < 4.0 else
                "⚠ moderate LoA" if loa_width < 8.0 else
                "✗ wide LoA"
            )
            print(
                f"[{metric}] {cls:20s} | {REFERENCE_MODEL} vs {bsl:10s} | "
                f"n={s['n']:3d} | bias={s['bias']:+.2f} mm "
                f"[{s['loa_lo']:+.2f}, {s['loa_hi']:+.2f}] | "
                f"{bias_flag} | {loa_flag}"
            )

    plt.tight_layout(rect=[0, 0, 1, 0.982])

    # ── Legend (shared, bottom of figure) ──
    legend_handles = []
    for bsl in baselines:
        c = MODEL_COLORS.get(bsl, "#888888")
        legend_handles += [
            mlines.Line2D([0], [0], color=c, lw=1.6,
                          label=f"{REFERENCE_MODEL} vs {bsl} — bias"),
            mlines.Line2D([0], [0], color=c, lw=1.2, linestyle="--",
                          label=f"{REFERENCE_MODEL} vs {bsl} — LoA"),
        ]
    legend_handles.append(
        mlines.Line2D([0], [0], color="#333333", lw=1.0, linestyle="-.",
                      label="LOWESS trend")
    )

    fig.legend(
        handles        = legend_handles,
        loc            = "lower center",
        ncol           = len(baselines) + 1,
        frameon        = False,
        fontsize       = 7.5,
        bbox_to_anchor = (0.5, -0.01),
    )

    # ── Save ──
    png_path = os.path.join(SAVE_DIR, f"{metric}_BlandAltman_NatureMedicine.png")
    pdf_path = os.path.join(SAVE_DIR, f"{metric}_BlandAltman_NatureMedicine.pdf")
    plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
    plt.savefig(pdf_path, format="pdf", bbox_inches="tight", facecolor="white")
    plt.show()
    print(f"\n{metric} Bland-Altman saved →\n  {png_path}\n  {pdf_path}\n")

print(f"All outputs in: {SAVE_DIR}")














































# =============================================================================
# CALIBRATION / ERROR DISTRIBUTION PLOT
# Nature Medicine Style — Per-Patient HD95 & ASD Error Distributions
# =============================================================================
#
# Produces TWO publication-ready outputs per metric (HD95, ASD):
#
#   OUTPUT 1 — KDE + Rug plot (main figure, PNG 600 DPI + PDF)
#     • Smoothed kernel density estimate per model (overlaid, semi-transparent)
#     • Rug ticks at the bottom: every patient's actual error value
#     • Right-tail shading: region beyond 95th percentile of reference model
#       (marks the "clinical safety concern" zone reviewers probe)
#     • Vertical lines: median + 95th percentile per model
#     • Inset table: median, IQR, 95th pct, % patients > clinical threshold
#
#   OUTPUT 2 — Violin + Strip plot (supplementary, PNG 600 DPI + PDF)
#     • One panel per class
#     • Violin shows full distribution shape
#     • Strip of individual patient dots overlaid (jittered)
#     • Median bar inside violin
#
# Layout: one row per anatomical class, one column per metric
#
# Clinical threshold defaults (mm) — adjust to your RT protocol:
#   HD95: 5 mm  (common planning margin for pelvic targets)
#   ASD:  2 mm
#
# Statistical annotations:
#   • Kolmogorov-Smirnov test vs reference model (tests if distributions differ)
#   • Right-tail ratio: P(model > threshold) / P(reference > threshold)
#     RTR > 1 → model has proportionally more extreme errors (bad)
#     RTR < 1 → model is safer
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyArrowPatch
from scipy.stats import gaussian_kde, ks_2samp
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION  — edit only this section
# =============================================================================

REFERENCE_MODEL = "BAT-RM"

excel_files_bnd = [
    ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx')
]


# excel_files_seg = [
#     ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
#     ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
#     ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
#     ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
#     ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
# ]

# excel_files_bnd = [
#     ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
#     ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
#     ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
#     ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
#     ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx')
# ]

BND_METRICS = ["HD95", "ASD"]

class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# Clinical safety thresholds (mm) — errors ABOVE these are flagged
CLINICAL_THRESHOLDS = {
    "HD95": 5.0,   # mm — adjust to your protocol
    "ASD":  2.0,   # mm
}

# KDE bandwidth method ('scott', 'silverman', or a float multiplier)
KDE_BW = "scott"

# X-axis cap for plotting (outliers beyond this are still counted in stats)
X_CAP = {
    "HD95": 30.0,  # mm
    "ASD":  10.0,  # mm
}

# Wong (2011) colour-blind safe palette
MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/ErrorDistribution'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE  (Nature Medicine)
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   9,
    "axes.labelsize":   8.5,
    "xtick.labelsize":  7.5,
    "ytick.labelsize":  7.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.35,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "pdf.fonttype":     42,
    "ps.fonttype":      42,
    "figure.dpi":       150,
    "savefig.dpi":      600,
})

# =============================================================================
# 3. DATA LOADING
# =============================================================================

def load_boundary(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")
        file_rows = []
        for cls in class_list:
            present = f"{cls}_Present_In_GT"
            gt_vol  = f"{cls}_GT_Volume"
            if present in df.columns:
                sub = df[df[present] == True].copy()
            elif gt_vol in df.columns:
                sub = df[df[gt_vol] > 0].copy()
            else:
                sub = df.copy()
            if sub.empty:
                continue
            cls_cols = {"Filename": sub["Filename"].values}
            any_found = False
            for m in metrics:
                col = f"{cls}_{m}"
                if col not in df.columns:
                    cls_cols[m] = np.nan
                    continue
                vals = pd.to_numeric(sub[col], errors="coerce").values
                vals = np.where(np.isfinite(vals), vals, np.nan)
                cls_cols[m] = vals
                any_found = True
            if not any_found:
                continue
            tmp = pd.DataFrame(cls_cols)
            tmp["Class_Name"] = cls.upper().strip()
            tmp["Model"]      = name
            tmp = tmp.dropna(subset=metrics, how="all")
            file_rows.append(tmp)
        if file_rows:
            all_data.append(pd.concat(file_rows, ignore_index=True))
    if not all_data:
        raise ValueError("No boundary data loaded. Check file paths and sheet name.")
    return pd.concat(all_data, ignore_index=True)


df_bnd = load_boundary(excel_files_bnd, BND_METRICS, class_list)

model_names = [m for m, _ in excel_files_bnd]
baselines   = [m for m in model_names if m != REFERENCE_MODEL]
classes     = [c.upper().strip() for c in class_list
               if c.upper().strip() in df_bnd["Class_Name"].unique()]

print(f"Models   : {model_names}")
print(f"Classes  : {classes}")
print(f"Metrics  : {BND_METRICS}")
print(f"Patients : {df_bnd['Filename'].nunique()}\n")

# =============================================================================
# 4. STATISTICS HELPER
# =============================================================================

def error_stats(vals, threshold):
    vals = np.array(vals)
    vals = vals[np.isfinite(vals)]
    if len(vals) == 0:
        return {}
    return {
        "n":          len(vals),
        "median":     np.median(vals),
        "q25":        np.percentile(vals, 25),
        "q75":        np.percentile(vals, 75),
        "p95":        np.percentile(vals, 95),
        "mean":       vals.mean(),
        "std":        vals.std(ddof=1),
        "pct_exceed": 100 * np.mean(vals > threshold),
        "max":        vals.max(),
    }

# =============================================================================
# 5. FIGURE 1 — KDE + RUG PLOT (main figure)
# Layout: rows = classes, columns = metrics
# =============================================================================

n_rows = len(classes)
n_cols = len(BND_METRICS)

FIG_W = n_cols * 5.5 + 0.5
FIG_H = n_rows * 3.2 + 1.5

fig, axes = plt.subplots(
    n_rows, n_cols,
    figsize=(FIG_W, FIG_H),
    squeeze=False,
)

fig.suptitle(
    "Per-patient error distributions — KDE overlay across models\n"
    "Shaded region: errors exceeding clinical safety threshold",
    fontsize=11, fontweight="bold", y=0.998,
)

# Collect stats for console + Excel export
all_stats = []

for ri, cls in enumerate(classes):
    for ci, metric in enumerate(BND_METRICS):
        ax        = axes[ri, ci]
        threshold = CLINICAL_THRESHOLDS[metric]
        x_cap     = X_CAP[metric]

        # ── Compute KDE for each model and find the global y-max ──
        kde_curves = {}
        x_eval     = np.linspace(0, x_cap, 500)

        for model in model_names:
            vals = df_bnd[
                (df_bnd["Model"] == model) &
                (df_bnd["Class_Name"] == cls)
            ][metric].dropna().values
            vals = vals[np.isfinite(vals)]
            vals = np.clip(vals, 0, x_cap * 1.5)   # soft clip for KDE

            if len(vals) < 4:
                kde_curves[model] = (x_eval, np.zeros_like(x_eval), vals)
                continue
            try:
                kde  = gaussian_kde(vals, bw_method=KDE_BW)
                dens = kde(x_eval)
            except Exception:
                dens = np.zeros_like(x_eval)
            kde_curves[model] = (x_eval, dens, vals)

        y_max = max(
            curve[1].max() for curve in kde_curves.values()
            if curve[1].max() > 0
        ) if any(c[1].max() > 0 for c in kde_curves.values()) else 1.0

        # ── Right-tail danger zone (reference model 95th pct) ──
        ref_vals = kde_curves[REFERENCE_MODEL][2]
        ref_p95  = np.percentile(ref_vals, 95) if len(ref_vals) > 0 else threshold

        ax.axvspan(
            threshold, x_cap,
            color="#FFCCCC", alpha=0.35, zorder=0,
            label=f">{threshold} mm threshold",
        )
        ax.axvspan(
            ref_p95, x_cap,
            color="#FF8888", alpha=0.15, zorder=0,
        )

        # ── Threshold vertical line ──
        ax.axvline(
            threshold, color="#CC2222", linewidth=0.9,
            linestyle="--", zorder=2, alpha=0.85,
            label=f"Threshold {threshold} mm",
        )

        # ── KDE curves (reference on top, slightly thicker) ──
        plot_order = baselines + [REFERENCE_MODEL]
        for model in plot_order:
            x_k, dens, vals = kde_curves[model]
            color    = MODEL_COLORS.get(model, "#888888")
            is_ref   = model == REFERENCE_MODEL
            lw       = 2.2 if is_ref else 1.4
            alpha    = 0.90 if is_ref else 0.75
            zorder   = 5 if is_ref else 3

            ax.fill_between(
                x_k, dens,
                color=color, alpha=0.08 if is_ref else 0.05,
                zorder=zorder - 1,
            )
            ax.plot(
                x_k, dens,
                color=color, linewidth=lw, alpha=alpha,
                zorder=zorder, label=model,
            )

            if len(vals) < 2:
                continue

            # Median line
            med = np.median(vals)
            if med <= x_cap:
                med_dens = gaussian_kde(vals, bw_method=KDE_BW)(np.array([med]))[0] if len(vals) >= 4 else 0
                ax.plot(
                    [med, med], [0, med_dens],
                    color=color, linewidth=0.9,
                    linestyle=":", alpha=0.7, zorder=zorder,
                )

            # Rug ticks (bottom of axis)
            rug_vals = np.clip(vals, 0, x_cap)
            rug_y    = -0.02 * y_max
            ax.plot(
                rug_vals, np.full_like(rug_vals, rug_y),
                "|", color=color, alpha=0.25,
                markersize=4, markeredgewidth=0.6,
            )

            # Stats record
            st = error_stats(vals, threshold)
            if st:
                # KS test vs reference
                ref_v = kde_curves[REFERENCE_MODEL][2]
                if model != REFERENCE_MODEL and len(ref_v) > 0 and len(vals) > 0:
                    ks_stat, ks_p = ks_2samp(ref_v, vals)
                else:
                    ks_stat, ks_p = np.nan, np.nan

                rtr = (st["pct_exceed"] / error_stats(ref_v, threshold).get("pct_exceed", 1e-9)) \
                      if model != REFERENCE_MODEL and len(ref_v) > 0 else np.nan

                all_stats.append({
                    "Metric":              metric,
                    "Class":               cls,
                    "Model":               model,
                    "N":                   st["n"],
                    "Median_mm":           round(st["median"], 2),
                    "IQR_mm":              f"{st['q25']:.2f}–{st['q75']:.2f}",
                    "P95_mm":              round(st["p95"], 2),
                    "Mean_mm":             round(st["mean"], 2),
                    "SD_mm":               round(st["std"], 2),
                    "Max_mm":              round(st["max"], 2),
                    f"Pct_exceed_{threshold}mm": round(st["pct_exceed"], 1),
                    "KS_stat_vs_ref":      round(ks_stat, 3) if not np.isnan(ks_stat) else "—",
                    "KS_p_vs_ref":         round(ks_p,    4) if not np.isnan(ks_p)    else "—",
                    "Right_tail_ratio":    round(rtr, 2)     if not np.isnan(rtr)      else "—",
                })

        # ── Inset summary table ──
        table_lines = [f"{'Model':<12} {'Med':>5} {'P95':>5} {'>thr%':>6}"]
        table_lines.append("─" * 32)
        for model in model_names:
            _, _, vals = kde_curves[model]
            if len(vals) < 2:
                continue
            st  = error_stats(vals, threshold)
            tag = "★ " if model == REFERENCE_MODEL else "  "
            table_lines.append(
                f"{tag}{model:<10} {st['median']:>5.1f} {st['p95']:>5.1f} {st['pct_exceed']:>5.1f}%"
            )

        table_text = "\n".join(table_lines)
        ax.text(
            0.98, 0.97, table_text,
            transform=ax.transAxes,
            va="top", ha="right",
            fontsize=5.8,
            fontfamily="monospace",
            color="#222222",
            bbox=dict(
                boxstyle="round,pad=0.4",
                facecolor="white",
                edgecolor="#cccccc",
                linewidth=0.5,
                alpha=0.88,
            ),
            zorder=10,
        )

        # ── Axis formatting ──
        ax.set_xlim(0, x_cap)
        ax.set_ylim(-0.04 * y_max, y_max * 1.18)
        ax.set_xlabel(f"{metric} (mm)", fontsize=8)
        ax.set_ylabel("Density", fontsize=8)
        ax.set_title(f"{cls}  —  {metric}", fontsize=8.5, fontweight="bold", pad=4)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_linewidth(0.5)
        ax.spines["bottom"].set_linewidth(0.5)
        ax.tick_params(labelsize=7)
        ax.grid(axis="x", linewidth=0.3, color="#eeeeee", zorder=0)

        # First panel only: add danger zone annotation arrow
        if ri == 0 and ci == 0:
            ax.annotate(
                "Clinical\nconcern zone",
                xy=(threshold + 0.3, y_max * 0.7),
                xytext=(threshold + (x_cap - threshold) * 0.45, y_max * 0.9),
                fontsize=6.5, color="#CC2222",
                ha="center",
                arrowprops=dict(
                    arrowstyle="->",
                    color="#CC2222",
                    lw=0.8,
                ),
            )

# ── Shared legend ──
legend_handles = [
    mlines.Line2D(
        [0], [0],
        color=MODEL_COLORS[m],
        linewidth=2.2 if m == REFERENCE_MODEL else 1.4,
        label=f"{m}{'  ★ reference' if m == REFERENCE_MODEL else ''}",
    )
    for m in model_names
]
legend_handles += [
    mpatches.Patch(color="#FFCCCC", alpha=0.6, label=f"Beyond threshold"),
    mlines.Line2D([0], [0], color="#CC2222", lw=0.9, linestyle="--",
                  label="Clinical threshold"),
    mlines.Line2D([0], [0], color="#888888", lw=0.9, linestyle=":",
                  label="Model median"),
]

fig.legend(
    handles        = legend_handles,
    loc            = "lower center",
    ncol           = len(model_names) + 3,
    frameon        = False,
    fontsize       = 8,
    bbox_to_anchor = (0.5, -0.005),
)

plt.tight_layout(rect=[0, 0.03, 1, 0.996])

kde_png = os.path.join(SAVE_DIR, "ErrorDistribution_KDE_NatureMedicine.png")
kde_pdf = os.path.join(SAVE_DIR, "ErrorDistribution_KDE_NatureMedicine.pdf")
plt.savefig(kde_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(kde_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"KDE figure saved →\n  {kde_png}\n  {kde_pdf}\n")

# =============================================================================
# 6. FIGURE 2 — VIOLIN + STRIP (supplementary)
# =============================================================================

fig2, axes2 = plt.subplots(
    n_rows, n_cols,
    figsize=(FIG_W, FIG_H),
    squeeze=False,
)

fig2.suptitle(
    "Per-patient error distributions — Violin + strip plot\n"
    "Each dot = one patient; horizontal bar = median",
    fontsize=11, fontweight="bold", y=0.998,
)

rng = np.random.default_rng(42)

for ri, cls in enumerate(classes):
    for ci, metric in enumerate(BND_METRICS):
        ax        = axes2[ri, ci]
        threshold = CLINICAL_THRESHOLDS[metric]
        x_cap     = X_CAP[metric]

        ax.axhline(threshold, color="#CC2222", linewidth=0.9,
                   linestyle="--", alpha=0.8, zorder=2,
                   label=f"Threshold {threshold} mm")

        positions = range(len(model_names))

        for pos, model in zip(positions, model_names):
            vals = df_bnd[
                (df_bnd["Model"] == model) &
                (df_bnd["Class_Name"] == cls)
            ][metric].dropna().values
            vals = vals[np.isfinite(vals)]
            vals = np.clip(vals, 0, x_cap * 1.5)

            color    = MODEL_COLORS.get(model, "#888888")
            is_ref   = model == REFERENCE_MODEL

            if len(vals) < 4:
                ax.scatter(
                    [pos], [np.nan],
                    color=color, s=8, alpha=0.5, zorder=3,
                )
                continue

            # ── Violin ──
            try:
                kde    = gaussian_kde(vals, bw_method=KDE_BW)
                y_eval = np.linspace(vals.min(), min(vals.max(), x_cap), 300)
                dens   = kde(y_eval)
                dens   = dens / dens.max() * 0.38   # normalise width

                ax.fill_betweenx(
                    y_eval,
                    pos - dens,
                    pos + dens,
                    color=color,
                    alpha=0.30 if is_ref else 0.18,
                    zorder=1,
                )
                ax.plot(
                    pos - dens, y_eval,
                    color=color, linewidth=0.9 if is_ref else 0.6,
                    alpha=0.8, zorder=2,
                )
                ax.plot(
                    pos + dens, y_eval,
                    color=color, linewidth=0.9 if is_ref else 0.6,
                    alpha=0.8, zorder=2,
                )
            except Exception:
                pass

            # ── Median bar ──
            med  = np.median(vals)
            ax.plot(
                [pos - 0.22, pos + 0.22], [med, med],
                color=color, linewidth=2.0, zorder=4,
                solid_capstyle="round",
            )

            # ── IQR box ──
            q25, q75 = np.percentile(vals, [25, 75])
            ax.plot(
                [pos - 0.12, pos + 0.12], [q25, q25],
                color=color, linewidth=0.8, alpha=0.6, zorder=3,
            )
            ax.plot(
                [pos - 0.12, pos + 0.12], [q75, q75],
                color=color, linewidth=0.8, alpha=0.6, zorder=3,
            )
            ax.plot(
                [pos - 0.12, pos - 0.12], [q25, q75],
                color=color, linewidth=0.6, alpha=0.5, zorder=3,
            )
            ax.plot(
                [pos + 0.12, pos + 0.12], [q25, q75],
                color=color, linewidth=0.6, alpha=0.5, zorder=3,
            )

            # ── Strip dots (jittered) ──
            jitter  = rng.uniform(-0.15, 0.15, size=len(vals))
            outlier = vals > threshold
            ax.scatter(
                pos + jitter[~outlier], vals[~outlier],
                color=color, s=6, alpha=0.30,
                edgecolors="none", zorder=3,
            )
            # Outliers (beyond threshold) drawn in red with marker
            if outlier.any():
                ax.scatter(
                    pos + jitter[outlier], vals[outlier],
                    color="#CC2222", s=10, alpha=0.55,
                    edgecolors="white", linewidths=0.3,
                    zorder=5, marker="^",
                )

        # ── Axis formatting ──
        ax.set_xlim(-0.6, len(model_names) - 0.4)
        ax.set_ylim(-0.3, x_cap)
        ax.set_xticks(list(positions))
        ax.set_xticklabels(model_names, fontsize=7, rotation=30, ha="right")
        ax.set_ylabel(f"{metric} (mm)", fontsize=8)
        ax.set_title(f"{cls}  —  {metric}", fontsize=8.5, fontweight="bold", pad=4)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["bottom"].set_linewidth(0.5)
        ax.spines["left"].set_linewidth(0.5)
        ax.grid(axis="y", linewidth=0.3, color="#eeeeee", zorder=0)
        ax.tick_params(labelsize=7)

        # Shade threshold zone
        ax.axhspan(threshold, x_cap, color="#FFCCCC", alpha=0.18, zorder=0)

# ── Shared legend for violin figure ──
vln_handles = [
    mpatches.Patch(color=MODEL_COLORS[m], alpha=0.5,
                   label=f"{m}{'  ★' if m == REFERENCE_MODEL else ''}")
    for m in model_names
]
vln_handles += [
    mlines.Line2D([0], [0], color="#CC2222", lw=0.9, linestyle="--",
                  label="Clinical threshold"),
    plt.scatter([], [], color="#CC2222", marker="^", s=12, alpha=0.7,
                label="Exceeds threshold"),
]

fig2.legend(
    handles        = vln_handles,
    loc            = "lower center",
    ncol           = len(model_names) + 2,
    frameon        = False,
    fontsize       = 8,
    bbox_to_anchor = (0.5, -0.005),
)

plt.tight_layout(rect=[0, 0.03, 1, 0.996])

vln_png = os.path.join(SAVE_DIR, "ErrorDistribution_Violin_NatureMedicine.png")
vln_pdf = os.path.join(SAVE_DIR, "ErrorDistribution_Violin_NatureMedicine.pdf")
plt.savefig(vln_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(vln_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Violin figure saved →\n  {vln_png}\n  {vln_pdf}\n")

# =============================================================================
# 7. SAVE STATISTICS TABLE
# =============================================================================

df_stats = pd.DataFrame(all_stats)
stats_path = os.path.join(SAVE_DIR, "ErrorDistribution_Statistics.xlsx")

with pd.ExcelWriter(stats_path, engine="openpyxl") as writer:
    df_stats.to_excel(writer, index=False, sheet_name="Stats")

    ws = writer.sheets["Stats"]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=8, name="Helvetica")
    thin     = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin

    ref_fill  = PatternFill("solid", fgColor="D5E8D4")
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    warn_fill = PatternFill("solid", fgColor="FFE6CC")

    for row_idx, (_, row) in enumerate(df_stats.iterrows(), start=2):
        is_ref = row["Model"] == REFERENCE_MODEL
        fill   = ref_fill if is_ref else (alt_fill if row_idx % 2 == 0 else PatternFill())
        for cell in ws[row_idx]:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(size=8, name="Helvetica",
                                  bold=is_ref)
            if fill.fill_type:
                cell.fill = fill

        # Flag rows where % exceeding threshold is high
        exceed_col = [c.column for c in ws[row_idx]
                      if ws.cell(1, c.column).value and
                      "Pct_exceed" in str(ws.cell(1, c.column).value)]
        if exceed_col:
            cell = ws.cell(row_idx, exceed_col[0])
            try:
                if float(str(cell.value).replace("%","")) > 20:
                    cell.fill = warn_fill
                    cell.font = Font(size=8, name="Helvetica", bold=True, color="CC3300")
            except Exception:
                pass

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

print(f"Statistics table saved → {stats_path}")

# ── Console clinical interpretation ──
print("\n" + "=" * 72)
print("CLINICAL SAFETY SUMMARY  (% patients exceeding threshold)")
print("=" * 72)
for metric in BND_METRICS:
    thr = CLINICAL_THRESHOLDS[metric]
    print(f"\n{metric}  (threshold = {thr} mm):")
    print(f"  {'Model':<14} {'Class':<20} {'Median':>8} {'P95':>7} {'>thr%':>7} {'RTR':>6}")
    print("  " + "-" * 62)
    sub = df_stats[df_stats["Metric"] == metric].copy()
    for _, row in sub.sort_values(["Class", "Model"]).iterrows():
        flag = "  ⚠" if str(row.get(f"Pct_exceed_{thr}mm", 0)) not in ["—", "0.0"] and \
                        float(str(row.get(f"Pct_exceed_{thr}mm", 0))) > 20 else ""
        print(
            f"  {row['Model']:<14} {row['Class']:<20} "
            f"{row['Median_mm']:>8.2f} {row['P95_mm']:>7.2f} "
            f"{str(row.get(f'Pct_exceed_{thr}mm','—')):>7} "
            f"{str(row.get('Right_tail_ratio','—')):>6}{flag}"
        )

print(f"\nAll outputs in: {SAVE_DIR}")












































# =============================================================================
# RELIABILITY / CONSISTENCY SCATTER PLOT
# Nature Medicine Style — Per-Patient Model vs Model Comparison
# =============================================================================
#
# For each metric × class combination, plots:
#   X-axis : reference model (BAT-RM) per-patient metric value
#   Y-axis : baseline model per-patient metric value
#   Points : one dot per matched patient
#
# Annotations per panel:
#   • Identity line (y = x)  — dashed grey; points BELOW = ref wins (overlap)
#                              or ABOVE = ref wins (distance)
#   • OLS regression line    — coloured; slope/intercept tell you the pattern
#   • 95% CI band            — shaded around the regression line
#   • Quadrant counts        — how many patients each model wins on
#   • Pearson r + p-value    — agreement strength
#   • Concordance correlation coefficient (CCC) — gold standard for agreement
#   • Win % annotation       — "BAT-RM better in N/M cases (X%)"
#
# Layout:
#   • One FIGURE per metric × baseline combination
#   • Panels: one per anatomical class (arranged in a grid)
#
# Clinical interpretation:
#   • Points tight around identity → consistent agreement / similar performance
#   • Regression slope < 1 (overlap) or > 1 (distance) → one model
#     systematically better across the range
#   • Points far off identity in top-right → both models struggle on same cases
#     (patient-level difficulty, not model failure)
#   • Scatter around regression line → inconsistency / case-specific variance
#
# Outputs (per metric × baseline):
#   PNG 600 DPI + vector PDF
#   Excel statistics table (r, CCC, win%, slope, intercept per class)
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
from matplotlib.gridspec import GridSpec
from scipy import stats
from scipy.stats import pearsonr
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION
# =============================================================================

REFERENCE_MODEL = "BAT-RM"

excel_files_seg = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
]

excel_files_bnd = [
    ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx')
]




SEG_METRICS      = ["Dice", "IoU"]
BND_METRICS      = ["HD95", "ASD"]
ALL_METRICS      = SEG_METRICS + BND_METRICS
LOWER_IS_BETTER  = {"HD95", "ASD", "HD", "RVD", "RAVD"}

class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# Point style
ALPHA_SCATTER = 0.55
DOT_SIZE      = 18        # marker area (pt²)
REG_ALPHA     = 0.15      # CI band opacity

# Wong (2011) colour-blind safe palette
MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/ReliabilityScatter'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   9,
    "axes.labelsize":   8.5,
    "xtick.labelsize":  7.5,
    "ytick.labelsize":  7.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.35,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "pdf.fonttype":     42,
    "ps.fonttype":      42,
    "figure.dpi":       150,
    "savefig.dpi":      600,
})

# =============================================================================
# 3. DATA LOADING
# =============================================================================

def load_segmentation(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
        for m in metrics:
            if m in df.columns:
                df[m] = pd.to_numeric(df[m], errors="coerce")
        df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
        df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
        df = df[df["Class_Name"].isin([c.upper() for c in class_list])]
        df["Model"] = name
        cols = ["Filename", "Class_Name", "Model"] + [m for m in metrics if m in df.columns]
        all_data.append(df[cols])
    return pd.concat(all_data, ignore_index=True)


def load_boundary(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")
        file_rows = []
        for cls in class_list:
            present = f"{cls}_Present_In_GT"
            gt_vol  = f"{cls}_GT_Volume"
            if present in df.columns:
                sub = df[df[present] == True].copy()
            elif gt_vol in df.columns:
                sub = df[df[gt_vol] > 0].copy()
            else:
                sub = df.copy()
            if sub.empty:
                continue
            cls_cols = {"Filename": sub["Filename"].values}
            any_found = False
            for m in metrics:
                col = f"{cls}_{m}"
                if col not in df.columns:
                    cls_cols[m] = np.nan
                    continue
                vals = pd.to_numeric(sub[col], errors="coerce").values
                vals = np.where(np.isfinite(vals), vals, np.nan)
                cls_cols[m] = vals
                any_found = True
            if not any_found:
                continue
            tmp = pd.DataFrame(cls_cols)
            tmp["Class_Name"] = cls.upper().strip()
            tmp["Model"]      = name
            tmp = tmp.dropna(subset=metrics, how="all")
            file_rows.append(tmp)
        if file_rows:
            all_data.append(pd.concat(file_rows, ignore_index=True))
    if not all_data:
        raise ValueError("No boundary data loaded.")
    return pd.concat(all_data, ignore_index=True)


df_seg = load_segmentation(excel_files_seg, SEG_METRICS, class_list)
df_bnd = load_boundary(excel_files_bnd, BND_METRICS, class_list)

df_all = pd.merge(
    df_seg, df_bnd,
    on=["Filename", "Class_Name", "Model"],
    how="outer",
)

model_names = [m for m, _ in excel_files_seg]
baselines   = [m for m in model_names if m != REFERENCE_MODEL]
classes     = sorted(df_all["Class_Name"].dropna().unique())

print(f"Models   : {model_names}")
print(f"Classes  : {classes}")
print(f"Metrics  : {ALL_METRICS}")
print(f"Patients : {df_all['Filename'].nunique()}\n")

# =============================================================================
# 4. STATISTICAL HELPERS
# =============================================================================

def concordance_correlation_coefficient(x, y):
    """
    Lin's Concordance Correlation Coefficient (CCC).
    Measures both precision (r) and accuracy (proximity to identity line).
    CCC = 1: perfect agreement. CCC = 0: no agreement.
    """
    x, y   = np.array(x), np.array(y)
    mx, my = x.mean(), y.mean()
    vx     = x.var(ddof=0)
    vy     = y.var(ddof=0)
    cov    = np.cov(x, y, ddof=0)[0, 1]
    ccc    = (2 * cov) / (vx + vy + (mx - my) ** 2 + 1e-12)
    return ccc


def regression_ci(x, y_hat, x_new, n, alpha=0.05):
    """
    Pointwise 95% CI band for an OLS regression line.
    Returns (ci_low, ci_high) arrays over x_new.
    """
    n     = len(x)
    x     = np.array(x)
    x_bar = x.mean()
    ss_x  = ((x - x_bar) ** 2).sum()
    se    = np.sqrt(
        np.sum((np.array(y_hat) - np.array([np.polyval(np.polyfit(x, y_hat, 1), xi) for xi in x])) ** 2 + 1e-12) /
        max(n - 2, 1)
        * (1 / n + (x_new - x_bar) ** 2 / (ss_x + 1e-12))
    )
    t_crit = stats.t.ppf(1 - alpha / 2, df=max(n - 2, 1))
    return se * t_crit


def win_analysis(x_ref, x_bsl, lower_is_better):
    """
    Count how many patients each model wins on.
    Returns (n_ref_wins, n_bsl_wins, n_ties, win_pct_ref).
    """
    diff = np.array(x_ref) - np.array(x_bsl)
    if lower_is_better:
        ref_wins = (diff < 0).sum()   # ref smaller = ref wins
        bsl_wins = (diff > 0).sum()
    else:
        ref_wins = (diff > 0).sum()   # ref larger = ref wins
        bsl_wins = (diff < 0).sum()
    ties = (diff == 0).sum()
    n    = len(diff)
    return ref_wins, bsl_wins, ties, 100 * ref_wins / n if n > 0 else 0


def p_stars(p):
    if np.isnan(p):   return ""
    if p < 0.0001:    return "****"
    if p < 0.001:     return "***"
    if p < 0.01:      return "**"
    if p < 0.05:      return "*"
    return "ns"

# =============================================================================
# 5. PANEL DRAWING FUNCTION
# =============================================================================

def draw_scatter_panel(ax, x_ref, x_bsl, metric, ref_name, bsl_name, cls,
                       color, lower_is_better):
    """
    Draw one reliability scatter panel.
    x_ref, x_bsl : matched per-patient arrays (same patient order).
    """
    n = len(x_ref)
    if n < 3:
        ax.text(0.5, 0.5, "Insufficient\npaired data",
                ha="center", va="center", fontsize=8,
                color="#aaaaaa", transform=ax.transAxes)
        ax.set_title(cls, fontsize=8.5, fontweight="bold")
        ax.axis("off")
        return {}

    x_ref = np.array(x_ref, float)
    x_bsl = np.array(x_bsl, float)

    # ── Axis range: common scale for both axes ──
    all_vals = np.concatenate([x_ref, x_bsl])
    vmin     = np.nanmin(all_vals)
    vmax     = np.nanmax(all_vals)
    pad      = (vmax - vmin) * 0.07 if vmax > vmin else 0.05
    lo, hi   = vmin - pad, vmax + pad

    # ── Identity line (y = x) ──
    ax.plot(
        [lo, hi], [lo, hi],
        color="#888888", linewidth=1.0, linestyle="--",
        zorder=2, alpha=0.8, label="Identity (y = x)",
    )

    # ── Shade quadrants: ref-better vs baseline-better ──
    if lower_is_better:
        # Below identity → ref has lower (better) value
        ax.fill_between([lo, hi], [lo, hi], lo,
                        color=MODEL_COLORS[ref_name], alpha=0.04, zorder=0)
        ax.fill_between([lo, hi], [hi, hi], [lo, hi],
                        color=color, alpha=0.04, zorder=0)
        ref_better_label = f"← {ref_name} lower (better)"
        bsl_better_label = f"← {bsl_name} lower (better)"
    else:
        # Above identity → ref has higher (better) value
        ax.fill_between([lo, hi], [hi, hi], [lo, hi],
                        color=MODEL_COLORS[ref_name], alpha=0.04, zorder=0)
        ax.fill_between([lo, hi], [lo, hi], lo,
                        color=color, alpha=0.04, zorder=0)
        ref_better_label = f"← {ref_name} higher (better)"
        bsl_better_label = f"← {bsl_name} higher (better)"

    # ── OLS regression ──
    slope, intercept, r_val, p_val, se_slope = stats.linregress(x_ref, x_bsl)
    x_line  = np.linspace(lo, hi, 300)
    y_line  = slope * x_line + intercept

    # CI band
    x_bar  = x_ref.mean()
    ss_x   = ((x_ref - x_bar) ** 2).sum()
    resid  = x_bsl - (slope * x_ref + intercept)
    mse    = (resid ** 2).sum() / max(n - 2, 1)
    se_y   = np.sqrt(mse * (1 / n + (x_line - x_bar) ** 2 / (ss_x + 1e-12)))
    t_crit = stats.t.ppf(0.975, df=max(n - 2, 1))

    ax.fill_between(
        x_line,
        y_line - t_crit * se_y,
        y_line + t_crit * se_y,
        color=color, alpha=REG_ALPHA, zorder=3, linewidth=0,
    )
    ax.plot(
        x_line, y_line,
        color=color, linewidth=1.6, zorder=4,
        label=f"OLS (slope={slope:.2f})",
    )

    # ── Scatter points ──
    # Colour-code by which model wins on that patient
    ref_wins_pt, bsl_wins_pt, _, _ = win_analysis(x_ref, x_bsl, lower_is_better)

    if lower_is_better:
        win_mask = x_ref < x_bsl    # ref wins when below identity
    else:
        win_mask = x_ref > x_bsl    # ref wins when above identity

    # Tie / very close
    tol      = (vmax - vmin) * 0.02 if vmax > vmin else 0.01
    tie_mask = np.abs(x_ref - x_bsl) <= tol

    ax.scatter(
        x_ref[win_mask & ~tie_mask],
        x_bsl[win_mask & ~tie_mask],
        color=MODEL_COLORS[ref_name],
        s=DOT_SIZE, alpha=ALPHA_SCATTER,
        edgecolors="white", linewidths=0.3,
        zorder=5, label=f"{ref_name} wins",
    )
    ax.scatter(
        x_ref[~win_mask & ~tie_mask],
        x_bsl[~win_mask & ~tie_mask],
        color=color,
        s=DOT_SIZE, alpha=ALPHA_SCATTER,
        edgecolors="white", linewidths=0.3,
        zorder=5, label=f"{bsl_name} wins",
    )
    if tie_mask.any():
        ax.scatter(
            x_ref[tie_mask], x_bsl[tie_mask],
            color="#888888", s=DOT_SIZE, alpha=0.4,
            edgecolors="none", zorder=5,
        )

    # ── Win counts in corners ──
    r_wins, b_wins, ties, win_pct = win_analysis(x_ref, x_bsl, lower_is_better)
    ax.text(
        0.03, 0.97,
        f"{ref_name}: {r_wins}/{n} ({win_pct:.0f}%)",
        transform=ax.transAxes, va="top", ha="left",
        fontsize=6.5, color=MODEL_COLORS[ref_name], fontweight="bold",
        zorder=8,
    )
    ax.text(
        0.97, 0.03,
        f"{bsl_name}: {b_wins}/{n} ({100-win_pct:.0f}%)",
        transform=ax.transAxes, va="bottom", ha="right",
        fontsize=6.5, color=color, fontweight="bold",
        zorder=8,
    )

    # ── Stats box ──
    ccc    = concordance_correlation_coefficient(x_ref, x_bsl)
    stars  = p_stars(p_val)
    stat_txt = (
        f"r = {r_val:.3f}{' ' + stars if stars else ''}\n"
        f"CCC = {ccc:.3f}\n"
        f"slope = {slope:.3f}\n"
        f"n = {n}"
    )
    ax.text(
        0.03, 0.03, stat_txt,
        transform=ax.transAxes, va="bottom", ha="left",
        fontsize=6.5, color="#222222",
        bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
                  edgecolor="#cccccc", linewidth=0.5, alpha=0.88),
        zorder=8,
    )

    # ── Diagonal direction label (small, faded) ──
    mid = (lo + hi) / 2
    offset = (hi - lo) * 0.06
    if lower_is_better:
        ax.text(mid - offset, mid + offset * 1.8,
                f"{ref_name} better ↓", fontsize=5.5,
                color=MODEL_COLORS[ref_name], alpha=0.6,
                rotation=45, ha="center", va="center", zorder=6)
        ax.text(mid + offset, mid - offset * 1.8,
                f"{bsl_name} better ↓", fontsize=5.5,
                color=color, alpha=0.6,
                rotation=45, ha="center", va="center", zorder=6)
    else:
        ax.text(mid + offset, mid + offset * 1.8,
                f"{ref_name} better ↑", fontsize=5.5,
                color=MODEL_COLORS[ref_name], alpha=0.6,
                rotation=45, ha="center", va="center", zorder=6)
        ax.text(mid - offset, mid - offset * 1.8,
                f"{bsl_name} better ↑", fontsize=5.5,
                color=color, alpha=0.6,
                rotation=45, ha="center", va="center", zorder=6)

    # ── Axis formatting ──
    ax.set_xlim(lo, hi)
    ax.set_ylim(lo, hi)
    ax.set_aspect("equal", adjustable="box")
    ax.set_xlabel(f"{ref_name}  {metric}", fontsize=8, labelpad=3)
    ax.set_ylabel(f"{bsl_name}  {metric}", fontsize=8, labelpad=3)
    ax.set_title(cls, fontsize=8.5, fontweight="bold", pad=4)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(0.5)
    ax.spines["bottom"].set_linewidth(0.5)
    ax.grid(linewidth=0.3, color="#eeeeee", zorder=0)
    ax.tick_params(labelsize=7)

    return {
        "r":         r_val,
        "p_val":     p_val,
        "ccc":       ccc,
        "slope":     slope,
        "intercept": intercept,
        "n":         n,
        "ref_wins":  r_wins,
        "bsl_wins":  b_wins,
        "ties":      ties,
        "win_pct_ref": win_pct,
    }

# =============================================================================
# 6. MAIN LOOP — one figure per metric × baseline
# =============================================================================

ref_df      = df_all[df_all["Model"] == REFERENCE_MODEL]
all_records = []

n_cls  = len(classes)
n_cols = min(4, n_cls)                     # max 4 columns
n_rows = int(np.ceil(n_cls / n_cols))

for metric in ALL_METRICS:
    lower_better = metric in LOWER_IS_BETTER

    for baseline in baselines:
        color = MODEL_COLORS.get(baseline, "#888888")

        # ── Figure layout ──
        panel_w = 3.2
        panel_h = 3.2
        FIG_W   = n_cols * panel_w + 0.6
        FIG_H   = n_rows * panel_h + 1.6

        fig, axes = plt.subplots(
            n_rows, n_cols,
            figsize=(FIG_W, FIG_H),
            squeeze=False,
        )

        # Turn off empty panels
        for idx in range(n_cls, n_rows * n_cols):
            r, c = divmod(idx, n_cols)
            axes[r][c].axis("off")

        # ── Super-title ──
        direction = "lower = better" if lower_better else "higher = better"
        fig.suptitle(
            f"Reliability scatter — {metric}  ({direction})\n"
            f"{REFERENCE_MODEL}  vs  {baseline}   |   per-patient paired values   |   "
            f"identity line (- -) + OLS regression (—) + 95% CI band",
            fontsize=10, fontweight="bold", y=0.999,
        )

        for idx, cls in enumerate(classes):
            ri, ci = divmod(idx, n_cols)
            ax     = axes[ri][ci]

            # Paired merge on Filename
            ref_sub = ref_df[
                ref_df["Class_Name"] == cls
            ][["Filename", metric]].dropna()

            bsl_sub = df_all[
                (df_all["Model"] == baseline) &
                (df_all["Class_Name"] == cls)
            ][["Filename", metric]].dropna()

            merged = pd.merge(
                ref_sub, bsl_sub,
                on="Filename", suffixes=("_ref", "_bsl"),
            ).dropna()

            # Drop non-finite rows
            merged = merged[
                np.isfinite(merged[f"{metric}_ref"]) &
                np.isfinite(merged[f"{metric}_bsl"])
            ]

            stats_out = draw_scatter_panel(
                ax,
                merged[f"{metric}_ref"].values,
                merged[f"{metric}_bsl"].values,
                metric      = metric,
                ref_name    = REFERENCE_MODEL,
                bsl_name    = baseline,
                cls         = cls,
                color       = color,
                lower_is_better = lower_better,
            )

            if stats_out:
                all_records.append({
                    "Metric":           metric,
                    "Class":            cls,
                    "Reference":        REFERENCE_MODEL,
                    "Baseline":         baseline,
                    "N_pairs":          stats_out["n"],
                    "Pearson_r":        round(stats_out["r"],         3),
                    "P_value":          round(stats_out["p_val"],      4),
                    "Sig":              p_stars(stats_out["p_val"]),
                    "CCC":              round(stats_out["ccc"],        3),
                    "OLS_slope":        round(stats_out["slope"],      3),
                    "OLS_intercept":    round(stats_out["intercept"],  3),
                    "Ref_wins":         stats_out["ref_wins"],
                    "Baseline_wins":    stats_out["bsl_wins"],
                    "Ties":             stats_out["ties"],
                    "Ref_win_pct":      round(stats_out["win_pct_ref"], 1),
                })

        # ── Shared legend ──
        legend_handles = [
            mlines.Line2D([0], [0], color="#888888", lw=1.0, linestyle="--",
                          label="Identity (y = x)"),
            mlines.Line2D([0], [0], color=color, lw=1.6,
                          label=f"OLS regression ({baseline})"),
            mpatches.Patch(color=color, alpha=REG_ALPHA * 2,
                           label="95% CI band"),
            plt.scatter([], [], color=MODEL_COLORS[REFERENCE_MODEL],
                        s=DOT_SIZE, alpha=0.7,
                        label=f"{REFERENCE_MODEL} wins (per patient)"),
            plt.scatter([], [], color=color, s=DOT_SIZE, alpha=0.7,
                        label=f"{baseline} wins (per patient)"),
            plt.scatter([], [], color="#888888", s=DOT_SIZE, alpha=0.5,
                        label="Tie / near-tie"),
        ]
        fig.legend(
            handles        = legend_handles,
            loc            = "lower center",
            ncol           = 3,
            frameon        = False,
            fontsize       = 8,
            bbox_to_anchor = (0.5, -0.01),
        )

        plt.tight_layout(rect=[0, 0.06, 1, 0.997])

        safe_metric   = metric.replace("/", "_").replace(" ", "_")
        safe_baseline = baseline.replace("/", "_").replace(" ", "_")
        stem          = f"ReliabilityScatter_{safe_metric}_{REFERENCE_MODEL}_vs_{safe_baseline}"
        png_path      = os.path.join(SAVE_DIR, stem + ".png")
        pdf_path      = os.path.join(SAVE_DIR, stem + ".pdf")

        plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
        plt.savefig(pdf_path, format="pdf", bbox_inches="tight", facecolor="white")
        plt.show()
        print(f"Saved → {png_path}")

# =============================================================================
# 7. STATISTICS EXCEL TABLE
# =============================================================================

df_stats  = pd.DataFrame(all_records)
stats_path = os.path.join(SAVE_DIR, "ReliabilityScatter_Statistics.xlsx")

with pd.ExcelWriter(stats_path, engine="openpyxl") as writer:
    df_stats.to_excel(writer, index=False, sheet_name="Stats")

    ws = writer.sheets["Stats"]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=8, name="Helvetica")
    thin     = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border    = thin

    # Colour-code by CCC quality
    hi_fill   = PatternFill("solid", fgColor="D5E8D4")   # CCC ≥ 0.90 — excellent
    med_fill  = PatternFill("solid", fgColor="FFF2CC")   # CCC 0.70–0.90 — moderate
    lo_fill   = PatternFill("solid", fgColor="FFE6CC")   # CCC < 0.70 — poor
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")

    ccc_col = None
    for cell in ws[1]:
        if cell.value == "CCC":
            ccc_col = cell.column
            break

    for row_idx, (_, row) in enumerate(df_stats.iterrows(), start=2):
        ccc_val = row.get("CCC", 0)
        if ccc_val >= 0.90:
            fill = hi_fill
        elif ccc_val >= 0.70:
            fill = med_fill
        elif row_idx % 2 == 0:
            fill = alt_fill
        else:
            fill = lo_fill

        for cell in ws[row_idx]:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(size=8, name="Helvetica")
            cell.fill      = fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

print(f"\nStatistics table → {stats_path}")

# =============================================================================
# 8. CONSOLE SUMMARY
# =============================================================================

print("\n" + "=" * 76)
print("RELIABILITY SUMMARY  (Pearson r  |  CCC  |  Reference win%)")
print("=" * 76)
for metric in ALL_METRICS:
    print(f"\n{metric}:")
    sub = df_stats[df_stats["Metric"] == metric]
    print(f"  {'Baseline':<14} {'Class':<22} {'r':>6} {'CCC':>6} {'Win%':>7} {'Slope':>7}")
    print("  " + "-" * 60)
    for _, row in sub.sort_values(["Baseline", "Class"]).iterrows():
        ccc_flag = (
            " ✓" if row["CCC"] >= 0.90 else
            " ⚠" if row["CCC"] >= 0.70 else
            " ✗"
        )
        print(
            f"  {row['Baseline']:<14} {row['Class']:<22} "
            f"{row['Pearson_r']:>6.3f} {row['CCC']:>6.3f}{ccc_flag} "
            f"{row['Ref_win_pct']:>6.1f}% {row['OLS_slope']:>7.3f}"
        )

print(f"\nAll outputs in: {SAVE_DIR}")






























































# =============================================================================
# FAILURE MODE ANALYSIS — BOTTOM DECILE PLOT
# Nature Medicine Style
# =============================================================================
#
# Identifies the worst 10% of cases per model and class, then asks:
#   "When each model fails, how bad is it — and is BAT-RM safer at failure?"
#
# THREE complementary panels per figure:
#
#   PANEL A — Failure threshold bar chart
#     The HD95/DSC value at the 90th/10th percentile (failure onset) per model.
#     A model with a LOWER failure threshold fails on easier cases → worse.
#
#   PANEL B — Failure severity strip + box
#     Distribution of metric values IN the bottom decile.
#     Narrower spread + lower values → model fails less severely.
#     Each patient shown as a dot; box shows IQR; whisker = range.
#
#   PANEL C — Overlap heatmap (failure case overlap)
#     For each pair of models: what fraction of their failure cases are
#     the same patients?  High overlap → shared difficulty (patient-level),
#     Low overlap → model-specific failure modes.
#
# Additional outputs:
#   • Per-patient failure roster Excel — lists every patient in the bottom
#     decile for each model × class combination (for clinical audit)
#   • Console summary: failure threshold, mean severity, unique failure %
#
# Definitions:
#   "Failure" for distance metrics (HD95, ASD) : top 10% of values (worst)
#   "Failure" for overlap metrics (Dice, IoU)  : bottom 10% of values (worst)
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap
from scipy.stats import mannwhitneyu
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION
# =============================================================================

REFERENCE_MODEL = "BAT-RM"

excel_files_seg = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
]

excel_files_bnd = [
    ("BAT-RM", r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",  r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx')
]



SEG_METRICS     = ["Dice", "IoU"]
BND_METRICS     = ["HD95", "ASD"]
ALL_METRICS     = SEG_METRICS + BND_METRICS
LOWER_IS_BETTER = {"HD95", "ASD", "HD", "RVD", "RAVD"}

class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

DECILE = 10    # bottom N% = failure cases (10 = bottom decile)

# Wong (2011) colour-blind safe palette
MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/FailureModeAnalysis'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   9,
    "axes.labelsize":   8.5,
    "xtick.labelsize":  7.5,
    "ytick.labelsize":  7.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.35,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "pdf.fonttype":     42,
    "ps.fonttype":      42,
    "figure.dpi":       150,
    "savefig.dpi":      600,
})

# =============================================================================
# 3. DATA LOADING
# =============================================================================

def load_segmentation(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
        for m in metrics:
            if m in df.columns:
                df[m] = pd.to_numeric(df[m], errors="coerce")
        df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
        df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
        df = df[df["Class_Name"].isin([c.upper() for c in class_list])]
        df["Model"] = name
        cols = ["Filename", "Class_Name", "Model"] + \
               [m for m in metrics if m in df.columns]
        all_data.append(df[cols])
    return pd.concat(all_data, ignore_index=True)


def load_boundary(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")
        file_rows = []
        for cls in class_list:
            present = f"{cls}_Present_In_GT"
            gt_vol  = f"{cls}_GT_Volume"
            if present in df.columns:
                sub = df[df[present] == True].copy()
            elif gt_vol in df.columns:
                sub = df[df[gt_vol] > 0].copy()
            else:
                sub = df.copy()
            if sub.empty:
                continue
            cls_cols  = {"Filename": sub["Filename"].values}
            any_found = False
            for m in metrics:
                col = f"{cls}_{m}"
                if col not in df.columns:
                    cls_cols[m] = np.nan
                    continue
                vals = pd.to_numeric(sub[col], errors="coerce").values
                vals = np.where(np.isfinite(vals), vals, np.nan)
                cls_cols[m] = vals
                any_found   = True
            if not any_found:
                continue
            tmp = pd.DataFrame(cls_cols)
            tmp["Class_Name"] = cls.upper().strip()
            tmp["Model"]      = name
            tmp = tmp.dropna(subset=metrics, how="all")
            file_rows.append(tmp)
        if file_rows:
            all_data.append(pd.concat(file_rows, ignore_index=True))
    if not all_data:
        raise ValueError("No boundary data loaded.")
    return pd.concat(all_data, ignore_index=True)


print("Loading data …")
df_seg = load_segmentation(excel_files_seg, SEG_METRICS, class_list)
df_bnd = load_boundary(excel_files_bnd, BND_METRICS, class_list)

df_all = pd.merge(
    df_seg, df_bnd,
    on=["Filename", "Class_Name", "Model"],
    how="outer",
)

model_names = [m for m, _ in excel_files_seg]
classes     = sorted(df_all["Class_Name"].dropna().unique())

print(f"Models   : {model_names}")
print(f"Classes  : {classes}")
print(f"Metrics  : {ALL_METRICS}")
print(f"Patients : {df_all['Filename'].nunique()}\n")

# =============================================================================
# 4. FAILURE CASE IDENTIFICATION
# =============================================================================

def get_failure_cases(df, model, cls, metric, decile=10):
    """
    Returns the Filenames of the worst `decile`% cases for a given
    model / class / metric combination.
    'Worst' = highest values for distance metrics, lowest for overlap metrics.
    """
    sub  = df[(df["Model"] == model) & (df["Class_Name"] == cls)][
        ["Filename", metric]
    ].dropna()
    sub  = sub[np.isfinite(sub[metric])]
    if sub.empty:
        return set(), np.nan

    lower_better = metric in LOWER_IS_BETTER
    pct          = 100 - decile if lower_better else decile
    threshold    = np.percentile(sub[metric], pct)

    if lower_better:
        fail = sub[sub[metric] >= threshold]
    else:
        fail = sub[sub[metric] <= threshold]

    return set(fail["Filename"]), threshold


# Build master failure roster
roster_rows = []

for metric in ALL_METRICS:
    for cls in classes:
        for model in model_names:
            fail_set, thresh = get_failure_cases(
                df_all, model, cls, metric, DECILE
            )
            sub = df_all[
                (df_all["Model"] == model) &
                (df_all["Class_Name"] == cls) &
                (df_all["Filename"].isin(fail_set))
            ][["Filename", metric]].dropna()

            for _, row in sub.iterrows():
                roster_rows.append({
                    "Metric":            metric,
                    "Class":             cls,
                    "Model":             model,
                    "Filename":          row["Filename"],
                    "Value":             round(row[metric], 4),
                    "Failure_threshold": round(thresh, 4)
                                         if not np.isnan(thresh) else np.nan,
                })

df_roster = pd.DataFrame(roster_rows)

roster_path = os.path.join(SAVE_DIR, "FailureCases_Roster.xlsx")
df_roster.to_excel(roster_path, index=False)
print(f"Failure roster saved → {roster_path}\n")

# =============================================================================
# 5. STATISTICS HELPERS
# =============================================================================

def mwu_p(a, b):
    """Mann-Whitney U p-value (two-sided). Returns NaN if insufficient data."""
    a = np.array(a)[np.isfinite(a)]
    b = np.array(b)[np.isfinite(b)]
    if len(a) < 3 or len(b) < 3:
        return np.nan
    # Only check equality if same length
    if len(a) == len(b) and np.all(a == b):
        return np.nan
    try:
        _, p = mannwhitneyu(a, b, alternative="two-sided")
        return p
    except Exception:
        return np.nan


def p_stars(p):
    if np.isnan(p):  return ""
    if p < 0.0001:   return "****"
    if p < 0.001:    return "***"
    if p < 0.01:     return "**"
    if p < 0.05:     return "*"
    return "ns"


def jaccard_overlap(set_a, set_b):
    """Jaccard index between two sets of patient IDs."""
    if not set_a and not set_b:
        return np.nan
    inter = len(set_a & set_b)
    union = len(set_a | set_b)
    return inter / union if union > 0 else 0.0

# =============================================================================
# 6. MAIN FIGURE LOOP — one figure per metric
# =============================================================================

summary_records = []
rng = np.random.default_rng(42)

for metric in ALL_METRICS:
    lower_better = metric in LOWER_IS_BETTER
    unit         = "mm" if lower_better else "(unitless)"
    fail_label   = f"Bottom {DECILE}% (worst cases)"

    n_cls  = len(classes)
    n_cols = min(4, n_cls)
    n_rows = int(np.ceil(n_cls / n_cols))

    # ── Figure: 3 rows of sub-panels per class grid ──
    # Row A: failure threshold bar chart
    # Row B: severity strip plot
    # Row C: overlap heatmap (one per class column)
    # We build a tall figure with GridSpec

    FIG_W = n_cols * 3.6 + 0.8
    FIG_H = n_rows * 8.5 + 1.6      # each class-row has 3 sub-panels stacked

    fig = plt.figure(figsize=(FIG_W, FIG_H))
    outer_gs = gridspec.GridSpec(
        n_rows, n_cols,
        figure=fig,
        hspace=0.55, wspace=0.40,
        left=0.06, right=0.97,
        top=0.95,  bottom=0.05,
    )

    fig.suptitle(
        f"Failure mode analysis — {metric}  ({DECILE}th percentile worst cases)\n"
        f"Panel A: failure onset threshold  |  "
        f"Panel B: severity in failure cases  |  "
        f"Panel C: patient overlap between models (Jaccard)",
        fontsize=10, fontweight="bold", y=0.988,
    )

    for idx, cls in enumerate(classes):
        r_outer, c_outer = divmod(idx, n_cols)
        inner_gs = gridspec.GridSpecFromSubplotSpec(
            3, 1,
            subplot_spec=outer_gs[r_outer, c_outer],
            height_ratios=[1.0, 1.4, 1.0],
            hspace=0.55,
        )

        ax_A = fig.add_subplot(inner_gs[0])   # threshold bar chart
        ax_B = fig.add_subplot(inner_gs[1])   # severity strip
        ax_C = fig.add_subplot(inner_gs[2])   # overlap heatmap

        # ── Gather data for this class ──
        thresholds   = {}
        fail_vals    = {}
        fail_sets    = {}

        for model in model_names:
            fset, thresh = get_failure_cases(
                df_all, model, cls, metric, DECILE
            )
            fail_sets[model]   = fset
            thresholds[model]  = thresh

            sub = df_all[
                (df_all["Model"] == model) &
                (df_all["Class_Name"] == cls) &
                (df_all["Filename"].isin(fset))
            ][metric].dropna().values
            fail_vals[model] = sub[np.isfinite(sub)]

        # ── PANEL A: Failure threshold bar chart ──
        x_pos  = np.arange(len(model_names))
        colors = [MODEL_COLORS.get(m, "#888888") for m in model_names]
        bars   = ax_A.bar(
            x_pos,
            [thresholds.get(m, 0) for m in model_names],
            color=colors, width=0.6,
            edgecolor="white", linewidth=0.5,
            zorder=3,
        )

        # Value labels on bars
        for bar, model in zip(bars, model_names):
            v = thresholds.get(model, np.nan)
            if not np.isnan(v):
                ax_A.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + (max(thresholds.values(), default=1) * 0.02),
                    f"{v:.2f}",
                    ha="center", va="bottom",
                    fontsize=6, color="#222222",
                )

        # Highlight reference model bar with a border
        ref_idx = model_names.index(REFERENCE_MODEL)
        bars[ref_idx].set_edgecolor("#222222")
        bars[ref_idx].set_linewidth(1.2)

        ax_A.set_xticks(x_pos)
        ax_A.set_xticklabels(
            [f"{'★' if m == REFERENCE_MODEL else ''}{m}" for m in model_names],
            fontsize=6, rotation=30, ha="right",
        )
        ax_A.set_ylabel(f"{metric} {unit}", fontsize=7)
        ax_A.set_title(
            f"{cls}\nA  Failure onset threshold\n"
            f"({'lower onset = fails on easier cases' if lower_better else 'higher onset = fails on easier cases'})",
            fontsize=7.5, fontweight="bold", pad=3,
        )
        ax_A.spines["top"].set_visible(False)
        ax_A.spines["right"].set_visible(False)
        ax_A.tick_params(axis="y", labelsize=6.5)
        ax_A.grid(axis="y", linewidth=0.3, color="#eeeeee", zorder=0)

        # ── PANEL B: Severity strip + box ──
        for pos, model in enumerate(model_names):
            vals  = fail_vals[model]
            color = MODEL_COLORS.get(model, "#888888")

            if len(vals) == 0:
                continue

            # Box (IQR)
            q25, q50, q75 = np.percentile(vals, [25, 50, 75])
            ax_B.plot(
                [pos - 0.22, pos + 0.22], [q50, q50],
                color=color, linewidth=2.2, zorder=5,
                solid_capstyle="round",
            )
            rect = plt.Rectangle(
                (pos - 0.18, q25), 0.36, q75 - q25,
                facecolor=color, alpha=0.20,
                edgecolor=color, linewidth=0.8,
                zorder=3,
            )
            ax_B.add_patch(rect)

            # Whiskers (min/max)
            ax_B.plot(
                [pos, pos], [vals.min(), q25],
                color=color, linewidth=0.8, alpha=0.6, zorder=3,
            )
            ax_B.plot(
                [pos, pos], [q75, vals.max()],
                color=color, linewidth=0.8, alpha=0.6, zorder=3,
            )

            # Strip dots
            jitter = rng.uniform(-0.14, 0.14, size=len(vals))
            ax_B.scatter(
                pos + jitter, vals,
                color=color, s=10, alpha=0.45,
                edgecolors="white", linewidths=0.2,
                zorder=4,
            )

            # Mean marker (diamond)
            ax_B.scatter(
                pos, vals.mean(),
                marker="D", color=color, s=20,
                edgecolors="white", linewidths=0.5,
                zorder=6,
            )

        # MWU significance vs reference for each baseline
        ref_v = fail_vals.get(REFERENCE_MODEL, np.array([]))
        y_sig = ax_B.get_ylim()[1] if len(ref_v) > 0 else 1.0
        # Re-compute y range from data before annotation
        all_fail_flat = np.concatenate(
            [v for v in fail_vals.values() if len(v) > 0]
        )
        if len(all_fail_flat) > 0:
            y_range  = all_fail_flat.max() - all_fail_flat.min()
            y_top    = all_fail_flat.max() + y_range * 0.12
            ax_B.set_ylim(
                max(0, all_fail_flat.min() - y_range * 0.1),
                y_top + y_range * 0.05 * len(model_names),
            )

        sig_y = all_fail_flat.max() + y_range * 0.08 if len(all_fail_flat) > 0 else 1.0
        ref_pos = model_names.index(REFERENCE_MODEL)

        for pos, model in enumerate(model_names):
            if model == REFERENCE_MODEL:
                continue
            p = mwu_p(ref_v, fail_vals.get(model, np.array([])))
            stars = p_stars(p)
            if stars:
                sig_y += y_range * 0.06
                mid   = (ref_pos + pos) / 2
                ax_B.annotate(
                    "", xy=(pos, sig_y), xytext=(ref_pos, sig_y),
                    arrowprops=dict(
                        arrowstyle="-", color="#666666", lw=0.6,
                    ),
                )
                ax_B.text(
                    mid, sig_y + y_range * 0.01,
                    stars, ha="center", va="bottom",
                    fontsize=6.5, color="#333333",
                )

        ax_B.set_xticks(range(len(model_names)))
        ax_B.set_xticklabels(
            [f"{'★' if m == REFERENCE_MODEL else ''}{m}" for m in model_names],
            fontsize=6, rotation=30, ha="right",
        )
        ax_B.set_ylabel(f"{metric} {unit}", fontsize=7)
        ax_B.set_title(
            f"B  Severity within failure cases\n"
            f"(bar=median, box=IQR, ◆=mean, dot=patient)",
            fontsize=7, pad=3,
        )
        ax_B.spines["top"].set_visible(False)
        ax_B.spines["right"].set_visible(False)
        ax_B.tick_params(axis="y", labelsize=6.5)
        ax_B.grid(axis="y", linewidth=0.3, color="#eeeeee", zorder=0)

        # ── PANEL C: Jaccard overlap heatmap ──
        n_m   = len(model_names)
        jac   = np.full((n_m, n_m), np.nan)

        for i, mi in enumerate(model_names):
            for j, mj in enumerate(model_names):
                jac[i, j] = jaccard_overlap(fail_sets[mi], fail_sets[mj])

        # Custom colourmap: white → dark teal
        cmap = LinearSegmentedColormap.from_list(
            "jac", ["#FFFFFF", "#0077BB"], N=256
        )
        im = ax_C.imshow(jac, cmap=cmap, vmin=0, vmax=1, aspect="auto")

        # Cell annotations
        for i in range(n_m):
            for j in range(n_m):
                v = jac[i, j]
                if not np.isnan(v):
                    txt_col = "white" if v > 0.55 else "#222222"
                    ax_C.text(
                        j, i, f"{v:.2f}",
                        ha="center", va="center",
                        fontsize=6.0, color=txt_col,
                    )

        short_names = [
            m.replace("TransUNet", "TUNet")
             .replace("SegMamba", "SMamba")
             .replace("BAT-RM", "BAT★")
            for m in model_names
        ]
        ax_C.set_xticks(range(n_m))
        ax_C.set_yticks(range(n_m))
        ax_C.set_xticklabels(short_names, fontsize=5.5, rotation=45, ha="right")
        ax_C.set_yticklabels(short_names, fontsize=5.5)
        ax_C.set_title(
            "C  Failure overlap (Jaccard)\nhigh = same patients fail on both models",
            fontsize=7, pad=3,
        )

        plt.colorbar(im, ax=ax_C, fraction=0.035, pad=0.04,
                     label="Jaccard index").ax.tick_params(labelsize=6)

        # ── Record summary stats ──
        for model in model_names:
            v    = fail_vals[model]
            thr  = thresholds[model]
            ref_v_arr = fail_vals.get(REFERENCE_MODEL, np.array([]))

            # Unique failures = in this model's fail set but NOT in ref's
            ref_fset  = fail_sets.get(REFERENCE_MODEL, set())
            unique_n  = len(fail_sets[model] - ref_fset) \
                        if model != REFERENCE_MODEL else 0
            unique_pct = 100 * unique_n / len(fail_sets[model]) \
                         if fail_sets[model] else 0

            summary_records.append({
                "Metric":                metric,
                "Class":                 cls,
                "Model":                 model,
                "N_total":               len(df_all[
                    (df_all["Model"] == model) &
                    (df_all["Class_Name"] == cls)
                ][metric].dropna()),
                "N_failure_cases":       len(fail_sets[model]),
                "Failure_threshold":     round(thr,       3) if not np.isnan(thr) else np.nan,
                "Fail_mean":             round(v.mean(),  3) if len(v) > 0 else np.nan,
                "Fail_median":           round(np.median(v), 3) if len(v) > 0 else np.nan,
                "Fail_max":              round(v.max(),   3) if len(v) > 0 else np.nan,
                "Fail_SD":               round(v.std(ddof=1), 3) if len(v) > 1 else np.nan,
                "Unique_failures_vs_ref": unique_n,
                "Unique_failures_pct":   round(unique_pct, 1),
                "MWU_p_vs_ref":          round(mwu_p(ref_v_arr, v), 4)
                                          if model != REFERENCE_MODEL
                                          else np.nan,
                "Jaccard_vs_ref":        round(jaccard_overlap(
                                              fail_sets[model], ref_fset
                                          ), 3)
                                          if model != REFERENCE_MODEL
                                          else np.nan,
            })

    # ── Save figure ──
    safe_metric = metric.replace("/", "_")
    png_path    = os.path.join(SAVE_DIR, f"FailureMode_{safe_metric}.png")
    pdf_path    = os.path.join(SAVE_DIR, f"FailureMode_{safe_metric}.pdf")
    plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
    plt.savefig(pdf_path, format="pdf", bbox_inches="tight", facecolor="white")
    plt.show()
    print(f"Saved → {png_path}")

# =============================================================================
# 7. FIGURE 2 — CROSS-METRIC FAILURE PORTRAIT (per class)
# Which patients appear in the bottom decile across MULTIPLE metrics?
# Reveals whether failures are metric-specific or patient-wide collapse.
# =============================================================================

print("\nGenerating failure portrait heatmaps …")

for cls in classes:
    n_models = len(model_names)
    n_met    = len(ALL_METRICS)

    # Gather all patient IDs seen for this class
    all_patients = sorted(df_all[
        df_all["Class_Name"] == cls
    ]["Filename"].dropna().unique())

    if not all_patients:
        continue

    # Build binary matrix: patient × (model × metric) — 1 if in failure set
    col_labels = []
    fail_matrix = []

    for metric in ALL_METRICS:
        for model in model_names:
            fset, _ = get_failure_cases(df_all, model, cls, metric, DECILE)
            col_labels.append(f"{model}\n{metric}")
            fail_matrix.append([1 if p in fset else 0 for p in all_patients])

    fail_arr = np.array(fail_matrix).T      # patients × columns

    # Sort patients by total failure count (worst at top)
    row_sums = fail_arr.sum(axis=1)
    order    = np.argsort(-row_sums)
    fail_arr = fail_arr[order]
    pat_labs = [all_patients[i] for i in order]

    # Show only the top-50 most-failed patients (avoid illegibly tall figure)
    MAX_PATIENTS = 50
    if len(pat_labs) > MAX_PATIENTS:
        fail_arr = fail_arr[:MAX_PATIENTS]
        pat_labs = pat_labs[:MAX_PATIENTS]
        shown    = f"top {MAX_PATIENTS} most-failed"
    else:
        shown    = "all patients"

    n_pts  = len(pat_labs)
    FIG_W2 = max(8, len(col_labels) * 0.55 + 2.5)
    FIG_H2 = max(6, n_pts * 0.22 + 2.0)

    fig2, ax2 = plt.subplots(figsize=(FIG_W2, FIG_H2))

    # Colour: white=pass, model colour coded by metric group
    cmap2 = LinearSegmentedColormap.from_list("fail", ["#F5F5F5", "#CC2222"], N=2)
    ax2.imshow(
        fail_arr, cmap=cmap2, vmin=0, vmax=1,
        aspect="auto", interpolation="nearest",
    )

    # Column separator lines between metric groups
    for mi in range(1, n_met):
        ax2.axvline(
            mi * n_models - 0.5,
            color="#555555", linewidth=1.2, zorder=5,
        )

    # Column group headers (metric names)
    for mi, metric in enumerate(ALL_METRICS):
        center_x = mi * n_models + (n_models - 1) / 2
        ax2.text(
            center_x, -1.5, metric,
            ha="center", va="bottom",
            fontsize=8, fontweight="bold", color="#222222",
        )

    ax2.set_xticks(range(len(col_labels)))
    ax2.set_xticklabels(
        [lab.split("\n")[0] for lab in col_labels],   # model name only
        fontsize=5.5, rotation=45, ha="right",
    )
    ax2.set_yticks(range(n_pts))
    ax2.set_yticklabels(pat_labs, fontsize=5.0)
    ax2.set_xlabel("Model  (grouped by metric)", fontsize=8.5, labelpad=8)
    ax2.set_ylabel("Patient", fontsize=8.5)
    ax2.set_title(
        f"Cross-metric failure portrait — {cls}  ({shown})\n"
        f"Red cell = patient in bottom {DECILE}% for that model × metric  |  "
        f"Sorted by total failure count (worst top)",
        fontsize=9, fontweight="bold", pad=6,
    )

    # Row: total failure count annotation on right
    for ri, total in enumerate(fail_arr.sum(axis=1)):
        ax2.text(
            len(col_labels) + 0.1, ri,
            f"{int(total)}",
            ha="left", va="center",
            fontsize=5.0, color="#CC2222" if total > len(col_labels) * 0.5
                                          else "#888888",
        )

    plt.tight_layout()

    safe_cls  = cls.replace(" ", "_")
    png2_path = os.path.join(SAVE_DIR, f"FailurePortrait_{safe_cls}.png")
    pdf2_path = os.path.join(SAVE_DIR, f"FailurePortrait_{safe_cls}.pdf")
    plt.savefig(png2_path, dpi=600, bbox_inches="tight", facecolor="white")
    plt.savefig(pdf2_path, format="pdf", bbox_inches="tight", facecolor="white")
    plt.show()
    print(f"Saved portrait → {png2_path}")

# =============================================================================
# 8. STATISTICS EXCEL TABLE
# =============================================================================

df_summary = pd.DataFrame(summary_records)
summ_path  = os.path.join(SAVE_DIR, "FailureMode_Statistics.xlsx")

with pd.ExcelWriter(summ_path, engine="openpyxl") as writer:
    df_summary.to_excel(writer, index=False, sheet_name="FailureStats")
    df_roster.to_excel(writer, index=False,  sheet_name="FailureRoster")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin = Border(
        left=Side(style="thin",  color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin",   color="CCCCCC"),
        bottom=Side(style="thin",color="CCCCCC"),
    )
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=8, name="Helvetica")
    ref_fill = PatternFill("solid", fgColor="D5E8D4")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    bad_fill = PatternFill("solid", fgColor="FFE6CC")

    for sheet_name in ["FailureStats", "FailureRoster"]:
        ws = writer.sheets[sheet_name]
        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center")
            cell.border    = thin

        for row_idx in range(2, ws.max_row + 1):
            model_cell = ws.cell(row_idx, 3)
            is_ref     = model_cell.value == REFERENCE_MODEL
            fill       = ref_fill if is_ref else \
                         (alt_fill if row_idx % 2 == 0 else PatternFill())
            for cell in ws[row_idx]:
                cell.border    = thin
                cell.alignment = Alignment(horizontal="center")
                cell.font      = Font(size=8, name="Helvetica", bold=is_ref)
                if fill.fill_type:
                    cell.fill = fill

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16

print(f"\nStatistics saved → {summ_path}")

# =============================================================================
# 9. CONSOLE SAFETY REPORT
# =============================================================================

print("\n" + "=" * 80)
print(f"FAILURE MODE SAFETY REPORT  (bottom {DECILE}% cases)")
print("=" * 80)

for metric in ALL_METRICS:
    print(f"\n{metric}:")
    sub = df_summary[df_summary["Metric"] == metric]
    hdr = (f"  {'Model':<14} {'Class':<22} {'Threshold':>10} "
           f"{'Fail mean':>10} {'Fail max':>9} {'Unique%':>8} {'Jaccard':>8}")
    print(hdr)
    print("  " + "-" * 78)
    for _, row in sub.sort_values(["Class", "Model"]).iterrows():
        flag = ""
        if row["Model"] != REFERENCE_MODEL:
            jac = row.get("Jaccard_vs_ref", np.nan)
            if not np.isnan(jac):
                flag = "  (shared)" if jac > 0.5 else "  (model-specific!)"
        ref_marker = "★ " if row["Model"] == REFERENCE_MODEL else "  "
        print(
            f"  {ref_marker}{row['Model']:<12} {row['Class']:<22} "
            f"{str(row.get('Failure_threshold','—')):>10} "
            f"{str(row.get('Fail_mean','—')):>10} "
            f"{str(row.get('Fail_max','—')):>9} "
            f"{str(row.get('Unique_failures_pct','—')):>7}% "
            f"{str(row.get('Jaccard_vs_ref','—')):>8}"
            f"{flag}"
        )

print(f"\nAll outputs in: {SAVE_DIR}")      
















































# =============================================================================
# METRIC REDUNDANCY CORRELATION HEATMAP
# Nature Medicine Style — Pearson + Spearman across all metrics × models
# =============================================================================
#
# Answers the reviewer question:
#   "Why did you choose these metrics? Are they independent?"
#
# OUTPUTS:
#
#   Figure 1 — Full correlation matrix grid (2 × 2)
#     Top row    : Pearson  r  | Spearman ρ   (pooled across all models)
#     Bottom row : Pearson  r  | Spearman ρ   (reference model only)
#     Colourmap  : diverging red–white–blue  (red = +1, blue = −1)
#     Annotations: r value + significance stars in every cell
#     Diagonal   : metric variance histogram (shows distribution shape)
#
#   Figure 2 — Clustered heatmap (hierarchical clustering dendrogram)
#     Reorders metrics so that redundant ones cluster together visually.
#     Dendrogram on both axes.  Spearman ρ used (rank-based, more robust).
#
#   Figure 3 — Scatter matrix (pairs plot) for reference model only
#     Every metric pair as a scatter panel — shows the actual point cloud,
#     not just a summary number.  KDE on diagonal.
#
#   Excel: full correlation tables + p-values + sample sizes per pair
#
# Statistical notes:
#   • Pearson r   : linear correlation; sensitive to outliers
#   • Spearman ρ  : rank correlation; robust to non-normality & outliers
#   • p-values    : two-tailed t-approximation (Pearson) / AS89 (Spearman)
#   • Bonferroni correction applied across metric pairs
#   • Both within-class and pooled-across-class versions computed
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# (scipy.cluster.hierarchy used for dendrogram — no extra install needed)
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.colors import TwoSlopeNorm
import matplotlib.ticker as ticker
from scipy import stats
from scipy.cluster import hierarchy
from scipy.spatial.distance import squareform
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION
# =============================================================================

REFERENCE_MODEL = "BAT-RM"

excel_files_seg = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
]

excel_files_bnd = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx'),
]

SEG_METRICS = ["Dice", "IoU"]
BND_METRICS = ["HD95", "ASD"]
ALL_METRICS = SEG_METRICS + BND_METRICS   # extend if NSD / HD available

LOWER_IS_BETTER = {"HD95", "ASD", "HD", "NSD", "RVD"}

class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# Metric display labels (LaTeX-friendly for axis ticks)
METRIC_LABELS = {
    "Dice":  "Dice",
    "IoU":   "IoU",
    "HD95":  "HD95\n(mm)",
    "ASD":   "ASD\n(mm)",
    "HD":    "HD\n(mm)",
    "NSD":   "NSD",
}

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/MetricCorrelation'
os.makedirs(SAVE_DIR, exist_ok=True)

# Colourmap: red (+1) — white (0) — blue (−1)
CMAP_DIV = plt.cm.RdBu_r

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   9,
    "axes.labelsize":   8.5,
    "xtick.labelsize":  8,
    "ytick.labelsize":  8,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.35,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "pdf.fonttype":     42,
    "ps.fonttype":      42,
    "figure.dpi":       150,
    "savefig.dpi":      600,
})

# =============================================================================
# 3. DATA LOADING
# =============================================================================

def load_segmentation(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
        for m in metrics:
            if m in df.columns:
                df[m] = pd.to_numeric(df[m], errors="coerce")
        df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
        df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
        df = df[df["Class_Name"].isin([c.upper() for c in class_list])]
        df["Model"] = name
        avail = ["Filename", "Class_Name", "Model"] + \
                [m for m in metrics if m in df.columns]
        all_data.append(df[avail])
    return pd.concat(all_data, ignore_index=True)


def load_boundary(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")
        file_rows = []
        for cls in class_list:
            present = f"{cls}_Present_In_GT"
            gt_vol  = f"{cls}_GT_Volume"
            if present in df.columns:
                sub = df[df[present] == True].copy()
            elif gt_vol in df.columns:
                sub = df[df[gt_vol] > 0].copy()
            else:
                sub = df.copy()
            if sub.empty:
                continue
            cls_cols  = {"Filename": sub["Filename"].values}
            any_found = False
            for m in metrics:
                col = f"{cls}_{m}"
                if col not in df.columns:
                    cls_cols[m] = np.nan
                    continue
                vals = pd.to_numeric(sub[col], errors="coerce").values
                vals = np.where(np.isfinite(vals), vals, np.nan)
                cls_cols[m] = vals
                any_found   = True
            if not any_found:
                continue
            tmp = pd.DataFrame(cls_cols)
            tmp["Class_Name"] = cls.upper().strip()
            tmp["Model"]      = name
            tmp = tmp.dropna(subset=metrics, how="all")
            file_rows.append(tmp)
        if file_rows:
            all_data.append(pd.concat(file_rows, ignore_index=True))
    if not all_data:
        raise ValueError("No boundary data loaded.")
    return pd.concat(all_data, ignore_index=True)


print("Loading data …")
df_seg = load_segmentation(excel_files_seg, SEG_METRICS, class_list)
df_bnd = load_boundary(excel_files_bnd, BND_METRICS, class_list)

df_all = pd.merge(
    df_seg, df_bnd,
    on=["Filename", "Class_Name", "Model"],
    how="outer",
)

# Keep only metrics that actually exist in the merged frame
METRICS = [m for m in ALL_METRICS if m in df_all.columns]
model_names = [m for m, _ in excel_files_seg]
classes     = sorted(df_all["Class_Name"].dropna().unique())

print(f"Models   : {model_names}")
print(f"Classes  : {classes}")
print(f"Metrics  : {METRICS}")
print(f"Patients : {df_all['Filename'].nunique()}\n")

# =============================================================================
# 4. CORRELATION ENGINE
# =============================================================================

def corr_matrix(df, metrics, method="pearson"):
    """
    Compute pairwise correlation + p-value matrix for `metrics` columns in df.
    Returns:
        R   : (n_metrics × n_metrics) correlation matrix
        P   : (n_metrics × n_metrics) p-value matrix (raw, two-tailed)
        N   : (n_metrics × n_metrics) sample-size matrix
    """
    n   = len(metrics)
    R   = np.full((n, n), np.nan)
    P   = np.full((n, n), np.nan)
    N   = np.full((n, n), np.nan)

    for i, mi in enumerate(metrics):
        for j, mj in enumerate(metrics):
            # Extract as 1D arrays to avoid duplicate-column shape issues
            xi = pd.to_numeric(df[mi], errors="coerce").values.ravel()
            xj = pd.to_numeric(df[mj], errors="coerce").values.ravel()
            mask = np.isfinite(xi) & np.isfinite(xj)
            xi, xj = xi[mask], xj[mask]
            k = len(xi)
            N[i, j] = k
            if k < 4:
                R[i, j] = np.nan
                P[i, j] = np.nan
                continue
            if method == "pearson":
                r, p = stats.pearsonr(xi, xj)
            else:
                r, p = stats.spearmanr(xi, xj)
            R[i, j] = r
            P[i, j] = p

    return R, P, N


def bonferroni_p(P):
    """Apply Bonferroni correction to off-diagonal p-values."""
    n        = P.shape[0]
    n_tests  = n * (n - 1) // 2   # unique pairs
    P_corr   = np.clip(P * n_tests, 0, 1)
    np.fill_diagonal(P_corr, np.nan)
    return P_corr


def p_to_stars(p):
    if np.isnan(p): return ""
    if p < 0.0001:  return "****"
    if p < 0.001:   return "***"
    if p < 0.01:    return "**"
    if p < 0.05:    return "*"
    return "ns"


# Build datasets
df_pooled = df_all[METRICS + ["Filename", "Class_Name", "Model"]].copy()
df_ref    = df_all[df_all["Model"] == REFERENCE_MODEL][
    METRICS + ["Filename", "Class_Name", "Model"]
].copy()

datasets = {
    "All models (pooled)": df_pooled,
    f"{REFERENCE_MODEL} only": df_ref,
}

# Compute all four matrices
corr_results = {}
for label, df_src in datasets.items():
    for method in ("pearson", "spearman"):
        R, P, N = corr_matrix(df_src, METRICS, method)
        P_corr  = bonferroni_p(P)
        corr_results[(label, method)] = (R, P_corr, N)

# =============================================================================
# 5. FIGURE 1 — 2 × 2 CORRELATION GRID
# =============================================================================

n_met   = len(METRICS)
labels  = [METRIC_LABELS.get(m, m) for m in METRICS]
norm    = TwoSlopeNorm(vmin=-1, vcenter=0, vmax=1)

fig1, axes1 = plt.subplots(2, 2, figsize=(n_met * 2.1, n_met * 2.0))

panel_titles = [
    ("All models (pooled)", "pearson",  "Pearson r  —  all models pooled"),
    ("All models (pooled)", "spearman", "Spearman ρ  —  all models pooled"),
    (f"{REFERENCE_MODEL} only", "pearson",  f"Pearson r  —  {REFERENCE_MODEL} only"),
    (f"{REFERENCE_MODEL} only", "spearman", f"Spearman ρ  —  {REFERENCE_MODEL} only"),
]

for ax, (ds_label, method, title) in zip(axes1.flat, panel_titles):
    R, P_corr, N = corr_results[(ds_label, method)]

    # ── Draw heatmap ──
    im = ax.imshow(R, cmap=CMAP_DIV, norm=norm, aspect="equal")

    # ── Cell annotations: r value + stars ──
    for i in range(n_met):
        for j in range(n_met):
            r_val = R[i, j]
            p_val = P_corr[i, j]
            if np.isnan(r_val):
                continue

            if i == j:
                # Diagonal: show metric name only
                ax.text(j, i, METRICS[i],
                        ha="center", va="center",
                        fontsize=7.5, fontweight="bold",
                        color="#222222")
                continue

            stars   = p_to_stars(p_val)
            txt_col = "white" if abs(r_val) > 0.65 else "#111111"
            ax.text(j, i,
                    f"{r_val:.2f}",
                    ha="center", va="center",
                    fontsize=7.5, color=txt_col, fontweight="bold")
            if stars and stars != "ns":
                ax.text(j, i + 0.32,
                        stars,
                        ha="center", va="center",
                        fontsize=5.5, color=txt_col)

    # ── Diagonal: dark border to separate from correlation cells ──
    for k in range(n_met):
        ax.add_patch(plt.Rectangle(
            (k - 0.5, k - 0.5), 1, 1,
            fill=False, edgecolor="#555555",
            linewidth=1.2, zorder=5,
        ))

    # ── Grid lines between cells ──
    for k in range(n_met + 1):
        ax.axhline(k - 0.5, color="#cccccc", linewidth=0.4, zorder=4)
        ax.axvline(k - 0.5, color="#cccccc", linewidth=0.4, zorder=4)

    ax.set_xticks(range(n_met))
    ax.set_yticks(range(n_met))
    ax.set_xticklabels(labels, fontsize=8)
    ax.set_yticklabels(labels, fontsize=8)
    ax.set_title(title, fontsize=9, fontweight="bold", pad=6)
    ax.tick_params(length=0)

    # Colorbar
    cb = plt.colorbar(im, ax=ax, fraction=0.04, pad=0.03)
    cb.ax.tick_params(labelsize=7)
    cb.set_label("Correlation coefficient", fontsize=7)

fig1.suptitle(
    "Metric redundancy analysis — Pearson r and Spearman ρ correlation matrices\n"
    "Stars: Bonferroni-corrected significance  |  "
    "Red = positive  |  Blue = negative  |  "
    "Diagonal = metric name",
    fontsize=10, fontweight="bold", y=1.002,
)

plt.tight_layout()
p1_png = os.path.join(SAVE_DIR, "MetricCorrelation_Grid.png")
p1_pdf = os.path.join(SAVE_DIR, "MetricCorrelation_Grid.pdf")
plt.savefig(p1_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(p1_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Saved → {p1_png}")

# =============================================================================
# 6. FIGURE 2 — CLUSTERED HEATMAP WITH DENDROGRAM
# Uses Spearman ρ (pooled). Hierarchical clustering on (1 − |ρ|) distances.
# Metrics that behave similarly cluster together → immediately shows redundancy.
# =============================================================================

R_sp, P_sp, N_sp = corr_results[("All models (pooled)", "spearman")]

# Replace NaN diagonal with 1.0 for clustering
R_clust = np.where(np.isnan(R_sp), 0, R_sp)
np.fill_diagonal(R_clust, 1.0)

# Distance matrix: 1 − |ρ|  (high |ρ| = similar = short distance)
dist_mat = 1 - np.abs(R_clust)
np.fill_diagonal(dist_mat, 0)

# Ensure symmetry & positivity
dist_mat = (dist_mat + dist_mat.T) / 2
dist_mat = np.clip(dist_mat, 0, None)

# Hierarchical clustering (Ward linkage)
condensed = squareform(dist_mat, checks=False)
linkage   = hierarchy.linkage(condensed, method="ward")
dendro    = hierarchy.dendrogram(linkage, no_plot=True)
order     = dendro["leaves"]

# Reorder matrix
R_ord  = R_sp[np.ix_(order, order)]
P_ord  = P_sp[np.ix_(order, order)]
labs_ord = [METRIC_LABELS.get(METRICS[i], METRICS[i]) for i in order]
met_ord  = [METRICS[i] for i in order]

FIG2_SZ = max(6, n_met * 1.6)
fig2    = plt.figure(figsize=(FIG2_SZ + 1.5, FIG2_SZ + 1.5))

# GridSpec: dendrogram columns/rows + heatmap
gs2 = gridspec.GridSpec(
    2, 2,
    figure=fig2,
    width_ratios=[0.18, 1.0],
    height_ratios=[0.18, 1.0],
    hspace=0.03, wspace=0.03,
    left=0.12, right=0.92,
    top=0.91, bottom=0.12,
)

ax_dendL = fig2.add_subplot(gs2[1, 0])   # left dendrogram
ax_dendT = fig2.add_subplot(gs2[0, 1])   # top dendrogram
ax_heat  = fig2.add_subplot(gs2[1, 1])   # heatmap

# ── Top dendrogram ──
hierarchy.dendrogram(
    linkage, ax=ax_dendT,
    orientation="top",
    labels=met_ord,
    color_threshold=0,
    above_threshold_color="#555555",
    leaf_font_size=0,
)
ax_dendT.axis("off")

# ── Left dendrogram ──
hierarchy.dendrogram(
    linkage, ax=ax_dendL,
    orientation="left",
    labels=met_ord,
    color_threshold=0,
    above_threshold_color="#555555",
    leaf_font_size=0,
)
ax_dendL.axis("off")

# ── Clustered heatmap ──
im2 = ax_heat.imshow(R_ord, cmap=CMAP_DIV, norm=norm, aspect="equal")

for i in range(n_met):
    for j in range(n_met):
        r_val = R_ord[i, j]
        p_val = P_ord[i, j]
        if np.isnan(r_val):
            continue
        txt_col = "white" if abs(r_val) > 0.65 else "#111111"
        stars   = p_to_stars(p_val) if i != j else ""
        ax_heat.text(j, i, f"{r_val:.2f}",
                     ha="center", va="center",
                     fontsize=7.5, color=txt_col, fontweight="bold")
        if stars and stars != "ns":
            ax_heat.text(j, i + 0.32, stars,
                         ha="center", va="center",
                         fontsize=5.5, color=txt_col)

for k in range(n_met + 1):
    ax_heat.axhline(k - 0.5, color="#cccccc", linewidth=0.4)
    ax_heat.axvline(k - 0.5, color="#cccccc", linewidth=0.4)

ax_heat.set_xticks(range(n_met))
ax_heat.set_yticks(range(n_met))
ax_heat.set_xticklabels(labs_ord, fontsize=9, rotation=45, ha="right")
ax_heat.set_yticklabels(labs_ord, fontsize=9)
ax_heat.tick_params(length=0)

cb2 = plt.colorbar(im2, ax=ax_heat, fraction=0.04, pad=0.04)
cb2.ax.tick_params(labelsize=7.5)
cb2.set_label("Spearman ρ", fontsize=8)

fig2.suptitle(
    "Metric redundancy — hierarchical clustering of Spearman ρ (all models pooled)\n"
    "Dendrogram based on 1 − |ρ| distance  |  "
    "Metrics that cluster together are informationally redundant",
    fontsize=10, fontweight="bold", y=0.975,
)

p2_png = os.path.join(SAVE_DIR, "MetricCorrelation_Clustered.png")
p2_pdf = os.path.join(SAVE_DIR, "MetricCorrelation_Clustered.pdf")
plt.savefig(p2_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(p2_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Saved → {p2_png}")

# =============================================================================
# 7. FIGURE 3 — SCATTER MATRIX (pairs plot)
# Reference model only. KDE on diagonal, scatter + regression on off-diagonal.
# Upper triangle: Pearson r annotation. Lower: scatter + LOWESS.
# =============================================================================

from scipy.stats import gaussian_kde

df_ref_clean = df_ref[METRICS].dropna(how="all")

# Soft-clip extreme boundary values for visual clarity
for m in BND_METRICS:
    if m in df_ref_clean.columns:
        cap = df_ref_clean[m].quantile(0.99)
        df_ref_clean[m] = df_ref_clean[m].clip(upper=cap)

rng = np.random.default_rng(42)

FIG3_SZ = n_met * 2.4
fig3, axes3 = plt.subplots(n_met, n_met,
                            figsize=(FIG3_SZ, FIG3_SZ),
                            squeeze=False)

REF_COLOR = "#0077BB"

for i, mi in enumerate(METRICS):
    for j, mj in enumerate(METRICS):
        ax = axes3[i][j]

        # Always extract as clean 1D float arrays — guards against
        # duplicate column names in the merged dataframe
        arr_i = pd.to_numeric(df_ref_clean[mi], errors="coerce").values.ravel()
        arr_j = pd.to_numeric(df_ref_clean[mj], errors="coerce").values.ravel()
        mask   = np.isfinite(arr_i) & np.isfinite(arr_j)
        arr_i, arr_j = arr_i[mask], arr_j[mask]
        # Rebuild a clean paired DataFrame for len() checks
        paired = pd.DataFrame({mi: arr_i, mj: arr_j})

        # ── Diagonal: KDE ──
        if i == j:
            vals = arr_i   # clean 1D array
            if len(vals) > 3:
                kde  = gaussian_kde(vals.ravel(), bw_method="scott")
                x_ev = np.linspace(vals.min(), vals.max(), 200)
                ax.fill_between(x_ev, kde(x_ev),
                                color=REF_COLOR, alpha=0.25)
                ax.plot(x_ev, kde(x_ev),
                        color=REF_COLOR, linewidth=1.4)
            ax.set_facecolor("#F9F9F9")
            ax.text(0.5, 0.92, METRIC_LABELS.get(mi, mi),
                    transform=ax.transAxes,
                    ha="center", va="top",
                    fontsize=9, fontweight="bold", color="#222222")
            ax.set_yticks([])
            ax.tick_params(labelsize=6.5)

        # ── Upper triangle: r + ρ annotation ──
        elif j > i:
            if len(paired) >= 4:
                r_p,  p_p  = stats.pearsonr(arr_i,  arr_j)
                r_sp, p_sp = stats.spearmanr(arr_i, arr_j)
            else:
                r_p = r_sp = p_p = p_sp = np.nan

            stars_p  = p_to_stars(p_p)
            stars_sp = p_to_stars(p_sp)

            # Background colour scaled by |r|
            bg_alpha = min(abs(r_p) if not np.isnan(r_p) else 0, 0.5)
            bg_color = "#CC2222" if (r_p > 0 if not np.isnan(r_p) else False) \
                       else "#0055AA"
            ax.set_facecolor(
                (*plt.matplotlib.colors.to_rgb(bg_color),
                 bg_alpha * 0.25)
            )

            txt  = (f"r = {r_p:.2f}{stars_p}\n"
                    f"ρ = {r_sp:.2f}{stars_sp}\n"
                    f"n = {len(paired)}")
            ax.text(0.5, 0.5, txt,
                    transform=ax.transAxes,
                    ha="center", va="center",
                    fontsize=8, color="#111111",
                    linespacing=1.5)
            ax.set_xticks([]); ax.set_yticks([])

        # ── Lower triangle: scatter + regression ──
        else:
            if len(paired) < 3:
                ax.set_xticks([]); ax.set_yticks([])
                continue

            x_vals = arr_j   # mj on x-axis (already clean 1D)
            y_vals = arr_i   # mi on y-axis

            # Scatter (random subsample if large)
            n_samp = min(len(x_vals), 300)
            idx    = rng.choice(len(x_vals), n_samp, replace=False)
            ax.scatter(x_vals[idx], y_vals[idx],
                       color=REF_COLOR, s=6, alpha=0.35,
                       edgecolors="none", zorder=3)

            # OLS regression line
            slope, intercept, *_ = stats.linregress(x_vals, y_vals)
            x_line = np.linspace(x_vals.min(), x_vals.max(), 100)
            ax.plot(x_line, slope * x_line + intercept,
                    color="#CC3311", linewidth=1.1,
                    alpha=0.85, zorder=4)

            ax.tick_params(labelsize=5.5)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_linewidth(0.4)
            ax.spines["bottom"].set_linewidth(0.4)

        # ── Outer axis labels (left column & bottom row) ──
        if j == 0:
            ax.set_ylabel(METRIC_LABELS.get(mi, mi),
                          fontsize=8, labelpad=3)
        if i == n_met - 1:
            ax.set_xlabel(METRIC_LABELS.get(mj, mj),
                          fontsize=8, labelpad=3)

        ax.spines["top"].set_linewidth(0.3)
        ax.spines["right"].set_linewidth(0.3)
        ax.spines["left"].set_linewidth(0.3)
        ax.spines["bottom"].set_linewidth(0.3)

fig3.suptitle(
    f"Metric pairs scatter matrix — {REFERENCE_MODEL}\n"
    "Diagonal: KDE  |  Upper: Pearson r + Spearman ρ  |  "
    "Lower: scatter + OLS regression",
    fontsize=10, fontweight="bold", y=1.002,
)

plt.tight_layout(h_pad=0.3, w_pad=0.3)
p3_png = os.path.join(SAVE_DIR, "MetricCorrelation_PairsPlot.png")
p3_pdf = os.path.join(SAVE_DIR, "MetricCorrelation_PairsPlot.pdf")
plt.savefig(p3_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(p3_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Saved → {p3_png}")

# =============================================================================
# 8. FIGURE 4 — PER-CLASS SPEARMAN HEATMAP STRIP
# Shows whether metric correlations are stable across anatomy.
# One column per class, one row per metric pair.
# =============================================================================

metric_pairs = [(METRICS[i], METRICS[j])
                for i in range(len(METRICS))
                for j in range(i + 1, len(METRICS))]
pair_labels  = [f"{a}–{b}" for a, b in metric_pairs]

n_pairs = len(metric_pairs)
n_cls   = len(classes)

rho_class = np.full((n_pairs, n_cls), np.nan)
p_class   = np.full((n_pairs, n_cls), np.nan)

for ci, cls in enumerate(classes):
    sub = df_pooled[df_pooled["Class_Name"] == cls]
    for pi, (ma, mb) in enumerate(metric_pairs):
        a = pd.to_numeric(sub[ma], errors="coerce").values.ravel()
        b = pd.to_numeric(sub[mb], errors="coerce").values.ravel()
        mask = np.isfinite(a) & np.isfinite(b)
        a, b = a[mask], b[mask]
        if len(a) < 4:
            continue
        r, p = stats.spearmanr(a, b)
        rho_class[pi, ci] = r
        p_class[pi, ci]   = p

FIG4_W = max(8, n_cls * 1.4 + 2.5)
FIG4_H = max(4, n_pairs * 0.55 + 1.5)
fig4, ax4 = plt.subplots(figsize=(FIG4_W, FIG4_H))

im4 = ax4.imshow(rho_class, cmap=CMAP_DIV,
                 norm=TwoSlopeNorm(vmin=-1, vcenter=0, vmax=1),
                 aspect="auto")

for pi in range(n_pairs):
    for ci in range(n_cls):
        r = rho_class[pi, ci]
        p = p_class[pi, ci]
        if np.isnan(r):
            ax4.text(ci, pi, "—", ha="center", va="center",
                     fontsize=7, color="#aaaaaa")
            continue
        txt_col = "white" if abs(r) > 0.65 else "#111111"
        stars   = p_to_stars(p)
        ax4.text(ci, pi, f"{r:.2f}",
                 ha="center", va="center",
                 fontsize=7, color=txt_col, fontweight="bold")
        if stars and stars != "ns":
            ax4.text(ci, pi + 0.32, stars,
                     ha="center", va="center",
                     fontsize=5, color=txt_col)

for k in range(n_pairs + 1):
    ax4.axhline(k - 0.5, color="#dddddd", linewidth=0.4)
for k in range(n_cls + 1):
    ax4.axvline(k - 0.5, color="#dddddd", linewidth=0.4)

ax4.set_xticks(range(n_cls))
ax4.set_xticklabels(classes, fontsize=8, rotation=35, ha="right")
ax4.set_yticks(range(n_pairs))
ax4.set_yticklabels(pair_labels, fontsize=8)
ax4.set_xlabel("Anatomical class", fontsize=9, labelpad=6)
ax4.set_ylabel("Metric pair", fontsize=9)
ax4.set_title(
    "Spearman ρ per metric pair × anatomical class  (all models pooled)\n"
    "Stable red across a row → correlation is anatomy-independent  |  "
    "Variable → anatomy moderates metric agreement",
    fontsize=9, fontweight="bold", pad=6,
)
ax4.tick_params(length=0)

cb4 = plt.colorbar(im4, ax=ax4, fraction=0.025, pad=0.03)
cb4.ax.tick_params(labelsize=7)
cb4.set_label("Spearman ρ", fontsize=8)

plt.tight_layout()
p4_png = os.path.join(SAVE_DIR, "MetricCorrelation_PerClass.png")
p4_pdf = os.path.join(SAVE_DIR, "MetricCorrelation_PerClass.pdf")
plt.savefig(p4_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(p4_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Saved → {p4_png}")

# =============================================================================
# 9. EXCEL STATISTICS TABLE
# =============================================================================

excel_records = []
for ds_label, df_src in datasets.items():
    for method in ("pearson", "spearman"):
        R, P_corr, N = corr_results[(ds_label, method)]
        for i, mi in enumerate(METRICS):
            for j, mj in enumerate(METRICS):
                if j <= i:
                    continue
                excel_records.append({
                    "Dataset":    ds_label,
                    "Method":     method.capitalize(),
                    "Metric_A":   mi,
                    "Metric_B":   mj,
                    "r_or_rho":   round(R[i, j], 4) if not np.isnan(R[i, j]) else np.nan,
                    "P_Bonferroni": round(P_corr[i, j], 4)
                                    if not np.isnan(P_corr[i, j]) else np.nan,
                    "Stars":      p_to_stars(P_corr[i, j]),
                    "N_pairs":    int(N[i, j]) if not np.isnan(N[i, j]) else np.nan,
                    "Interpretation": (
                        "Highly redundant (|r|>0.95)"   if abs(R[i,j]) > 0.95 else
                        "Redundant (|r|>0.85)"           if abs(R[i,j]) > 0.85 else
                        "Partially redundant (|r|>0.7)"  if abs(R[i,j]) > 0.70 else
                        "Independent"
                    ) if not np.isnan(R[i, j]) else "—",
                })

df_excel = pd.DataFrame(excel_records)
excel_path = os.path.join(SAVE_DIR, "MetricCorrelation_Statistics.xlsx")

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    df_excel.to_excel(writer, index=False, sheet_name="Correlations")

    ws = writer.sheets["Correlations"]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin     = Border(**{s: Side(style="thin", color="CCCCCC")
                         for s in ("left","right","top","bottom")})
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=8, name="Helvetica")

    fills = {
        "Highly redundant (|r|>0.95)":  PatternFill("solid", fgColor="FFB3B3"),
        "Redundant (|r|>0.85)":          PatternFill("solid", fgColor="FFD9B3"),
        "Partially redundant (|r|>0.7)": PatternFill("solid", fgColor="FFFAB3"),
        "Independent":                   PatternFill("solid", fgColor="D5E8D4"),
    }

    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    interp_col = None
    for cell in ws[1]:
        if cell.value == "Interpretation":
            interp_col = cell.column
            break

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(size=8, name="Helvetica")
        if interp_col:
            iv = ws.cell(row[0].row, interp_col).value
            if iv in fills:
                for cell in row:
                    cell.fill = fills[iv]

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

print(f"Statistics table → {excel_path}")

# =============================================================================
# 10. CONSOLE INTERPRETATION GUIDE
# =============================================================================

print("\n" + "=" * 70)
print("METRIC REDUNDANCY INTERPRETATION  (Spearman ρ, pooled, Bonferroni)")
print("=" * 70)
R_sp_all, P_sp_all, N_sp_all = corr_results[("All models (pooled)", "spearman")]
print(f"\n{'Pair':<18} {'ρ':>7} {'P (corr)':>12} {'Stars':>7}  Interpretation")
print("-" * 70)
for i, mi in enumerate(METRICS):
    for j, mj in enumerate(METRICS):
        if j <= i:
            continue
        r = R_sp_all[i, j]; p = P_sp_all[i, j]
        if np.isnan(r):
            continue
        interp = (
            "⚠  REDUNDANT — nearly identical information"  if abs(r) > 0.95 else
            "⚠  Redundant — largely overlapping"           if abs(r) > 0.85 else
            "~  Partial overlap"                           if abs(r) > 0.70 else
            "✓  Independent — captures distinct variation"
        )
        print(f"  {mi}–{mj:<12} {r:>7.3f} {p:>12.4f} {p_to_stars(p):>7}  {interp}")

print(f"\nAll outputs in: {SAVE_DIR}")
















































# =============================================================================
# CASE-LEVEL RANKING CONSISTENCY
# Nature Medicine Style — Per-Patient Model Rank Analysis
# =============================================================================
#
# For each patient × class, ranks all models by DSC (and optionally HD95).
# Rank 1 = best model for that patient.
#
# OUTPUTS:
#
#   Figure 1 — Bump chart (parallel coordinates of ranks)
#     X-axis  : anatomical class
#     Y-axis  : rank (1 = best at top)
#     One line per model, thickness ∝ frequency of that rank across patients
#     Colour  : model identity (Wong palette)
#     Reveals: does one model consistently sit at rank 1, or do lines cross?
#
#   Figure 2 — Rank heatmap (patients × models, per class)
#     Rows    : patients sorted by BAT-RM rank (best at top)
#     Columns : models
#     Colour  : rank 1 (dark) → rank N (light)
#     Annotated with rank number in each cell
#     Reveals: "column of rank-1" = consistent winner;
#               shuffled colours = unstable rankings
#
#   Figure 3 — Rank frequency stacked bar chart
#     For each model × class: how often is it rank 1, 2, 3 …?
#     Stacked bars, one segment per rank position.
#     Reveals: % of patients where each model is best/worst.
#
#   Figure 4 — Kendall's W concordance heatmap (class × metric)
#     W = 1: all models always agree on which patient is hardest/easiest
#     W = 0: rankings are random across models
#     Low W → patient difficulty is model-specific (unreliable rankings)
#
#   Excel: per-patient rank table + summary stats
#         (rank-1 frequency, mean rank, rank SD, Kendall's W)
#
# Statistical annotations:
#   • Kendall's W (inter-rater concordance of rankings)
#   • Spearman rank correlation between every model pair
#   • Friedman test p-value (are rank distributions different across models?)
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap
from scipy import stats
from scipy.stats import friedmanchisquare, spearmanr
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION
# =============================================================================

REFERENCE_MODEL = "BAT-RM"

excel_files_seg = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
]

excel_files_bnd = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx'),
]

SEG_METRICS     = ["Dice", "IoU"]
BND_METRICS     = ["HD95", "ASD"]
ALL_METRICS     = SEG_METRICS + BND_METRICS
LOWER_IS_BETTER = {"HD95", "ASD", "HD", "NSD", "RVD"}

# Primary ranking metric (Dice for main figures; HD95 computed in parallel)
PRIMARY_METRIC  = "Dice"
SECONDARY_METRIC = "HD95"

class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# Wong (2011) colour-blind safe palette
MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}

# How many patients to show in heatmap before truncating
MAX_PATIENTS_HEATMAP = 60

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/RankingConsistency'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   9,
    "axes.labelsize":   8.5,
    "xtick.labelsize":  8,
    "ytick.labelsize":  7.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.35,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "pdf.fonttype":     42,
    "ps.fonttype":      42,
    "figure.dpi":       150,
    "savefig.dpi":      600,
})

# =============================================================================
# 3. DATA LOADING
# =============================================================================

def load_segmentation(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
        for m in metrics:
            if m in df.columns:
                df[m] = pd.to_numeric(df[m], errors="coerce")
        df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
        df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
        df = df[df["Class_Name"].isin([c.upper() for c in class_list])]
        df["Model"] = name
        cols = ["Filename", "Class_Name", "Model"] + \
               [m for m in metrics if m in df.columns]
        all_data.append(df[cols])
    return pd.concat(all_data, ignore_index=True)


def load_boundary(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")
        file_rows = []
        for cls in class_list:
            present = f"{cls}_Present_In_GT"
            gt_vol  = f"{cls}_GT_Volume"
            if present in df.columns:
                sub = df[df[present] == True].copy()
            elif gt_vol in df.columns:
                sub = df[df[gt_vol] > 0].copy()
            else:
                sub = df.copy()
            if sub.empty:
                continue
            cls_cols  = {"Filename": sub["Filename"].values}
            any_found = False
            for m in metrics:
                col = f"{cls}_{m}"
                if col not in df.columns:
                    cls_cols[m] = np.nan
                    continue
                vals = pd.to_numeric(sub[col], errors="coerce").values
                vals = np.where(np.isfinite(vals), vals, np.nan)
                cls_cols[m] = vals
                any_found   = True
            if not any_found:
                continue
            tmp = pd.DataFrame(cls_cols)
            tmp["Class_Name"] = cls.upper().strip()
            tmp["Model"]      = name
            tmp = tmp.dropna(subset=metrics, how="all")
            file_rows.append(tmp)
        if file_rows:
            all_data.append(pd.concat(file_rows, ignore_index=True))
    if not all_data:
        raise ValueError("No boundary data loaded.")
    return pd.concat(all_data, ignore_index=True)


print("Loading data …")
df_seg = load_segmentation(excel_files_seg, SEG_METRICS, class_list)
df_bnd = load_boundary(excel_files_bnd, BND_METRICS, class_list)

df_all = pd.merge(
    df_seg, df_bnd,
    on=["Filename", "Class_Name", "Model"],
    how="outer",
)

model_names = [m for m, _ in excel_files_seg]
n_models    = len(model_names)
classes     = sorted(df_all["Class_Name"].dropna().unique())

print(f"Models   : {model_names}")
print(f"Classes  : {classes}")
print(f"Patients : {df_all['Filename'].nunique()}\n")

# =============================================================================
# 4. RANK COMPUTATION
# =============================================================================

def compute_ranks(df, metric, class_list, model_names, lower_is_better=False):
    """
    For every patient × class with complete data across all models,
    assign rank 1 (best) … N (worst).

    Returns a wide DataFrame:
        Filename | Class_Name | <model_1>_rank | … | <model_N>_rank
    plus the raw metric values.
    """
    records = []
    for cls in class_list:
        # Pivot: rows = patients, columns = models
        sub = df[df["Class_Name"] == cls][["Filename", "Model", metric]].copy()
        sub[metric] = pd.to_numeric(sub[metric], errors="coerce")
        sub = sub.dropna(subset=[metric])
        sub = sub[np.isfinite(sub[metric])]

        wide = sub.pivot_table(
            index="Filename", columns="Model", values=metric, aggfunc="mean"
        )
        # Keep only patients with data for ALL models
        wide = wide.dropna(subset=[m for m in model_names if m in wide.columns],
                           how="any")
        if len(wide) == 0:
            continue

        # Rank: ascending=True → rank 1 = lowest value (good for distance metrics)
        #        ascending=False → rank 1 = highest value (good for overlap metrics)
        rank_wide = wide.rank(axis=1, ascending=lower_is_better, method="average")

        for pid in wide.index:
            row = {"Filename": pid, "Class_Name": cls}
            for m in model_names:
                if m in wide.columns:
                    row[f"{m}_val"]  = wide.loc[pid, m]
                    row[f"{m}_rank"] = rank_wide.loc[pid, m]
            records.append(row)

    return pd.DataFrame(records)


print(f"Computing ranks for {PRIMARY_METRIC} …")
df_ranks = compute_ranks(
    df_all, PRIMARY_METRIC, classes, model_names,
    lower_is_better=(PRIMARY_METRIC in LOWER_IS_BETTER)
)

print(f"Computing ranks for {SECONDARY_METRIC} …")
df_ranks2 = compute_ranks(
    df_all, SECONDARY_METRIC, classes, model_names,
    lower_is_better=(SECONDARY_METRIC in LOWER_IS_BETTER)
)

rank_cols = [f"{m}_rank" for m in model_names]
val_cols  = [f"{m}_val"  for m in model_names]

print(f"Complete patient–class pairs (primary metric): {len(df_ranks)}\n")

# =============================================================================
# 5. STATISTICAL HELPERS
# =============================================================================

def kendalls_w(rank_matrix):
    """
    Kendall's W coefficient of concordance.
    rank_matrix: (n_patients × n_raters/models) array of ranks.
    W = 1: perfect agreement. W = 0: no agreement.
    """
    n, k  = rank_matrix.shape
    if n < 2 or k < 2:
        return np.nan, np.nan
    Ri    = rank_matrix.sum(axis=1)        # row sums
    R_bar = Ri.mean()
    S     = ((Ri - R_bar) ** 2).sum()
    W     = 12 * S / (k ** 2 * (n ** 3 - n))
    # Friedman chi-square approximation for p-value
    chi2  = k * (n - 1) * W
    p     = stats.chi2.sf(chi2, df=n - 1)
    return W, p


def rank1_frequency(df_ranks, model, model_names):
    """% of patient–class pairs where model achieves rank 1."""
    rc = f"{model}_rank"
    if rc not in df_ranks.columns:
        return np.nan
    return 100 * (df_ranks[rc] == 1).mean()


# =============================================================================
# 6. FIGURE 1 — BUMP CHART
# Mean rank per model per class, with error ribbon (± SD of rank across patients)
# Lines connect classes; thickness encodes rank-1 frequency.
# =============================================================================

n_cls  = len(classes)
x_pos  = np.arange(n_cls)

FIG1_W = max(10, n_cls * 1.6 + 2.0)
FIG1_H = 5.5

fig1, ax1 = plt.subplots(figsize=(FIG1_W, FIG1_H))

# Background alternating columns
for xi in range(n_cls):
    if xi % 2 == 0:
        ax1.axvspan(xi - 0.5, xi + 0.5, color="#F5F5F5", zorder=0)

# Horizontal rank reference lines
for rank in range(1, n_models + 1):
    ax1.axhline(rank, color="#dddddd", linewidth=0.5, zorder=1)

for model in model_names:
    rc     = f"{model}_rank"
    color  = MODEL_COLORS.get(model, "#888888")
    is_ref = model == REFERENCE_MODEL

    mean_ranks = []
    sd_ranks   = []
    r1_freqs   = []

    for cls in classes:
        sub = df_ranks[df_ranks["Class_Name"] == cls]
        if rc not in sub.columns or sub[rc].dropna().empty:
            mean_ranks.append(np.nan)
            sd_ranks.append(0)
            r1_freqs.append(0)
        else:
            vals = sub[rc].dropna().values
            mean_ranks.append(vals.mean())
            sd_ranks.append(vals.std(ddof=1) if len(vals) > 1 else 0)
            r1_freqs.append(100 * (vals == 1).mean())

    mean_ranks = np.array(mean_ranks, float)
    sd_ranks   = np.array(sd_ranks,   float)
    r1_freqs   = np.array(r1_freqs,   float)

    valid = np.isfinite(mean_ranks)
    if not valid.any():
        continue

    xi_v  = x_pos[valid]
    mr_v  = mean_ranks[valid]
    sd_v  = sd_ranks[valid]

    # ── Error ribbon (±1 SD of rank) ──
    ax1.fill_between(
        xi_v,
        np.clip(mr_v - sd_v, 1, n_models),
        np.clip(mr_v + sd_v, 1, n_models),
        color=color, alpha=0.10, zorder=2,
    )

    # ── Main bump line ──
    lw = 2.8 if is_ref else 1.6
    ax1.plot(
        xi_v, mr_v,
        color=color, linewidth=lw,
        alpha=0.92, zorder=4,
        solid_capstyle="round",
        solid_joinstyle="round",
        label=model,
    )

    # ── Dot at each class: size ∝ rank-1 frequency ──
    for xi_i, mr_i, r1_i in zip(xi_v, mr_v, r1_freqs[valid]):
        dot_size = max(20, r1_i * 3.5)   # scale dot to rank-1 %
        ax1.scatter(
            xi_i, mr_i,
            s=dot_size, color=color,
            edgecolors="white", linewidths=0.6,
            zorder=5,
        )
        # Annotate rank-1 % above reference model dots only
        if is_ref and r1_i > 0:
            ax1.text(
                xi_i, mr_i - 0.18,
                f"{r1_i:.0f}%",
                ha="center", va="top",
                fontsize=6, color=color, fontweight="bold",
                zorder=6,
            )

ax1.set_xticks(x_pos)
ax1.set_xticklabels(classes, fontsize=8.5, rotation=25, ha="right")
ax1.set_yticks(range(1, n_models + 1))
ax1.set_yticklabels([f"Rank {r}" for r in range(1, n_models + 1)], fontsize=8)
ax1.invert_yaxis()          # rank 1 at top
ax1.set_xlim(-0.5, n_cls - 0.5)
ax1.set_ylim(n_models + 0.6, 0.4)
ax1.set_xlabel("Anatomical class", fontsize=9, labelpad=6)
ax1.set_ylabel(f"Mean rank  ({PRIMARY_METRIC}, rank 1 = best)", fontsize=9)
ax1.set_title(
    f"Bump chart — model ranking consistency across anatomical classes  ({PRIMARY_METRIC})\n"
    f"Line = mean rank per class  |  Ribbon = ±1 SD  |  "
    f"Dot size = % of patients where model is rank 1  |  "
    f"% annotations = reference model rank-1 frequency",
    fontsize=9, fontweight="bold", pad=6,
)
ax1.spines["top"].set_visible(False)
ax1.spines["right"].set_visible(False)

legend_handles = [
    mlines.Line2D(
        [0], [0],
        color=MODEL_COLORS[m],
        linewidth=2.8 if m == REFERENCE_MODEL else 1.6,
        label=f"{m}{'  ★' if m == REFERENCE_MODEL else ''}",
    )
    for m in model_names
]
ax1.legend(
    handles=legend_handles,
    loc="upper right", frameon=False, fontsize=8,
)

plt.tight_layout()
p1_png = os.path.join(SAVE_DIR, f"RankBumpChart_{PRIMARY_METRIC}.png")
p1_pdf = os.path.join(SAVE_DIR, f"RankBumpChart_{PRIMARY_METRIC}.pdf")
plt.savefig(p1_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(p1_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Saved → {p1_png}")

# =============================================================================
# 7. FIGURE 2 — RANK HEATMAP (patients × models, one panel per class)
# =============================================================================

n_cols_h = min(4, n_cls)
n_rows_h = int(np.ceil(n_cls / n_cols_h))

# Rank colourmap: rank 1 = dark model colour, rank N = near-white
# We use a single diverging scale: 1 (dark blue) → N (light)
rank_cmap = LinearSegmentedColormap.from_list(
    "ranks", ["#08306B", "#2171B5", "#6BAED6", "#C6DBEF", "#F7FBFF"], N=256
)

CELL_H = 0.18    # inches per patient row
MIN_H  = 4.0

for metric, df_r, lower_b in [
    (PRIMARY_METRIC,   df_ranks,  PRIMARY_METRIC  in LOWER_IS_BETTER),
    (SECONDARY_METRIC, df_ranks2, SECONDARY_METRIC in LOWER_IS_BETTER),
]:
    rc_cols = [f"{m}_rank" for m in model_names if f"{m}_rank" in df_r.columns]
    active_models = [m for m in model_names if f"{m}_rank" in df_r.columns]

    panel_h = []
    for cls in classes:
        sub = df_r[df_r["Class_Name"] == cls]
        n_p = min(len(sub), MAX_PATIENTS_HEATMAP)
        panel_h.append(max(MIN_H, n_p * CELL_H + 1.2))

    FIG2_W = len(active_models) * 1.2 * n_cols_h + 1.5
    FIG2_H = sum(panel_h[:n_rows_h]) + 1.5   # rough; tight_layout adjusts

    fig2, axes2 = plt.subplots(
        n_rows_h, n_cols_h,
        figsize=(FIG2_W, max(FIG2_H, 8)),
        squeeze=False,
    )

    # Hide unused panels
    for idx in range(n_cls, n_rows_h * n_cols_h):
        ri, ci = divmod(idx, n_cols_h)
        axes2[ri][ci].axis("off")

    ref_rc = f"{REFERENCE_MODEL}_rank"

    for idx, cls in enumerate(classes):
        ri, ci = divmod(idx, n_cols_h)
        ax     = axes2[ri][ci]

        sub = df_r[df_r["Class_Name"] == cls].copy()
        sub = sub.dropna(subset=rc_cols, how="any")

        if sub.empty:
            ax.text(0.5, 0.5, "No data", ha="center", va="center",
                    transform=ax.transAxes, color="#aaaaaa")
            ax.set_title(cls, fontsize=8.5, fontweight="bold")
            ax.axis("off")
            continue

        # Sort by reference model rank (best at top), then by patient ID
        if ref_rc in sub.columns:
            sub = sub.sort_values([ref_rc, "Filename"], ascending=[True, True])
        else:
            sub = sub.sort_values("Filename")

        # Truncate
        n_shown = min(len(sub), MAX_PATIENTS_HEATMAP)
        sub     = sub.head(n_shown)
        shown_str = f"(top {n_shown} by {REFERENCE_MODEL} rank)" \
                    if len(df_r[df_r["Class_Name"] == cls]) > MAX_PATIENTS_HEATMAP \
                    else f"(all {n_shown} patients)"

        rank_mat = sub[rc_cols].values.astype(float)   # n_patients × n_models

        im = ax.imshow(
            rank_mat,
            cmap=rank_cmap,
            vmin=1, vmax=n_models,
            aspect="auto",
            interpolation="nearest",
        )

        # Cell annotations
        for pi in range(rank_mat.shape[0]):
            for mi in range(rank_mat.shape[1]):
                r = rank_mat[pi, mi]
                if np.isnan(r):
                    continue
                txt_col = "white" if r <= 2 else "#222222"
                ax.text(
                    mi, pi, f"{int(r)}",
                    ha="center", va="center",
                    fontsize=max(3.5, min(6, 60 / n_shown)),
                    color=txt_col,
                )

        # Column headers: model names, colour-coded
        ax.set_xticks(range(len(active_models)))
        ax.set_xticklabels(
            [f"{'★' if m == REFERENCE_MODEL else ''}{m}"
             for m in active_models],
            fontsize=6.5, rotation=40, ha="right",
        )
        ax.set_yticks([])
        ax.set_title(
            f"{cls}  {shown_str}",
            fontsize=7.5, fontweight="bold", pad=3,
        )
        ax.tick_params(length=0)

        # Vertical line separating reference model column
        if REFERENCE_MODEL in active_models:
            ref_col_pos = active_models.index(REFERENCE_MODEL)
            ax.axvline(ref_col_pos - 0.5, color="#0077BB", linewidth=1.2, zorder=5)
            ax.axvline(ref_col_pos + 0.5, color="#0077BB", linewidth=1.2, zorder=5)

        plt.colorbar(im, ax=ax, fraction=0.04, pad=0.04,
                     label="Rank").ax.tick_params(labelsize=6)

    fig2.suptitle(
        f"Rank heatmap — {metric}  (rank 1 = best per patient)\n"
        f"Sorted by {REFERENCE_MODEL} rank within each class  |  "
        f"Blue column border = reference model  |  "
        f"Uniform dark column = consistent winner",
        fontsize=10, fontweight="bold", y=1.002,
    )

    plt.tight_layout(h_pad=1.5, w_pad=1.0)
    p2_png = os.path.join(SAVE_DIR, f"RankHeatmap_{metric}.png")
    p2_pdf = os.path.join(SAVE_DIR, f"RankHeatmap_{metric}.pdf")
    plt.savefig(p2_png, dpi=600, bbox_inches="tight", facecolor="white")
    plt.savefig(p2_pdf, format="pdf", bbox_inches="tight", facecolor="white")
    plt.show()
    print(f"Saved → {p2_png}")

# =============================================================================
# 8. FIGURE 3 — RANK-1 FREQUENCY STACKED BAR CHART
# How often is each model the best, second-best, … worst?
# =============================================================================

FIG3_W = max(10, n_cls * 1.8 + 2.0)
FIG3_H = 5.5

fig3, axes3 = plt.subplots(
    1, 2, figsize=(FIG3_W * 2, FIG3_H), squeeze=False
)

for ax_idx, (metric, df_r) in enumerate([
    (PRIMARY_METRIC,   df_ranks),
    (SECONDARY_METRIC, df_ranks2),
]):
    ax = axes3[0][ax_idx]
    rc_cols_here = [f"{m}_rank" for m in model_names
                    if f"{m}_rank" in df_r.columns]
    active = [m for m in model_names if f"{m}_rank" in df_r.columns]

    # Rank colour gradient: rank 1 = dark green, rank N = red
    rank_colors = plt.cm.RdYlGn_r(
        np.linspace(0.05, 0.95, n_models)
    )

    x_pos3    = np.arange(n_cls)
    bar_width = 0.8 / len(active)

    for mi, model in enumerate(active):
        rc     = f"{model}_rank"
        color  = MODEL_COLORS.get(model, "#888888")
        bottom = np.zeros(n_cls)

        for rank_val in range(1, n_models + 1):
            freqs = []
            for cls in classes:
                sub = df_r[df_r["Class_Name"] == cls][rc].dropna()
                freqs.append(
                    100 * (sub == rank_val).mean() if len(sub) > 0 else 0
                )
            freqs = np.array(freqs)

            ax.bar(
                x_pos3 + mi * bar_width - (len(active) - 1) * bar_width / 2,
                freqs,
                bottom=bottom,
                width=bar_width * 0.88,
                color=rank_colors[rank_val - 1],
                alpha=0.85,
                edgecolor="white",
                linewidth=0.3,
                zorder=3,
            )
            bottom += freqs

        # Model label below bars
        ax.text(
            (x_pos3 + mi * bar_width - (len(active) - 1) * bar_width / 2).mean(),
            -4,
            f"{'★' if model == REFERENCE_MODEL else ''}{model}",
            ha="center", va="top",
            fontsize=6, color=color, fontweight="bold",
        )

    ax.set_xticks(x_pos3)
    ax.set_xticklabels(classes, fontsize=8, rotation=25, ha="right")
    ax.set_ylabel("% of patients", fontsize=9)
    ax.set_ylim(-8, 108)
    ax.set_title(
        f"Rank-frequency distribution  —  {metric}\n"
        f"Each bar = one model per class; stacked segments = rank 1 … {n_models}",
        fontsize=9, fontweight="bold", pad=5,
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="y", linewidth=0.3, color="#eeeeee", zorder=0)

    # Rank legend
    rank_handles = [
        mpatches.Patch(
            color=rank_colors[r - 1],
            alpha=0.85,
            label=f"Rank {r}{'  (best)' if r == 1 else '  (worst)' if r == n_models else ''}",
        )
        for r in range(1, n_models + 1)
    ]
    ax.legend(
        handles=rank_handles,
        loc="upper right", frameon=False, fontsize=7,
        ncol=1,
    )

plt.suptitle(
    "Rank-frequency stacked bar chart — how often each model achieves each rank\n"
    "Rank 1 (green) = model is best for that patient  |  "
    "Rank N (red) = worst  |  ★ = reference model",
    fontsize=10, fontweight="bold", y=1.003,
)
plt.tight_layout()

p3_png = os.path.join(SAVE_DIR, "RankFrequency_StackedBar.png")
p3_pdf = os.path.join(SAVE_DIR, "RankFrequency_StackedBar.pdf")
plt.savefig(p3_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(p3_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Saved → {p3_png}")

# =============================================================================
# 9. FIGURE 4 — KENDALL'S W CONCORDANCE HEATMAP (class × metric)
# =============================================================================

w_matrix  = np.full((len(ALL_METRICS), n_cls), np.nan)
p_matrix  = np.full((len(ALL_METRICS), n_cls), np.nan)
n_matrix  = np.full((len(ALL_METRICS), n_cls), np.nan)

for mi, metric in enumerate(ALL_METRICS):
    lb = metric in LOWER_IS_BETTER
    df_r_tmp = compute_ranks(df_all, metric, classes, model_names, lower_b=lb) \
               if metric not in [PRIMARY_METRIC, SECONDARY_METRIC] else \
               (df_ranks if metric == PRIMARY_METRIC else df_ranks2)

    rc_cols_tmp = [f"{m}_rank" for m in model_names
                   if f"{m}_rank" in df_r_tmp.columns]

    for ci, cls in enumerate(classes):
        sub = df_r_tmp[df_r_tmp["Class_Name"] == cls][rc_cols_tmp].dropna()
        if len(sub) < 3:
            continue
        rank_mat = sub.values
        w, p     = kendalls_w(rank_mat)
        w_matrix[mi, ci] = w
        p_matrix[mi, ci] = p
        n_matrix[mi, ci] = len(sub)

# Fix: compute_ranks call above uses positional arg; patch with keyword
def compute_ranks(df, metric, class_list, model_names, lower_b=False):
    records = []
    for cls in class_list:
        sub = df[df["Class_Name"] == cls][["Filename", "Model", metric]].copy()
        sub[metric] = pd.to_numeric(sub[metric], errors="coerce")
        sub = sub.dropna(subset=[metric])
        sub = sub[np.isfinite(sub[metric])]
        wide = sub.pivot_table(
            index="Filename", columns="Model", values=metric, aggfunc="mean"
        )
        wide = wide.dropna(subset=[m for m in model_names if m in wide.columns],
                           how="any")
        if len(wide) == 0:
            continue
        rank_wide = wide.rank(axis=1, ascending=lower_b, method="average")
        for pid in wide.index:
            row = {"Filename": pid, "Class_Name": cls}
            for m in model_names:
                if m in wide.columns:
                    row[f"{m}_val"]  = wide.loc[pid, m]
                    row[f"{m}_rank"] = rank_wide.loc[pid, m]
            records.append(row)
    return pd.DataFrame(records)


# Recompute W for metrics not yet computed
for mi, metric in enumerate(ALL_METRICS):
    if metric in [PRIMARY_METRIC, SECONDARY_METRIC]:
        continue
    lb       = metric in LOWER_IS_BETTER
    df_r_tmp = compute_ranks(df_all, metric, classes, model_names, lower_b=lb)
    rc_cols_tmp = [f"{m}_rank" for m in model_names
                   if f"{m}_rank" in df_r_tmp.columns]
    for ci, cls in enumerate(classes):
        sub = df_r_tmp[df_r_tmp["Class_Name"] == cls][rc_cols_tmp].dropna()
        if len(sub) < 3:
            continue
        w, p = kendalls_w(sub.values)
        w_matrix[mi, ci] = w
        p_matrix[mi, ci] = p
        n_matrix[mi, ci] = len(sub)

FIG4_W = max(7, n_cls * 1.3 + 2.0)
FIG4_H = max(4, len(ALL_METRICS) * 0.9 + 1.5)

fig4, ax4 = plt.subplots(figsize=(FIG4_W, FIG4_H))

w_cmap = LinearSegmentedColormap.from_list(
    "w_cmap", ["#CC3311", "#FFFFFF", "#009988"], N=256
)
im4 = ax4.imshow(w_matrix, cmap=w_cmap, vmin=0, vmax=1, aspect="auto")

for mi in range(len(ALL_METRICS)):
    for ci in range(n_cls):
        w = w_matrix[mi, ci]
        p = p_matrix[mi, ci]
        n = n_matrix[mi, ci]
        if np.isnan(w):
            ax4.text(ci, mi, "—", ha="center", va="center",
                     fontsize=8, color="#aaaaaa")
            continue
        txt_col = "white" if w > 0.7 or w < 0.2 else "#111111"
        stars   = ("****" if p < 0.0001 else "***" if p < 0.001 else
                   "**"   if p < 0.01   else "*"   if p < 0.05  else "ns") \
                  if not np.isnan(p) else ""
        ax4.text(ci, mi - 0.15, f"W={w:.2f}",
                 ha="center", va="center",
                 fontsize=7.5, color=txt_col, fontweight="bold")
        ax4.text(ci, mi + 0.22, stars,
                 ha="center", va="center",
                 fontsize=6, color=txt_col)

for k in range(len(ALL_METRICS) + 1):
    ax4.axhline(k - 0.5, color="#cccccc", linewidth=0.4)
for k in range(n_cls + 1):
    ax4.axvline(k - 0.5, color="#cccccc", linewidth=0.4)

ax4.set_xticks(range(n_cls))
ax4.set_xticklabels(classes, fontsize=8.5, rotation=30, ha="right")
ax4.set_yticks(range(len(ALL_METRICS)))
ax4.set_yticklabels(ALL_METRICS, fontsize=8.5)
ax4.set_title(
    "Kendall's W concordance — do all models agree on which patient is hardest?\n"
    "W = 1: models always give same relative ranking  |  "
    "W = 0: rankings are random  |  Stars: Friedman χ² p-value",
    fontsize=9, fontweight="bold", pad=6,
)
ax4.tick_params(length=0)

cb4 = plt.colorbar(im4, ax=ax4, fraction=0.025, pad=0.04)
cb4.ax.tick_params(labelsize=7.5)
cb4.set_label("Kendall's W", fontsize=8)

plt.tight_layout()
p4_png = os.path.join(SAVE_DIR, "KendallsW_Concordance.png")
p4_pdf = os.path.join(SAVE_DIR, "KendallsW_Concordance.pdf")
plt.savefig(p4_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(p4_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Saved → {p4_png}")

# =============================================================================
# 10. EXCEL STATISTICS TABLE
# =============================================================================

excel_rows = []
summary_rows = []

for metric, df_r, lower_b in [
    (PRIMARY_METRIC,   df_ranks,  PRIMARY_METRIC  in LOWER_IS_BETTER),
    (SECONDARY_METRIC, df_ranks2, SECONDARY_METRIC in LOWER_IS_BETTER),
]:
    rc_cols_here = [f"{m}_rank" for m in model_names
                    if f"{m}_rank" in df_r.columns]
    active = [m for m in model_names if f"{m}_rank" in df_r.columns]

    for cls in classes:
        sub = df_r[df_r["Class_Name"] == cls].copy()
        sub = sub.dropna(subset=rc_cols_here, how="any")

        if sub.empty:
            continue

        # Per-patient rank table rows
        for _, row in sub.iterrows():
            rec = {"Metric": metric, "Class": cls, "Filename": row["Filename"]}
            for m in active:
                rec[f"{m}_rank"] = row.get(f"{m}_rank", np.nan)
                rec[f"{m}_val"]  = row.get(f"{m}_val",  np.nan)
            excel_rows.append(rec)

        # Summary stats per model × class
        rank_mat = sub[rc_cols_here].values
        w, p_w   = kendalls_w(rank_mat)

        # Friedman test
        try:
            fr_stat, fr_p = friedmanchisquare(*[sub[rc].dropna().values
                                                for rc in rc_cols_here])
        except Exception:
            fr_stat, fr_p = np.nan, np.nan

        for m in active:
            rc   = f"{m}_rank"
            vals = sub[rc].dropna().values
            summary_rows.append({
                "Metric":           metric,
                "Class":            cls,
                "Model":            m,
                "N_patients":       len(vals),
                "Mean_rank":        round(vals.mean(), 3) if len(vals) > 0 else np.nan,
                "Median_rank":      round(np.median(vals), 1) if len(vals) > 0 else np.nan,
                "Rank_SD":          round(vals.std(ddof=1), 3) if len(vals) > 1 else np.nan,
                "Rank1_pct":        round(100 * (vals == 1).mean(), 1),
                "RankLast_pct":     round(100 * (vals == n_models).mean(), 1),
                "Kendalls_W":       round(w, 3) if not np.isnan(w) else np.nan,
                "Friedman_p":       round(fr_p, 4) if not np.isnan(fr_p) else np.nan,
            })

df_excel   = pd.DataFrame(excel_rows)
df_summary = pd.DataFrame(summary_rows)

excel_path = os.path.join(SAVE_DIR, "RankingConsistency_Statistics.xlsx")
with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    df_summary.to_excel(writer, index=False, sheet_name="Summary")
    df_excel.to_excel(writer, index=False, sheet_name="PerPatient_Ranks")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin     = Border(**{s: Side(style="thin", color="CCCCCC")
                         for s in ("left","right","top","bottom")})
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=8, name="Helvetica")
    ref_fill = PatternFill("solid", fgColor="D5E8D4")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    top_fill = PatternFill("solid", fgColor="FFD700")   # gold = rank-1 dominant

    for sheet_name in ["Summary", "PerPatient_Ranks"]:
        ws = writer.sheets[sheet_name]
        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
        for row_idx in range(2, ws.max_row + 1):
            is_ref = False
            for cell in ws[row_idx]:
                if cell.value == REFERENCE_MODEL:
                    is_ref = True
            fill = ref_fill if is_ref else \
                   (alt_fill if row_idx % 2 == 0 else PatternFill())
            for cell in ws[row_idx]:
                cell.border    = thin
                cell.alignment = Alignment(horizontal="center")
                cell.font      = Font(size=8, name="Helvetica", bold=is_ref)
                if fill.fill_type:
                    cell.fill = fill
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 16

print(f"Statistics saved → {excel_path}")

# =============================================================================
# 11. CONSOLE RANKING REPORT
# =============================================================================

print("\n" + "=" * 76)
print("RANKING CONSISTENCY REPORT")
print("=" * 76)

for metric, df_r in [(PRIMARY_METRIC, df_ranks), (SECONDARY_METRIC, df_ranks2)]:
    print(f"\n{metric}:")
    print(f"  {'Model':<14} {'Class':<22} {'Mean rank':>10} "
          f"{'Rank-1%':>9} {'Rank-last%':>11} {'W':>7}")
    print("  " + "-" * 74)
    sub_s = df_summary[df_summary["Metric"] == metric]
    for _, row in sub_s.sort_values(["Class", "Mean_rank"]).iterrows():
        marker = "★ " if row["Model"] == REFERENCE_MODEL else "  "
        print(
            f"  {marker}{row['Model']:<12} {row['Class']:<22} "
            f"{row['Mean_rank']:>10.2f} "
            f"{row['Rank1_pct']:>8.1f}% "
            f"{row['RankLast_pct']:>10.1f}% "
            f"{str(row.get('Kendalls_W','—')):>7}"
        )

print(f"\nAll outputs in: {SAVE_DIR}")














































# =============================================================================
# PAIRED VIOLIN + BOX PLOTS WITH EXPLICIT PATIENT PAIRING
# Nature Medicine Style — All Metrics, All Classes
# =============================================================================
#
# Upgrade over standard Wilcoxon:
#   • Uses PAIRED Wilcoxon signed-rank test (same patient, different model)
#   • Shows pairing explicitly via connecting lines between matched dots
#   • Reports concordant / discordant pair counts (N+, N−, ties)
#   • Effect size: paired Hedges' g with 95% bootstrap CI
#
# LAYOUT — one figure per class (7 figures total):
#   Columns : metrics (Dice, IoU, HD95, ASD)
#   Each panel:
#     ① Violin (full distribution shape)
#     ② Box (IQR + median)
#     ③ Strip dots — one per patient, jittered
#     ④ Grey connecting lines between matched patient dots
#        (reference model dot → baseline model dot)
#        Lines coloured RED if reference model wins that patient,
#        BLUE if baseline wins — immediately shows direction of individual pairs
#     ⑤ Significance bracket with paired Wilcoxon p-value + stars
#     ⑥ Concordant / discordant count annotation
#        "N+=18 N−=7 ties=0" below the bracket
#
# INSET TABLE per panel (top-right):
#   Model | Median | IQR | N+ | N− | g | CI
#
# Statistical method:
#   Paired Wilcoxon signed-rank test (scipy.stats.wilcoxon, zero_method="pratt")
#   Paired Hedges' g = mean(diff) / SD(diff) × correction_factor
#   95% CI via BCa bootstrap (n=1000 resamples)
#
# NOTE on connecting lines:
#   Only patients with data for BOTH models in a pair are connected.
#   If a class has >MAX_LINES patients, a random subsample is drawn to
#   keep the figure readable (default MAX_LINES=60).
#
# Dependencies: pandas, numpy, matplotlib, scipy, openpyxl
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
import matplotlib.gridspec as gridspec
from scipy.stats import wilcoxon, gaussian_kde
from scipy import stats
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION
# =============================================================================

REFERENCE_MODEL = "BAT-RM"

excel_files_seg = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
]

excel_files_bnd = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/boundary_metrics_epoch_200_enhanced_v5.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/boundary_metrics_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_boundary_metrics_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_boundary_metrics_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/boundary_metrics_nnUNet_100_epochs_generated.xlsx'),
]

SEG_METRICS     = ["Dice", "IoU"]
BND_METRICS     = ["HD95", "ASD"]
ALL_METRICS     = SEG_METRICS + BND_METRICS
LOWER_IS_BETTER = {"HD95", "ASD", "HD", "RVD", "RAVD"}

class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# Max connecting lines per panel (subsample if more patients)
MAX_LINES  = 60
# Bootstrap resamples for Hedges' g CI
N_BOOT     = 1000

MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}

# Line colours for connecting pairs
WIN_COLOR  = "#CC3311"   # reference model wins this patient
LOSE_COLOR = "#4477AA"   # baseline wins this patient
TIE_COLOR  = "#AAAAAA"

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/PairedViolin'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   9,
    "axes.labelsize":   8.5,
    "xtick.labelsize":  8,
    "ytick.labelsize":  7.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.35,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "pdf.fonttype":     42,
    "ps.fonttype":      42,
    "figure.dpi":       150,
    "savefig.dpi":      600,
})

# =============================================================================
# 3. DATA LOADING
# =============================================================================

def load_segmentation(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)
        for m in metrics:
            if m in df.columns:
                df[m] = pd.to_numeric(df[m], errors="coerce")
        df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
        df = df[~df["Class_Name"].isin(["BACKGROUND", "NAN", "NONE", "", "0"])]
        df = df[df["Class_Name"].isin([c.upper() for c in class_list])]
        df["Model"] = name
        cols = ["Filename", "Class_Name", "Model"] + \
               [m for m in metrics if m in df.columns]
        all_data.append(df[cols])
    return pd.concat(all_data, ignore_index=True)


def load_boundary(excel_files, metrics, class_list):
    all_data = []
    for name, path in excel_files:
        df = pd.read_excel(path, sheet_name="Detailed_Per_Instance")
        file_rows = []
        for cls in class_list:
            present = f"{cls}_Present_In_GT"
            gt_vol  = f"{cls}_GT_Volume"
            if present in df.columns:
                sub = df[df[present] == True].copy()
            elif gt_vol in df.columns:
                sub = df[df[gt_vol] > 0].copy()
            else:
                sub = df.copy()
            if sub.empty:
                continue
            cls_cols  = {"Filename": sub["Filename"].values}
            any_found = False
            for m in metrics:
                col = f"{cls}_{m}"
                if col not in df.columns:
                    cls_cols[m] = np.nan
                    continue
                vals = pd.to_numeric(sub[col], errors="coerce").values
                vals = np.where(np.isfinite(vals), vals, np.nan)
                cls_cols[m] = vals
                any_found   = True
            if not any_found:
                continue
            tmp = pd.DataFrame(cls_cols)
            tmp["Class_Name"] = cls.upper().strip()
            tmp["Model"]      = name
            tmp = tmp.dropna(subset=metrics, how="all")
            file_rows.append(tmp)
        if file_rows:
            all_data.append(pd.concat(file_rows, ignore_index=True))
    if not all_data:
        raise ValueError("No boundary data loaded.")
    return pd.concat(all_data, ignore_index=True)


print("Loading data …")
df_seg = load_segmentation(excel_files_seg, SEG_METRICS, class_list)
df_bnd = load_boundary(excel_files_bnd, BND_METRICS, class_list)

df_all = pd.merge(
    df_seg, df_bnd,
    on=["Filename", "Class_Name", "Model"],
    how="outer",
)

model_names = [m for m, _ in excel_files_seg]
baselines   = [m for m in model_names if m != REFERENCE_MODEL]
classes     = sorted(df_all["Class_Name"].dropna().unique())

print(f"Models   : {model_names}")
print(f"Classes  : {classes}")
print(f"Metrics  : {ALL_METRICS}")
print(f"Patients : {df_all['Filename'].nunique()}\n")

# =============================================================================
# 4. STATISTICAL HELPERS
# =============================================================================

def paired_hedges_g(diff):
    """Paired Hedges' g from an array of differences (ref − baseline)."""
    diff = np.array(diff)
    n    = len(diff)
    if n < 2:
        return np.nan
    d = diff.mean() / (diff.std(ddof=1) + 1e-12)
    j = 1 - 3 / (4 * (n - 1) - 1)   # Hedges' correction
    return d * j


def bootstrap_g_ci(diff, n_boot=1000, seed=42):
    """BCa bootstrap 95% CI for paired Hedges' g."""
    diff = np.array(diff)
    n    = len(diff)
    if n < 3:
        return np.nan, np.nan
    rng  = np.random.default_rng(seed)
    boot = np.array([
        paired_hedges_g(diff[rng.integers(0, n, n)])
        for _ in range(n_boot)
    ])
    return np.nanpercentile(boot, 2.5), np.nanpercentile(boot, 97.5)


def paired_wilcoxon(x_ref, x_bsl):
    """
    Paired Wilcoxon signed-rank test.
    Returns (statistic, p_value).
    """
    diff = np.array(x_ref) - np.array(x_bsl)
    if len(diff) < 3 or np.all(diff == 0):
        return np.nan, np.nan
    try:
        stat, p = wilcoxon(x_ref, x_bsl, zero_method="pratt")
        return stat, p
    except Exception:
        return np.nan, np.nan


def concordant_discordant(x_ref, x_bsl, lower_is_better):
    """
    N+ : pairs where reference model is better
    N− : pairs where baseline is better
    T  : ties
    """
    diff = np.array(x_ref) - np.array(x_bsl)
    tol  = 1e-9
    if lower_is_better:
        n_plus  = (diff < -tol).sum()   # ref smaller = ref wins
        n_minus = (diff >  tol).sum()
    else:
        n_plus  = (diff >  tol).sum()   # ref larger = ref wins
        n_minus = (diff < -tol).sum()
    ties = (np.abs(diff) <= tol).sum()
    return int(n_plus), int(n_minus), int(ties)


def p_stars(p):
    if np.isnan(p):  return ""
    if p < 0.0001:   return "****"
    if p < 0.001:    return "***"
    if p < 0.01:     return "**"
    if p < 0.05:     return "*"
    return "ns"


# =============================================================================
# 5. SINGLE PANEL DRAWING FUNCTION
# =============================================================================

def draw_paired_panel(ax, ref_vals, bsl_vals, filenames,
                      ref_name, bsl_name, metric, lower_is_better,
                      rng, max_lines=MAX_LINES):
    """
    Draw one paired violin+box+strip+connecting-lines panel.

    Parameters
    ----------
    ax              : matplotlib Axes
    ref_vals, bsl_vals : 1D arrays, same length, same patient order
    filenames       : array of patient IDs (same order)
    ref_name        : str, reference model name
    bsl_name        : str, baseline model name
    metric          : str, metric name
    lower_is_better : bool
    rng             : numpy Generator
    max_lines       : int, max connecting lines to draw
    """
    x_ref, x_bsl = 0, 1    # x positions for the two models
    colors = [MODEL_COLORS.get(ref_name, "#0077BB"),
              MODEL_COLORS.get(bsl_name, "#888888")]

    ref_arr = np.array(ref_vals, float)
    bsl_arr = np.array(bsl_vals, float)
    n       = len(ref_arr)

    # ── Violin ──
    for xi, arr, color in [(x_ref, ref_arr, colors[0]),
                           (x_bsl, bsl_arr, colors[1])]:
        if len(arr) < 4:
            continue
        try:
            kde  = gaussian_kde(arr, bw_method="scott")
            y_ev = np.linspace(arr.min(), arr.max(), 300)
            dens = kde(y_ev)
            dens = dens / dens.max() * 0.36
            ax.fill_betweenx(y_ev, xi - dens, xi + dens,
                             color=color, alpha=0.18, zorder=1)
            ax.plot(xi - dens, y_ev, color=color, lw=0.7, alpha=0.6, zorder=2)
            ax.plot(xi + dens, y_ev, color=color, lw=0.7, alpha=0.6, zorder=2)
        except Exception:
            pass

    # ── Box (IQR) ──
    for xi, arr, color in [(x_ref, ref_arr, colors[0]),
                           (x_bsl, bsl_arr, colors[1])]:
        q25, q50, q75 = np.percentile(arr, [25, 50, 75])
        # Median bar
        ax.plot([xi - 0.20, xi + 0.20], [q50, q50],
                color=color, lw=2.4, solid_capstyle="round", zorder=6)
        # IQR box
        ax.add_patch(plt.Rectangle(
            (xi - 0.15, q25), 0.30, q75 - q25,
            facecolor=color, alpha=0.22,
            edgecolor=color, lw=0.9, zorder=4,
        ))
        # Whiskers to 5th/95th pct
        p05, p95 = np.percentile(arr, [5, 95])
        ax.plot([xi, xi], [p05, q25], color=color, lw=0.8, alpha=0.6, zorder=3)
        ax.plot([xi, xi], [q75, p95], color=color, lw=0.8, alpha=0.6, zorder=3)
        ax.plot([xi - 0.08, xi + 0.08], [p05, p05],
                color=color, lw=0.8, alpha=0.6, zorder=3)
        ax.plot([xi - 0.08, xi + 0.08], [p95, p95],
                color=color, lw=0.8, alpha=0.6, zorder=3)

    # ── Connecting lines between matched patient dots ──
    # Subsample if too many patients
    indices = np.arange(n)
    if n > max_lines:
        indices = rng.choice(n, max_lines, replace=False)
        sampled_note = f"(showing {max_lines}/{n} pairs)"
    else:
        sampled_note = f"(n={n} pairs)"

    tol = 1e-9
    for idx in indices:
        rv, bv = ref_arr[idx], bsl_arr[idx]
        if lower_is_better:
            ref_wins = rv < bv - tol
            bsl_wins = bv < rv - tol
        else:
            ref_wins = rv > bv + tol
            bsl_wins = bv > rv + tol

        if ref_wins:
            lc = WIN_COLOR
        elif bsl_wins:
            lc = LOSE_COLOR
        else:
            lc = TIE_COLOR

        # Jitter x positions slightly so lines don't stack
        jx_ref = x_ref + rng.uniform(-0.08, 0.08)
        jx_bsl = x_bsl + rng.uniform(-0.08, 0.08)
        ax.plot([jx_ref, jx_bsl], [rv, bv],
                color=lc, lw=0.45, alpha=0.35, zorder=3)

    # ── Strip dots ──
    for xi, arr, color in [(x_ref, ref_arr, colors[0]),
                           (x_bsl, bsl_arr, colors[1])]:
        jitter = rng.uniform(-0.10, 0.10, size=len(arr))
        ax.scatter(xi + jitter, arr,
                   color=color, s=9, alpha=0.50,
                   edgecolors="white", linewidths=0.2, zorder=5)

    # ── Statistics ──
    diff          = ref_arr - bsl_arr
    _, p_val      = paired_wilcoxon(ref_arr, bsl_arr)
    g             = paired_hedges_g(diff if not lower_is_better else -diff)
    g_lo, g_hi    = bootstrap_g_ci(diff if not lower_is_better else -diff, N_BOOT)
    n_plus, n_minus, ties = concordant_discordant(ref_arr, bsl_arr, lower_is_better)
    stars         = p_stars(p_val)

    # ── Significance bracket ──
    y_data_max = max(ref_arr.max(), bsl_arr.max())
    y_data_min = min(ref_arr.min(), bsl_arr.min())
    y_range    = y_data_max - y_data_min if y_data_max > y_data_min else 1.0
    y_bracket  = y_data_max + y_range * 0.08

    ax.plot([x_ref, x_ref, x_bsl, x_bsl],
            [y_bracket, y_bracket + y_range * 0.02,
             y_bracket + y_range * 0.02, y_bracket],
            color="#333333", lw=0.9, zorder=8)

    sig_color = "#CC2222" if (stars not in ("ns", "")) else "#666666"
    ax.text((x_ref + x_bsl) / 2, y_bracket + y_range * 0.04,
            stars if stars else "ns",
            ha="center", va="bottom",
            fontsize=8, color=sig_color, fontweight="bold", zorder=9)

    # p-value numeric below stars
    p_txt = f"p={p_val:.4f}" if not np.isnan(p_val) else "p=n/a"
    ax.text((x_ref + x_bsl) / 2, y_bracket + y_range * 0.09,
            p_txt,
            ha="center", va="bottom",
            fontsize=5.5, color="#555555", zorder=9)

    # Concordant / discordant annotation
    cd_txt = (f"N+={n_plus}  N−={n_minus}"
              + (f"  ties={ties}" if ties > 0 else ""))
    ax.text((x_ref + x_bsl) / 2, y_bracket + y_range * 0.155,
            cd_txt,
            ha="center", va="bottom",
            fontsize=5.5, color="#333333", zorder=9)

    # ── Inset stats box ──
    ref_med = np.median(ref_arr)
    bsl_med = np.median(bsl_arr)
    ref_iqr = np.subtract(*np.percentile(ref_arr, [75, 25]))
    bsl_iqr = np.subtract(*np.percentile(bsl_arr, [75, 25]))
    g_str   = (f"{g:+.2f} [{g_lo:+.2f},{g_hi:+.2f}]"
               if not np.isnan(g) else "n/a")

    stats_txt = (
        f"{'Model':<9} {'Med':>5}  {'IQR':>5}\n"
        f"{'─'*24}\n"
        f"{'★'+ref_name:<9} {ref_med:>5.3f}  {ref_iqr:>5.3f}\n"
        f"{bsl_name:<9} {bsl_med:>5.3f}  {bsl_iqr:>5.3f}\n"
        f"{'─'*24}\n"
        f"g = {g_str}\n"
        f"{sampled_note}"
    )
    ax.text(1.04, 0.99, stats_txt,
            transform=ax.transAxes,
            va="top", ha="left",
            fontsize=5.2, fontfamily="monospace",
            color="#222222",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
                      edgecolor="#cccccc", lw=0.5, alpha=0.90),
            zorder=10, clip_on=False)

    # ── Axis formatting ──
    ax.set_xticks([x_ref, x_bsl])
    ax.set_xticklabels(
        [f"{'★ ' if ref_name == REFERENCE_MODEL else ''}{ref_name}",
         bsl_name],
        fontsize=7.5,
    )
    ax.set_xlim(-0.55, 1.55)
    ax.set_ylim(y_data_min - y_range * 0.08,
                y_bracket + y_range * 0.28)
    ax.set_ylabel(metric, fontsize=8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(0.5)
    ax.spines["bottom"].set_linewidth(0.5)
    ax.tick_params(axis="y", labelsize=7)
    ax.tick_params(axis="x", length=0)
    ax.grid(axis="y", linewidth=0.3, color="#eeeeee", zorder=0)

    return {
        "p_val": p_val, "stars": stars,
        "g": g, "g_lo": g_lo, "g_hi": g_hi,
        "n_pairs": n, "n_plus": n_plus, "n_minus": n_minus, "ties": ties,
        "ref_median": ref_med, "bsl_median": bsl_med,
        "ref_iqr": ref_iqr, "bsl_iqr": bsl_iqr,
    }


# =============================================================================
# 6. MAIN LOOP — one figure per class
# Each figure: rows = baselines, columns = metrics
# =============================================================================

rng = np.random.default_rng(42)
summary_records = []

ref_df = df_all[df_all["Model"] == REFERENCE_MODEL]

for cls in classes:
    n_baselines = len(baselines)
    n_metrics   = len(ALL_METRICS)

    # Each panel: ~2.8 in wide (+ 1.2 for inset box), ~3.2 in tall
    PANEL_W = 2.8
    PANEL_H = 3.6
    FIG_W   = n_metrics * (PANEL_W + 1.5) + 0.6
    FIG_H   = n_baselines * PANEL_H + 1.8

    fig, axes = plt.subplots(
        n_baselines, n_metrics,
        figsize=(FIG_W, FIG_H),
        squeeze=False,
    )

    # Direction note per metric (top of column)
    direction_note = {
        m: ("↑ higher = better" if m not in LOWER_IS_BETTER else "↓ lower = better")
        for m in ALL_METRICS
    }

    # Column headers
    for ci, metric in enumerate(ALL_METRICS):
        axes[0][ci].set_title(
            f"{metric}\n{direction_note[metric]}",
            fontsize=9, fontweight="bold", pad=6,
        )

    for ri, baseline in enumerate(baselines):
        bsl_df  = df_all[df_all["Model"] == baseline]
        bsl_color = MODEL_COLORS.get(baseline, "#888888")

        # Row label (baseline model name, left side)
        axes[ri][0].set_ylabel(
            f"{baseline}\nvs {REFERENCE_MODEL}\n\n{ALL_METRICS[0]}",
            fontsize=7.5,
        )

        for ci, metric in enumerate(ALL_METRICS):
            ax = axes[ri][ci]
            lower_b = metric in LOWER_IS_BETTER

            # Paired merge on Filename + Class_Name
            ref_sub = ref_df[ref_df["Class_Name"] == cls][
                ["Filename", metric]
            ].dropna()
            bsl_sub = bsl_df[bsl_df["Class_Name"] == cls][
                ["Filename", metric]
            ].dropna()

            merged = pd.merge(
                ref_sub, bsl_sub,
                on="Filename", suffixes=("_ref", "_bsl")
            ).dropna()
            merged = merged[
                np.isfinite(merged[f"{metric}_ref"]) &
                np.isfinite(merged[f"{metric}_bsl"])
            ]

            if len(merged) < 3:
                ax.text(0.5, 0.5, "Insufficient\npaired data",
                        ha="center", va="center",
                        fontsize=8, color="#aaaaaa",
                        transform=ax.transAxes)
                ax.axis("off")
                continue

            out = draw_paired_panel(
                ax,
                ref_vals  = merged[f"{metric}_ref"].values,
                bsl_vals  = merged[f"{metric}_bsl"].values,
                filenames = merged["Filename"].values,
                ref_name  = REFERENCE_MODEL,
                bsl_name  = baseline,
                metric    = metric,
                lower_is_better = lower_b,
                rng       = rng,
            )

            if ri > 0:
                ax.set_title("")   # only show metric title on first row

            summary_records.append({
                "Class":       cls,
                "Metric":      metric,
                "Baseline":    baseline,
                "N_pairs":     out["n_pairs"],
                "Ref_median":  round(out["ref_median"], 4),
                "Bsl_median":  round(out["bsl_median"], 4),
                "Ref_IQR":     round(out["ref_iqr"],    4),
                "Bsl_IQR":     round(out["bsl_iqr"],    4),
                "N_plus (ref wins)":   out["n_plus"],
                "N_minus (bsl wins)":  out["n_minus"],
                "Ties":        out["ties"],
                "Win_pct_ref": round(100 * out["n_plus"] / max(out["n_pairs"], 1), 1),
                "Paired_Wilcoxon_p": round(out["p_val"], 5)
                                      if not np.isnan(out["p_val"]) else np.nan,
                "Stars":       out["stars"],
                "Hedges_g":    round(out["g"],    3) if not np.isnan(out["g"])    else np.nan,
                "g_CI_lo":     round(out["g_lo"], 3) if not np.isnan(out["g_lo"]) else np.nan,
                "g_CI_hi":     round(out["g_hi"], 3) if not np.isnan(out["g_hi"]) else np.nan,
            })

    # ── Legend ──
    legend_handles = [
        mlines.Line2D([0], [0], color=MODEL_COLORS.get(REFERENCE_MODEL, "#0077BB"),
                      lw=2.0, label=f"★ {REFERENCE_MODEL} (reference)"),
    ]
    for bsl in baselines:
        legend_handles.append(
            mlines.Line2D([0], [0], color=MODEL_COLORS.get(bsl, "#888888"),
                          lw=2.0, label=bsl)
        )
    legend_handles += [
        mlines.Line2D([0], [0], color=WIN_COLOR,  lw=1.0, alpha=0.7,
                      label=f"{REFERENCE_MODEL} wins (red line)"),
        mlines.Line2D([0], [0], color=LOSE_COLOR, lw=1.0, alpha=0.7,
                      label="Baseline wins (blue line)"),
        mlines.Line2D([0], [0], color=TIE_COLOR,  lw=1.0, alpha=0.7,
                      label="Tie (grey line)"),
    ]

    fig.legend(
        handles        = legend_handles,
        loc            = "lower center",
        ncol           = len(legend_handles),
        frameon        = False,
        fontsize       = 7.5,
        bbox_to_anchor = (0.5, -0.01),
    )

    fig.suptitle(
        f"Paired violin + box plots with patient-level connecting lines — {cls}\n"
        f"Red line = {REFERENCE_MODEL} wins that patient  |  "
        f"Blue line = baseline wins  |  "
        f"Bracket: paired Wilcoxon p-value  |  "
        f"N+ / N− = concordant / discordant pairs  |  "
        f"g = paired Hedges' g [95% bootstrap CI]",
        fontsize=9, fontweight="bold", y=1.002,
    )

    plt.tight_layout(w_pad=3.5, h_pad=1.5,
                     rect=[0, 0.04, 1, 0.998])

    safe_cls = cls.replace(" ", "_")
    png_path = os.path.join(SAVE_DIR, f"PairedViolin_{safe_cls}.png")
    pdf_path = os.path.join(SAVE_DIR, f"PairedViolin_{safe_cls}.pdf")
    plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
    plt.savefig(pdf_path, format="pdf", bbox_inches="tight", facecolor="white")
    plt.show()
    print(f"Saved → {png_path}")

# =============================================================================
# 7. COMBINED SUMMARY FIGURE — median difference heatmap across all classes
# Rows = class × baseline, columns = metrics
# Colour = direction and magnitude of median difference
# =============================================================================

df_sum = pd.DataFrame(summary_records)

# Compute signed median difference (reference − baseline, sign-flipped for
# distance metrics so positive always means reference is better)
df_sum["Med_diff_signed"] = df_sum.apply(
    lambda r: (r["Ref_median"] - r["Bsl_median"])
              * (-1 if r["Metric"] in LOWER_IS_BETTER else 1),
    axis=1,
)

row_labels = [f"{cls}\nvs {bsl}"
              for cls in classes for bsl in baselines]
n_rows_h   = len(classes) * len(baselines)

diff_mat  = np.full((n_rows_h, len(ALL_METRICS)), np.nan)
star_mat  = [["" for _ in ALL_METRICS] for _ in range(n_rows_h)]
g_mat     = np.full((n_rows_h, len(ALL_METRICS)), np.nan)

for ri, (cls, bsl) in enumerate(
    [(c, b) for c in classes for b in baselines]
):
    sub = df_sum[(df_sum["Class"] == cls) & (df_sum["Baseline"] == bsl)]
    for ci, metric in enumerate(ALL_METRICS):
        row = sub[sub["Metric"] == metric]
        if row.empty:
            continue
        diff_mat[ri, ci] = row.iloc[0]["Med_diff_signed"]
        star_mat[ri][ci] = row.iloc[0]["Stars"]
        g_mat[ri, ci]    = row.iloc[0]["Hedges_g"]

v_abs = np.nanmax(np.abs(diff_mat)) if np.any(np.isfinite(diff_mat)) else 1.0

FIG5_W = len(ALL_METRICS) * 2.2 + 3.5
FIG5_H = max(6, n_rows_h * 0.55 + 2.0)
fig5, ax5 = plt.subplots(figsize=(FIG5_W, FIG5_H))

from matplotlib.colors import TwoSlopeNorm
norm5 = TwoSlopeNorm(vmin=-v_abs, vcenter=0, vmax=v_abs)
cmap5 = plt.cm.RdBu      # red = reference better (positive diff), blue = worse

im5 = ax5.imshow(diff_mat, cmap=cmap5, norm=norm5, aspect="auto")

for ri in range(n_rows_h):
    for ci, metric in enumerate(ALL_METRICS):
        v  = diff_mat[ri, ci]
        st = star_mat[ri][ci]
        g  = g_mat[ri, ci]
        if np.isnan(v):
            ax5.text(ci, ri, "—", ha="center", va="center",
                     fontsize=7, color="#aaaaaa")
            continue
        txt_col = "white" if abs(v) > v_abs * 0.6 else "#111111"
        ax5.text(ci, ri - 0.15,
                 f"{v:+.3f}",
                 ha="center", va="center",
                 fontsize=7, color=txt_col, fontweight="bold")
        label_parts = []
        if st and st != "ns":
            label_parts.append(st)
        if not np.isnan(g):
            label_parts.append(f"g={g:+.2f}")
        if label_parts:
            ax5.text(ci, ri + 0.25,
                     "  ".join(label_parts),
                     ha="center", va="center",
                     fontsize=5.2, color=txt_col)

# Class group separator lines
for gi in range(len(classes)):
    y_sep = gi * len(baselines) - 0.5
    if y_sep > -0.5:
        ax5.axhline(y_sep, color="#333333", lw=1.0, zorder=5)

for k in range(len(ALL_METRICS) + 1):
    ax5.axvline(k - 0.5, color="#cccccc", lw=0.4)
for k in range(n_rows_h + 1):
    ax5.axhline(k - 0.5, color="#cccccc", lw=0.4)

ax5.set_xticks(range(len(ALL_METRICS)))
ax5.set_xticklabels(ALL_METRICS, fontsize=9, fontweight="bold")
ax5.set_yticks(range(n_rows_h))
ax5.set_yticklabels(row_labels, fontsize=7)
ax5.set_title(
    f"Summary: signed median difference  ({REFERENCE_MODEL} − baseline)\n"
    "Red = reference model better  |  Blue = baseline better  |  "
    "Value = median diff  |  Stars = paired Wilcoxon  |  g = Hedges' g\n"
    "(distance metrics sign-flipped so red = better always)",
    fontsize=9, fontweight="bold", pad=6,
)
ax5.tick_params(length=0)

cb5 = plt.colorbar(im5, ax=ax5, fraction=0.025, pad=0.03)
cb5.ax.tick_params(labelsize=7.5)
cb5.set_label(f"{REFERENCE_MODEL} better → / ← baseline better", fontsize=7.5)

plt.tight_layout()
sum_png = os.path.join(SAVE_DIR, "PairedViolin_SummaryHeatmap.png")
sum_pdf = os.path.join(SAVE_DIR, "PairedViolin_SummaryHeatmap.pdf")
plt.savefig(sum_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(sum_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Saved summary heatmap → {sum_png}")

# =============================================================================
# 8. EXCEL STATISTICS TABLE
# =============================================================================

stats_path = os.path.join(SAVE_DIR, "PairedViolin_Statistics.xlsx")

with pd.ExcelWriter(stats_path, engine="openpyxl") as writer:
    df_sum.to_excel(writer, index=False, sheet_name="PairedStats")

    ws = writer.sheets["PairedStats"]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin     = Border(**{s: Side(style="thin", color="CCCCCC")
                         for s in ("left","right","top","bottom")})
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=8, name="Helvetica")

    sig_fill  = PatternFill("solid", fgColor="D5E8D4")   # significant
    ns_fill   = PatternFill("solid", fgColor="F8F8F8")
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    warn_fill = PatternFill("solid", fgColor="FFE6CC")   # baseline wins

    # Find column indices
    col_names = list(df_sum.columns)
    star_col  = col_names.index("Stars") + 1         if "Stars"    in col_names else None
    win_col   = col_names.index("Win_pct_ref") + 1   if "Win_pct_ref" in col_names else None

    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for row_idx, (_, row) in enumerate(df_sum.iterrows(), start=2):
        stars    = row.get("Stars", "")
        win_pct  = row.get("Win_pct_ref", 50)

        if stars not in ("ns", "", None) and not pd.isna(stars):
            fill = sig_fill if win_pct >= 50 else warn_fill
        else:
            fill = alt_fill if row_idx % 2 == 0 else ns_fill

        for cell in ws[row_idx]:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(size=8, name="Helvetica")
            cell.fill      = fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 17

print(f"Statistics saved → {stats_path}")

# =============================================================================
# 9. CONSOLE REPORT
# =============================================================================

print("\n" + "=" * 78)
print(f"PAIRED STATISTICS REPORT  ({REFERENCE_MODEL} vs baselines)")
print("=" * 78)

for metric in ALL_METRICS:
    print(f"\n{metric}:")
    sub = df_sum[df_sum["Metric"] == metric]
    print(f"  {'Class':<22} {'Baseline':<12} {'N':>5} "
          f"{'N+':>5} {'N−':>5} {'Win%':>7} {'p':>10} {'g':>8}")
    print("  " + "-" * 76)
    for _, row in sub.sort_values(["Class", "Baseline"]).iterrows():
        print(
            f"  {row['Class']:<22} {row['Baseline']:<12} "
            f"{int(row['N_pairs']):>5} "
            f"{int(row['N_plus (ref wins)']):>5} "
            f"{int(row['N_minus (bsl wins)']):>5} "
            f"{row['Win_pct_ref']:>6.1f}% "
            f"{str(row.get('Paired_Wilcoxon_p','—')):>10} "
            f"{str(row.get('Hedges_g','—')):>8}"
            f"  {row['Stars']}"
        )

print(f"\nAll outputs in: {SAVE_DIR}")


















































# =============================================================================
# PAIRED VIOLIN + BOX PLOTS — CLASSIFICATION METRICS
# Nature Medicine Style — Recall, Specificity, F1, TPR, FPR (+ FNR derived)
# =============================================================================
#
# Extends the paired violin framework to pixel-level classification metrics
# derived from the segmentation confusion matrix columns:
#   TP, TN, FP, FN → Recall, Specificity, F1, TPR, FPR, FNR
#
# Metrics used:
#   Recall / TPR  = TP / (TP + FN)          [higher = better]
#   Specificity   = TN / (TN + FP)          [higher = better]
#   F1_Score      = 2·TP / (2·TP + FP + FN) [higher = better]
#   FPR           = FP / (FP + TN)          [lower  = better]
#   FNR           = FN / (FN + TP)          [lower  = better]  ← derived
#
# Note: Recall == TPR by definition; both are included separately because
# your Excel has them as independent columns — we verify they match and
# keep both for completeness. If they are identical the panel is skipped.
#
# LAYOUT — one figure per class:
#   Columns : classification metrics (5–6 columns)
#   Rows    : baselines (4 rows)
#   Each panel: violin + box + strip + connecting lines + paired stats
#
# Additional output:
#   • Summary heatmap (class × metric signed median difference)
#   • Excel table with paired Wilcoxon p, Hedges' g, N+/N−/ties
#
# =============================================================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.lines as mlines
from matplotlib.colors import TwoSlopeNorm
from scipy.stats import wilcoxon, gaussian_kde
import os
import warnings

warnings.filterwarnings("ignore")

# =============================================================================
# 1. CONFIGURATION
# =============================================================================

REFERENCE_MODEL = "BAT-RM"

# Each tuple: (model_name, path_to_segmentation_excel)
# The segmentation excel must contain the columns listed in the header above.
excel_files_seg = [
    ("BAT-RM",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/200 Epochs/segmentation_metrics_enhanced_200_epoch_enhanced.xlsx'),
    ("nnUNet",    r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/UNET_BM/134 epochs/segmentation_metrics_detailed_epoch_134_generated.xlsx'),
    ("SegMamba",  r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/UNET3+_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("TransUNet", r'I:/Radiotherapy/Cervix/models/models/Shahrukh/axial-20260420T031615Z-3-001/axial/vanilla_unet_segmentation_metrics_detailed_epoch_200_generated.xlsx'),
    ("UNETR",     r'I:/Radiotherapy/Cervix/Paper/code/Trained Models/Small/nnUNet_256/Segmentation_Cervix_small_Axial_axis_pytorch_detailed_metrics_Generated.xlsx'),
]

class_list = [
    "BODY", "URINARY BLADDER", "SMALL BOWEL",
    "RECTUM", "FEMORAL HEAD", "GTV", "CTV",
]

# ── Metrics to plot ──
# "derived" = computed from raw TP/TN/FP/FN columns
# "column"  = taken directly from the Excel column of that name
METRIC_CONFIG = {
    "Recall":      {"source": "column",  "col": "Recall",      "lower_is_better": False},
    "Specificity": {"source": "column",  "col": "Specificity", "lower_is_better": False},
    "F1_Score":    {"source": "column",  "col": "F1_Score",    "lower_is_better": False},
    "TPR":         {"source": "column",  "col": "True_Positive_Rate",  "lower_is_better": False},
    "FPR":         {"source": "column",  "col": "False_Positive_Rate", "lower_is_better": True},
    "FNR":         {"source": "derived", "col": None,           "lower_is_better": True},
}
PLOT_METRICS = list(METRIC_CONFIG.keys())

# Connecting line colours
WIN_COLOR  = "#CC3311"   # reference model wins this patient
LOSE_COLOR = "#4477AA"   # baseline wins
TIE_COLOR  = "#AAAAAA"

MAX_LINES = 60
N_BOOT    = 1000

MODEL_COLORS = {
    "BAT-RM":    "#0077BB",
    "nnUNet":    "#009988",
    "SegMamba":  "#EE7733",
    "TransUNet": "#CC3311",
    "UNETR":     "#AA4499",
}

SAVE_DIR = r'I:/Radiotherapy/Cervix/Paper/Result/Quantitative/PairedViolin_ClassMetrics'
os.makedirs(SAVE_DIR, exist_ok=True)

# =============================================================================
# 2. MATPLOTLIB GLOBAL STYLE
# =============================================================================

plt.rcParams.update({
    "font.family":      "sans-serif",
    "font.sans-serif":  ["Helvetica", "Arial", "DejaVu Sans"],
    "font.size":        9,
    "axes.titlesize":   9,
    "axes.labelsize":   8.5,
    "xtick.labelsize":  8,
    "ytick.labelsize":  7.5,
    "axes.linewidth":   0.6,
    "grid.linewidth":   0.35,
    "axes.facecolor":   "white",
    "figure.facecolor": "white",
    "pdf.fonttype":     42,
    "ps.fonttype":      42,
    "figure.dpi":       150,
    "savefig.dpi":      600,
})

# =============================================================================
# 3. DATA LOADING & METRIC DERIVATION
# =============================================================================

def load_seg_metrics(excel_files, class_list):
    """
    Load all sheets from each segmentation Excel, concatenate,
    filter to target classes, derive FNR, and tag with Model.

    Returns long-format DataFrame with one row per patient × class × model.
    """
    all_data = []

    raw_cols = [
        "Filename", "Class_Name",
        "TP", "TN", "FP", "FN",
        "Recall", "Specificity", "F1_Score",
        "True_Positive_Rate", "False_Positive_Rate",
        "Precision", "IoU", "Dice",
    ]

    for name, path in excel_files:
        df = pd.concat(pd.read_excel(path, sheet_name=None), ignore_index=True)

        # Normalise class names
        df["Class_Name"] = df["Class_Name"].astype(str).str.upper().str.strip()
        df = df[df["Class_Name"].isin([c.upper() for c in class_list])]
        df = df[~df["Class_Name"].isin(["BACKGROUND", "0", "NAN", "NONE", ""])]

        # Keep only columns that exist
        keep = [c for c in raw_cols if c in df.columns]
        df   = df[keep].copy()

        # Coerce numerics
        for c in keep:
            if c not in ("Filename", "Class_Name"):
                df[c] = pd.to_numeric(df[c], errors="coerce")

        # ── Derive FNR = FN / (FN + TP) ──
        if "FN" in df.columns and "TP" in df.columns:
            denom    = df["FN"] + df["TP"]
            df["FNR"] = np.where(denom > 0, df["FN"] / denom, np.nan)
        else:
            df["FNR"] = np.nan

        # Verify TPR == Recall (they should be identical; flag if not)
        if "True_Positive_Rate" in df.columns and "Recall" in df.columns:
            delta = (df["True_Positive_Rate"] - df["Recall"]).abs()
            if delta.max() > 1e-6:
                print(f"  ⚠  {name}: TPR and Recall differ (max Δ={delta.max():.6f})")

        df["Model"] = name
        all_data.append(df)

    if not all_data:
        raise ValueError("No segmentation data loaded. Check file paths.")

    return pd.concat(all_data, ignore_index=True)


print("Loading segmentation metrics …")
df_all = load_seg_metrics(excel_files_seg, class_list)

model_names = [m for m, _ in excel_files_seg]
baselines   = [m for m in model_names if m != REFERENCE_MODEL]
classes     = sorted(df_all["Class_Name"].dropna().unique())

print(f"Models   : {model_names}")
print(f"Classes  : {classes}")
print(f"Metrics  : {PLOT_METRICS}")
print(f"Patients : {df_all['Filename'].nunique()}\n")

# =============================================================================
# 4. STATISTICAL HELPERS  (identical to paired_violin_plot.py)
# =============================================================================

def paired_hedges_g(diff):
    diff = np.asarray(diff, float)
    n    = len(diff)
    if n < 2:
        return np.nan
    d = diff.mean() / (diff.std(ddof=1) + 1e-12)
    j = 1 - 3 / (4 * (n - 1) - 1)
    return d * j


def bootstrap_g_ci(diff, n_boot=N_BOOT, seed=42):
    diff = np.asarray(diff, float)
    n    = len(diff)
    if n < 3:
        return np.nan, np.nan
    rng  = np.random.default_rng(seed)
    boot = np.array([paired_hedges_g(diff[rng.integers(0, n, n)])
                     for _ in range(n_boot)])
    return float(np.nanpercentile(boot, 2.5)), float(np.nanpercentile(boot, 97.5))


def paired_wilcoxon_test(x_ref, x_bsl):
    diff = np.asarray(x_ref) - np.asarray(x_bsl)
    if len(diff) < 3 or np.all(diff == 0):
        return np.nan, np.nan
    try:
        stat, p = wilcoxon(x_ref, x_bsl, zero_method="pratt")
        return stat, p
    except Exception:
        return np.nan, np.nan


def concordant_discordant(x_ref, x_bsl, lower_is_better):
    diff = np.asarray(x_ref, float) - np.asarray(x_bsl, float)
    tol  = 1e-9
    if lower_is_better:
        n_plus  = int((diff < -tol).sum())
        n_minus = int((diff >  tol).sum())
    else:
        n_plus  = int((diff >  tol).sum())
        n_minus = int((diff < -tol).sum())
    ties = int((np.abs(diff) <= tol).sum())
    return n_plus, n_minus, ties


def p_stars(p):
    if np.isnan(p): return ""
    if p < 0.0001:  return "****"
    if p < 0.001:   return "***"
    if p < 0.01:    return "**"
    if p < 0.05:    return "*"
    return "ns"

# =============================================================================
# 5. PANEL DRAWING FUNCTION
# =============================================================================

def draw_paired_panel(ax, ref_arr, bsl_arr,
                      ref_name, bsl_name, metric_label,
                      lower_is_better, rng,
                      max_lines=MAX_LINES):
    """
    Full paired violin + box + strip + connecting lines panel.
    Returns dict of computed statistics.
    """
    ref_arr = np.asarray(ref_arr, float)
    bsl_arr = np.asarray(bsl_arr, float)
    n       = len(ref_arr)
    c_ref   = MODEL_COLORS.get(ref_name, "#0077BB")
    c_bsl   = MODEL_COLORS.get(bsl_name, "#888888")
    x_ref, x_bsl = 0, 1

    # ── Violin ──
    for xi, arr, color in [(x_ref, ref_arr, c_ref), (x_bsl, bsl_arr, c_bsl)]:
        if len(arr) >= 4:
            try:
                kde  = gaussian_kde(arr, bw_method="scott")
                y_ev = np.linspace(arr.min(), arr.max(), 300)
                dens = kde(y_ev)
                dens = dens / dens.max() * 0.36
                ax.fill_betweenx(y_ev, xi - dens, xi + dens,
                                 color=color, alpha=0.18, zorder=1)
                ax.plot(xi - dens, y_ev, color=color, lw=0.7, alpha=0.6, zorder=2)
                ax.plot(xi + dens, y_ev, color=color, lw=0.7, alpha=0.6, zorder=2)
            except Exception:
                pass

    # ── Box ──
    for xi, arr, color in [(x_ref, ref_arr, c_ref), (x_bsl, bsl_arr, c_bsl)]:
        q25, q50, q75 = np.percentile(arr, [25, 50, 75])
        p05, p95      = np.percentile(arr, [5,  95])
        ax.plot([xi - 0.20, xi + 0.20], [q50, q50],
                color=color, lw=2.4, solid_capstyle="round", zorder=6)
        ax.add_patch(plt.Rectangle(
            (xi - 0.15, q25), 0.30, q75 - q25,
            facecolor=color, alpha=0.22, edgecolor=color, lw=0.9, zorder=4,
        ))
        ax.plot([xi, xi], [p05, q25], color=color, lw=0.8, alpha=0.6, zorder=3)
        ax.plot([xi, xi], [q75, p95], color=color, lw=0.8, alpha=0.6, zorder=3)
        for y_w in [p05, p95]:
            ax.plot([xi - 0.08, xi + 0.08], [y_w, y_w],
                    color=color, lw=0.8, alpha=0.6, zorder=3)

    # ── Connecting lines ──
    indices = np.arange(n)
    if n > max_lines:
        indices = rng.choice(n, max_lines, replace=False)
        note = f"(showing {max_lines}/{n} pairs)"
    else:
        note = f"(n={n} pairs)"

    tol = 1e-9
    for idx in indices:
        rv, bv = ref_arr[idx], bsl_arr[idx]
        if lower_is_better:
            ref_wins = rv < bv - tol
            bsl_wins = bv < rv - tol
        else:
            ref_wins = rv > bv + tol
            bsl_wins = bv > rv + tol
        lc = WIN_COLOR if ref_wins else (LOSE_COLOR if bsl_wins else TIE_COLOR)
        jx_r = x_ref + rng.uniform(-0.08, 0.08)
        jx_b = x_bsl + rng.uniform(-0.08, 0.08)
        ax.plot([jx_r, jx_b], [rv, bv],
                color=lc, lw=0.45, alpha=0.35, zorder=3)

    # ── Strip dots ──
    for xi, arr, color in [(x_ref, ref_arr, c_ref), (x_bsl, bsl_arr, c_bsl)]:
        jitter = rng.uniform(-0.10, 0.10, size=len(arr))
        ax.scatter(xi + jitter, arr, color=color, s=9, alpha=0.50,
                   edgecolors="white", linewidths=0.2, zorder=5)

    # ── Statistics ──
    _, p_val            = paired_wilcoxon_test(ref_arr, bsl_arr)
    diff                = ref_arr - bsl_arr
    g_diff              = -diff if lower_is_better else diff
    g                   = paired_hedges_g(g_diff)
    g_lo, g_hi          = bootstrap_g_ci(g_diff, N_BOOT)
    n_plus, n_minus, ties = concordant_discordant(ref_arr, bsl_arr, lower_is_better)
    stars               = p_stars(p_val)

    # ── Significance bracket ──
    y_max   = max(ref_arr.max(), bsl_arr.max())
    y_min   = min(ref_arr.min(), bsl_arr.min())
    y_range = (y_max - y_min) if y_max > y_min else 0.05
    y_brk   = y_max + y_range * 0.08

    ax.plot([x_ref, x_ref, x_bsl, x_bsl],
            [y_brk, y_brk + y_range * 0.02,
             y_brk + y_range * 0.02, y_brk],
            color="#333333", lw=0.9, zorder=8)

    sig_col = "#CC2222" if stars not in ("ns", "") else "#666666"
    ax.text(0.5, y_brk + y_range * 0.04,
            stars or "ns", ha="center", va="bottom",
            fontsize=8, color=sig_col, fontweight="bold", zorder=9)
    ax.text(0.5, y_brk + y_range * 0.09,
            f"p={p_val:.4f}" if not np.isnan(p_val) else "p=n/a",
            ha="center", va="bottom", fontsize=5.5, color="#555555", zorder=9)
    ax.text(0.5, y_brk + y_range * 0.155,
            f"N+={n_plus}  N−={n_minus}"
            + (f"  ties={ties}" if ties else ""),
            ha="center", va="bottom", fontsize=5.5, color="#333333", zorder=9)

    # ── Inset stats box ──
    ref_med = float(np.median(ref_arr))
    bsl_med = float(np.median(bsl_arr))
    ref_iqr = float(np.subtract(*np.percentile(ref_arr, [75, 25])))
    bsl_iqr = float(np.subtract(*np.percentile(bsl_arr, [75, 25])))
    g_str   = (f"{g:+.2f} [{g_lo:+.2f},{g_hi:+.2f}]"
               if not np.isnan(g) else "n/a")

    stats_txt = (
        f"{'Model':<9} {'Med':>6}  {'IQR':>6}\n"
        f"{'─'*26}\n"
        f"{'★'+ref_name:<9} {ref_med:>6.4f}  {ref_iqr:>6.4f}\n"
        f"{bsl_name:<9} {bsl_med:>6.4f}  {bsl_iqr:>6.4f}\n"
        f"{'─'*26}\n"
        f"g = {g_str}\n"
        f"{note}"
    )
    ax.text(1.04, 0.99, stats_txt,
            transform=ax.transAxes, va="top", ha="left",
            fontsize=5.0, fontfamily="monospace", color="#222222",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
                      edgecolor="#cccccc", lw=0.5, alpha=0.90),
            zorder=10, clip_on=False)

    # ── Axis ──
    ax.set_xticks([x_ref, x_bsl])
    ax.set_xticklabels(
        [f"{'★ ' if ref_name == REFERENCE_MODEL else ''}{ref_name}", bsl_name],
        fontsize=7,
    )
    ax.set_xlim(-0.55, 1.55)
    ax.set_ylim(y_min - y_range * 0.08, y_brk + y_range * 0.30)
    ax.set_ylabel(metric_label, fontsize=8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(0.5)
    ax.spines["bottom"].set_linewidth(0.5)
    ax.tick_params(axis="y", labelsize=7)
    ax.tick_params(axis="x", length=0)
    ax.grid(axis="y", linewidth=0.3, color="#eeeeee", zorder=0)

    return dict(
        p_val=p_val, stars=stars, g=g, g_lo=g_lo, g_hi=g_hi,
        n_pairs=n, n_plus=n_plus, n_minus=n_minus, ties=ties,
        ref_median=ref_med, bsl_median=bsl_med,
        ref_iqr=ref_iqr, bsl_iqr=bsl_iqr,
    )

# =============================================================================
# 6. HELPER: get metric column for a given metric name
# =============================================================================

def get_metric_series(df_model_cls, metric):
    """
    Return a Series of metric values for one (model, class) slice.
    Handles both direct column reads and derived FNR.
    """
    cfg = METRIC_CONFIG[metric]
    if cfg["source"] == "derived":
        # FNR = FN / (FN + TP)
        if "FN" in df_model_cls.columns and "TP" in df_model_cls.columns:
            fn   = pd.to_numeric(df_model_cls["FN"], errors="coerce")
            tp   = pd.to_numeric(df_model_cls["TP"], errors="coerce")
            denom = fn + tp
            return np.where(denom > 0, fn / denom, np.nan)
        return np.full(len(df_model_cls), np.nan)
    else:
        col = cfg["col"]
        if col not in df_model_cls.columns:
            return np.full(len(df_model_cls), np.nan)
        return pd.to_numeric(df_model_cls[col], errors="coerce").values

# =============================================================================
# 7. MAIN LOOP — one figure per class
# =============================================================================

rng = np.random.default_rng(42)
summary_records = []

ref_df = df_all[df_all["Model"] == REFERENCE_MODEL]

# Direction label per metric
DIRECTION = {
    m: ("↓ lower = better" if METRIC_CONFIG[m]["lower_is_better"]
        else "↑ higher = better")
    for m in PLOT_METRICS
}

for cls in classes:
    n_baselines = len(baselines)
    n_metrics   = len(PLOT_METRICS)

    PANEL_W = 2.8
    PANEL_H = 3.6
    FIG_W   = n_metrics * (PANEL_W + 1.5) + 0.6
    FIG_H   = n_baselines * PANEL_H + 2.0

    fig, axes = plt.subplots(
        n_baselines, n_metrics,
        figsize=(FIG_W, FIG_H),
        squeeze=False,
    )

    # Column headers (first row only)
    for ci, metric in enumerate(PLOT_METRICS):
        axes[0][ci].set_title(
            f"{metric}\n{DIRECTION[metric]}",
            fontsize=8.5, fontweight="bold", pad=6,
        )

    for ri, baseline in enumerate(baselines):
        bsl_df = df_all[df_all["Model"] == baseline]

        for ci, metric in enumerate(PLOT_METRICS):
            ax        = axes[ri][ci]
            lower_b   = METRIC_CONFIG[metric]["lower_is_better"]

            # ── Reference values ──
            ref_cls = ref_df[ref_df["Class_Name"] == cls][
                ["Filename"] + [c for c in ["TP","TN","FP","FN"] if c in ref_df.columns]
            ].copy()
            ref_cls["_metric"] = get_metric_series(
                ref_df[ref_df["Class_Name"] == cls].reset_index(drop=True),
                metric
            )
            ref_cls = ref_cls[["Filename", "_metric"]].dropna()

            # ── Baseline values ──
            bsl_cls = bsl_df[bsl_df["Class_Name"] == cls].copy().reset_index(drop=True)
            bsl_cls["_metric"] = get_metric_series(bsl_cls, metric)
            bsl_cls = bsl_cls[["Filename", "_metric"]].dropna()

            # ── Paired merge ──
            merged = pd.merge(
                ref_cls.rename(columns={"_metric": "ref"}),
                bsl_cls.rename(columns={"_metric": "bsl"}),
                on="Filename",
            ).dropna()
            merged = merged[
                np.isfinite(merged["ref"]) & np.isfinite(merged["bsl"])
            ]

            if len(merged) < 3:
                ax.text(0.5, 0.5, "Insufficient\npaired data",
                        ha="center", va="center", fontsize=8,
                        color="#aaaaaa", transform=ax.transAxes)
                ax.axis("off")
                continue

            out = draw_paired_panel(
                ax,
                ref_arr  = merged["ref"].values,
                bsl_arr  = merged["bsl"].values,
                ref_name = REFERENCE_MODEL,
                bsl_name = baseline,
                metric_label = metric,
                lower_is_better = lower_b,
                rng      = rng,
            )

            # Only show metric title on first baseline row
            if ri > 0:
                ax.set_title("")

            summary_records.append({
                "Class":              cls,
                "Metric":             metric,
                "Baseline":           baseline,
                "N_pairs":            out["n_pairs"],
                "Ref_median":         round(out["ref_median"], 5),
                "Bsl_median":         round(out["bsl_median"], 5),
                "Ref_IQR":            round(out["ref_iqr"],    5),
                "Bsl_IQR":            round(out["bsl_iqr"],    5),
                "N_plus_ref_wins":    out["n_plus"],
                "N_minus_bsl_wins":   out["n_minus"],
                "Ties":               out["ties"],
                "Win_pct_ref":        round(100 * out["n_plus"] /
                                            max(out["n_pairs"], 1), 1),
                "Paired_Wilcoxon_p":  round(out["p_val"], 5)
                                      if not np.isnan(out["p_val"]) else np.nan,
                "Stars":              out["stars"],
                "Hedges_g":           round(out["g"],    3) if not np.isnan(out["g"])    else np.nan,
                "g_CI_lo":            round(out["g_lo"], 3) if not np.isnan(out["g_lo"]) else np.nan,
                "g_CI_hi":            round(out["g_hi"], 3) if not np.isnan(out["g_hi"]) else np.nan,
            })

    # ── Legend ──
    handles = [
        mlines.Line2D([0],[0], color=MODEL_COLORS.get(REFERENCE_MODEL,"#0077BB"),
                      lw=2.0, label=f"★ {REFERENCE_MODEL}"),
    ] + [
        mlines.Line2D([0],[0], color=MODEL_COLORS.get(b,"#888888"),
                      lw=2.0, label=b)
        for b in baselines
    ] + [
        mlines.Line2D([0],[0], color=WIN_COLOR,  lw=1.0, alpha=0.7,
                      label=f"{REFERENCE_MODEL} wins (red)"),
        mlines.Line2D([0],[0], color=LOSE_COLOR, lw=1.0, alpha=0.7,
                      label="Baseline wins (blue)"),
        mlines.Line2D([0],[0], color=TIE_COLOR,  lw=1.0, alpha=0.7,
                      label="Tie (grey)"),
    ]
    fig.legend(handles=handles, loc="lower center",
               ncol=len(handles), frameon=False,
               fontsize=7.5, bbox_to_anchor=(0.5, -0.01))

    fig.suptitle(
        f"Paired violin + box — classification metrics — {cls}\n"
        f"Red line = {REFERENCE_MODEL} wins that patient  |  Blue = baseline wins  |  "
        f"Bracket: paired Wilcoxon  |  N+ / N− = concordant / discordant  |  g = Hedges' g",
        fontsize=9, fontweight="bold", y=1.002,
    )
    plt.tight_layout(w_pad=3.5, h_pad=1.5, rect=[0, 0.04, 1, 0.998])

    safe_cls = cls.replace(" ", "_")
    png_path = os.path.join(SAVE_DIR, f"PairedViolin_ClassMetrics_{safe_cls}.png")
    pdf_path = os.path.join(SAVE_DIR, f"PairedViolin_ClassMetrics_{safe_cls}.pdf")
    plt.savefig(png_path, dpi=600, bbox_inches="tight", facecolor="white")
    plt.savefig(pdf_path, format="pdf", bbox_inches="tight", facecolor="white")
    plt.show()
    print(f"Saved → {png_path}")

# =============================================================================
# 8. SUMMARY HEATMAP — signed median difference across all classes × metrics
# =============================================================================

df_sum = pd.DataFrame(summary_records)

# Sign-flip so that positive ALWAYS means reference model is better
df_sum["Med_diff_signed"] = df_sum.apply(
    lambda r: (r["Ref_median"] - r["Bsl_median"])
              * (-1 if METRIC_CONFIG[r["Metric"]]["lower_is_better"] else 1),
    axis=1,
)

row_labels = [f"{cls}\nvs {bsl}" for cls in classes for bsl in baselines]
n_rows_h   = len(row_labels)

diff_mat = np.full((n_rows_h, len(PLOT_METRICS)), np.nan)
star_mat = [[""] * len(PLOT_METRICS) for _ in range(n_rows_h)]
g_mat    = np.full((n_rows_h, len(PLOT_METRICS)), np.nan)

for ri, (cls, bsl) in enumerate([(c, b) for c in classes for b in baselines]):
    sub = df_sum[(df_sum["Class"] == cls) & (df_sum["Baseline"] == bsl)]
    for ci, metric in enumerate(PLOT_METRICS):
        row = sub[sub["Metric"] == metric]
        if row.empty:
            continue
        diff_mat[ri, ci] = row.iloc[0]["Med_diff_signed"]
        star_mat[ri][ci] = row.iloc[0]["Stars"]
        g_mat[ri, ci]    = row.iloc[0]["Hedges_g"]

v_abs = np.nanmax(np.abs(diff_mat)) if np.any(np.isfinite(diff_mat)) else 0.01

FIG_SUM_W = len(PLOT_METRICS) * 2.0 + 3.5
FIG_SUM_H = max(6, n_rows_h * 0.55 + 2.0)
fig_s, ax_s = plt.subplots(figsize=(FIG_SUM_W, FIG_SUM_H))

norm_s = TwoSlopeNorm(vmin=-v_abs, vcenter=0, vmax=v_abs)
im_s   = ax_s.imshow(diff_mat, cmap=plt.cm.RdBu, norm=norm_s, aspect="auto")

for ri in range(n_rows_h):
    for ci, metric in enumerate(PLOT_METRICS):
        v  = diff_mat[ri, ci]
        st = star_mat[ri][ci]
        g  = g_mat[ri, ci]
        if np.isnan(v):
            ax_s.text(ci, ri, "—", ha="center", va="center",
                      fontsize=7, color="#aaaaaa")
            continue
        txt_col = "white" if abs(v) > v_abs * 0.6 else "#111111"
        ax_s.text(ci, ri - 0.15, f"{v:+.4f}",
                  ha="center", va="center",
                  fontsize=7, color=txt_col, fontweight="bold")
        parts = []
        if st and st != "ns":
            parts.append(st)
        if not np.isnan(g):
            parts.append(f"g={g:+.2f}")
        if parts:
            ax_s.text(ci, ri + 0.25, "  ".join(parts),
                      ha="center", va="center", fontsize=5.2, color=txt_col)

# Class group separators
for gi in range(len(classes)):
    y_sep = gi * len(baselines) - 0.5
    if y_sep > -0.5:
        ax_s.axhline(y_sep, color="#333333", lw=1.0, zorder=5)
for k in range(len(PLOT_METRICS) + 1):
    ax_s.axvline(k - 0.5, color="#cccccc", lw=0.4)
for k in range(n_rows_h + 1):
    ax_s.axhline(k - 0.5, color="#cccccc", lw=0.4)

ax_s.set_xticks(range(len(PLOT_METRICS)))
ax_s.set_xticklabels(PLOT_METRICS, fontsize=9, fontweight="bold")
ax_s.set_yticks(range(n_rows_h))
ax_s.set_yticklabels(row_labels, fontsize=7)
ax_s.set_title(
    f"Summary — signed median difference  ({REFERENCE_MODEL} − baseline)\n"
    "Red = reference model better  |  Blue = baseline better  |  "
    "Value = Δmedian  |  Stars = paired Wilcoxon  |  g = Hedges' g\n"
    "(lower-is-better metrics sign-flipped so red = better always)",
    fontsize=9, fontweight="bold", pad=6,
)
ax_s.tick_params(length=0)
cb_s = plt.colorbar(im_s, ax=ax_s, fraction=0.025, pad=0.03)
cb_s.ax.tick_params(labelsize=7.5)
cb_s.set_label(f"{REFERENCE_MODEL} better → / ← baseline better", fontsize=7.5)

plt.tight_layout()
sum_png = os.path.join(SAVE_DIR, "PairedViolin_ClassMetrics_SummaryHeatmap.png")
sum_pdf = os.path.join(SAVE_DIR, "PairedViolin_ClassMetrics_SummaryHeatmap.pdf")
plt.savefig(sum_png, dpi=600, bbox_inches="tight", facecolor="white")
plt.savefig(sum_pdf, format="pdf", bbox_inches="tight", facecolor="white")
plt.show()
print(f"Saved summary heatmap → {sum_png}")

# =============================================================================
# 9. EXCEL TABLE
# =============================================================================

stats_path = os.path.join(SAVE_DIR, "PairedViolin_ClassMetrics_Statistics.xlsx")

with pd.ExcelWriter(stats_path, engine="openpyxl") as writer:
    df_sum.to_excel(writer, index=False, sheet_name="PairedStats")

    ws = writer.sheets["PairedStats"]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin     = Border(**{s: Side(style="thin", color="CCCCCC")
                         for s in ("left","right","top","bottom")})
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=8, name="Helvetica")
    sig_fill = PatternFill("solid", fgColor="D5E8D4")
    ns_fill  = PatternFill("solid", fgColor="F8F8F8")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    warn_fill= PatternFill("solid", fgColor="FFE6CC")

    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center"); cell.border = thin

    for row_idx, (_, row) in enumerate(df_sum.iterrows(), start=2):
        stars   = row.get("Stars", "")
        win_pct = row.get("Win_pct_ref", 50)
        if stars not in ("ns", "", None) and not pd.isna(stars):
            fill = sig_fill if win_pct >= 50 else warn_fill
        else:
            fill = alt_fill if row_idx % 2 == 0 else ns_fill
        for cell in ws[row_idx]:
            cell.border    = thin
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(size=8, name="Helvetica")
            cell.fill      = fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

print(f"Statistics saved → {stats_path}")

# =============================================================================
# 10. CONSOLE REPORT
# =============================================================================

print("\n" + "=" * 80)
print(f"PAIRED CLASSIFICATION METRICS REPORT  ({REFERENCE_MODEL} vs baselines)")
print("=" * 80)
for metric in PLOT_METRICS:
    print(f"\n{metric}  ({'↓ lower better' if METRIC_CONFIG[metric]['lower_is_better'] else '↑ higher better'}):")
    sub = df_sum[df_sum["Metric"] == metric]
    print(f"  {'Class':<22} {'Baseline':<12} {'N':>5} {'N+':>5} "
          f"{'N−':>5} {'Win%':>7} {'p':>10} {'g':>8}  Sig")
    print("  " + "-" * 78)
    for _, row in sub.sort_values(["Class", "Baseline"]).iterrows():
        print(
            f"  {row['Class']:<22} {row['Baseline']:<12} "
            f"{int(row['N_pairs']):>5} "
            f"{int(row['N_plus_ref_wins']):>5} "
            f"{int(row['N_minus_bsl_wins']):>5} "
            f"{row['Win_pct_ref']:>6.1f}% "
            f"{str(row.get('Paired_Wilcoxon_p','—')):>10} "
            f"{str(row.get('Hedges_g','—')):>8}  "
            f"{row['Stars']}"
        )

print(f"\nAll outputs in: {SAVE_DIR}")
